import tempfile
import unittest
from unittest import mock

import alr_quote_verifier as verifier


class QuoteFragmentTests(unittest.TestCase):
    def test_text_fragment_encoding_obeys_directive_grammar(self):
        encoded = verifier._encode_text_fragment("\u2013 , - & .")

        self.assertEqual(encoded, "%E2%80%93%20%2C%20%2D%20%26%20.")
        self.assertTrue(encoded.isascii())

    def test_scc_quote_link_prefers_a2aj_official_page(self):
        canlii = (
            "https://www.canlii.org/en/ca/scc/doc/2001/2001scc24/"
            "2001scc24.html#par47"
        )
        official = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1861/"
            "index.do"
        )

        preferred = verifier._prefer_scc_official_quote_link(
            {
                "_a2aj_url_reconciled": True,
                "_a2aj_dataset": "SCC",
                "_a2aj_source_url": official,
            },
            canlii,
        )
        built = verifier._append_text_fragment_directives(
            preferred,
            [verifier._text_fragment_directive("only voluntary conduct \u2013 behaviour")],
        )

        self.assertEqual(preferred, official + "#par47")
        self.assertEqual(
            built,
            official
            + "?iframe=true&site_preference=mobile"
            + "#par47:~:text=only%20voluntary%20conduct%20%E2%80%93%20behaviour",
        )

    def test_scc_quote_link_uses_existing_identity_lock_after_canlii_text_match(self):
        canlii = (
            "https://www.canlii.org/en/ca/scc/doc/2015/2015scc33/"
            "2015scc33.html#par42"
        )
        official = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/15397/"
            "index.do"
        )
        document = verifier.a2aj_client.A2AJDocument(
            dataset="SCC",
            citation="2015 SCC 33",
            alternate_citation="",
            name="R. v. Tatton",
            date="2015-06-05",
            url=official,
            text="[42] Identity-locked source text.",
            language="en",
            scraped_timestamp="",
            upstream_license="",
            raw={},
        )
        row = {"quote_match_pinpoint": "par42"}

        with mock.patch.dict(
            verifier._A2AJ_LOCKED_DOCUMENTS,
            {canlii.split("#", 1)[0].lower(): document},
            clear=True,
        ), mock.patch.object(
            verifier,
            "_normalized_a2aj_document_text",
            return_value=document.text,
        ):
            preferred = verifier._prefer_scc_official_quote_link(row, canlii)

        self.assertEqual(preferred, official + "#par42")
        self.assertEqual(row["_a2aj_dataset"], "SCC")
        self.assertEqual(row["_a2aj_source_url"], official)
        self.assertIs(row["_a2aj_url_reconciled"], True)

    def test_scc_original_anchor_match_promotes_citation_fallback_only_for_quote(self):
        quote = "the exact quotation appears in the cited paragraph"
        canlii_base = (
            "https://www.canlii.org/en/ca/scc/doc/2015/2015scc33/"
            "2015scc33.html"
        )
        canlii = canlii_base + "#par42"
        official = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/15397/"
            "index.do"
        )
        document = verifier.a2aj_client.A2AJDocument(
            dataset="SCC",
            citation="2015 SCC 33",
            alternate_citation="",
            name="R. v. Tatton",
            date="2015-06-05",
            url=official,
            text=f"[42] {quote}",
            language="en",
            scraped_timestamp="",
            upstream_license="",
            raw={},
        )
        row = {
            "footnote_id": 1,
            "citation_part_index": 1,
            "citation_part_kind": "case",
            "citation_part_link": canlii,
            "citation_part_anchor_text": quote,
            "pinpoint_fragments": ["par42"],
        }
        quotes = {
            1: [{
                "quote_inner": quote,
                "quote_raw": f'"{quote}"',
                "quote_delimiter_style": "STRAIGHT",
            }]
        }

        with mock.patch.object(verifier, "USE_A2AJ", False), mock.patch.dict(
            verifier._A2AJ_LOCKED_DOCUMENTS,
            {canlii_base.lower(): document},
            clear=True,
        ), mock.patch.object(
            verifier,
            "_normalized_a2aj_document_text",
            return_value=document.text,
        ):
            verifier._apply_quote_checks([row], quotes)

        self.assertEqual(row["quote_check_status"], "OG_PINPOINT_MATCH")
        self.assertEqual(row["citation_part_link"], canlii)
        self.assertEqual(row["quote_match_link"], official + "#par42")

    def test_scc_quote_link_does_not_replace_non_scc_canlii_link(self):
        link = (
            "https://www.canlii.org/en/on/onca/doc/2001/2001canlii1/"
            "2001canlii1.html#par47"
        )
        row = {
            "_a2aj_url_reconciled": True,
            "_a2aj_dataset": "SCC",
            "_a2aj_source_url": (
                "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1861/"
                "index.do"
            ),
        }

        self.assertEqual(verifier._prefer_scc_official_quote_link(row, link), link)

    def test_scc_quote_link_without_paragraph_stays_on_canlii(self):
        link = (
            "https://www.canlii.org/en/ca/scc/doc/2001/2001scc24/"
            "2001scc24.html"
        )
        row = {
            "_a2aj_url_reconciled": True,
            "_a2aj_dataset": "SCC",
            "_a2aj_source_url": (
                "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1861/"
                "index.do"
            ),
        }

        self.assertEqual(verifier._prefer_scc_official_quote_link(row, link), link)

    def test_ruzic_fragment_is_ascii_and_survives_xlsx_round_trip(self):
        import openpyxl

        canlii = (
            "https://www.canlii.org/en/ca/scc/doc/2001/2001scc24/"
            "2001scc24.html#par47"
        )
        official = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1861/"
            "index.do"
        )
        fragment = (
            "only voluntary conduct \u2013 behaviour that is the product of a free will and "
            "controlled body, unhindered by external constraints \u2013 should attract the "
            "penalty and stigma of criminal liability"
        )
        source_context = (
            "47 It is a principle of fundamental justice that "
            f"{fragment}. Depriving a person of liberty and branding her with the stigma "
            "of criminal liability."
        )
        root = verifier.etree.HTML(
            "<html><body>"
            f"<p>It is a principle of fundamental justice that {fragment}. Depriving a "
            "person of liberty and branding him or her with the stigma of criminal liability.</p>"
            f"<p data-viibes-parag='47'>47 It is a principle of fundamental justice that "
            f"{fragment}. Depriving a person of liberty and branding her with the stigma "
            "of criminal liability.</p>"
            "</body></html>"
        )
        preferred = verifier._prefer_scc_official_quote_link(
            {
                "_a2aj_url_reconciled": True,
                "_a2aj_dataset": "SCC",
                "_a2aj_source_url": official,
            },
            canlii,
        )

        with mock.patch.dict(verifier._FRAGMENT_DOC_TEXT_CACHE, clear=True):
            verifier._register_fragment_document_text(
                official,
                (
                    "It is a principle of fundamental justice that "
                    f"{fragment}. Depriving a person of liberty and branding him or her "
                    "with the stigma of criminal liability.\n"
                    + source_context
                ),
            )
            result = verifier._build_quote_check_fragment_url(
                preferred, fragment, source_context
            )

        expected = (
            official
            + "?iframe=true&site_preference=mobile#par47:~:text=only%20voluntary%20"
            "conduct%20%E2%80%93%20behaviour%20that%20is%20the%20product%20of%20a%20"
            "free%20will%20and%20controlled%20body%2C%20unhindered%20by%20external%20"
            "constraints%20%E2%80%93%20should%20attract%20the%20penalty%20and%20stigma%20"
            "of%20criminal%20liability.,-Depriving%20a%20person%20of%20liberty%20and%20"
            "branding%20her"
        )
        self.assertEqual(result.url, expected)
        self.assertTrue(result.url.isascii())

        with tempfile.TemporaryDirectory() as directory:
            path = f"{directory}\\ruzic.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active["A1"] = "Ruzic quote"
            workbook.active["A1"].hyperlink = result.url
            workbook.save(path)
            workbook.close()

            workbook = openpyxl.load_workbook(path)
            self.assertEqual(workbook.active["A1"].hyperlink.target, expected)
            workbook.close()



    def test_missing_source_warning_requires_grabbable_source_class(self):
        eligible = [
            {"citation_part_kind": "case", "citation_part_link": "https://example.test/case"},
            {"citation_part_kind": "unreported", "citation_part_link": "https://example.test/case"},
            {"citation_part_kind": "statute", "citation_part_link": "https://example.test/act"},
            {"citation_part_kind": "regulation", "citation_part_link": "https://example.test/reg"},
            {"citation_part_kind": "gazette", "citation_part_link": "https://example.test/gazette"},
            {
                "citation_part_kind": "journal",
                "citation_part_link": "https://example.test/article.pdf",
                "_journal_link_resolved": True,
            },
        ]
        ineligible = [
            {"citation_part_kind": "book", "citation_part_link": "https://example.test/book"},
            {"citation_part_kind": "report", "citation_part_link": "https://example.test/report"},
            {"citation_part_kind": "journal", "citation_part_link": "https://example.test/article"},
            {"citation_part_kind": "case", "citation_part_link": "other"},
            {"citation_part_kind": "case", "citation_part_link": "not a URL"},
            {"citation_part_kind": "case", "citation_part_link": "ftp://example.test/case"},
        ]

        for row in eligible:
            with self.subTest(eligible=row):
                self.assertTrue(verifier._should_report_source_text_failure(row))
        for row in ineligible:
            with self.subTest(ineligible=row):
                self.assertFalse(verifier._should_report_source_text_failure(row))

    def test_scc_text_fragment_uses_scrollable_mobile_inner_view(self):
        url = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1705/index.do"
            "#par191"
        )

        built = verifier._append_text_fragment_directives(
            url, [verifier._text_fragment_directive("motiveless act")]
        )

        self.assertEqual(
            built,
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1705/"
            "index.do?iframe=true&site_preference=mobile"
            "#par191:~:text=motiveless%20act",
        )

    def test_scc_text_fragment_does_not_duplicate_existing_inner_parameters(self):
        url = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1705/"
            "index.do?iframe=true#par191"
        )

        built = verifier._append_text_fragment_directives(
            url, [verifier._text_fragment_directive("motiveless act")]
        )

        self.assertIn("?iframe=true&site_preference=mobile#par191:~:text=", built)
        self.assertEqual(built.count("iframe="), 1)
        self.assertEqual(built.count("site_preference="), 1)

    def test_scc_text_fragment_overrides_broken_inner_parameters(self):
        url = (
            "https://decisions.scc-csc.ca/scc-csc/scc-csc/en/item/1705/"
            "index.do?foo=bar&iframe=false&site_preference=desktop#par191"
        )

        built = verifier._append_text_fragment_directives(
            url, [verifier._text_fragment_directive("motiveless act")]
        )

        self.assertIn(
            "?foo=bar&iframe=true&site_preference=mobile#par191:~:text=", built
        )
        self.assertNotIn("iframe=false", built)
        self.assertNotIn("site_preference=desktop", built)

    def test_non_scc_text_fragment_url_is_unchanged(self):
        url = "https://example.com/case#par4"

        built = verifier._append_text_fragment_directives(
            url, [verifier._text_fragment_directive("quoted words")]
        )

        self.assertEqual(
            built,
            "https://example.com/case#par4:~:text=quoted%20words",
        )


    def test_source_side_fragment_does_not_absorb_intro_for_bracketed_suffix(self):
        source = 'the issue or as some have put it, a "trial in a box": Young, at p'
        quote = "trial[s] in a box,"

        fragment = verifier._source_side_quote_fragment_text(source, quote)

        self.assertEqual(fragment, "trial in a box")

    def test_multi_quote_fragment_building_keeps_multiple_directives(self):
        corrected = '"first quoted phrase"\n\n"second quoted phrase"'
        target = verifier._build_quote_only_text_fragment_url(
            "https://example.com/source.html",
            corrected,
        )

        self.assertIn("first%20quoted%20phrase", target)
        self.assertIn("second%20quoted%20phrase", target)
        self.assertIn("&text=", target)

    def test_multi_quote_targeted_fragment_building_keeps_multiple_directives(self):
        corrected = '"first quoted phrase"\n\n"second quoted phrase"'
        source = "Before first quoted phrase after. Later second quoted phrase end."

        target, count = verifier._build_targeted_text_fragment_url(
            "https://example.com/source.html",
            source,
            corrected,
        )

        self.assertEqual(count, 2)
        self.assertIn("first%20quoted%20phrase", target)
        self.assertIn("second%20quoted%20phrase", target)
        self.assertIn("&text=", target)

    def test_hyphen_variants_score_as_exact_and_display_tightly(self):
        quote = (
            "the principle that a self-regulating profession must regulate in the public "
            "interest applies regardless of whether there is an express public interest "
            "clause in the enabling act."
        )
        source = (
            "In my view, the principle that a self\u2011regulating profession must regulate "
            "in the public interest applies regardless of whether there is an express "
            "public interest clause in the enabling act."
        )

        self.assertEqual(verifier._quote_match_score(quote, source), 1.0)
        corrected = verifier._build_corrected_citation(quote, source, prefer_single_quotes=False)
        self.assertIn("self-regulating", corrected)
        self.assertNotIn("self - regulating", corrected)

    def test_apply_quote_checks_treats_hyphen_variant_as_perfect_match(self):
        rows = [
            {
                "footnote_id": 1,
                "citation_part_index": 1,
                "citation_part_kind": "case",
                "citation_part_anchor_text": (
                    "[89] In my view, the principle that a self\u2011regulating profession "
                    "must regulate in the public interest applies regardless of whether "
                    "there is an express public interest clause in the enabling act."
                ),
            }
        ]
        quote = (
            "the principle that a self-regulating profession must regulate in the public "
            "interest applies regardless of whether there is an express public interest "
            "clause in the enabling act."
        )
        quotes = {
            1: [
                {
                    "quote_inner": quote,
                    "quote_raw": f"\u201c{quote}\u201d",
                    "quote_delimiter_style": "SMART",
                }
            ]
        }

        verifier._apply_quote_checks(rows, quotes)

        self.assertEqual(rows[0]["quote_check_status"], "OG_PINPOINT_MATCH")
        self.assertIn("self-regulating", rows[0]["quote_corrected_citation"])
        self.assertNotIn("self - regulating", rows[0]["quote_corrected_citation"])

    def test_terminal_punctuation_and_source_wrapper_quote_do_not_make_partial(self):
        rows = [
            {
                "footnote_id": 1,
                "citation_part_index": 1,
                "citation_part_kind": "case",
                "citation_part_anchor_text": (
                    'The components are "the law is supreme over the acts of both '
                    'government and private persons"; "an actual order of positive laws"; '
                    'and "the exercise of all public power must find its ultimate source '
                    'in a legal rule".'
                ),
            }
        ]
        q1 = "the law is supreme over the acts of both government and private persons"
        q2 = "an actual order of positive laws"
        q3 = "the exercise of all public power must find its ultimate source in a legal rule."
        quotes = {
            1: [
                {"quote_inner": q1, "quote_raw": f"\u201c{q1}\u201d", "quote_delimiter_style": "SMART"},
                {"quote_inner": q2, "quote_raw": f"\u201c{q2}\u201d", "quote_delimiter_style": "SMART"},
                {"quote_inner": q3, "quote_raw": f"\u201c{q3}\u201d", "quote_delimiter_style": "SMART"},
            ]
        }

        verifier._apply_quote_checks(rows, quotes)

        self.assertEqual(rows[0]["quote_check_status"], "OG_PINPOINT_MATCH")
        self.assertIn("the exercise of all public power must find its ultimate source in a legal rule.", rows[0]["quote_corrected_citation"])
        self.assertNotIn('rule ".', rows[0]["quote_corrected_citation"])
        self.assertNotIn('rule ".”', rows[0]["quote_corrected_citation"])

    def test_bracketed_suffix_fragment_url_does_not_absorb_intro(self):
        url = "https://www.canlii.org/en/on/onca/doc/2023/2023onca129/2023onca129.html#par38"
        source = 'the issue or as some have put it, a "trial in a box": Young, at p'
        quote = "trial[s] in a box,"
        fragment = verifier._source_side_quote_fragment_text(source, quote)

        target, count = verifier._build_source_side_text_fragment_url(url, fragment, source)

        self.assertEqual(count, 1)
        self.assertIn("#par38:~:text=trial%20in%20a%20box", target)
        self.assertNotIn("the%20issue%20or", target)

    def test_bracketed_suffix_can_align_to_expanded_source_word(self):
        source = 'yet some lawyers consider anti-SLAPP motions to be "trials in a box" rather than screening motions'
        quote = "trial[s] in a box,"

        fragment = verifier._source_side_quote_fragment_text(source, quote)

        self.assertEqual(fragment, "trials in a box")

    def test_bracketed_suffix_exact_match_scores_as_exact(self):
        source = 'the issue or as some have put it, a "trial in a box": Young, at p'
        quote = "trial[s] in a box,"

        self.assertEqual(verifier._quote_match_score(quote, source), 1.0)

    def test_bracketed_suffix_apply_quote_check_is_perfect_without_ellipsis(self):
        rows = [
            {
                "footnote_id": 1,
                "citation_part_index": 1,
                "citation_part_kind": "case",
                "citation_part_link": "https://www.canlii.org/en/on/onca/doc/2023/2023onca129/2023onca129.html#par38",
                "citation_part_anchor_text": 'the issue or as some have put it, a "trial in a box": Young, at p',
            }
        ]
        quotes = {
            1: [
                {
                    "quote_inner": "trial[s] in a box,",
                    "quote_raw": '"trial[s] in a box,"',
                    "quote_delimiter_style": "STRAIGHT",
                }
            ]
        }

        verifier._apply_quote_checks(rows, quotes)

        self.assertEqual(rows[0]["quote_check_status"], "OG_PINPOINT_MATCH")
        self.assertEqual(rows[0]["matched_source_fragment"], "trial in a box")
        self.assertNotIn("...", rows[0]["quote_corrected_citation"])

    def test_belnavis_canlii_fragment_keeps_full_stable_target(self):
        url = "https://www.canlii.org/en/ca/scc/doc/1997/1997canlii320/1997canlii320.html#par65"
        fragment = (
            "The courts have little \u201cfeel\u201d for what this means to persons who have committed "
            "no wrong or any idea of the number of such people who may be harassed by the "
            "overly zealous elements in any police force"
        )
        context = (
            "most cases that come before them relate to someone who has already been convicted. "
            + fragment
            + ". If such a draconian regime is to be imposed, it should be done by Parliament."
        )

        target, count = verifier._build_source_side_text_fragment_url(url, fragment, context)

        self.assertEqual(count, 1)
        self.assertIn("The%20courts%20have%20little", target)
        self.assertIn("overly%20zealous%20elements%20in%20any%20police%20force", target)
        self.assertNotIn("has%20already%20been%20convicted.-", target)

    def test_supply_chains_sec3_keeps_full_stable_target(self):
        url = "https://www.canlii.org/en/ca/laws/stat/sc-2023-c-9/latest/sc-2023-c-9.html#sec3"
        fragment = (
            "to implement Canada\u2019s international commitment to contribute to the fight "
            "against forced labour and child labour"
        )
        context = (
            "Purpose 3 The purpose of this Act is "
            + fragment
            + " through the imposition of reporting obligations on entities."
        )

        target, count = verifier._build_source_side_text_fragment_url(url, fragment, context)

        self.assertEqual(count, 1)
        self.assertIn("to%20implement%20Canada%E2%80%99s%20international%20commitment", target)
        self.assertIn("forced%20labour%20and%20child%20labour", target)
        self.assertNotIn("of%20this%20Act%20is-,", target)

    def test_short_pdf_page_fragment_uses_context_to_avoid_first_pdf_match(self):
        url = "https://digitalcommons.schulichlaw.dal.ca/cgi/viewcontent.cgi?article=1955&context=dlj#page=19"
        fragment = "defender of the rule of law"
        context = (
            'described as "the guardian of the public interest"67 or "the '
            'defender of the rule of law."68 despite kent roach\'s assertion that'
        )

        target, count = verifier._build_source_side_text_fragment_url(url, fragment, context)

        self.assertEqual(count, 1)
        self.assertIn("#page=19:~:text=", target)
        self.assertIn("-,defender%20of%20the%20rule%20of%20law", target)
        self.assertNotIn("public%20interest%2267%20or%20%22the%20defender", target)
        self.assertNotIn("68%20despite", target)

    def test_quote_check_fragment_uses_existing_context_scoping_for_short_pdf_phrase(self):
        url = "https://digitalcommons.schulichlaw.dal.ca/cgi/viewcontent.cgi?article=1955&context=dlj#page=19"
        fragment = "defender of the rule of law"
        context = (
            'described as "the guardian of the public interest"67 or "the '
            'defender of the rule of law."68 despite kent roach\'s assertion that'
        )

        result = verifier._build_quote_check_fragment_url(url, fragment, context)

        self.assertTrue(result.verified)
        self.assertEqual(result.fragment_count, 1)
        self.assertIn("#page=19:~:text=", result.url)
        self.assertIn("-,defender%20of%20the%20rule%20of%20law", result.url)
        self.assertNotIn("public%20interest%2267%20or%20%22the%20defender", result.url)
        self.assertNotIn("68%20despite", result.url)

    def test_verified_fragment_rejects_unscoped_repeated_short_phrase(self):
        url = "https://digitalcommons.schulichlaw.dal.ca/cgi/viewcontent.cgi?article=1955&context=dlj#page=19"
        context = (
            "This article describes the defender of the rule of law in one place. "
            "Later it discusses the defender of the rule of law again."
        )
        candidate = url + ":~:text=defender%20of%20the%20rule%20of%20law"

        self.assertFalse(
            verifier._verify_text_fragment_url(candidate, context, "defender of the rule of law")
        )

    def test_quote_check_fragment_keeps_multiple_directives_for_multi_quote_context(self):
        url = "https://example.test/statute#sec6"
        fragment = (
            "is frivolous or vexatious,\n\n"
            "was made in bad faith or for an improper purpose or motive,\n\n"
            "is entirely without merit."
        )
        context = (
            "6 The Minister may refuse to consider a request if it is frivolous or vexatious, "
            "was made in bad faith or for an improper purpose or motive, or is entirely "
            "without merit."
        )

        result = verifier._build_quote_check_fragment_url(url, fragment, context)

        self.assertTrue(result.verified)
        self.assertEqual(result.fragment_count, 3)
        self.assertEqual(result.url.count("text="), 3)
        self.assertIn("is%20frivolous%20or%20vexatious", result.url)
        self.assertIn("was%20made%20in%20bad%20faith", result.url)
        self.assertIn("is%20entirely%20without%20merit", result.url)

    def test_quote_check_fragment_without_context_does_not_export_todo(self):
        url = "https://example.com/source.html"
        fragment = '"first quote"\n\n"second quote"'

        result = verifier._build_quote_check_fragment_url(url, fragment)

        self.assertFalse(result.verified)
        self.assertEqual(result.fragment_count, 0)
        self.assertEqual(result.reason, "unmatched_multi_part")
        self.assertNotIn("TODO", result.url)
        self.assertNotIn("TODO", result.reason)

    def test_quote_check_fragment_uses_first_duplicate_source_part(self):
        url = "https://www.canlii.org/en/ab/abpc/doc/2018/2018abpc302/2018abpc302.html"
        fragment = "living separate and apart,\n\nliving separate and apart"
        context = (
            "one or both of the adult interdependent partners have obtained a declaration "
            "of irreconcilability under section 83, (ii) the adult interdependent partners "
            "are living separate and apart, or (iii) although the adult interdependent "
            "partners are not living separate and apart, the adult interdependent partners "
            "have entered into an agreement."
        )

        result = verifier._build_quote_check_fragment_url(url, fragment, context)

        self.assertTrue(result.verified)
        self.assertEqual(result.fragment_count, 1)
        self.assertEqual(result.url.count("text="), 1)
        self.assertIn("partners%20are-,living%20separate%20and%20apart", result.url)
        parsed = verifier._parse_text_fragment_directive(
            verifier._text_fragment_directives_from_url(result.url)[0]
        )
        self.assertTrue(parsed.target_start.endswith("apart,"))
        self.assertFalse(parsed.suffix)







    def test_spaced_suffix_punctuation_stays_in_context(self):
        built = verifier._build_source_side_text_fragment_directive(
            "target",
            "target (see Smith)",
            force_context=True,
        )

        parsed = verifier._parse_text_fragment_directive(built.directive)
        self.assertEqual(parsed.target_start, "target")
        self.assertEqual(parsed.suffix, "(see Smith")

        period_built = verifier._build_source_side_text_fragment_directive(
            "criminal liability",
            "criminal liability. Depriving a person of liberty",
            force_context=True,
        )
        period_parsed = verifier._parse_text_fragment_directive(period_built.directive)
        self.assertEqual(period_parsed.target_start, "criminal liability.")

    def test_fragment_context_strips_canlii_paragraph_markers(self):
        for marker in ("[28] A taking", "28] A taking", "(d) A taking", "d. A taking"):
            with self.subTest(marker=marker):
                self.assertEqual(
                    verifier._strip_fragment_context_list_marker(marker),
                    "A taking",
                )

    def test_fragment_context_does_not_start_mid_decimal_citation(self):
        source = "Section 33.1 Cr. C. are mutually exclusive."
        start = source.index("mutually")

        prefix, _suffix = verifier._text_fragment_context_window(
            source,
            start,
            start + len("mutually exclusive"),
            ["mutually", "exclusive"],
            4,
        )

        self.assertEqual(prefix, "C. are")
        self.assertFalse(prefix.startswith("1 "))

    def test_comma_not_absorbed_when_quote_omits_it(self):
        source = "At the first stage of the Anns test, two questions arise."
        built = verifier._build_source_side_text_fragment_directive(
            "Anns test",
            source,
            force_context=True,
            document_text=source,
            require_document_unique=True,
        )

        parsed = verifier._parse_text_fragment_directive(built.directive)
        self.assertEqual(parsed.target_start, "Anns test")
        self.assertEqual(parsed.prefix, "first stage of the")
        self.assertFalse(parsed.suffix)


    def test_fragment_directive_preserves_internal_nonbreaking_space(self):
        suffix = "the rule of law.\u00a0 As Hogg explains"

        directive = verifier._text_fragment_directive("target", suffix=suffix)

        self.assertIn("law.%C2%A0%20As", directive)
        parsed = verifier._parse_text_fragment_directive(directive)
        self.assertIn("law.\u00a0 As", parsed.suffix)

    def test_range_target_keeps_adjacent_trailing_punctuation(self):
        source = "one two three four five six seven eight nine ten. Afterward"
        end = source.index(" Afterward")

        _start_target, end_target = verifier._text_fragment_range_targets(
            source, 0, end
        )

        self.assertEqual(end_target, "eight nine ten.")






    def test_quote_check_fragment_handles_accented_source_word(self):
        url = "https://example.test/case#par2"
        source = (
            "[ 2 ] Strategic lawsuits against public participation are known as SLAPPs. "
            "In a SLAPP, the claim is merely a façade for the plaintiff, who is in fact "
            "manipulating the judicial system in order to limit the effectiveness of the "
            "opposing party's speech and deter that party from participating in public affairs."
        )
        quote = (
            "In a SLAPP, the claim is merely a facade for the plaintiff, who is in fact "
            "manipulating the judicial system in order to limit the effectiveness of the "
            "opposing party's speech and deter that party from participating in public affairs."
        )

        fragment = verifier._source_side_quote_fragment_text(source, quote)
        result = verifier._build_quote_check_fragment_url(url, fragment, source)

        self.assertIn("façade", fragment)
        self.assertTrue(result.verified)
        self.assertIn("fa%C3%A7ade", result.url)




    def test_source_side_fragment_strips_canlii_section_marker(self):
        source = (
            "2 (1) A Member breaches this Act if the Member takes part in a decision "
            "knowing that the decision might further a private interest of the Member."
        )
        quote = (
            "[a] Member breaches this Act if the Member takes part in a decision "
            "knowing that the decision might further a private interest of the Member."
        )

        fragment = verifier._source_side_quote_fragment_text(source, quote)

        self.assertTrue(fragment.startswith("A Member breaches this Act"))
        self.assertFalse(fragment.startswith("2 (1)"))





    def test_apply_quote_checks_uses_matching_anchor_segment_pinpoint(self):
        rows = [
            {
                "footnote_id": 1,
                "citation_part_index": 1,
                "citation_part_kind": "case",
                "citation_part_link": "https://www.canlii.org/en/ca/scc/doc/2001/2001scc67/2001scc67.html#par38",
                "citation_part_anchor_text": (
                    "38 Representation before a tribunal concerns legal rights. "
                    "42 While provinces may regulate professions as part of their jurisdiction "
                    "over property and civil rights, the legal profession is also part of the "
                    "administration of justice."
                ),
                "_citation_part_anchor_segments": [
                    {
                        "fragment": "par38",
                        "text": "38 Representation before a tribunal concerns legal rights.",
                    },
                    {
                        "fragment": "par42",
                        "text": (
                            "42 While provinces may regulate professions as part of their "
                            "jurisdiction over property and civil rights, the legal profession "
                            "is also part of the administration of justice."
                        ),
                    },
                ],
                "pinpoint_fragments": '["par38", "par42"]',
            }
        ]
        quotes = {
            1: [
                {
                    "quote_inner": "property and civil rights",
                    "quote_raw": '"property and civil rights"',
                    "quote_delimiter_style": "STRAIGHT",
                }
            ]
        }

        verifier._apply_quote_checks(rows, quotes)

        self.assertEqual(rows[0]["quote_check_status"], "OG_PINPOINT_MATCH")
        self.assertEqual(rows[0]["quote_match_pinpoint"], "par42")
        self.assertEqual(
            rows[0]["quote_match_link"],
            "https://www.canlii.org/en/ca/scc/doc/2001/2001scc67/2001scc67.html#par42",
        )
        self.assertEqual(rows[0]["matched_source_fragment"], "property and civil rights")

    def test_apply_quote_checks_dedupes_repeated_quote_within_one_cell(self):
        rows = [
            {
                "footnote_id": 1,
                "citation_part_index": 1,
                "citation_part_kind": "case",
                "citation_part_anchor_text": (
                    "the adult interdependent partners are living separate and apart, "
                    "or although the adult interdependent partners are not living separate "
                    "and apart, they have entered into an agreement."
                ),
            }
        ]
        quotes = {
            1: [
                {
                    "quote_inner": "living separate and apart,",
                    "quote_raw": '"living separate and apart,"',
                    "quote_delimiter_style": "STRAIGHT",
                },
                {
                    "quote_inner": "living separate and apart",
                    "quote_raw": '"living separate and apart"',
                    "quote_delimiter_style": "STRAIGHT",
                },
            ]
        }

        verifier._apply_quote_checks(rows, quotes)

        self.assertEqual(rows[0]["quote_check_status"], "OG_PINPOINT_MATCH")
        self.assertNotIn("\n\n", rows[0]["quote_corrected_citation"])
        self.assertNotIn("\n\n", rows[0]["matched_source_fragment"])
        self.assertEqual(
            rows[0]["quote_corrected_citation"].count("living separate and apart"),
            1,
        )

    def test_tanovich_bracket_initial_pdf_fragment_does_not_get_page_context(self):
        url = "https://digitalcommons.osgoode.yorku.ca/cgi/viewcontent.cgi?article=1446&context=ohlj#page=35"
        fragment = (
            "[t]he courts have little feel' for what [unconstitutional misconduct by the police] "
            "means to persons who have committed no wrong or any idea of the number of such people "
            "who may be harassed by the overly zealous elements in any police force."
        )
        context = (
            'as justice laforest 38 observed in his dissenting opinion in belnavis,196 "'
            + fragment
            + '" second, the vast majority of accused persons plead guilty.'
        )

        target, count = verifier._build_source_side_text_fragment_url(url, fragment, context)

        self.assertEqual(count, 1)
        self.assertIn("#page=35:~:text=%5Bt%5Dhe%20courts%20have%20little", target)
        self.assertNotIn("opinion%20in%20belnavis", target)
        self.assertNotIn(",-feel%27%20for%20what", target)

    def test_partial_overlap_rejects_only_obvious_noncontent_words(self):
        self.assertFalse(
            verifier._has_plausible_partial_content_overlap(
                "defender of the rule of law,",
                "of the parties is the Government of Alberta or a Minister or Provincial agency",
            )
        )
        self.assertFalse(
            verifier._has_plausible_partial_content_overlap(
                "the rule of the law of the land",
                "the parties of the government of the province",
            )
        )

    def test_partial_overlap_allows_real_content_overlap(self):
        self.assertTrue(
            verifier._has_plausible_partial_content_overlap(
                "adequate and proper testing",
                "the claim had to be supported by adequate and proper test results",
            )
        )
        self.assertTrue(
            verifier._has_plausible_partial_content_overlap(
                "[t]he courts have little feel for what [police action] means to persons",
                "The courts have little feel for what this means to persons who have committed no wrong",
            )
        )

    def test_apply_quote_checks_rejects_stopword_only_partial(self):
        rows = [
            {
                "footnote_id": 1,
                "citation_part_index": 1,
                "citation_part_kind": "statute",
                "citation_part_anchor_text": "the parties of the government of the province",
            }
        ]
        quotes = {
            1: [
                {
                    "quote_inner": "the rule of the law of the land",
                    "quote_raw": '"the rule of the law of the land"',
                    "quote_delimiter_style": "STRAIGHT",
                }
            ]
        }

        verifier._apply_quote_checks(rows, quotes)

        self.assertEqual(rows[0]["quote_check_status"], "NO_MATCH")
        self.assertEqual(rows[0]["quote_corrected_citation"], '"the rule of the law of the land"')
        self.assertNotIn("[rule]", rows[0]["quote_corrected_citation"])
        self.assertFalse(rows[0].get("matched_source"))



class BlindFragmentBuildTests(unittest.TestCase):
    """Blind (A2AJ-sourced) fragment building for pages we never fetched."""

    FAKE_URL = "https://www.canlii.org/en/ca/scc/doc/2099/2099scc99/2099scc99.html#par5"

    def setUp(self):
        verifier._FRAGMENT_DOC_TEXT_CACHE.clear()
        verifier._A2AJ_LOCKED_DOCUMENTS.clear()
        verifier._A2AJ_LOCKED_STRUCTURES.clear()

    tearDown = setUp

    def test_a2aj_paragraph_index_uses_consecutive_observed_sequence(self):
        text = "\n".join(
            [
                "2024 SCC 99",
                "[1] First paragraph contains enough ordinary decision prose to establish substantive structure.",
                "[2] Second paragraph contains enough ordinary decision prose to establish substantive structure.",
                "19. (1) An embedded statutory provision.",
                "[4] Fourth paragraph contains the distinctive quotation and enough substantive decision prose.",
                "[6] Sixth paragraph after another missing marker contains substantive decision prose.",
                "[7] Seventh paragraph contains enough ordinary decision prose to establish substantive structure.",
                "[8] Eighth paragraph contains enough ordinary decision prose to establish substantive structure.",
            ]
        )
        index = verifier._a2aj_paragraph_index(text)
        self.assertEqual([number for number, *_rest in index], [1, 2, 4, 6, 7, 8])
        self.assertIn("19. (1)", index[1][3])
        located = verifier._locate_a2aj_paragraph(
            text,
            "distinctive quotation",
            self.FAKE_URL,
        )
        self.assertEqual(located["label"], "par4")
        self.assertTrue(located["link"].endswith("#par4"))

    def test_a2aj_paragraph_index_rejects_merely_increasing_numbers(self):
        text = "\n".join(f"{number}. Citation or list item" for number in [2, 10, 20, 30, 40, 50])
        self.assertEqual(verifier._a2aj_paragraph_index(text), [])

    def test_a2aj_paragraph_index_rejects_mixed_numbered_lists(self):
        long_item = "This list entry contains enough ordinary words to make its individual block look deceptively substantive."
        for marker in ("{}", "{}."):
            with self.subTest(marker=marker):
                text = "\n".join([
                    marker.format(1) + " SHORT CODE",
                    marker.format(2) + " ANOTHER CODE",
                    marker.format(3) + " " + long_item,
                    marker.format(4) + " " + long_item,
                    marker.format(5) + " " + long_item,
                ])
                self.assertEqual(verifier._a2aj_paragraph_index(text), [])

    def test_a2aj_paragraph_index_prefers_brackets_over_competing_style(self):
        bracketed = [f"[{number}] Bracket paragraph" for number in range(1, 6)]
        dotted = [f"{number}. Dotted paragraph" for number in range(20, 25)]
        text = "\n".join(f"{line} with enough substantive decision words to form a real paragraph."
                         for line in bracketed + dotted)
        self.assertEqual([item[0] for item in verifier._a2aj_paragraph_index(text)], list(range(1, 6)))

    def test_a2aj_paragraph_index_rejects_monotone_endnotes(self):
        body = "Decision reasons occupy most of this document. " * 80
        notes = "\n".join(f"{n} . See note {n - 1} and R.S.C. 1985, c. C-1." for n in range(1, 10))
        self.assertEqual(verifier._a2aj_paragraph_index(body + "\n" + notes), [])

    def test_a2aj_paragraph_index_resumes_around_quoted_counter(self):
        lines = [
            "[1] First substantive judgment paragraph with enough ordinary words to pass structure checks.",
            "[2] Second substantive judgment paragraph with enough ordinary words to pass structure checks.",
            "[3] The court quotes another numbered decision in the following passage with context.",
            "[25] Quoted decision paragraph with enough words to look superficially like judgment text.",
            "[26] Another quoted decision paragraph with enough words to look superficially like judgment text.",
            "[4] The primary judgment resumes with enough ordinary substantive words for structure checks.",
            "[5] The primary judgment continues with enough ordinary substantive words for structure checks.",
        ]
        text = "\n".join(lines)
        self.assertEqual([item[0] for item in verifier._a2aj_paragraph_index(text)], [1, 2, 3, 4, 5])
        self.assertIsNone(verifier._locate_a2aj_paragraph(text, "Another quoted decision paragraph", self.FAKE_URL))

    def test_a2aj_explicit_page_locator(self):
        text = (
            "page 514\nOpening reporter-page text.\n"
            "page 515\nThe distinctive reporter quotation appears here.\n"
            "page 516\nClosing reporter-page text."
        )
        located = verifier._locate_a2aj_page(text, "distinctive reporter quotation")
        self.assertEqual(located["label"], "page 515")

    def test_a2aj_explicit_page_locator_accepts_observed_scc_variants(self):
        text = (
            "[Page 514]\nOpening reporter-page text ends here [Page515]"
            "The distinctive reporter quotation appears here.\n"
            "Page 516]\nClosing reporter-page text."
        )
        located = verifier._locate_a2aj_page(text, "distinctive reporter quotation")
        self.assertEqual(located["label"], "page 515")

    def test_a2aj_page_structure_rejects_p_abbreviation_and_unanchored_scc_page(self):
        self.assertEqual(verifier._a2aj_page_structure("[p. 21]\na\n[p. 22]\nb\n[p. 23]\nc"), [])
        text = "[page 21]\na\n[page 22]\nb\n[page 23]\nc"
        structure = verifier._a2aj_structure.analyze(text, "case", "2008 SCC 61", "", "SCC")
        self.assertEqual(structure["pages"], [])

    def test_a2aj_page_structure_uses_only_exact_consecutive_intervals(self):
        text = (
            "[Page 3]\npage three\n[Page 4]\npage four\n"
            "[Page 5]\nambiguous pages five through eight\n[Page 9]\npage nine"
        )
        pages = verifier._a2aj_structure.page_structure(text, 2)
        self.assertEqual([page[0] for page in pages], [2, 3, 4])
        self.assertNotIn("ambiguous pages", " ".join(page[3] for page in pages))

    def test_a2aj_page_structure_rejects_inversion(self):
        text = "page 10\na\npage 12\nb\npage 11\nc\npage 13\nd"
        self.assertEqual(verifier._a2aj_page_structure(text), [])

    def test_a2aj_page_structure_rejects_sparse_references(self):
        text = "page 10\na\npage 108\nb\npage 1087\nc"
        self.assertEqual(verifier._a2aj_page_structure(text), [])

    def test_a2aj_legislation_section_locator(self):
        text = (
            "1 Short title and introductory words governing this enactment.\n"
            "2 Definitions and interpretive provisions used throughout this enactment.\n"
            "3 Duties of the Minister under this enactment.\n"
            "(1) The Minister must prepare an annual report.\n"
            "(2) The Minister must publish reports as follows.\n"
            "(a) A public report must include prescribed information.\n"
            "(i) The distinctive annual report must be published every year.\n"
            "4 Regulations may prescribe further procedural requirements under this enactment."
        )
        url = "https://www.canlii.org/en/ca/laws/stat/example/latest/example.html"
        located = verifier._locate_a2aj_section(text, "distinctive annual report", url)
        self.assertEqual(located["label"], "sec3(2)(a)(i)")
        self.assertTrue(located["link"].endswith("#sec3"))

    def test_a2aj_rule_numbering_is_gated_by_exact_instrument_name(self):
        text = (
            "1-1 First rule text.\n1-2 Second rule text.\n"
            "1-3 Distinctive third rule text.\n2-1 Fourth rule text."
        )
        rules = verifier._a2aj_structure.analyze(
            text, "law", "BC Reg 168/2009", "", "REGULATIONS-BC",
            "Supreme Court Civil Rules",
        )
        ordinary_regulation = verifier._a2aj_structure.analyze(
            text, "law", "O Reg 170/03", "", "REGULATIONS-ON",
            "Drinking Water Systems",
        )
        self.assertEqual([item[0] for item in rules["sections"]], ["1-1", "1-2", "1-3", "2-1"])
        self.assertEqual(ordinary_regulation["status"], "unavailable")

    def test_a2aj_inline_first_subrule_is_indexed(self):
        text = (
            "11.9 First rule text.\n"
            "11.10(1) Distinctive first subrule text.\n"
            "(2) Distinctive second subrule text.\n"
            "11.11 Next rule text."
        )
        labels = [item[1] for item in verifier._a2aj_structure.legislation_blocks(text)]
        self.assertIn("sec11.10(1)", labels)
        self.assertIn("sec11.10(2)", labels)

    def test_a2aj_section_locator_uses_precomputed_gated_blocks(self):
        text = "1-1 First rule.\n1-2 Distinctive rule quotation.\n1-3 Third rule."
        blocks = verifier._a2aj_structure.legislation_blocks(text, allow_hyphen=True)
        structure = {"type": "section", "blocks": blocks}
        url = "https://www.canlii.org/en/bc/laws/regu/example/latest/example.html"
        with mock.patch.object(
            verifier._a2aj_structure,
            "legislation_blocks",
            side_effect=AssertionError("recomputed ungated structure"),
        ):
            located = verifier._locate_a2aj_pinpoint(
                text, "Distinctive rule quotation", url, structure, min_score=0.98
            )
        self.assertEqual(located["label"], "sec1-2")
        self.assertTrue(located["link"].endswith("#sec1-2"))

    def test_a2aj_repeated_quote_in_multiple_subrules_falls_back_to_rule(self):
        text = (
            "11.9 First rule text.\n"
            "11.10(1) The repeated phrase applies here.\n"
            "(2) The repeated phrase applies here too.\n"
            "11.11 Next rule text."
        )
        blocks = verifier._a2aj_structure.legislation_blocks(text)
        located = verifier._locate_a2aj_section(
            text, "repeated phrase applies here", blocks=blocks
        )
        self.assertEqual(located["label"], "sec11.10")

    def test_decimal_and_hyphenated_canlii_section_anchors_are_preserved(self):
        base = "https://www.canlii.org/en/ab/laws/regu/example/latest/example.html"
        for fragment in ("sec11.10", "sec4-1", "sec1457"):
            with self.subTest(fragment=fragment):
                link = verifier._append_first_pinpoint_fragment(base, [fragment])
                self.assertTrue(link.endswith("#" + fragment))
                self.assertTrue(verifier._url_has_section_anchor(link))

    def test_a2aj_locked_document_precomputes_structure(self):
        text = "\n".join(
            f"[{number}] Substantive decision paragraph number {number} contains enough ordinary words for validation."
            for number in range(1, 6)
        )
        document = verifier.a2aj_client.A2AJDocument(
            dataset="SCC", citation="2024 SCC 1", alternate_citation="", name="Example",
            date="2024-01-01", url="", text=text, language="en", scraped_timestamp="",
            upstream_license="", raw={},
        )
        base = "https://www.canlii.org/en/ca/scc/doc/2024/2024scc1/2024scc1.html"
        verifier._register_a2aj_document(base, document, "case")
        self.assertEqual(verifier._A2AJ_LOCKED_STRUCTURES[base]["status"], "usable")
        self.assertEqual(verifier._A2AJ_LOCKED_STRUCTURES[base]["type"], "paragraph")


    def test_alt_a2aj_export_preserves_citation_link_and_uses_quote_match_link(self):
        import openpyxl

        author_link = "https://author.example/decision"
        match_link = (
            "https://www.canlii.org/en/ca/scc/doc/2024/2024scc1/"
            "2024scc1.html#par3"
        )
        with tempfile.TemporaryDirectory() as directory:
            path = f"{directory}\\result.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "FootnoteReferences"
            sheet.append([
                "Footnote #", "Footnote Text", "Quotes and proposition", "Citation",
                "Automatic  Checking  System \u25ba", "Corrected quote",
                "citation_part_link", "quote_corrected_citation", "quote_check_status",
                "quote_match_link", "quote_check_notes",
            ])
            sheet.append([
                1, "", "", "Example", "", '"exact quote"',
                author_link, '"exact quote"', "ALT_PINPOINT_MATCH_A2AJ",
                match_link, "par3",
            ])
            workbook.save(path)
            workbook.close()

            verifier.apply_cell_formatting(path)

            workbook = openpyxl.load_workbook(path)
            sheet = workbook["FootnoteReferences"]
            self.assertEqual(sheet["D2"].hyperlink.target, author_link)
            self.assertEqual(sheet["F2"].hyperlink.target, match_link)
            workbook.close()

    def test_a2aj_exact_identity_replaces_model_link_without_browser(self):
        document = verifier.a2aj_client.A2AJDocument(
            dataset="SCC", citation="2024 SCC 1", alternate_citation="",
            name="R v Example", date="2024-01-01", url="https://example.test",
            text="[1] Decision text", language="en", scraped_timestamp="",
            upstream_license="", raw={},
        )
        lookup = verifier.a2aj_client.A2AJLookup("found", document, "exact_citation")
        with mock.patch.object(verifier, "USE_A2AJ", True), \
             mock.patch.object(verifier.a2aj_client, "lookup_document", return_value=lookup), \
             mock.patch.object(verifier.a2aj_client.get_client(), "coverage", return_value={"SCC"}), \
             mock.patch.object(verifier.LINK_RESOLVER, "resolve_url") as browser:
            link = verifier._resolve_footnote_part_link_unlocked(
                verbatim="R v Example, 2024 SCC 1 at para 2",
                citation_with_style="R v Example, 2024 SCC 1 at para 2",
                kind="case",
                link_candidate="https://www.canlii.org/en/on/onca/doc/2024/2024onca9/2024onca9.html",
                pinpoint_fragments=["par2"],
            )
        self.assertEqual(link, "https://www.canlii.org/en/ca/scc/doc/2024/2024scc1/2024scc1.html#par2")
        browser.assert_not_called()

    def test_a2aj_current_law_canlii_urls_are_constructed_from_locked_fields(self):
        cases = (
            (
                "LEGISLATION-FED",
                "RSC 1985, c 1 (2nd Supp)",
                "https://laws-lois.justice.gc.ca/eng/XML/C-52.6.xml",
                "ca", "stat", "rsc-1985-c-1-2nd-supp",
            ),
            (
                "LEGISLATION-FED",
                "SC 2019, c 28, s 1",
                "https://laws-lois.justice.gc.ca/eng/XML/I-2.5.xml",
                "ca", "stat", "sc-2019-c-28-s-1",
            ),
            (
                "LEGISLATION-AB",
                "SA 2002, c A-4.5",
                "https://kings-printer.alberta.ca/1266.cfm?page=A04P5.cfm&leg_type=Acts",
                "ab", "stat", "sa-2002-c-a-4.5",
            ),
            (
                "LEGISLATION-MB",
                "CCSM c L10",
                "https://web2.gov.mb.ca/laws/statutes/ccsm/l010.php?lang=en",
                "mb", "stat", "ccsm-c-l10",
            ),
            (
                "LEGISLATION-ON",
                "SO 2006, c 11, Sched A",
                "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/06c11",
                "on", "stat", "so-2006-c-11-sch-a",
            ),
            (
                "LEGISLATION-ON",
                "SO 2010, c 16 , Sched 4",
                "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/10c16b",
                "on", "stat", "so-2010-c-16-sch-4",
            ),
            (
                "LEGISLATION-FED",
                "SC 1980-81-82-83, c 108",
                "https://laws-lois.justice.gc.ca/eng/XML/C-46.xml",
                "ca", "stat", "sc-1980-81-82-83-c-108",
            ),
            (
                "REGULATIONS-AB",
                "Alta Reg 272/96",
                "https://kings-printer.alberta.ca/1266.cfm?page=1996_272.cfm&leg_type=Regs",
                "ab", "regu", "alta-reg-272-1996",
            ),
            (
                "REGULATIONS-FED",
                "SOR/86-304",
                "https://laws-lois.justice.gc.ca/eng/XML/SOR-86-304.xml",
                "ca", "regu", "sor-86-304",
            ),
            (
                "REGULATIONS-MB",
                "Man Reg 555/88 R",
                "https://web2.gov.mb.ca/laws/regs/current/555-88r.php?lang=en",
                "mb", "regu", "man-reg-555-88-r",
            ),
            (
                "REGULATIONS-NB",
                "NB Reg 82-73, r 11",
                "https://laws.gnb.ca/en/document/cr/82-73.11",
                "nb", "regu", "nb-reg-82-73",
            ),
            (
                "REGULATIONS-NL",
                "NLR 105/03",
                "https://www.assembly.nl.ca/Legislation/sr/regulations/rc030105.htm",
                "nl", "regu", "nlr-105-03",
            ),
            (
                "REGULATIONS-NT",
                "RRNWT 1990, c A-1",
                "https://www.justice.gov.nt.ca/en/files/legislation/example/example.r1.pdf",
                "nt", "regu", "rrnwt-1990-c-a-1",
            ),
            (
                "REGULATIONS-ON",
                "O Reg 114/99",
                "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/regulation/990114",
                "on", "regu", "o-reg-114-99",
            ),
            (
                "REGULATIONS-YT",
                "YOIC 2005/222",
                "https://laws.yukon.ca/cms/images/LEGISLATION/SUBORDINATE/2005/2005-0222/2005-0222_1.pdf",
                "yk", "regu", "yoic-2005-222",
            ),
        )
        for dataset, citation, source_url, jurisdiction, family, slug in cases:
            with self.subTest(dataset=dataset, citation=citation):
                expected = (
                    f"https://www.canlii.org/en/{jurisdiction}/laws/{family}/"
                    f"{slug}/latest/{slug}.html"
                )
                self.assertEqual(
                    verifier._a2aj_canlii_law_url(dataset, citation, source_url, "en"),
                    expected,
                )

    def test_a2aj_current_law_constructor_abstains_on_unproven_inputs(self):
        valid_federal_source = "https://laws-lois.justice.gc.ca/eng/XML/C-46.xml"
        cases = (
            ("LEGISLATION-SK", "SS 2013, c S-15.1", valid_federal_source, "en"),
            ("LEGISLATION-FED", "RSC 1985, c C-46", valid_federal_source, "fr"),
            ("LEGISLATION-ON", "SO 2000, c 36, Sched", "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/00p36", "en"),
            ("LEGISLATION-ON", "SO 2006, c 33, Sched Z2", "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/06p33", "en"),
            ("LEGISLATION-ON", "SO 2006, c 33 , Sched Z7", "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/06p33", "en"),
            ("LEGISLATION-ON", "SO 2009, c 18 , Sched 19", "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/09p18", "en"),
            ("LEGISLATION-ON", "SO 2009, c 18 , Sched 21", "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/09p18", "en"),
            ("LEGISLATION-ON", "SO 2010, c 16 , Sched 4", "https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/statute/wrong", "en"),
            ("LEGISLATION-FED", "RSC 1985, c 41 (4th Supp)", valid_federal_source, "en"),
            ("LEGISLATION-FED", "SC 1980-81-82-83, c 44, Part IV", valid_federal_source, "en"),
            ("LEGISLATION-FED", "SC 2018, c 12, s 49.1", valid_federal_source, "en"),
            ("REGULATIONS-FED", "PC 1945-7421", valid_federal_source, "en"),
            ("REGULATIONS-BC", "", "https://www.bclaws.gov.bc.ca/civix/document/id/complete/statreg/example/xml", "en"),
            ("LEGISLATION-FED", "RSC 1985, c C-46", "https://example.test/eng/XML/C-46.xml", "en"),
            ("REGULATIONS-AB", "Alta Reg 272/96", "https://kings-printer.alberta.ca/1266.cfm?page=1996_271.cfm&leg_type=Regs", "en"),
            ("REGULATIONS-AB", "Alta Reg 272/96", "https://kings-printer.alberta.ca/1266.cfm?leg_type=Regs", "en"),
        )
        for dataset, citation, source_url, language in cases:
            with self.subTest(dataset=dataset, citation=citation, source_url=source_url):
                self.assertEqual(
                    verifier._a2aj_canlii_law_url(
                        dataset, citation, source_url, language
                    ),
                    "",
                )

    def test_a2aj_current_law_identity_replaces_model_link_without_browser(self):
        probe = {
            "_a2aj_dataset": "LEGISLATION-FED",
            "_a2aj_citation": "RSC 1985, c C-46",
            "_a2aj_source_url": "https://laws-lois.justice.gc.ca/eng/XML/C-46.xml",
            "_a2aj_language": "en",
        }
        expected = (
            "https://www.canlii.org/en/ca/laws/stat/rsc-1985-c-c-46/"
            "latest/rsc-1985-c-c-46.html#sec33.1"
        )
        with mock.patch.object(
            verifier, "_a2aj_has_law_before_browser", return_value=probe
        ), mock.patch.object(verifier.LINK_RESOLVER, "resolve_url") as browser:
            link = verifier._resolve_footnote_part_link_unlocked(
                verbatim="Criminal Code, RSC 1985, c C-46, s 33.1",
                citation_with_style="Criminal Code, RSC 1985, c C-46, s 33.1",
                kind="statute",
                link_candidate="https://laws-lois.justice.gc.ca/eng/acts/C-46/",
                pinpoint_fragments=["sec33.1"],
                bare_citation="RSC 1985, c C-46",
            )
        self.assertEqual(link, expected)
        browser.assert_not_called()

    def test_a2aj_constructed_law_url_reconciles_exactly(self):
        document = verifier.a2aj_client.A2AJDocument(
            dataset="REGULATIONS-ON",
            citation="O Reg 114/99",
            alternate_citation="",
            name="Family Law Rules",
            date="",
            url="https://www.ontario.ca/laws/api/v2/legislation/en/doc-search/regulation/990114",
            text="1 Current rule text.",
            language="en",
            scraped_timestamp="",
            upstream_license="",
            raw={},
        )
        expected = (
            "https://www.canlii.org/en/on/laws/regu/o-reg-114-99/"
            "latest/o-reg-114-99.html"
        )
        self.assertTrue(verifier._a2aj_url_matches_document(expected, document, "law"))
        self.assertFalse(
            verifier._a2aj_url_matches_document(
                expected.replace("o-reg-114-99", "o-reg-114-98"), document, "law"
            )
        )
        self.assertFalse(
            verifier._a2aj_url_matches_document(
                expected.replace("/regu/", "/hstat/"), document, "law"
            )
        )


    def test_normalize_a2aj_source_text_strips_markdown_and_artifacts(self):
        raw = (
            "### Duty of person in control\n"
            "*federal law* means the whole of an Act , read together.\n"
            "( 1 ) The Minister may act."
        )
        out = verifier._normalize_a2aj_source_text(raw)
        self.assertEqual(
            out,
            "Duty of person in control\n"
            "federal law means the whole of an Act, read together.\n"
            "(1) The Minister may act.",
        )
        self.assertEqual(out.count("\n"), 2)

    def test_document_registry_keys_by_link_base(self):
        verifier._register_fragment_document_text(self.FAKE_URL, "some text")
        same_base = self.FAKE_URL.split("#", 1)[0] + "#par9"
        self.assertEqual(verifier._fragment_document_text_for_url(same_base), "some text")
        self.assertEqual(verifier._fragment_document_text_for_url("https://elsewhere.example/x"), "")

    def test_blind_build_disambiguates_headnote_copy_with_context(self):
        body_sentence = (
            "the safety of the community would not be endangered "
            "by the offender serving the sentence"
        )
        document = (
            f"Held: {body_sentence}.\n"
            "Some unrelated paragraph about other matters entirely.\n"
            f"In our respectful view, {body_sentence}, and nothing more.\n"
        )
        verifier._register_fragment_document_text(self.FAKE_URL, document)
        result = verifier._build_quote_check_fragment_url(
            self.FAKE_URL,
            body_sentence,
            f"In our respectful view, {body_sentence}, and nothing more.",
        )
        self.assertEqual(result.fragment_count, 1)
        self.assertEqual(result.builder, "source_document")
        # must carry disambiguating context (prefix and/or suffix directive)
        self.assertTrue("-," in result.url or ",-" in result.url)

    def test_blind_long_fragment_uses_range_around_apostrophe_drift(self):
        url = "https://www.canlii.org/en/on/onca/doc/2020/2020onca333/2020onca333.html#par137"
        source = (
            "Parliament has accepted, alcohol intoxication is not capable, on its own, "
            "of inducing a state of automatism: see Preamble of Bill C-72. Had similar "
            "evidence been presented and accepted at Mr. Daviault’s retrial, he would "
            "have been convicted."
        )
        verifier._register_fragment_document_text(url, source)

        result = verifier._build_quote_check_fragment_url(
            url, source, source, prefer_range=True
        )

        directive = verifier._text_fragment_directives_from_url(result.url)[0]
        parsed = verifier._parse_text_fragment_directive(directive)
        self.assertTrue(result.verified)
        self.assertTrue(parsed.target_end)
        self.assertNotIn("Daviault’s", parsed.target_start + parsed.target_end)
        browser_text = source.replace("Daviault’s", "Daviault's")
        self.assertTrue(
            verifier._text_fragment_directive_matches(browser_text, parsed)
        )

    def test_blind_build_suppresses_fragment_when_ambiguous(self):
        sentence = (
            "the safety of the community would not be endangered "
            "by the offender serving the sentence"
        )
        line = f"In our respectful view, {sentence}, and nothing more."
        document = f"{line}\nAnother paragraph sits here.\n{line}\n"
        verifier._register_fragment_document_text(self.FAKE_URL, document)
        result = verifier._build_quote_check_fragment_url(
            self.FAKE_URL,
            sentence,
            line,
        )
        self.assertEqual(result.fragment_count, 0)
        self.assertNotIn(":~:text=", result.url)

    def test_blind_build_rejects_target_spanning_paragraphs(self):
        document = (
            "The first paragraph ends with these exact words\n"
            "and the second paragraph starts with those exact words instead.\n"
        )
        verifier._register_fragment_document_text(self.FAKE_URL, document)
        # context presents the two paragraphs flattened into one line, as a
        # trimmed match window would
        flattened = (
            "The first paragraph ends with these exact words "
            "and the second paragraph starts with those exact words instead."
        )
        result = verifier._build_quote_check_fragment_url(
            self.FAKE_URL,
            "ends with these exact words and the second paragraph starts",
            flattened,
        )
        self.assertEqual(result.fragment_count, 0)

    def test_unregistered_document_keeps_legacy_window_behavior(self):
        sentence = (
            "the safety of the community would not be endangered "
            "by the offender serving the sentence"
        )
        result = verifier._build_quote_check_fragment_url(
            self.FAKE_URL,
            sentence,
            f"In our respectful view, {sentence}, and nothing more.",
        )
        self.assertEqual(result.fragment_count, 1)
        self.assertEqual(result.builder, "source_window")

    def test_fetch_for_row_normalizes_and_registers(self):
        row = {
            "bare_citation": "2099 SCC 99",
            "citation_part_kind": "case",
            "citation_part_link": self.FAKE_URL,
        }
        raw = "### Heading\n*term* means a thing , clearly.\n"
        document = verifier.a2aj_client.A2AJDocument(
            dataset="SCC", citation="2099 SCC 99", alternate_citation="",
            name="Example", date="2099-01-01", url="https://example.test/case",
            text=raw, language="en", scraped_timestamp="", upstream_license="", raw={},
        )
        with mock.patch.object(
            verifier.a2aj_client, "lookup_document",
            return_value=verifier.a2aj_client.A2AJLookup("found", document, "exact_citation"),
        ) as lookup:
            text = verifier._fetch_a2aj_source_text_for_row(row)
        lookup.assert_called_once_with("2099 SCC 99", "case", language="en")
        self.assertEqual(text, "Heading\nterm means a thing, clearly.\n")
        self.assertEqual(
            verifier._fragment_document_text_for_url(self.FAKE_URL), text
        )
        base = self.FAKE_URL.split("#", 1)[0]
        self.assertEqual(verifier._A2AJ_LOCKED_DOCUMENTS[base].citation, "2099 SCC 99")

    def test_fetch_for_row_does_not_bind_text_to_mismatched_case_url(self):
        wrong_url = (
            "https://www.canlii.org/en/ca/scc/doc/2099/2099scc98/"
            "2099scc98.html#par5"
        )
        row = {
            "bare_citation": "2099 SCC 99",
            "citation_part_kind": "case",
            "citation_part_link": wrong_url,
        }
        raw = "\n".join(
            f"[{number}] Decision paragraph {number} contains enough substantive words for structure."
            for number in range(1, 6)
        )
        document = verifier.a2aj_client.A2AJDocument(
            dataset="SCC", citation="2099 SCC 99", alternate_citation="",
            name="Example", date="2099-01-01", url="https://official.example/case",
            text=raw, language="en", scraped_timestamp="", upstream_license="", raw={},
        )
        with mock.patch.object(
            verifier.a2aj_client, "lookup_document",
            return_value=verifier.a2aj_client.A2AJLookup("found", document, "exact_citation"),
        ):
            text = verifier._fetch_a2aj_source_text_for_row(row)
        self.assertTrue(text)
        self.assertTrue(row["_a2aj_identity_locked"])
        self.assertFalse(row["_a2aj_url_reconciled"])
        self.assertIn("status", row["_a2aj_structure"])
        self.assertEqual(verifier._fragment_document_text_for_url(wrong_url), "")
        self.assertEqual(verifier._A2AJ_LOCKED_DOCUMENTS, {})

    def test_a2aj_law_pinpoint_is_row_local_when_url_is_unproven(self):
        link = (
            "https://www.canlii.org/en/ab/laws/hstat/rsa-1980-c-a-1/"
            "latest/rsa-1980-c-a-1.html"
        )
        quote = "The distinctive public duty applies in every proceeding"
        text = (
            "1 Introductory provision contains enough ordinary words for structure.\n"
            f"2 {quote} before the tribunal.\n"
            "3 Concluding provision contains enough ordinary words for structure."
        )
        row = {
            "footnote_id": 1,
            "citation_part_index": 1,
            "citation_part_kind": "statute",
            "citation_part_link": link,
            "bare_citation": "RSA 1980, c A-1",
        }
        document = verifier.a2aj_client.A2AJDocument(
            dataset="LEGISLATION-AB", citation="RSA 1980, c A-1", alternate_citation="",
            name="Example Act", date="", url="https://official.example/act",
            text=text, language="en", scraped_timestamp="", upstream_license="", raw={},
        )
        lookup = verifier.a2aj_client.A2AJLookup("found", document, "exact_citation")
        quotes = {1: [{
            "quote_inner": quote,
            "quote_raw": f'"{quote}"',
            "quote_delimiter_style": "STRAIGHT",
        }]}
        with mock.patch.object(verifier, "USE_A2AJ", True), mock.patch.object(
            verifier.a2aj_client, "lookup_document", return_value=lookup
        ):
            verifier._apply_quote_checks([row], quotes)
        self.assertEqual(row["citation_part_link"], link)
        self.assertEqual(row["quote_match_pinpoint"], "sec2")
        self.assertNotIn("quote_match_link", row)
        self.assertFalse(row["_a2aj_url_reconciled"])

    def test_law_link_resolution_does_not_rewrite_astat_or_hstat_to_stat(self):
        for family in ("astat", "hstat"):
            with self.subTest(family=family):
                link = (
                    f"https://www.canlii.org/en/ab/laws/{family}/sa-2022-c-8/"
                    "latest/sa-2022-c-8.html"
                )
                with mock.patch.object(
                    verifier.LINK_RESOLVER, "resolve_url", return_value=link
                ), mock.patch.object(verifier.a2aj_client, "lookup_document") as lookup:
                    resolved = verifier._resolve_footnote_part_link_unlocked(
                        verbatim="Example Act, SA 2022, c 8",
                        citation_with_style="Example Act, SA 2022, c 8",
                        kind="statute",
                        link_candidate=link,
                        pinpoint_fragments=[],
                    )
                self.assertEqual(resolved, link)
                lookup.assert_not_called()

    def test_fetch_for_row_rejects_non_exact_lookup(self):
        row = {
            "bare_citation": "2099 SCC 99",
            "citation_part_kind": "case",
            "citation_part_link": self.FAKE_URL,
        }
        for status in ("not_found", "ambiguous", "network_error"):
            with self.subTest(status=status), mock.patch.object(
                verifier.a2aj_client, "lookup_document",
                return_value=verifier.a2aj_client.A2AJLookup(status),
            ):
                self.assertEqual(verifier._fetch_a2aj_source_text_for_row(row), "")
                self.assertEqual(verifier._fragment_document_text_for_url(self.FAKE_URL), "")

    def test_fetch_for_row_preserves_canonical_rule_document_suffix(self):
        row = {
            "bare_citation": "NB Reg 82-73, r 2",
            "citation_part_kind": "statute",
            "citation_part_link": "other",
        }
        document = verifier.a2aj_client.A2AJDocument(
            dataset="REGULATIONS-NB", citation="NB Reg 82-73, r 2",
            alternate_citation="", name="NON-COMPLIANCE WITH THE RULES",
            date="", url="https://laws.gnb.ca/en/document/cr/Rule-2",
            text="2.01 The court may dispense with compliance.", language="en",
            scraped_timestamp="", upstream_license="", raw={},
        )
        with mock.patch.object(
            verifier.a2aj_client, "lookup_document",
            return_value=verifier.a2aj_client.A2AJLookup("found", document, "exact_citation"),
        ) as lookup:
            self.assertTrue(verifier._fetch_a2aj_source_text_for_row(row))
        lookup.assert_called_once_with("NB Reg 82-73, r 2", "statute", language="en")
        self.assertEqual(row["_a2aj_citation"], "NB Reg 82-73, r 2")

    def test_blind_build_tolerates_repeated_target_in_context_window(self):
        # Overlapping match regions repeat the quote inside the context, but
        # the document has it exactly once — the doc-line check must govern.
        sentence = (
            "never presented himself at the Canadian border and therefore "
            "never requested a determination"
        )
        document = (
            "An unrelated opening paragraph sits here.\n"
            f"The respondent {sentence} regarding his eligibility to claim.\n"
            "A closing paragraph follows the passage.\n"
        )
        context = (
            f"The respondent {sentence} regarding his eligibility to claim. "
            f"The respondent {sentence} regarding his eligibility to claim."
        )
        verifier._register_fragment_document_text(self.FAKE_URL, document)
        result = verifier._build_quote_check_fragment_url(
            self.FAKE_URL, sentence, context
        )
        self.assertEqual(result.fragment_count, 1)
        self.assertEqual(result.builder, "source_document")

    def test_a2aj_query_citation_strips_pinpoints_and_history(self):
        cases = {
            "2005 SCC 57 at para 68.": "2005 SCC 57",
            "[1962] SCR 746 at 763–64": "[1962] SCR 746",
            "2001 SCC 24 at para 47, [2001] 1 SCR 687": "2001 SCC 24",
            "RSC 1985, c C-46, s 33.1": "RSC 1985, c C-46",
            "RSA 2000 c P-7, ss 3(2)(b), 4(f)": "RSA 2000 c P-7",
            "RSC 1985, c 1 (2nd Supp)": "RSC 1985, c 1 (2nd Supp)",
            "1999 CanLII 4334 (rev'd on other grounds, 2001 NWTCA 1) at para 30":
                "1999 CanLII 4334",
            "2020 SCC 10 at para 5": "2020 SCC 10",
        }
        for raw, want in cases.items():
            self.assertEqual(verifier._a2aj_query_citation(raw), want, raw)

    def test_canlii_doc_citation_from_url(self):
        cases = {
            "https://www.canlii.org/en/ca/scc/doc/2023/2023scc17/2023scc17.html#par148":
                "2023 SCC 17",
            "https://www.canlii.org/en/ca/irb/doc/2011/2011canlii67655/2011canlii67655.html":
                "2011 CanLII 67655",
            "https://www.canlii.org/en/ca/laws/stat/rsc-1985-c-c-46/latest/rsc-1985-c-c-46.html": "",
            "https://example.com/doc/2023/2023scc17/2023scc17.html": "",
            "": "",
        }
        for url, want in cases.items():
            self.assertEqual(verifier._canlii_doc_citation_from_url(url), want, url)

    def test_normalize_a2aj_source_text_removes_dot_leaders_keeps_ellipses(self):
        raw = "Safe Third Country Designation.................... 5\nHe said \"stop...\" and left."
        out = verifier._normalize_a2aj_source_text(raw)
        self.assertNotIn("....", out)
        self.assertIn('"stop..."', out)

    def test_fetch_for_row_skips_unsupported_kinds(self):
        row = {
            "bare_citation": "Some Book",
            "citation_part_kind": "secondary",
            "citation_part_link": self.FAKE_URL,
        }
        with mock.patch.object(verifier.a2aj_client, "lookup_document") as fetch:
            self.assertEqual(verifier._fetch_a2aj_source_text_for_row(row), "")
        fetch.assert_not_called()
        self.assertEqual(verifier._fragment_document_text_for_url(self.FAKE_URL), "")


class CanliiPdfSanitizerTests(unittest.TestCase):
    def test_pdf_with_par_anchor_rewrites_to_html(self):
        self.assertEqual(
            verifier._sanitize_url_candidate(
                "https://www.canlii.org/en/ca/scc/doc/1985/1985canlii69/1985canlii69.pdf#par80"
            ),
            "https://www.canlii.org/en/ca/scc/doc/1985/1985canlii69/1985canlii69.html#par80",
        )

    def test_bare_pdf_rewrites_to_html(self):
        self.assertEqual(
            verifier._sanitize_url_candidate(
                "https://www.canlii.org/en/ca/scc/doc/1985/1985canlii69/1985canlii69.pdf"
            ),
            "https://www.canlii.org/en/ca/scc/doc/1985/1985canlii69/1985canlii69.html",
        )

    def test_pdf_page_pinpoint_is_preserved(self):
        url = "https://www.canlii.org/en/ca/scc/doc/1985/1985canlii69/1985canlii69.pdf#page=12"
        self.assertEqual(verifier._sanitize_url_candidate(url), url)

    def test_non_canlii_pdf_untouched(self):
        url = "https://example.com/reports/decision.pdf#par80"
        self.assertEqual(verifier._sanitize_url_candidate(url), url)


if __name__ == "__main__":
    unittest.main()
