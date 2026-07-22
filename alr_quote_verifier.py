#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ALR Footnote + Quote Verification (local runner)

What this does
- Extracts footnotes from a DOCX.
- Splits multi-citation footnotes using an OpenAI model.
- Resolves ibid/supra chains and propagates originating links.
- Produces an Excel workbook for audit/review.

Quick start (macOS)
1) Create a virtual environment:
     python3 -m venv .venv
     source .venv/bin/activate
2) Install dependencies:
     pip install -r requirements.txt
3) Set your OpenAI key:
     export OPENAI_API_KEY="YOUR_KEY"
4) Run on a folder of .docx files:
     python alr_quote_verifier.py --input "/path/to/folder" --output-name "CHECKED_EDITS"

Notes
- The script does not embed credentials; it reads OPENAI_API_KEY from the environment.

"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import statistics
import sys
import threading
import time
import unicodedata
import zipfile
from dataclasses import dataclass, field, replace
from functools import lru_cache
from typing import Any, Dict, Iterable, List, Optional, Tuple

import a2aj_client
import journal_search
from verifier_core import deterministic_splitter as _deterministic_splitter
from verifier_core import supra_fallbacks as _supra_fallbacks

from lxml import etree

# openai and openpyxl are deferred: together they were ~1.5s of GUI launch
# time and neither is needed before the first model call / first export.
# _ensure_openpyxl() binds the names below into module globals; every
# function that uses them is reached only via write_workbook,
# apply_cell_formatting, or finalize_workbook_export, which call it first.


def _ensure_openpyxl() -> None:
    global Workbook, Alignment, Font, PatternFill
    global get_column_letter, DataValidation, Hyperlink
    if "Workbook" in globals():
        return
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Alignment as _Alignment, Font as _Font, PatternFill as _PatternFill
    from openpyxl.utils import get_column_letter as _get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation as _DataValidation
    from openpyxl.worksheet.hyperlink import Hyperlink as _Hyperlink
    Workbook, Alignment, Font, PatternFill = _Workbook, _Alignment, _Font, _PatternFill
    get_column_letter, DataValidation, Hyperlink = _get_column_letter, _DataValidation, _Hyperlink

MAX_LOOKAHEAD_CHARS_FOR_QUOTE_TO_FOOTNOTE = int(os.environ.get("MAX_LOOKAHEAD", "400"))

import atexit
import hashlib
import json
import logging
import random
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit, quote, unquote

def _split_url(url: str) -> Tuple[str, str]:
    """
    Returns (base_url_without_fragment, fragment_without_hash).
    """
    u = (url or "").strip()
    if not u:
        return "", ""
    parts = urlsplit(u)
    base = urlunsplit((parts.scheme, parts.netloc, parts.path, parts.query, ""))
    frag = parts.fragment or ""
    return base, frag


def _recombine_url(base_url: str, fragment: str) -> str:
    if not base_url:
        return ""
    frag = (fragment or "").lstrip("#")
    return f"{base_url}#{frag}" if frag else base_url


def _canlii_html_variant_for_text(url: str) -> str:
    """Use CanLII HTML siblings for source-text work while preserving exported links."""
    base, frag = _split_url(url)
    if not base or "canlii.org" not in base.lower():
        return url
    if not urlsplit(base).path.lower().endswith(".pdf"):
        return url
    html_base = re.sub(r"\.pdf$", ".html", base, flags=re.IGNORECASE)
    return _recombine_url(html_base, frag)


def _canlii_source_lookup_url(url: str) -> str:
    """Normalize CanLII URLs for HTML/source lookup, stripping PDF page markers."""
    candidate = _canlii_html_variant_for_text(url)
    base, frag = _split_url(candidate)
    if not base or "canlii.org" not in base.lower():
        return candidate
    if re.match(r"^page=-?\d+$", (frag or "").strip(), flags=re.IGNORECASE):
        return base
    return candidate


def _scc_scrollable_text_url(url: str) -> str:
    """Use SCC's complete, scrollable inner view for generated text links."""
    base, frag = _split_url(url)
    parts = urlsplit(base)
    if (
        parts.hostname != "decisions.scc-csc.ca"
        or not parts.path.lower().endswith("/index.do")
    ):
        return url
    params = [
        (key, value)
        for key, value in parse_qsl(parts.query, keep_blank_values=True)
        if key.lower() not in {"iframe", "site_preference"}
    ]
    params.extend((("iframe", "true"), ("site_preference", "mobile")))
    base = urlunsplit(
        (parts.scheme, parts.netloc, parts.path, urlencode(params), "")
    )
    return _recombine_url(base, frag)


def _prefer_scc_official_quote_link(row: Dict[str, Any], link: str) -> str:
    """Use A2AJ's official SCC page for quote links, not citation links."""
    link = str(link or "").strip()
    link_base, link_fragment = _split_url(link)
    link_parts = urlsplit(link_base)
    is_canlii_scc = (
        link_parts.hostname in {"canlii.org", "www.canlii.org"}
        and re.search(r"/(?:ca/scc|ca/csc-scc)/", link_parts.path, flags=re.IGNORECASE)
    )
    is_official_scc = (
        link_parts.hostname == "decisions.scc-csc.ca"
        and link_parts.path.lower().endswith("/index.do")
    )
    if not (is_canlii_scc or is_official_scc):
        return link

    reconciled = bool(row.get("_a2aj_url_reconciled"))
    dataset = str(row.get("_a2aj_dataset") or "").strip()
    official = str(row.get("_a2aj_source_url") or "").strip()
    locked = None
    if not reconciled and is_canlii_scc:
        locked = _A2AJ_LOCKED_DOCUMENTS.get(link_base.lower())
        if locked:
            reconciled = True
            dataset = locked.dataset
            official = locked.url
    if not reconciled or dataset.upper() != "SCC":
        return link

    official_base, _official_fragment = _split_url(official)
    official_parts = urlsplit(official_base)
    if (
        official_parts.scheme.lower() not in {"http", "https"}
        or official_parts.hostname != "decisions.scc-csc.ca"
        or not official_parts.path.lower().endswith("/index.do")
    ):
        return link
    if locked:
        row["_a2aj_dataset"] = locked.dataset
        row["_a2aj_source_url"] = locked.url
        row["_a2aj_url_reconciled"] = True
        _register_fragment_document_text(
            locked.url, _normalized_a2aj_document_text(locked, "case")
        )
    anchor = _normalize_anchor_fragment(_fragment_anchor_only(link_fragment))
    if not re.fullmatch(r"par\d{1,4}", anchor, flags=re.IGNORECASE):
        anchor = _first_pinpoint_from_summary(row.get("quote_match_pinpoint") or "")
    if not re.fullmatch(r"par\d{1,4}", anchor, flags=re.IGNORECASE):
        return link
    return _recombine_url(official_base, anchor)


_SEC_FRAGMENT_INT_RE = re.compile(r"^sec(\d+)$", re.IGNORECASE)
_SEC_ANCHOR_RE = re.compile(r"sec\d+(?:[.-]\d+){0,3}[A-Za-z]?", re.IGNORECASE)
_CANLII_ANCHOR_RE = re.compile(
    r"(?:par\d+|sec\d+(?:[.-]\d+){0,3}[A-Za-z]?)", re.IGNORECASE
)


def _repair_decimal_section_fragments(
    fragments: Optional[List[str]], verbatim: str
) -> List[str]:
    """Upgrade truncated section anchors to their decimal form.

    Defensively upgrade a model-truncated "s 672.54(b)" -> "sec672" because
    CanLII keeps the decimal (#sec672.54) and #sec672 is a different provision.
    When the citation text shows exactly one decimal extension for the
    integer, restore it; parenthetical subsection letters stay out (not
    reliable in CanLII anchors).
    """
    out: List[str] = []
    for raw in fragments or []:
        m = _SEC_FRAGMENT_INT_RE.match(str(raw or "").strip())
        if m:
            decimals = set(re.findall(
                rf"\b{re.escape(m.group(1))}\.(\d+)", verbatim or ""))
            if len(decimals) == 1:
                out.append(f"sec{m.group(1)}.{next(iter(decimals))}")
                continue
        out.append(raw)
    return out


def _pinpoints_for_source_kind(
    kind: str,
    fragments: Optional[List[str]],
    pages: Optional[List[int]],
) -> Tuple[List[str], List[int]]:
    """Keep only pinpoint forms that can belong to this source type."""
    source_kind = (kind or "").strip().casefold()
    case_source = source_kind in {"case", "unreported"}
    law_source = source_kind in {"statute", "regulation", "legislation"}
    unresolved_reference = source_kind in {"", "other"}

    paragraph_pattern = r"par\d{1,4}"
    section_pattern = r"sec\d{1,8}(?:[.-]\d{1,8}){0,3}(?:\([^)]+\))*"
    if case_source:
        fragment_pattern = paragraph_pattern
    elif law_source:
        fragment_pattern = section_pattern
    elif unresolved_reference:
        fragment_pattern = rf"(?:{paragraph_pattern}|{section_pattern})"
    else:
        fragment_pattern = ""
    kept_fragments = []
    for fragment in fragments or []:
        value = str(fragment or "").strip()
        if fragment_pattern and re.fullmatch(
            fragment_pattern, value.lstrip("#"), flags=re.IGNORECASE
        ):
            kept_fragments.append(value)
    kept_pages: List[int] = []
    if not law_source:
        for page in pages or []:
            try:
                number = int(page)
            except (TypeError, ValueError):
                continue
            if number > 0 and number not in kept_pages:
                kept_pages.append(number)
    return kept_fragments, kept_pages


def _append_first_pinpoint_fragment(link: str, pinpoint_fragments: Optional[List[str]]) -> str:
    candidate = _sanitize_url_candidate(link)
    if not candidate or candidate.lower() == "other":
        return link
    base, existing_frag = _split_url(candidate)
    if not base or existing_frag or "canlii.org" not in base.lower():
        return link
    if urlsplit(base).path.lower().endswith(".pdf"):
        return link

    first = ""
    for raw in pinpoint_fragments or []:
        norm = _normalize_anchor_fragment(raw)
        if norm:
            first = norm
            break
    if not first:
        return link

    path = urlsplit(base).path.lower()
    if first.startswith("par") and "/doc/" in path:
        return _recombine_url(base, first)
    if first.startswith("sec") and "/laws/" in path:
        return _recombine_url(base, first)
    return link


def _strip_invalid_page_fragment(url: str) -> str:
    base, frag = _split_url(url)
    m = re.match(r"^page=(-?\d+)$", (frag or "").strip(), flags=re.IGNORECASE)
    if not m:
        return url
    try:
        page = int(m.group(1))
    except ValueError:
        return url
    if page < 1:
        return base
    return url


def _is_canlii_source_lookup_variant(original_url: str, source_url: str) -> bool:
    return bool(
        original_url
        and source_url
        and original_url != source_url
        and "canlii.org" in original_url.lower()
    )


def _encode_text_fragment(text: str) -> str:
    """Encode text for use in a W3C Text Fragment (:~:text=...)."""
    encoded = quote(text, safe="")
    # The text fragment spec treats -, &, and , as special characters
    for char in "-&,":
        encoded = encoded.replace(quote(char, safe=""), f"%{ord(char):02X}")
    return encoded


def _normalize_text_fragment_whitespace(text: str) -> str:
    """Collapse rendered ASCII whitespace without erasing an internal NBSP."""
    value = str(text or "").strip()
    return re.sub(r"[ \t\r\n\f\v]+", " ", value).strip()




def _node_class_contains(node: Any, token: str) -> bool:
    return token in (node.get("class") or "").split()




def _extract_fragment_boundaries(corrected: str) -> tuple[str, str]:
    """From a difflib-corrected quote extract the first and last unbracketed
    content words (up to 3 each) for use as text fragment boundaries.

    E.g. 'In ... this Court [, special] ... [a] question.' → ('In', 'question')
    E.g. 'only voluntary conduct – behaviour ... criminal liability.'
          → ('only voluntary conduct', 'criminal liability')
    """
    s = corrected.strip().strip("\"'").strip("\u201c\u201d")
    parts = re.split(r"(\[.*?\]|\.\.\.+)", s)

    start_words: List[str] = []
    for part in parts:
        if part.startswith("[") or part.startswith("."):
            break
        found = re.findall(r"[A-Za-z0-9]+(?:['\u2019][A-Za-z0-9]+)*", part)
        for w in found:
            if len(start_words) >= 3:
                break
            start_words.append(w)
        if len(start_words) >= 3:
            break

    end_words: List[str] = []
    for part in reversed(parts):
        if part.startswith("[") or part.startswith("."):
            break
        found = re.findall(r"[A-Za-z0-9]+(?:['\u2019][A-Za-z0-9]+)*", part)
        for w in reversed(found):
            if len(end_words) >= 3:
                break
            end_words.append(w)
        if len(end_words) >= 3:
            break
    end_words.reverse()

    start_phrase = " ".join(start_words)
    end_phrase = " ".join(end_words)
    # Fallback: if no brackets/dots, use single first/last word
    if start_phrase == end_phrase:
        all_w = start_phrase.split()
        if len(all_w) >= 2:
            return (all_w[0], all_w[-1])
    return (start_phrase, end_phrase)


def _phrase_regex_for_fragment(phrase: str) -> Optional[re.Pattern[str]]:
    words = re.findall(r"[A-Za-z0-9]+(?:['\u2019][A-Za-z0-9]+)*", phrase or "")
    if not words:
        return None
    pattern = r"\b" + r"[\s\u00a0]+".join(re.escape(w) for w in words) + r"\b"
    return re.compile(pattern, re.IGNORECASE)


def _text_fragment_quote_text(text: str) -> str:
    s = str(text or "").strip().strip("\"'").strip("\u201c\u201d")
    s = re.sub(r"\[([A-Za-z])\](?=[A-Za-z])", r"\1", s)
    s = re.sub(r"\[([^\]]+)\]", r"\1", s)
    s = s.replace("...", " ")
    return re.sub(r"\s+", " ", s).strip()


def _iter_text_fragment_quote_candidates(corrected: str) -> List[str]:
    text = str(corrected or "").strip()
    if not text:
        return []
    parts = [part.strip() for part in re.split(r"\n\s*\n+", text) if part.strip()]
    candidates: List[str] = []
    seen: set[str] = set()
    for part in parts or [text]:
        candidate = _text_fragment_quote_text(part)
        key = candidate.lower()
        if candidate and key not in seen:
            candidates.append(candidate)
            seen.add(key)
    return candidates


def _iter_single_text_fragment_quote_candidate(corrected: str) -> List[str]:
    candidates = _iter_text_fragment_quote_candidates(corrected)
    # TODO: Multi-quote text fragments are disabled for now. The old path
    # returned every candidate and produced one URL with multiple text=
    # directives, but that built bad fragments in multi-quote cells and needs
    # tighter scoping before it is re-enabled.
    # return candidates
    return candidates[:1]


def _text_fragment_word_spans(text: str) -> List[Tuple[str, int, int]]:
    return [(m.group(0), m.start(), m.end()) for m in _WORD_TOKEN_RE.finditer(text or "")]


def _text_fragment_phrase_spans(source_text: str, words: List[str]) -> List[Tuple[int, int]]:
    if not source_text or not words:
        return []
    pattern = (
        r"(?<![^\W_])"
        + r"[\W_]+".join(re.escape(w) for w in words)
        + r"(?![^\W_])"
    )
    return [(m.start(), m.end()) for m in re.finditer(pattern, source_text, flags=re.IGNORECASE)]


def _starts_with_words(words: List[str], prefix: List[str]) -> bool:
    if not words or not prefix or len(words) < len(prefix):
        return False
    return [w.lower() for w in words[:len(prefix)]] == [w.lower() for w in prefix]


def _choose_text_fragment_span(source_text: str, quote_text: str) -> Tuple[int, int]:
    words = _quote_word_tokens(quote_text)
    spans = _text_fragment_phrase_spans(source_text, words)
    if not spans:
        return (-1, -1)

    word_spans = _text_fragment_word_spans(source_text)
    best: Optional[Tuple[int, int]] = None
    best_score = -10**9
    for start, end in spans:
        raw_target = source_text[start:end]
        before = [w.lower() for w, _s, e in word_spans if e <= start][-len(words):]
        after = [w.lower() for w, s, _e in word_spans if s >= end][:len(words)]
        score = 0
        if raw_target == quote_text:
            score += 75
        elif raw_target.lower() == quote_text.lower():
            score += 25
        if _starts_with_words(after, words):
            score -= 100
        if after[:1] == ["means"]:
            score += 25
        score += len([w for w, _s, e in word_spans if e <= start][-4:])
        score += len([w for w, s, _e in word_spans if s >= end][:4])
        if score > best_score:
            best_score = score
            # Keep terminal punctuation with the target so it cannot become
            # impossible leading suffix context (``target,-. Next``).  A
            # comma is deliberately excluded: Chromium can reject a comma
            # that visually follows text but lives across an inline-node
            # boundary (for example ``<i>Anns </i>test,``).
            terminal_chars = r",.!?;:…'\"’”»)\]" if quote_text.rstrip().endswith(",") else r".!?;:…'\"’”»)\]"
            trailing = re.match(rf"[{terminal_chars}]+", source_text[end:])
            if trailing:
                end += trailing.end()
            best = (_expand_editorial_initial_span(source_text, start), end)
    return best or (-1, -1)


def _expand_editorial_initial_span(source_text: str, start: int) -> int:
    if start <= 0:
        return start
    if re.match(r"\[[A-Za-z]\]", source_text[start - 1:start + 2] or ""):
        return start - 1
    return start


def _source_side_text_fragment_candidates(source_fragment: str) -> List[str]:
    source_fragment = re.sub(r"\s+", " ", str(source_fragment or "")).strip()
    if not source_fragment:
        return []

    candidates: List[str] = []
    marker_stripped = _strip_source_leading_marker(source_fragment)
    if marker_stripped and marker_stripped != source_fragment:
        candidates.append(marker_stripped)
    candidates.append(source_fragment)
    if re.match(r"^\[[A-Za-z]\]", source_fragment):
        word_spans = _text_fragment_word_spans(source_fragment)
        if len(word_spans) >= 5:
            start = _expand_editorial_initial_span(source_fragment, word_spans[0][1])
            end = word_spans[4][2]
            shortened = re.sub(r"\s+", " ", source_fragment[start:end]).strip()
            if shortened:
                candidates.append(shortened)
    deduped: List[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = candidate.lower()
        if candidate and key not in seen:
            deduped.append(candidate)
            seen.add(key)
    return deduped


def _text_from_context_words(
    source_text: str,
    word_spans: List[Tuple[str, int, int]],
    indexes: List[int],
    *,
    start_boundary: Optional[int] = None,
    end_boundary: Optional[int] = None,
) -> str:
    if not indexes:
        return ""
    start = word_spans[indexes[0]][1] if start_boundary is None else start_boundary
    end = word_spans[indexes[-1]][2] if end_boundary is None else end_boundary
    return _normalize_text_fragment_whitespace(source_text[start:end])


def _browser_safe_prefix_indexes(
    source_text: str,
    word_spans: List[Tuple[str, int, int]],
    indexes: List[int],
) -> List[int]:
    """Do not begin browser context in the tail of a decimal citation."""
    indexes = list(indexes)
    if not indexes:
        return indexes
    start = word_spans[indexes[0]][1]
    if (
        start >= 2
        and source_text[start - 1] == "."
        and source_text[start - 2].isdigit()
        and word_spans[indexes[0]][0][:1].isdigit()
    ):
        indexes.pop(0)
    return indexes


def _text_fragment_context(
    source_text: str,
    start: int,
    end: int,
    target_words: List[str],
) -> Tuple[str, str]:
    word_spans = _text_fragment_word_spans(source_text)
    before_indexes = [i for i, (_w, _s, e) in enumerate(word_spans) if e <= start][-4:]
    before_indexes = _browser_safe_prefix_indexes(source_text, word_spans, before_indexes)
    after_indexes = [i for i, (_w, s, _e) in enumerate(word_spans) if s >= end][:4]
    prefix = _text_from_context_words(source_text, word_spans, before_indexes, end_boundary=start)
    suffix = _text_from_context_words(source_text, word_spans, after_indexes, start_boundary=end)

    prefix_words = _quote_word_tokens(prefix)
    suffix_words = _quote_word_tokens(suffix)
    if target_words and prefix_words[-len(target_words):] == target_words:
        prefix = ""
    if target_words and _starts_with_words(suffix_words, target_words):
        suffix = ""
    return _strip_fragment_context_list_marker(prefix), suffix


_TEXT_FRAGMENT_CONTEXT_WORD_WINDOWS = (0, 2, 4, 8, 12, 16, 24, 32)
_TEXT_FRAGMENT_LONG_TARGET_WORDS = 30
_TEXT_FRAGMENT_LONG_TARGET_CHARS = 220
_TEXT_FRAGMENT_RANGE_BOUNDARY_WORDS = 5


def _trim_text_fragment_prefix_to_near_sentence(prefix: str) -> str:
    prefix = _normalize_text_fragment_whitespace(prefix)
    if not prefix:
        return ""
    matches = list(re.finditer(r"[.!?]\s+", prefix))
    if not matches:
        return prefix
    candidate = prefix[matches[-1].end():].strip()
    if len(_quote_word_tokens(candidate)) >= 2:
        return candidate
    return prefix


def _text_fragment_context_window(
    source_text: str,
    start: int,
    end: int,
    target_words: List[str],
    words_each_side: int,
) -> Tuple[str, str]:
    if words_each_side <= 0:
        return "", ""
    word_spans = _text_fragment_word_spans(source_text)
    before_indexes = [
        i for i, (_w, _s, e) in enumerate(word_spans) if e <= start
    ][-words_each_side:]
    before_indexes = _browser_safe_prefix_indexes(source_text, word_spans, before_indexes)
    after_indexes = [
        i for i, (_w, s, _e) in enumerate(word_spans) if s >= end
    ][:words_each_side]
    prefix = _text_from_context_words(
        source_text,
        word_spans,
        before_indexes,
        end_boundary=start,
    )
    suffix = _text_from_context_words(
        source_text,
        word_spans,
        after_indexes,
        start_boundary=end,
    )
    prefix_words = _quote_word_tokens(prefix)
    suffix_words = _quote_word_tokens(suffix)
    if target_words and prefix_words[-len(target_words):] == target_words:
        prefix = ""
    if target_words and _starts_with_words(suffix_words, target_words):
        suffix = ""
    prefix = _trim_text_fragment_prefix_to_near_sentence(prefix)
    return _strip_fragment_context_list_marker(prefix), suffix


def _text_fragment_range_targets(
    source_text: str,
    start: int,
    end: int,
    boundary_words: int = _TEXT_FRAGMENT_RANGE_BOUNDARY_WORDS,
) -> Tuple[str, str]:
    word_spans = _text_fragment_word_spans(source_text)
    inside = [
        i
        for i, (_w, s, e) in enumerate(word_spans)
        if s >= start and e <= end
    ]
    if len(inside) < boundary_words * 2:
        return "", ""
    count = min(boundary_words, max(3, len(inside) // 3))
    start_indexes = inside[:count]
    end_indexes = inside[-count:]
    if start_indexes[-1] >= end_indexes[0]:
        return "", ""
    start_target = _text_from_context_words(source_text, word_spans, start_indexes)
    end_boundary = word_spans[end_indexes[-1]][2]
    trailing = re.match(r"[.!?;:…'\"’”»)\]]+", source_text[end_boundary:end])
    if trailing:
        end_boundary += trailing.end()
    end_target = _text_from_context_words(
        source_text,
        word_spans,
        end_indexes,
        end_boundary=end_boundary,
    )
    if not start_target or not end_target:
        return "", ""
    if _quote_dedupe_key(start_target) == _quote_dedupe_key(end_target):
        return "", ""
    return start_target, end_target


def _strip_fragment_context_list_marker(text: str) -> str:
    """Avoid using list markers like '(d)' as fragment context."""
    marker = r"(?:\[\s*[A-Za-z0-9]{1,4}\s*\]|[A-Za-z0-9]{1,4}\s*\]|\(?[A-Za-z0-9]{1,4}\)?[\).])"
    return re.sub(rf"^{marker}\s+", "", text or "").strip()


def _is_leading_definition_term_span(source_text: str, start: int, after_words: List[str]) -> bool:
    if after_words[:1] not in (["means"], ["includes"]):
        return False
    leading = str(source_text or "")[:start].strip()
    leading = re.sub(r"^\(?[A-Za-z0-9]{1,4}\)?[\).]?\s+", "", leading).strip()
    return bool(re.fullmatch(r"[\[({\"'“‘]*", leading))






def _fragment_anchor_only(fragment: str) -> str:
    return (fragment or "").split(":~:", 1)[0]


def _url_has_page_anchor(url: str) -> bool:
    _base, frag = _split_url(url)
    anchor = _fragment_anchor_only(frag).strip()
    return bool(re.fullmatch(r"page=\d{1,5}", anchor, flags=re.IGNORECASE))


def _url_has_section_anchor(url: str) -> bool:
    _base, frag = _split_url(url)
    anchor = _fragment_anchor_only(frag).strip()
    return bool(_SEC_ANCHOR_RE.fullmatch(anchor))














def _text_fragment_directive(target: str, prefix: str = "", suffix: str = "") -> str:
    target = _normalize_text_fragment_whitespace(target)
    prefix = _normalize_text_fragment_whitespace(prefix)
    suffix = _normalize_text_fragment_whitespace(suffix)
    if not target:
        return ""
    encoded_target = _encode_text_fragment(target)
    if prefix and suffix:
        return f"text={_encode_text_fragment(prefix)}-,{encoded_target},-{_encode_text_fragment(suffix)}"
    if prefix:
        return f"text={_encode_text_fragment(prefix)}-,{encoded_target}"
    if suffix:
        return f"text={encoded_target},-{_encode_text_fragment(suffix)}"
    return f"text={encoded_target}"


def _text_fragment_range_directive(
    target_start: str,
    target_end: str,
    prefix: str = "",
    suffix: str = "",
) -> str:
    target_start = _normalize_text_fragment_whitespace(target_start)
    target_end = _normalize_text_fragment_whitespace(target_end)
    prefix = _normalize_text_fragment_whitespace(prefix)
    suffix = _normalize_text_fragment_whitespace(suffix)
    if not target_start or not target_end:
        return ""
    encoded_target = (
        f"{_encode_text_fragment(target_start)},{_encode_text_fragment(target_end)}"
    )
    if prefix and suffix:
        return f"text={_encode_text_fragment(prefix)}-,{encoded_target},-{_encode_text_fragment(suffix)}"
    if prefix:
        return f"text={_encode_text_fragment(prefix)}-,{encoded_target}"
    if suffix:
        return f"text={encoded_target},-{_encode_text_fragment(suffix)}"
    return f"text={encoded_target}"


def _append_text_fragment_directives(url: str, directives: List[str]) -> str:
    directives = [d for d in directives if d]
    if not directives or not url or url.lower() == "other":
        return url
    url = _scc_scrollable_text_url(_canlii_source_lookup_url(url))
    separator = "&" if ":~:" in url else ":~:"
    directive_text = "&".join(directives)
    if "#" in url:
        return f"{url}{separator}{directive_text}"
    return f"{url}#:~:{directive_text}"


@dataclass
class TextFragmentBuildResult:
    url: str
    fragment_count: int
    builder: str
    verified: bool
    reason: str = ""


@dataclass
class _BuiltTextDirective:
    directive: str
    start: int
    end: int
    target: str
    prefix: str
    suffix: str


@dataclass
class _ParsedTextFragmentDirective:
    prefix: str
    target_start: str
    target_end: str
    suffix: str


def _text_fragment_anchor_url(url: str) -> str:
    base, frag = _split_url(_sanitize_url_candidate(url))
    if not base:
        return url
    anchor = _fragment_anchor_only(frag)
    return _recombine_url(base, anchor)


def _text_fragment_directives_from_url(url: str) -> List[str]:
    _base, frag = _split_url(url)
    if ":~:" not in frag:
        return []
    directive_text = frag.split(":~:", 1)[1]
    directives: List[str] = []
    for raw in directive_text.split("&"):
        if raw.startswith("text="):
            directives.append(raw)
    return directives


def _parse_text_fragment_directive(directive: str) -> Optional[_ParsedTextFragmentDirective]:
    if not directive.startswith("text="):
        return None
    value = directive[len("text="):]
    prefix = suffix = ""
    if "-," in value:
        prefix, value = value.split("-,", 1)
    if ",-" in value:
        value, suffix = value.rsplit(",-", 1)

    target_start = value
    target_end = ""
    if "," in value:
        target_start, target_end = value.split(",", 1)

    target_start = unquote(target_start).strip()
    target_end = unquote(target_end).strip()
    prefix = unquote(prefix).strip()
    suffix = unquote(suffix).strip()
    if not target_start:
        return None
    return _ParsedTextFragmentDirective(
        prefix=prefix,
        target_start=target_start,
        target_end=target_end,
        suffix=suffix,
    )


def _word_indexes_around_span(
    source_text: str,
    start: int,
    end: int,
) -> Tuple[List[str], List[str]]:
    words = _text_fragment_word_spans(source_text)
    before = [w.lower() for w, _s, e in words if e <= start]
    after = [w.lower() for w, s, _e in words if s >= end]
    return before, after


def _prefix_suffix_match(
    source_text: str,
    start: int,
    end: int,
    parsed: _ParsedTextFragmentDirective,
) -> bool:
    before_words, after_words = _word_indexes_around_span(source_text, start, end)
    prefix_words = [w.lower() for w in _quote_word_tokens(parsed.prefix)]
    suffix_words = [w.lower() for w in _quote_word_tokens(parsed.suffix)]
    if prefix_words and before_words[-len(prefix_words):] != prefix_words:
        return False
    if suffix_words and after_words[:len(suffix_words)] != suffix_words:
        return False
    return True


def _text_fragment_directive_matches(
    source_text: str,
    parsed: _ParsedTextFragmentDirective,
) -> List[Tuple[int, int]]:
    if not source_text:
        return []
    start_words = _quote_word_tokens(parsed.target_start)
    if not start_words:
        return []

    matches: List[Tuple[int, int]] = []
    start_spans = _text_fragment_phrase_spans(source_text, start_words)
    for start, end in start_spans:
        if parsed.target_end:
            end_words = _quote_word_tokens(parsed.target_end)
            if not end_words:
                continue
            end_candidates = [
                (s2, e2)
                for s2, e2 in _text_fragment_phrase_spans(source_text[end:], end_words)
            ]
            if not end_candidates:
                continue
            _rel_start, rel_end = end_candidates[0]
            end = end + rel_end
        if _prefix_suffix_match(source_text, start, end, parsed):
            matches.append((start, end))
    return matches


def _source_fragment_parts_for_text_fragment(source_fragment: str) -> List[str]:
    text = str(source_fragment or "").strip()
    if not text:
        return []
    parts = [re.sub(r"\s+", " ", part).strip() for part in re.split(r"\n\s*\n+", text)]
    return [part for part in parts if part and not _is_only_bracketed_quote_token(part)]


def _source_fragment_reference_spans(
    source_text: str,
    source_fragment: str,
) -> List[Tuple[int, int]]:
    spans: List[Tuple[int, int]] = []
    for part in _source_fragment_parts_for_text_fragment(source_fragment):
        words = _quote_word_tokens(part)
        if words:
            spans.extend(_text_fragment_phrase_spans(source_text, words))
    return spans


def _spans_overlap(a: Tuple[int, int], b: Tuple[int, int]) -> bool:
    return max(a[0], b[0]) < min(a[1], b[1])


def _verified_directive_in_text(
    directive: str,
    source_text: str,
    source_fragment: str,
) -> bool:
    parsed = _parse_text_fragment_directive(directive)
    if parsed is None:
        return False

    matches = _text_fragment_directive_matches(source_text, parsed)
    if len(matches) != 1:
        return False

    source_spans = _source_fragment_reference_spans(source_text, source_fragment)
    if source_spans:
        return any(_spans_overlap(matches[0], span) for span in source_spans)

    target_text = parsed.target_start
    if parsed.target_end:
        target_text = f"{parsed.target_start} {parsed.target_end}"
    return _quote_match_score(target_text, source_fragment) >= 0.75


def _text_fragment_verification_texts(url: str, source_context: str) -> List[str]:
    texts: List[str] = []

    context = re.sub(r"\s+", " ", str(source_context or "")).strip()
    if context and not texts:
        texts.append(context)

    deduped: List[str] = []
    seen: set[str] = set()
    for text in texts:
        key = text.lower()
        if text and key not in seen:
            deduped.append(text)
            seen.add(key)
    return deduped


def _verify_text_fragment_url(
    url: str,
    source_context: str,
    source_fragment: str,
) -> bool:
    directives = _text_fragment_directives_from_url(url)
    if not directives:
        return False
    texts = _text_fragment_verification_texts(url, source_context)
    if not texts:
        return False

    for directive in directives:
        if not any(_verified_directive_in_text(directive, text, source_fragment) for text in texts):
            return False
    return True


def _directive_verified_for_texts(
    directive: str,
    source_context: str,
    source_fragment: str,
    document_text: str = "",
    *,
    require_document_unique: bool = False,
) -> bool:
    if not directive:
        return False
    if require_document_unique and document_text and "\n" in document_text:
        # Full-source build (registered text with line structure):
        # the line-scoped document check subsumes the context check — it
        # enforces global uniqueness, paragraph scoping, and overlap with the
        # source fragment. The context window is only a scaffold here, and
        # overlapping match regions can legitimately repeat the target
        # inside it.
        return _verified_directive_in_document_lines(
            directive, document_text, source_fragment
        )
    if not _verified_directive_in_text(
        directive,
        source_context,
        source_fragment,
    ):
        return False
    if require_document_unique:
        if not document_text:
            return False
        return _verified_directive_in_text(directive, document_text, source_fragment)
    return True


def _verified_directive_in_document_lines(
    directive: str,
    document_text: str,
    source_fragment: str,
) -> bool:
    """Document-level verification with paragraph scoping.

    Source texts that keep their line structure (A2AJ full texts) use one
    line per paragraph/block, and a browser text fragment cannot match
    across block boundaries — so the directive must match exactly once in
    the whole document with every component (prefix, target, suffix)
    contained in a single line, and that match must sit on the same line
    as the source fragment it points at."""
    parsed = _parse_text_fragment_directive(directive)
    if parsed is None:
        return False
    scan_words = _quote_word_tokens(parsed.prefix or parsed.target_start)
    if not scan_words:
        return False
    scan_word = scan_words[0]
    matched_line = ""
    matched_span: Optional[Tuple[int, int]] = None
    for line in document_text.splitlines():
        if len(line) < 8 or scan_word not in line.lower():
            continue
        for span in _text_fragment_directive_matches(line, parsed):
            if matched_span is not None:
                return False
            matched_line, matched_span = line, span
    if matched_span is None:
        return False
    source_spans = _source_fragment_reference_spans(matched_line, source_fragment)
    if source_spans:
        return any(_spans_overlap(matched_span, span) for span in source_spans)
    target_text = parsed.target_start
    if parsed.target_end:
        target_text = f"{parsed.target_start} {parsed.target_end}"
    return _quote_match_score(target_text, source_fragment) >= 0.75


# ---------------------------------------------------------------------------
# Full source text fetched for verification is registered here by link base
# so the export-time builder can enforce document-level uniqueness and
# paragraph scoping.
_FRAGMENT_DOC_TEXT_CACHE: Dict[str, str] = {}


def _fragment_doc_key(url: str) -> str:
    base, _frag = _split_url(_canlii_source_lookup_url(_sanitize_url_candidate(url)))
    return (base or "").strip().lower()


def _register_fragment_document_text(url: str, source_text: str) -> None:
    key = _fragment_doc_key(url)
    if key and (source_text or "").strip():
        _FRAGMENT_DOC_TEXT_CACHE.setdefault(key, source_text)


def _fragment_document_text_for_url(url: str) -> str:
    return _FRAGMENT_DOC_TEXT_CACHE.get(_fragment_doc_key(url), "")


def _a2aj_registered_paragraph_text(url: str, pinpoint: str) -> str:
    """Text of the ``parN`` paragraph from the registered A2AJ full document
    text for ``url``'s base, or "". Full-text availability is what lets a
    fragment directive gather disambiguating context and be verified unique
    against the whole document."""
    par_match = re.match(
        r"^par(\d+)$", str(pinpoint or "").strip(), flags=re.IGNORECASE
    )
    if not par_match:
        return ""
    document_text = _fragment_document_text_for_url(url)
    if not document_text:
        return ""
    number = int(par_match.group(1))
    for item in _a2aj_paragraph_structure(document_text):
        num, start, end = item[:3]
        if num == number:
            paragraph = item[3] if len(item) > 3 else document_text[start:end]
            return re.sub(r"\s+", " ", paragraph).strip()
    return ""


_A2AJ_MD_HEADING_RE = re.compile(r"^\s*#{1,6}\s+", re.MULTILINE)
_A2AJ_MD_EMPHASIS_RE = re.compile(r"[*`]+")
_A2AJ_DOT_LEADER_RE = re.compile(r"\.{4,}")  # TOC leaders; ellipses are 3 dots
_A2AJ_SPACE_BEFORE_PUNCT_RE = re.compile(r"[ \t]+([,.;:!?%)\]”’])")
_A2AJ_SPACE_AFTER_OPEN_RE = re.compile(r"([(\[“‘])[ \t]+")
_A2AJ_NORMALIZED_CASE_CACHE_DIR = os.path.join("cache", "a2aj_normalized_cases")
_A2AJ_NORMALIZATION_VERSION = "v1"


@lru_cache(maxsize=16)
def _normalize_a2aj_source_text(text: str) -> str:
    """Normalize A2AJ text toward what a browser renders on the target page:
    strip the markdown markup A2AJ uses for laws texts, and the extraction
    artifacts (spaces hugging punctuation at inline-markup boundaries) that
    do not exist in rendered text. Newlines are preserved — A2AJ lines are
    paragraphs, and a text fragment cannot match across paragraphs."""
    text = str(text or "")
    if not text:
        return text
    text = _A2AJ_MD_HEADING_RE.sub("", text)
    text = _A2AJ_MD_EMPHASIS_RE.sub("", text)
    text = _A2AJ_DOT_LEADER_RE.sub(" ", text)
    text = _A2AJ_SPACE_BEFORE_PUNCT_RE.sub(r"\1", text)
    text = _A2AJ_SPACE_AFTER_OPEN_RE.sub(r"\1", text)
    return text


def _normalized_a2aj_document_text(
    document: a2aj_client.A2AJDocument, source_kind: str
) -> str:
    """Reuse normalized case text across runs; changing laws stay memory-only."""
    raw = str(document.text or "")
    if source_kind != "case" or not raw:
        return _normalize_a2aj_source_text(raw)
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    path = os.path.join(
        _A2AJ_NORMALIZED_CASE_CACHE_DIR,
        f"{_A2AJ_NORMALIZATION_VERSION}_{digest}.txt",
    )
    try:
        with open(path, "r", encoding="utf-8") as handle:
            return handle.read()
    except OSError:
        pass
    normalized = _normalize_a2aj_source_text(raw)
    try:
        os.makedirs(_A2AJ_NORMALIZED_CASE_CACHE_DIR, exist_ok=True)
        temporary = f"{path}.{os.getpid()}.{threading.get_ident()}.tmp"
        with open(temporary, "w", encoding="utf-8", newline="") as handle:
            handle.write(normalized)
        os.replace(temporary, path)
    except OSError:
        pass
    return normalized


def _a2aj_query_citation(bare: str) -> str:
    """Trim a bare citation to what A2AJ's /fetch matcher accepts: it rejects
    citations carrying pinpoints, so drop the pinpoint (and anything after
    it), statute section lists, and history parentheticals like (rev'd ...).
    Structural parentheticals such as (2nd Supp) are part of the citation
    and must stay."""
    text = re.sub(r"\s+", " ", str(bare or "")).strip()
    text = re.sub(
        r"\s+at\s+(?:(?:p{1,2}|paras?|(?:sub)?sections?|ss?|rules?|rr?|"
        r"arts?|articles?)\.?\s*)?\d.*$",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r",\s*(?:(?:sub)?sections?|ss?\.?|rules?|rr?\.?|arts?\.?|articles?)"
        r"\s+\d[\w().,\s–-]*$",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\s*\((?:rev|aff|var|leave|appeal|overrul)[^()]*\)",
        "",
        text,
        flags=re.IGNORECASE,
    )
    return text.strip().rstrip(",;.")


_CANLII_DOC_SLUG_RE = re.compile(r"/doc/\d{4}/((\d{4})([a-z]+)(\d+))/", re.IGNORECASE)


def _canlii_doc_citation_from_url(url: str) -> str:
    """Derive the neutral (or CanLII) citation from a canlii.org decision URL
    slug: .../doc/2023/2023scc17/... -> "2023 SCC 17"."""
    candidate = _canlii_source_lookup_url(_sanitize_url_candidate(url or ""))
    if "canlii.org" not in candidate.lower():
        return ""
    m = _CANLII_DOC_SLUG_RE.search(candidate)
    if not m:
        return ""
    _slug, year, court, num = m.groups()
    court = "CanLII" if court.lower() == "canlii" else court.upper()
    return f"{year} {court} {num}"


def _fetch_a2aj_source_text_for_row(
    row: Dict[str, Any],
    bare: str = "",
    kind: str = "",
    register_url: Optional[str] = None,
) -> str:
    """Fetch exact A2AJ text and retain its evidence on the row.

    URL-keyed fragment state is populated only when the supplied CanLII URL
    independently identifies the same document.
    """
    raw_bare = re.sub(
        r"\s+", " ", str(bare or row.get("bare_citation") or "")
    ).strip()
    bare = _a2aj_query_citation(raw_bare)
    kind = (kind or row.get("citation_part_kind") or "").strip()
    if not bare or kind not in ("case", "unreported", "statute", "gazette"):
        return ""
    _pause_gate()
    if register_url is None:
        register_url = (row.get("citation_part_link") or "").strip()
    language = "fr" if "/fr/" in (register_url or "").lower() else "en"
    query = bare
    if kind in ("statute", "gazette") and re.search(
        r",\s*(?:r(?:ule)?s?\.?)\s*\d+(?:\.\d+)*\s*$", raw_bare, flags=re.I
    ):
        # Some A2AJ corpora model one rule as one document (for example
        # ``NB Reg 82-73, r 11``).  Preserve that canonical suffix for the
        # exact lookup; ordinary regulation pinpoints still fall back to the
        # stripped parent citation below.
        query = raw_bare.rstrip(".;")
    lookup = a2aj_client.lookup_document(query, kind, language=language)
    if lookup.status == "not_found" and query != bare:
        lookup = a2aj_client.lookup_document(bare, kind, language=language)
    if (
        (lookup.status != "found" or lookup.document is None)
        and kind in ("case", "unreported")
    ):
        canonical = _adapter_identity_citation(raw_bare)
        if canonical:
            lookup = a2aj_client.lookup_document(canonical, kind, language=language)
    if lookup.status != "found" or lookup.document is None:
        return ""
    document = lookup.document
    source_kind = "law" if kind in ("statute", "gazette") else "case"
    text, structure = _a2aj_document_evidence(document, source_kind)
    if text:
        row["_a2aj_identity_locked"] = True
        row["_a2aj_structure"] = structure
        row["_a2aj_structure_status"] = structure.get("status", "unavailable")
        row["_a2aj_structure_type"] = structure.get("type", "")
        row["_a2aj_structure_count"] = structure.get("count", 0)
        row["_a2aj_dataset"] = document.dataset
        row["_a2aj_citation"] = document.citation
        row["_a2aj_source_url"] = document.url
        row["_a2aj_language"] = document.language
        base, _fragment = _split_url(_canlii_source_lookup_url(register_url or ""))
        reconciled = bool(
            base and _a2aj_url_matches_document(base, document, source_kind)
        )
        row["_a2aj_url_reconciled"] = reconciled
        if reconciled:
            _register_a2aj_document(
                base,
                document,
                source_kind,
                structure=structure,
                evidence_text=text,
            )
            _register_fragment_document_text(register_url or "", text)
            _register_fragment_document_text(document.url, text)
    return text


def _build_source_side_text_fragment_directive(
    source_fragment: str,
    source_context: str,
    url: str = "",
    *,
    force_context: bool = False,
    document_text: str = "",
    require_document_unique: bool = False,
    prefer_range: bool = False,
) -> Optional[_BuiltTextDirective]:
    source_fragment = re.sub(r"\s+", " ", str(source_fragment or "")).strip()
    if not source_fragment or _is_only_bracketed_quote_token(source_fragment):
        return None

    context = str(source_context or "").strip()
    if not context:
        return None

    for candidate in _source_side_text_fragment_candidates(source_fragment):
        start, end = _choose_text_fragment_span(context, candidate)
        if start < 0 or end <= start:
            continue

        target = _normalize_text_fragment_whitespace(context[start:end])
        target_words = _quote_word_tokens(target)
        if not target_words:
            continue
        starts_with_editorial_initial = bool(re.match(r"^\[[A-Za-z]\]", source_fragment))
        repeated = len(_text_fragment_phrase_spans(context, target_words)) > 1
        context_needed = (
            force_context
            or len(target_words) <= 3
            or (
                _url_has_page_anchor(url)
                and len(target_words) <= 8
                and not starts_with_editorial_initial
            )
            or repeated
        )
        exact_document_repeats = (
            bool(document_text)
            and len(_text_fragment_phrase_spans(document_text, target_words)) > 1
        )

        modes = ["exact"]
        if prefer_range:
            if (
                not starts_with_editorial_initial
                and not _url_has_section_anchor(url)
                and not repeated
                and not exact_document_repeats
            ):
                modes = ["range", "exact"]
            else:
                modes = ["exact", "range"]

        if context_needed:
            if _url_has_page_anchor(url):
                windows = [w for w in _TEXT_FRAGMENT_CONTEXT_WORD_WINDOWS if w > 0]
            else:
                windows = [4, 2, 8, 12, 16, 24, 32]
        elif require_document_unique or exact_document_repeats:
            windows = list(_TEXT_FRAGMENT_CONTEXT_WORD_WINDOWS)
        else:
            windows = [0]

        for mode in modes:
            range_start = range_end = ""
            if mode == "range":
                range_start, range_end = _text_fragment_range_targets(context, start, end)
                if not range_start or not range_end:
                    continue
                directive_words = _quote_word_tokens(f"{range_start} {range_end}")
            else:
                directive_words = target_words

            for window in windows:
                prefix = suffix = ""
                if window:
                    prefix, suffix = _text_fragment_context_window(
                        context,
                        start,
                        end,
                        directive_words,
                        window,
                    )
                    prefix = _strip_fragment_context_list_marker(prefix)

                context_options = [(prefix, suffix)]
                if window:
                    context_options = [(prefix, ""), ("", suffix), (prefix, suffix)]
                    if _quote_word_tokens(suffix)[:1] in (["means"], ["includes"]):
                        context_options = [("", suffix), (prefix, suffix), (prefix, "")]
                    if force_context:
                        context_options = [
                            option for option in context_options if option[0] or option[1]
                        ]

                for prefix_candidate, suffix_candidate in context_options:
                    if mode == "range":
                        directive = _text_fragment_range_directive(
                            range_start,
                            range_end,
                            prefix_candidate,
                            suffix_candidate,
                        )
                        built_target = f"{range_start} {range_end}"
                    else:
                        directive = _text_fragment_directive(
                            target,
                            prefix_candidate,
                            suffix_candidate,
                        )
                        built_target = target

                    if _directive_verified_for_texts(
                        directive,
                        context,
                        source_fragment,
                        document_text,
                        require_document_unique=require_document_unique,
                    ):
                        return _BuiltTextDirective(
                            directive=directive,
                            start=start,
                            end=end,
                            target=built_target,
                            prefix=prefix_candidate,
                            suffix=suffix_candidate,
                        )
    return None










def _build_quote_check_fragment_url(
    url: str,
    source_fragment: str,
    source_context: str = "",
    corrected_text: str = "",
    *,
    prefer_range: bool = False,
) -> TextFragmentBuildResult:
    base_url = _text_fragment_anchor_url(url)
    parts = _source_fragment_parts_for_text_fragment(source_fragment)
    if not base_url or not parts:
        return TextFragmentBuildResult(base_url or url, 0, "", False, "no_source_fragment")

    # Multi-directive text fragments can be fragile in browsers when the source
    # context is loose. Keep construction centralized here so source scoping and
    # verification apply consistently to every directive.
    deduped_parts: List[str] = []
    seen_part_keys: set[str] = set()
    for part in parts:
        key = _quote_dedupe_key(part)
        if key and key in seen_part_keys:
            continue
        if key:
            seen_part_keys.add(key)
        deduped_parts.append(part)
    parts = deduped_parts

    source_context = re.sub(r"\s+", " ", str(source_context or "")).strip()
    fragment_context = re.sub(r"\s+", " ", str(source_fragment or "")).strip()

    build_url = base_url
    # If the full source text used for quote verification was registered,
    # require every directive to be unique in it with paragraph scoping;
    # unresolvable ambiguity falls back to the anchor-only link rather than
    # risk sending the browser to the wrong copy (e.g. a headnote duplicate).
    registered_document = _fragment_document_text_for_url(build_url)
    directives: List[str] = []
    builders: List[str] = []
    seen: set[str] = set()

    for part in parts:
        built: Optional[_BuiltTextDirective] = None
        builder = ""
        if source_context:
            built = _build_source_side_text_fragment_directive(
                part,
                source_context,
                build_url,
                document_text=registered_document,
                require_document_unique=bool(registered_document),
                prefer_range=prefer_range,
            )
            if built:
                builder = "source_document" if registered_document else "source_window"

        if not built:
            reason = "unmatched_multi_part" if len(parts) > 1 else "unmatched_source_fragment"
            return TextFragmentBuildResult(base_url, 0, "", False, reason)

        if built.directive not in seen:
            directives.append(built.directive)
            seen.add(built.directive)
        builders.append(builder)

    if not directives:
        return TextFragmentBuildResult(base_url, 0, "", False, "unverified")

    result_url = _append_text_fragment_directives(build_url, directives)
    if registered_document:
        if not all(
            _verified_directive_in_document_lines(d, registered_document, source_fragment)
            for d in directives
        ):
            return TextFragmentBuildResult(base_url, 0, "", False, "verification_failed")
    elif not _verify_text_fragment_url(result_url, source_context or fragment_context, source_fragment):
        return TextFragmentBuildResult(base_url, 0, "", False, "verification_failed")

    return TextFragmentBuildResult(
        url=result_url,
        fragment_count=len(directives),
        builder="+".join(dict.fromkeys(builders)),
        verified=True,
        reason="",
    )


def _build_targeted_text_fragment_url(url: str, source_text: str, corrected: str) -> Tuple[str, int]:
    source = str(source_text or "")
    directives_with_order: List[Tuple[int, str]] = []
    seen: set[str] = set()
    # Multi-directive text fragments can be fragile; this targeted path only
    # emits directives that are found in the supplied source text.
    for candidate in _iter_text_fragment_quote_candidates(corrected):
        words = _quote_word_tokens(candidate)
        if not words:
            continue
        start, end = _choose_text_fragment_span(source, candidate)
        if start < 0 or end <= start:
            continue
        target = _normalize_text_fragment_whitespace(source[start:end])
        target_words = _quote_word_tokens(target)
        repeated = len(_text_fragment_phrase_spans(source, target_words)) > 1
        prefix = suffix = ""
        if len(target_words) <= 3 or repeated:
            prefix, suffix = _text_fragment_context(source, start, end, target_words)
        directive = _text_fragment_directive(target, prefix, suffix)
        if directive and directive not in seen:
            directives_with_order.append((start, directive))
            seen.add(directive)

    if not directives_with_order:
        return (url, 0)

    directives = [d for _start, d in sorted(directives_with_order, key=lambda item: item[0])]
    return (_append_text_fragment_directives(url, directives), len(directives))


def _build_quote_only_text_fragment_url(url: str, corrected: str, source_text: str = "") -> str:
    candidate_url = _canlii_source_lookup_url(_sanitize_url_candidate(url))
    if not candidate_url or candidate_url.lower() == "other":
        return ""

    if source_text:
        targeted, targeted_count = _build_targeted_text_fragment_url(candidate_url, source_text, corrected)
        if targeted_count:
            return targeted

    directives: List[str] = []
    seen: set[str] = set()
    # Last-resort quote-only construction is less reliable than source-scoped
    # construction, but keep multi-directive support available for legacy paths.
    for candidate in _iter_text_fragment_quote_candidates(corrected):
        directive = _text_fragment_directive(candidate)
        if directive and directive not in seen:
            directives.append(directive)
            seen.add(directive)
    if not directives:
        return ""
    return _append_text_fragment_directives(candidate_url, directives)


def _find_text_fragment_snippet(source_text: str, corrected: str) -> str:
    """Find the matched quote span in source_text for a text-fragment URL."""
    src_text = str(source_text or "")
    corr_text = str(corrected or "")
    if not src_text or not corr_text:
        return ""
    start_w, end_w = _extract_fragment_boundaries(corr_text)
    start_re = _phrase_regex_for_fragment(start_w)
    end_re = _phrase_regex_for_fragment(end_w)
    if start_re and end_re:
        start_m = start_re.search(src_text)
        if start_m:
            end_m = end_re.search(src_text, start_m.end())
            if end_m:
                return src_text[start_m.start():end_m.end()]

    exact_corr = re.sub(
        r"\s+",
        " ",
        corr_text.strip().strip("\"'").strip("\u201c\u201d"),
    ).strip()
    if not exact_corr:
        return ""
    words = re.findall(r"[A-Za-z0-9]+(?:['\u2019][A-Za-z0-9]+)*", exact_corr)
    if not words:
        return ""
    exact_re = re.compile(r"\b" + r"[\s\u00a0]+".join(re.escape(w) for w in words) + r"\b", re.IGNORECASE)
    m = exact_re.search(src_text)
    return src_text[m.start():m.end()] if m else ""


def _build_fragment_url(url: str, snippet: str) -> str:
    """Append a W3C text fragment (:~:text=...) encoding a snippet of source text.
    Handles URLs that already have a fragment (#par42).
    """
    if not snippet or not url or url.lower() == "other":
        return url
    url = _canlii_source_lookup_url(url)
    encoded = _encode_text_fragment(snippet)
    suffix = f":~:text={encoded}"
    if "#" in url:
        return f"{url}{suffix}"
    return f"{url}#{suffix}"


def _sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)
TRAILING_URL_PUNCT = ".,;:!?)]}>\"'“”’‘"


def _canlii_pdf_to_html_sibling(url: str) -> str:
    """Rewrite CanLII .pdf links to their .html sibling: par/sec anchors only
    exist on the HTML page, and the bare PDF is the scanned original of a
    document whose item page is the canonical link. #page=N is the one
    exception — a PDF page pinpoint is functional in the browser's viewer."""
    base, frag = _split_url(url)
    if not base or "canlii.org" not in base.lower():
        return url
    if not urlsplit(base).path.lower().endswith(".pdf"):
        return url
    if re.match(r"^page=-?\d+$", (frag or "").strip(), flags=re.IGNORECASE):
        return url
    return _recombine_url(re.sub(r"\.pdf$", ".html", base, flags=re.IGNORECASE), frag)


def _sanitize_url_candidate(raw: str) -> str:
    """
    Extract a URL from a mixed string and strip trailing sentence punctuation.
    """
    s = (raw or "").strip()
    if not s:
        return ""
    m = URL_RE.search(s)
    if m:
        s = m.group(0)
    else:
        s = s.split()[0]
    return _canlii_pdf_to_html_sibling(s.rstrip(TRAILING_URL_PUNCT))


def _ts_print(message: str = "") -> None:
    stamp = time.strftime("%Y-%m-%d %H:%M:%S")
    if message:
        print(f"{stamp} {message}", flush=True)
    else:
        print(stamp, flush=True)


TIMING_LOG_PATH = ""
TIMING_RUN_ID = f"{int(time.time())}-{os.getpid()}"


def _set_timing_log_path(path: str) -> None:
    global TIMING_LOG_PATH
    TIMING_LOG_PATH = (path or "").strip()
    if TIMING_LOG_PATH:
        log_path = Path(TIMING_LOG_PATH).expanduser()
        if not log_path.is_absolute():
            log_path = (Path.cwd() / log_path).resolve()
        log_path.parent.mkdir(parents=True, exist_ok=True)
        TIMING_LOG_PATH = str(log_path)


def _timing_event(event: str, **fields: Any) -> None:
    if not TIMING_LOG_PATH:
        return
    payload = {
        "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
        "run_id": TIMING_RUN_ID,
        "event": event,
    }
    payload.update(fields)
    with open(TIMING_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")


class _TimingSpan:
    def __init__(self, event: str, **fields: Any) -> None:
        self.event = event
        self.fields = fields
        self.started = 0.0

    def __enter__(self) -> "_TimingSpan":
        self.started = time.perf_counter()
        _timing_event(f"{self.event}:start", **self.fields)
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        elapsed = time.perf_counter() - self.started
        fields = dict(self.fields)
        fields["elapsed_s"] = round(elapsed, 3)
        fields["ok"] = exc_type is None
        if exc is not None:
            fields["error"] = repr(exc)
        _timing_event(f"{self.event}:end", **fields)


def _timing_span(event: str, **fields: Any) -> _TimingSpan:
    return _TimingSpan(event, **fields)


# When this file runs as a script, register it under its module name too so
# other modules bind to this instance instead of executing it twice.
sys.modules.setdefault("alr_quote_verifier", sys.modules[__name__])

# URL resolution uses deterministic links, A2AJ source text, and the
# configured US/UK case URL providers.
from verifier_core.protocols import UrlResolver  # noqa: E402
from verifier_core import registry as _provider_registry  # noqa: E402
from verifier_core import a2aj_structure as _a2aj_structure  # noqa: E402
from verifier_core import a2aj_pinpoint_scope as _a2aj_pinpoint_scope  # noqa: E402

def _parse_footnote_ids(raw: Optional[str]) -> Optional[set[int]]:
    if not raw or not raw.strip():
        return None
    ids: set[int] = set()
    raw_clean = raw.strip()
    for part in raw_clean.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            try:
                a, b = part.split("-", 1)
                start, end = int(a.strip()), int(b.strip())
                if start <= end:
                    ids.update(range(start, end + 1))
                else:
                    ids.update(range(end, start + 1))
            except Exception:
                pass
        else:
            try:
                ids.add(int(part))
            except Exception:
                pass
    if ids:
        _ts_print(f"  FN filter parsed '{raw_clean}' -> sorted IDs: {sorted(ids)}")
    return ids if ids else None


def _normalize_selection_key(value: Any) -> str:
    return str(value or "").strip().replace("\\", "/").lower()


def _selection_keys_for_doc(docx_path: Path, input_root: Path) -> set[str]:
    keys = {
        _normalize_selection_key(docx_path.name),
        _normalize_selection_key(docx_path.stem),
    }
    try:
        keys.add(_normalize_selection_key(docx_path.relative_to(input_root)))
    except Exception:
        pass
    try:
        keys.add(_normalize_selection_key(docx_path.resolve()))
    except Exception:
        pass
    return {k for k in keys if k}


def _doc_key_for_manifest(docx_path: Path, input_root: Path) -> str:
    try:
        return docx_path.relative_to(input_root).as_posix()
    except Exception:
        return docx_path.name


def _parse_footnote_id_cell(value: Any) -> set[int]:
    raw = str(value or "").strip()
    if not raw:
        return set()
    ids: set[int] = set()
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            try:
                a, b = part.split("-", 1)
                start, end = int(a.strip()), int(b.strip())
                if start <= end:
                    ids.update(range(start, end + 1))
                else:
                    ids.update(range(end, start + 1))
            except Exception:
                continue
        else:
            try:
                ids.add(int(part))
            except Exception:
                continue
    return ids


def _load_footnote_selection_file(path: str | Path) -> Dict[str, set[int]]:
    selection_path = Path(path)
    selections: Dict[str, set[int]] = {}
    with selection_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return selections
        fields = {name.strip().lower(): name for name in reader.fieldnames if name}
        id_field = (
            fields.get("footnote_id")
            or fields.get("footnote_ids")
            or fields.get("footnote_display_id")
            or fields.get("display_footnote_id")
            or fields.get("id")
        )
        if not id_field:
            raise ValueError(f"Footnote selection file has no footnote_id column: {selection_path}")
        for row in reader:
            ids = _parse_footnote_id_cell(row.get(id_field))
            if not ids:
                continue
            keys: set[str] = set()
            doc_key_field = fields.get("doc_key")
            doc_key = row.get(doc_key_field) if doc_key_field else None
            if doc_key:
                keys.add(_normalize_selection_key(doc_key))
            else:
                for col in ("doc_name", "doc_path", "document", "file", "filename"):
                    field = fields.get(col)
                    raw = row.get(field) if field else None
                    if not raw:
                        continue
                    keys.add(_normalize_selection_key(raw))
                    try:
                        p = Path(str(raw))
                        keys.add(_normalize_selection_key(p.name))
                        keys.add(_normalize_selection_key(p.stem))
                    except Exception:
                        pass
            keys = {k for k in keys if k}
            if not keys:
                raise ValueError(f"Selection row is missing a document key/name/path: {row}")
            for key in keys:
                selections.setdefault(key, set()).update(ids)
    return selections


def _selection_ids_for_doc(
    footnote_selection: Optional[Dict[str, set[int]]],
    docx_path: Path,
    input_root: Path,
) -> Optional[set[int]]:
    if not footnote_selection:
        return None
    selected: set[int] = set()
    for key in _selection_keys_for_doc(docx_path, input_root):
        selected.update(footnote_selection.get(key, set()))
    return selected


def _merge_selection_maps(*maps: Optional[Dict[str, set[int]]]) -> Dict[str, set[int]]:
    merged: Dict[str, set[int]] = {}
    for mapping in maps:
        if not mapping:
            continue
        for key, ids in mapping.items():
            merged.setdefault(key, set()).update(ids)
    return merged


# Deterministic URLs are retained while A2AJ supplies source text.
LINK_RESOLVER = UrlResolver()

# Court/tribunal slug table: shared, provider-agnostic knowledge; lives in
# verifier_core so every build constructs canlii.org URLs deterministically.
from verifier_core.canlii_urls import COURT_MAP as _COURT_MAP  # noqa: E402

SUPRA_MODE: str = "aggressive"
USE_DB_SEARCH: bool = True
USE_A2AJ: bool = True
LOCAL_ONLY: bool = False
SEARCH_ALT_PINPOINTS: bool = True
TEXT_FRAGMENT_MODE: str = "off"  # "all", "pinpointless", "off"
EXPORT_DETAIL_MODE: str = "diagnostic-hidden"  # "display", "display-json", "diagnostic-hidden", or "diagnostic"
LLM_CACHE_ENABLED: bool = True

# --- Fallback CanLII URL generation from neutral citations ---

_NEUTRAL_RE = re.compile(r'(\d{4})\s+(' + '|'.join(sorted(_COURT_MAP.keys(), key=lambda x: -len(x))) + r')\s+(\d+)', re.IGNORECASE)
_CANLII_RE = re.compile(
    r'(\d{4})\s+CanLII\s+(\d+)(?:\s+at\s+[^()]*)?\s*\(([A-Za-z ]+)\)', re.IGNORECASE
)
_PARA_RE = re.compile(
    r'(?:at\s+)?(?:para(?:graph)?s?\.?|¶)\s*(\d+)', re.IGNORECASE
)
_FRENCH_PREFIXES = {"qc", "nb"}


# Statute citations with dotted or French-style abbreviations ("R.C.S. 1985
# c. C-46", "L.R.O. 1990") don't parse in the legislation DB's candidate
# builder; normalize them to the plain McGill forms first.
_STATUTE_DOTTED_ABBR_RE = re.compile(r"\b(?:[A-Z]\.){2,}(?=\s|,|$)")
_STATUTE_C_DOT_RE = re.compile(r"\bc\.\s*(?=[A-Z0-9])")
_STATUTE_FRENCH_MAP = {"RCS": "RSC", "LRC": "RSC", "LRO": "RSO", "LRQ": "RSQ",
                       "RLRQ": "CQLR", "LRM": "RSM", "LRA": "RSA"}
_STATUTE_FRENCH_RE = re.compile(r"\b(RCS|LRC|LRO|LRQ|RLRQ|LRM|LRA)\b(?=\s*\d{4})")


def _normalize_statute_text(text: str) -> str:
    seg = _STATUTE_DOTTED_ABBR_RE.sub(lambda m: m.group(0).replace(".", ""), str(text or ""))
    seg = _STATUTE_C_DOT_RE.sub("c ", seg)
    seg = _STATUTE_FRENCH_RE.sub(lambda m: _STATUTE_FRENCH_MAP[m.group(1)], seg)
    return seg


def _generate_fallback_url(verbatim: str, first_pinpoint: Optional[str] = None, kind: str = "other") -> str:
    m = _NEUTRAL_RE.search(verbatim)
    if m:
        year, court, num = m.groups()
        abbr = court.upper()
        slug_court = _COURT_MAP.get(abbr)
        if slug_court:
            slug = f"{year}{court.lower()}{num}"
            anch = ""
            if first_pinpoint and first_pinpoint.startswith("par"):
                anch = "#" + first_pinpoint
            elif not anch:
                para_m = _PARA_RE.search(verbatim)
                if para_m:
                    anch = "#par" + para_m.group(1)
            prefix = slug_court.split("/", 1)[0]
            lang = "fr" if prefix.lower() in _FRENCH_PREFIXES else "en"
            return f"https://www.canlii.org/{lang}/{slug_court}/doc/{year}/{slug}/{slug}.html{anch}"

    m = _CANLII_RE.search(verbatim)
    if m:
        year, num, court = m.groups()
        normalized = court.upper().replace(" ", "").strip()
        slug_court = _COURT_MAP.get(normalized)
        if slug_court:
            slug = f"{year}canlii{num}"
            anch = ""
            if first_pinpoint and first_pinpoint.startswith("par"):
                anch = "#" + first_pinpoint
            return f"https://www.canlii.org/en/{slug_court}/doc/{year}/{slug}/{slug}.html{anch}"

    if kind in ("statute", "gazette"):
        statute_text = _normalize_statute_text(verbatim)
        rsc = re.search(
            r'\bRSC\s*,?\s*(\d{4})\s*,?\s*c\.?\s*([A-Za-z0-9.-]+)',
            statute_text,
            flags=re.IGNORECASE,
        )
        if rsc:
            year, chapter = rsc.groups()
            chapter = chapter.lower().replace('.', '-').strip('-')
            legislation_id = f"rsc-{year}-c-{chapter}"
            return (
                "https://www.canlii.org/en/ca/laws/stat/"
                f"{legislation_id}/latest/{legislation_id}.html"
            )

    citation_db = _provider_registry.get_citation_db()
    if USE_DB_SEARCH and citation_db is not None:
        if kind in ("case", "unreported"):
            return citation_db.search_case_db(verbatim, first_pinpoint)
        if kind in ("statute", "gazette"):
            return citation_db.search_legislation_db(_normalize_statute_text(verbatim))

    return ""




def _normalize_anchor_fragment(fragment: str) -> str:
    frag = (fragment or "").strip()
    if not frag:
        return ""
    frag = frag.lstrip("#")
    frag = frag.strip(TRAILING_URL_PUNCT)
    frag = re.sub(r"\s+", "", frag)
    return frag
















def _extract_provider_anchor_text_segments(
    link: str,
    pinpoint_fragments: Optional[List[str]] = None,
) -> List[Dict[str, str]]:
    if not USE_DB_SEARCH or not link or link.lower() == "other":
        return []
    citation_db = _provider_registry.get_citation_db()
    fetch = getattr(citation_db, "fetch_pinpoint_segments", None)
    if not callable(fetch):
        return []
    try:
        return fetch(link, pinpoint_fragments) or []
    except Exception as exc:
        _ts_print(f"  [CASE URL] provider pinpoint lookup failed: {exc}")
        return []



# -----------------------------
# DOCX parsing
# -----------------------------

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"w": W_NS, "r": R_NS}
_EXPLICIT_FOOTNOTE_URL_RE = re.compile(
    r"(?i)(?<![\w@])(?:[a-z][a-z0-9+.-]*://|www\.|(?:[a-z0-9-]+\.)+[a-z]{2,}/)[^\s<>\"']+"
)
_HYPERLINK_FIELD_RE = re.compile(r'\bHYPERLINK\s+[\"\']([^\"\']+)[\"\']', re.IGNORECASE)
MARKER_RE = re.compile(r"⟦FN:(\d+)⟧")

OPEN_SMART_DQ = "“"
CLOSE_SMART_DQ = "”"
STRAIGHT_DQ = '"'
_CLOSE_DOUBLE_QUOTES = {STRAIGHT_DQ, CLOSE_SMART_DQ, "\u00bb"}

# Reference detection
IBID_RE = re.compile(r"\bibid\b\.?", re.IGNORECASE)
SUPRA_NOTE_RE = re.compile(r"\bsupra\s+note\s+(\d+)\b", re.IGNORECASE)
SUPRA_N_RE = re.compile(r"\bsupra\s+n\.?\s+(\d+)\b", re.IGNORECASE)
SUPRA_NN_RE = re.compile(r"\bsupra\s+nn\.?\s+(\d+)\b", re.IGNORECASE)

REF_FIELDS = [
    "ref_kind",
    "ref_target_footnote_id",
    "ref_target_citation_part_index",
    "ref_target_citation_part_text",
    "ref_target_footnote_full",
    "ref_chain_origin_footnote_id",
    "ref_chain_origin_citation_part_index",
    "ref_chain_origin_citation_part_text",
    "ref_resolution_notes",
    "ref_chain_path",
]

QUOTE_STYLE_FIELDS = [
    "quote_delimiter_style",  # SMART / STRAIGHT / MIXED / BLOCK
]


def _blank_ref_fields() -> Dict[str, str]:
    return {k: "" for k in REF_FIELDS}


def _get_zip_xml(zf: zipfile.ZipFile, name: str):
    try:
        data = zf.read(name)
    except KeyError:
        return None
    return etree.fromstring(data)


def iter_paragraph_tokens(p) -> Iterable[Tuple[str, Any]]:
    for node in p.iter():
        tag = etree.QName(node).localname
        if tag == "t":
            yield ("text", node.text or "")
        elif tag == "tab":
            yield ("text", "\t")
        elif tag in ("br", "cr"):
            yield ("text", "\n")
        elif tag == "footnoteReference":
            fid = node.get(f"{{{W_NS}}}id")
            if fid is not None:
                yield ("anchor", int(fid))


def parse_styles(styles_root):
    styles: Dict[str, Dict[str, Optional[str]]] = {}
    if styles_root is None:
        return styles

    for st in styles_root.findall("w:style", namespaces=NS):
        st_type = st.get(f"{{{W_NS}}}type")
        if st_type != "paragraph":
            continue
        style_id = st.get(f"{{{W_NS}}}styleId")
        name_el = st.find("w:name", namespaces=NS)
        name = name_el.get(f"{{{W_NS}}}val") if name_el is not None else None

        ind_left = None
        pPr = st.find("w:pPr", namespaces=NS)
        if pPr is not None:
            ind = pPr.find("w:ind", namespaces=NS)
            if ind is not None:
                ind_left = ind.get(f"{{{W_NS}}}left")

        styles[style_id] = {"name": name, "indent_left": ind_left}
    return styles


def extract_doc_stream_with_styles(doc_root, styles_root=None) -> List[Dict[str, Any]]:
    style_map = parse_styles(styles_root)
    body = doc_root.find("w:body", namespaces=NS)
    if body is None:
        raise ValueError("DOCX is missing word/document.xml w:body")

    paragraphs: List[Dict[str, Any]] = []
    for p in body.findall("w:p", namespaces=NS):
        pPr = p.find("w:pPr", namespaces=NS)
        style_id = None
        indent_left_direct = None

        if pPr is not None:
            pStyle = pPr.find("w:pStyle", namespaces=NS)
            if pStyle is not None:
                style_id = pStyle.get(f"{{{W_NS}}}val")
            ind = pPr.find("w:ind", namespaces=NS)
            if ind is not None:
                indent_left_direct = ind.get(f"{{{W_NS}}}left")

        style_name = style_map.get(style_id, {}).get("name") if style_id else None
        eff_indent_left = (
            indent_left_direct
            if indent_left_direct is not None
            else style_map.get(style_id, {}).get("indent_left")
        )

        text_parts: List[str] = []
        anchors: List[Dict[str, int]] = []
        offset = 0

        for kind, val in iter_paragraph_tokens(p):
            if kind == "text":
                text_parts.append(val)
                offset += len(val)
            elif kind == "anchor":
                anchors.append({"footnote_id": val, "offset": offset})
                marker = f"⟦FN:{val}⟧"
                text_parts.append(marker)
                offset += len(marker)

        paragraphs.append(
            {
                "style_id": style_id,
                "style_name": style_name,
                "effective_indent_left": eff_indent_left,
                "text": "".join(text_parts),
                "anchors": anchors,
            }
        )

    return paragraphs


def build_global_text(paragraphs: List[Dict[str, Any]], sep: str = "\n\n"):
    global_parts: List[str] = []
    para_starts: List[int] = []
    anchors: List[Dict[str, int]] = []
    pos = 0

    for i, p in enumerate(paragraphs):
        para_starts.append(pos)
        global_parts.append(p["text"])

        for a in p["anchors"]:
            anchors.append(
                {
                    "footnote_id": a["footnote_id"],
                    "para_index": i,
                    "para_offset": a["offset"],
                    "global_pos": pos + a["offset"],
                }
            )

        pos += len(p["text"])
        if i != len(paragraphs) - 1:
            global_parts.append(sep)
            pos += len(sep)

    return "".join(global_parts), para_starts, anchors


def _trim_url_tail_punct(url: str) -> str:
    """Strip sentence punctuation from a literal URL's tail, including a
    closing ")" that has no opening "(" inside the URL (footnotes wrap
    URLs in parentheticals; URLs may legitimately contain balanced ones)."""
    trimmed = (url or "").rstrip(".,;:!?]}>\"")
    while trimmed.endswith(")") and trimmed.count("(") < trimmed.count(")"):
        trimmed = trimmed[:-1].rstrip(".,;:!?]}>\"")
    return trimmed


def _explicit_footnote_url_target(value: str) -> str:
    target = _trim_url_tail_punct(value)
    if target and not re.match(r"^[a-z][a-z0-9+.-]*://", target, flags=re.IGNORECASE):
        target = "https://" + target
    return target


def extract_footnotes(
    footnotes_root,
    relationships_root=None,
    author_links_out: Optional[Dict[int, List[Dict[str, Any]]]] = None,
) -> Dict[int, str]:
    footnotes: Dict[int, str] = {}
    if footnotes_root is None:
        return footnotes

    relationship_targets: Dict[str, str] = {}
    if relationships_root is not None:
        for rel in relationships_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
            rel_id = rel.get("Id") or ""
            target = rel.get("Target") or ""
            if rel_id and target:
                relationship_targets[rel_id] = target

    for fn in footnotes_root.findall("w:footnote", namespaces=NS):
        fn_type = fn.get(f"{{{W_NS}}}type")
        if fn_type is not None:
            continue

        fid = fn.get(f"{{{W_NS}}}id")
        if fid is None:
            continue
        fid = int(fid)
        if fid <= 0:
            continue

        parts: List[str] = []
        hyperlink_runs: List[Tuple[int, int, str]] = []
        raw_pos = 0
        for node in fn.iter():
            tag = etree.QName(node).localname
            value = ""
            if tag == "t":
                value = node.text or ""
            elif tag == "tab":
                value = "\t"
            elif tag in ("br", "cr"):
                value = "\n"
            if not value:
                continue

            target = ""
            for ancestor in node.iterancestors():
                ancestor_tag = etree.QName(ancestor).localname
                if ancestor_tag == "hyperlink":
                    target = relationship_targets.get(ancestor.get(f"{{{R_NS}}}id") or "", "")
                    break
                if ancestor_tag == "fldSimple":
                    field_match = _HYPERLINK_FIELD_RE.search(
                        ancestor.get(f"{{{W_NS}}}instr") or ""
                    )
                    if field_match:
                        target = field_match.group(1)
                        break
            parts.append(value)
            if target:
                hyperlink_runs.append((raw_pos, raw_pos + len(value), target))
            raw_pos += len(value)

        raw_text = "".join(parts)
        text = re.sub(r"\s+", " ", raw_text).strip()
        footnotes[fid] = text

        if author_links_out is not None:
            links: List[Dict[str, Any]] = []
            search_from = 0
            for raw_start, raw_end, target in hyperlink_runs:
                anchor = re.sub(r"\s+", " ", raw_text[raw_start:raw_end]).strip()
                if not anchor:
                    continue
                start = text.find(anchor, search_from)
                if start < 0:
                    start = text.find(anchor)
                if start < 0:
                    continue
                end = start + len(anchor)
                search_from = end
                links.append({
                    "start": start,
                    "end": end,
                    "target": target,
                    "text": anchor,
                    "source": "hyperlink",
                })

            for match in _EXPLICIT_FOOTNOTE_URL_RE.finditer(text):
                raw_url = match.group(0)
                target = _explicit_footnote_url_target(raw_url)
                if not target:
                    continue
                visible_len = len(_trim_url_tail_punct(raw_url))
                start, end = match.start(), match.start() + visible_len
                if any(
                    item["start"] == start and item["end"] == end
                    and item["target"].rstrip("/") == target.rstrip("/")
                    for item in links
                ):
                    continue
                links.append({
                    "start": start,
                    "end": end,
                    "target": target,
                    "text": text[start:end],
                    "source": "literal",
                })
            if links:
                links.sort(key=lambda item: (int(item["start"]), int(item["end"])))
                author_links_out[fid] = links

    return footnotes


def build_clean_text_and_index_map(raw_text: str) -> Tuple[str, List[int]]:
    raw_to_clean = [0] * (len(raw_text) + 1)
    clean_chars: List[str] = []
    clean_idx = 0
    i = 0
    while i < len(raw_text):
        m = MARKER_RE.match(raw_text, i)
        if m:
            span_len = m.end() - m.start()
            for j in range(span_len):
                raw_to_clean[i + j] = clean_idx
            i += span_len
            continue
        clean_chars.append(raw_text[i])
        raw_to_clean[i] = clean_idx
        clean_idx += 1
        i += 1
    raw_to_clean[len(raw_text)] = clean_idx
    return "".join(clean_chars), raw_to_clean


# -----------------------------
# Quote extraction
# -----------------------------


def _quote_delimiter_style(open_ch: str, close_ch: str) -> str:
    if open_ch == OPEN_SMART_DQ and close_ch == CLOSE_SMART_DQ:
        return "SMART"
    if open_ch == STRAIGHT_DQ and close_ch == STRAIGHT_DQ:
        return "STRAIGHT"
    return "MIXED"


def find_inline_quotes(text: str) -> List[Dict[str, Any]]:
    quotes: List[Dict[str, Any]] = []
    open_pos = None
    open_ch = None

    for i, ch in enumerate(text):
        if open_pos is None:
            if ch in (OPEN_SMART_DQ, STRAIGHT_DQ):
                open_pos = i
                open_ch = ch
            continue

        if ch in (CLOSE_SMART_DQ, STRAIGHT_DQ):
            close_pos = i
            close_ch = ch

            inner = text[open_pos + 1 : close_pos]
            if inner.strip():
                raw = text[open_pos : close_pos + 1]
                style = _quote_delimiter_style(open_ch, close_ch)
                qtype = (
                    "inline_smart_quotes"
                    if style == "SMART"
                    else ("inline_straight_quotes" if style == "STRAIGHT" else "inline_mixed_quotes")
                )

                quotes.append(
                    {
                        "quote_type": qtype,
                        "start": open_pos,
                        "end": close_pos + 1,
                        "raw": raw,
                        "inner": inner,
                        "quote_delimiter_style": style,
                    }
                )

            open_pos = None
            open_ch = None

    quotes.sort(key=lambda q: q["start"])
    return quotes


def extract_sentence(clean_text: str, start: int, end: int, max_window: int = 600) -> str:
    left_limit = max(0, start - max_window)
    right_limit = min(len(clean_text), end + max_window)
    segment = clean_text[left_limit:right_limit]
    rel_start = start - left_limit
    rel_end = end - left_limit

    left_candidates = [segment.rfind(c, 0, rel_start) for c in ".!?"]
    left_nl = segment.rfind("\n", 0, rel_start)
    left_idx = max(left_candidates + [left_nl])
    left_idx = 0 if left_idx == -1 else left_idx + 1

    right_idxs = [segment.find(c, rel_end) for c in ".!?"]
    right_nl = segment.find("\n", rel_end)
    right_candidates = [i for i in (right_idxs + [right_nl]) if i != -1]
    right_idx = (min(right_candidates) + 1) if right_candidates else len(segment)

    sent = segment[left_idx:right_idx].strip()
    return re.sub(r"\s+", " ", sent)


def _sentence_bounds(clean_text: str, pos: int, max_window: int = 1200) -> Tuple[int, int]:
    """Return (start, end) indices for the sentence containing `pos`.
    Sentence boundaries are approximated using . ! ? and newlines.
    """
    pos = max(0, min(len(clean_text), int(pos)))
    left_limit = max(0, pos - max_window)
    right_limit = min(len(clean_text), pos + max_window)

    segment = clean_text[left_limit:right_limit]
    rel = pos - left_limit

    left_candidates = [segment.rfind(c, 0, rel) for c in ".!?"]
    left_nl = segment.rfind("\n", 0, rel)
    left_idx = max(left_candidates + [left_nl])
    start = left_limit + (0 if left_idx == -1 else left_idx + 1)

    # If we split on terminal punctuation that is immediately followed by a closing double quote,
    # skip that quote so it stays with the preceding sentence.
    if start > 0 and start < len(clean_text):
        if clean_text[start - 1] in ".!?" and clean_text[start] in _CLOSE_DOUBLE_QUOTES:
            start += 1

    # skip leading whitespace
    while start < right_limit and clean_text[start].isspace():
        start += 1

    right_idxs = [segment.find(c, rel) for c in ".!?"]
    right_nl = segment.find("\n", rel)
    right_candidates = [i for i in (right_idxs + [right_nl]) if i != -1]
    end = left_limit + ((min(right_candidates) + 1) if right_candidates else len(segment))

    # Include trailing closing double quotes that immediately follow terminal punctuation.
    while 0 < end < len(clean_text) and clean_text[end - 1] in ".!?" and clean_text[end] in _CLOSE_DOUBLE_QUOTES:
        end += 1

    return (start, end)


def _anchor_sentence_pos(clean_text: str, anchor_pos: int) -> int:
    """Pick an index that falls inside the intended proposition sentence.
    If the anchor is at the *end* of a sentence (typically after terminal punctuation and/or a closer),
    return the terminal punctuation index so sentence extraction returns the preceding sentence.
    Otherwise return anchor_pos.
    """
    if not clean_text:
        return 0
    j = min(max(anchor_pos - 1, 0), len(clean_text) - 1)

    # Skip whitespace just before the anchor
    while j > 0 and clean_text[j].isspace():
        j -= 1

    if clean_text[j] in ".!?":
        return j

    closers = set(')"\']}]' + "”’")
    if clean_text[j] in closers:
        k = j
        # step back over consecutive closers
        while k > 0 and clean_text[k] in closers:
            k -= 1
        # skip whitespace
        while k > 0 and clean_text[k].isspace():
            k -= 1
        if clean_text[k] in ".!?":
            return k

    return anchor_pos


def _introduction_cutoff(clean_text: str) -> int:
    """Return an index after the first occurrence of the 'Introduction' heading, if present."""
    m = re.search(r"\bIntroduction\b", clean_text, flags=re.IGNORECASE)
    if not m:
        return 0
    cut = m.end()
    while cut < len(clean_text) and clean_text[cut].isspace():
        cut += 1
    return cut


def assign_quotes_to_anchors(
    quotes: List[Dict[str, Any]],
    anchors: List[Dict[str, Any]],
    max_lookahead: int = 400,
) -> List[Dict[str, Any]]:
    anchors_sorted = sorted(anchors, key=lambda a: a["global_pos"])
    quotes_sorted = sorted(quotes, key=lambda q: q["start"])

    assignments: List[Dict[str, Any]] = []
    anchor_idx = 0

    for q in quotes_sorted:
        while anchor_idx < len(anchors_sorted) and anchors_sorted[anchor_idx]["global_pos"] < q["end"]:
            anchor_idx += 1

        assigned = None
        if anchor_idx < len(anchors_sorted):
            dist = anchors_sorted[anchor_idx]["global_pos"] - q["end"]
            if 0 <= dist <= max_lookahead:
                assigned = {**anchors_sorted[anchor_idx], "distance": dist}

        assignments.append({"quote": q, "anchor": assigned})

    return assignments


# -----------------------------
# Rule-Based Footnote splitting
# -----------------------------


# def split_footnote_parts(text: str) -> List[str]:
#     """Split on top-level semicolons; ignore semicolons inside (), [], {}, or quotes."""

#     parts: List[str] = []
#     buf: List[str] = []
#     depth_paren = depth_brack = depth_brace = 0
#     in_smart_quote = False
#     in_straight_quote = False

#     i = 0
#     while i < len(text):
#         ch = text[i]

#         # Quote-state handling that tolerates mismatched pairs.
#         if ch == OPEN_SMART_DQ:
#             in_smart_quote = True
#         elif ch == CLOSE_SMART_DQ:
#             in_smart_quote = False
#             if in_straight_quote:
#                 in_straight_quote = False
#         elif ch == STRAIGHT_DQ:
#             if in_smart_quote:
#                 in_smart_quote = False
#             else:
#                 in_straight_quote = not in_straight_quote

#         if not in_smart_quote and not in_straight_quote:
#             if ch == "(":
#                 depth_paren += 1
#             elif ch == ")" and depth_paren > 0:
#                 depth_paren -= 1
#             elif ch == "[":
#                 depth_brack += 1
#             elif ch == "]" and depth_brack > 0:
#                 depth_brack -= 1
#             elif ch == "{":
#                 depth_brace += 1
#             elif ch == "}" and depth_brace > 0:
#                 depth_brace -= 1

#         if (
#             ch == ";"
#             and depth_paren == 0
#             and depth_brack == 0
#             and depth_brace == 0
#             and not in_smart_quote
#             and not in_straight_quote
#         ):
#             part = "".join(buf).strip()
#             if part:
#                 parts.append(part)
#             buf = []
#             i += 1
#             while i < len(text) and text[i] == " ":
#                 i += 1
#             continue

#         buf.append(ch)
#         i += 1

#     last = "".join(buf).strip()
#     if last:
#         parts.append(last)
#     return parts


# -----------------------------
# GPT footnote splitting
# -----------------------------

# OpenAI is imported lazily in _ensure_llm_client (and gui's key validator):
# the sdk costs ~1s at import and isn't needed to show the window.

# LLM provider configuration
LLM_MODEL: str = "gpt-5.2"
LLM_API_KEY: str = ""  # overrides everything when set
# Optional OpenAI-compatible endpoint override (e.g. an OpenRouter/gateway
# Responses endpoint). Empty = api.openai.com. Falls back to LLM_BASE_URL in
# the environment so callers can point elsewhere
# without code changes.
LLM_BASE_URL: str = os.environ.get("LLM_BASE_URL", "")

def _get_key(env_var: str) -> str:
    key = os.environ.get(env_var, "")
    if key:
        return key
    try:
        import keys as _keys
        return getattr(_keys, env_var, "")
    except Exception:
        return ""

client: OpenAI | None = None


def _resolve_api_key() -> str:
    """Key resolution order: explicit LLM_API_KEY (test override) -> keys.py / ALT_OPENAI_API_KEY env -> OPENAI_API_KEY env -> the per-user
    encrypted store (verifier_core.api_key_store, populated from the GUI)."""
    key = LLM_API_KEY or _get_key("ALT_OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY", "")
    if key:
        return key
    try:
        from verifier_core import api_key_store
        return api_key_store.get_key()
    except Exception:
        return ""


_LLM_CLIENT_LOCK = threading.Lock()


def _ensure_llm_client() -> OpenAI:
    global client
    if client is not None:
        return client
    from openai import OpenAI
    with _LLM_CLIENT_LOCK:
        if client is None:
            client = OpenAI(api_key=_resolve_api_key(), base_url=(LLM_BASE_URL or None))
    return client


# Optional pacing hook, set by the GUI when several documents verify in
# parallel: a verifier_core.rate_governor.RateLimitGovernor that reads the
# account's live OpenAI rate limits from response headers and holds new
# requests when quota runs low. None everywhere else (CLI, benches), where
# _llm_call degrades to a plain responses.create.
_LLM_GOVERNOR = None


def _llm_call(**kwargs):
    """Every model call funnels through here. The request payload is built
    by the caller and passed through untouched (cache fingerprints depend
    on it); only the transport differs when a governor is installed."""
    llm = _ensure_llm_client()
    gov = _LLM_GOVERNOR
    raw_capable = getattr(getattr(llm, "responses", None), "with_raw_response", None)
    if gov is None or raw_capable is None:
        return llm.responses.create(**kwargs)
    gov.before_request(pause_gate=_pause_gate)
    raw = raw_capable.create(**kwargs)
    try:
        gov.observe(raw.headers)
    except Exception:
        pass
    return raw.parse()

@dataclass(frozen=True)
class FootnotePart:
    verbatim: str
    corrected: str
    kind: str  # e.g., "statute", "case", "journal", "book", "report", "other"
    link: str
    pinpoint_fragments: List[str]
    bare_citation: str = ""  # clean citation without style of cause, e.g. "[1962] SCR 746"
    citation_with_style: str = ""  # full citation with style of cause, e.g. "The Queen v. King, [1962] SCR 746"
    short_form: str = ""
    page_pinpoints: List[int] = field(default_factory=list)  # numeric page pinpoints from citation, e.g. [763, 764] for "at 763-64"
    author_provided_link: str = ""
    author_provided_links: List[str] = field(default_factory=list)
    pre_provider_link: Optional[str] = None


_SCR_REPORTED_CITATION_RE = re.compile(
    r"\[\s*(?P<year>\d{4})\s*\]\s+(?:(?P<volume>\d{1,4})\s+)?SCR\s+(?P<first_page>\d{1,4})",
    re.IGNORECASE,
)
_PAGE_LIKE_PINPOINT_RE = re.compile(
    r"\bat\s+(?:p{1,2}\.?\s*)?(?P<pinpoint>\d{1,4})(?:\s*[-–]\s*(?P<end>\d{1,4}))?",
    re.IGNORECASE,
)


def _has_scr_reported_citation(text: str) -> bool:
    return bool(_SCR_REPORTED_CITATION_RE.search(text or ""))


def _extract_scr_reported_citation_info(text: str) -> Optional[Tuple[int, int]]:
    match = _SCR_REPORTED_CITATION_RE.search(text or "")
    if not match:
        return None
    try:
        return int(match.group("year")), int(match.group("first_page"))
    except (TypeError, ValueError):
        return None


def _extract_page_like_pinpoint(text: str) -> Optional[int]:
    cleaned = text or ""
    for match in _PAGE_LIKE_PINPOINT_RE.finditer(cleaned):
        before = cleaned[max(0, match.start() - 12):match.start()].lower()
        if re.search(r"\bpara(?:graph)?s?\.?\s*$", before):
            continue
        try:
            return int(match.group("pinpoint"))
        except (TypeError, ValueError):
            continue
    return None




def _build_footnote_history_entries(parts: List["FootnotePart"]) -> str:
    entries = []
    for part in parts:
        entries.append(
            "- Citation: " + (part.verbatim or "")
            + " --> Link: " + (
                part.pre_provider_link
                if part.pre_provider_link is not None
                else (part.link or "")
            )
            + " --> short_form: " + ((part.short_form or "").strip() or "N/A") + "\n"
        )
    return "".join(entries)


# Serializes HTML/URL resolution when footnotes are split in parallel: a
# resolver's browser/HTTP caches may not be thread-safe.
_LINK_RESOLVER_LOCK = threading.Lock()

# How many run_audit calls are in flight (the GUI runs documents in
# parallel); the first run of a batch owns the per-document cache reset.
_ACTIVE_RUNS = 0
_ACTIVE_RUNS_LOCK = threading.Lock()


# Hook for the desktop GUI's pause button: the GUI sets this to a callable
# that blocks while the user has the run paused. Checked before every slow
# operation (browser fetch, LLM call, A2AJ fetch, per-row quote/journal work)
# so pausing stops activity promptly, not just between footnotes. None
# everywhere else (CLI, benches), where it costs one attribute read per check.
PAUSE_GATE = None


def _pause_gate() -> None:
    gate = PAUSE_GATE
    if gate is not None:
        gate()


_A2AJ_LOCKED_DOCUMENTS: Dict[str, a2aj_client.A2AJDocument] = {}
_A2AJ_LOCKED_STRUCTURES: Dict[str, Dict[str, Any]] = {}
_A2AJ_LOCKED_TEXTS: Dict[str, str] = {}


def _a2aj_mapped_law_evidence(
    document: a2aj_client.A2AJDocument,
) -> Tuple[str, Dict[str, Any]]:
    """Build exact law blocks from the local corpus's authoritative section map."""
    raw_sections = document.raw.get(f"unofficial_sections_{document.language}")
    if isinstance(raw_sections, str):
        try:
            raw_sections = json.loads(raw_sections)
        except (TypeError, ValueError):
            raw_sections = None
    if not isinstance(raw_sections, dict):
        return "", {}

    pieces: List[str] = []
    blocks: List[Tuple[str, str, int, int]] = []
    position = 0
    section_count = 0
    for raw_label, raw_text in raw_sections.items():
        label = str(raw_label or "").strip()
        section_text = _normalize_a2aj_source_text(str(raw_text or "")).strip()
        if not section_text:
            continue
        if pieces:
            pieces.append("\n")
            position += 1
        pieces.append(section_text)
        if re.fullmatch(r"\d{1,8}(?:[.-]\d{1,8}){0,3}", label):
            blocks.extend(
                _a2aj_structure.single_section_blocks(
                    section_text, label, start=position
                )
            )
            section_count += 1
        position += len(section_text)
    if not pieces:
        return "", {}
    return "".join(pieces), {
        "status": "usable" if blocks else "unavailable",
        "type": "section" if blocks else "",
        "source": "section_map",
        "blocks": blocks,
        "count": section_count,
    }


def _a2aj_document_evidence(
    document: a2aj_client.A2AJDocument, source_kind: str
) -> Tuple[str, Dict[str, Any]]:
    if source_kind == "law":
        mapped_text, mapped_structure = _a2aj_mapped_law_evidence(document)
        if mapped_text:
            return mapped_text, mapped_structure

    text = _normalized_a2aj_document_text(document, source_kind)
    structure = _a2aj_structure.analyze(
        text,
        source_kind,
        document.citation,
        document.alternate_citation,
        document.dataset,
        document.name,
    )
    # Rows can outlive quote checking.  Keep offsets/locators, not a second
    # copy of every paragraph/page/section body in each row.
    compact = {
        "status": structure.get("status", "unavailable"),
        "type": structure.get("type", ""),
        "count": structure.get("count", 0),
    }
    if structure.get("type") == "section":
        compact["blocks"] = structure.get("blocks") or []
    else:
        compact["paragraphs"] = [
            item[:3] for item in (structure.get("paragraphs") or [])
        ]
        compact["pages"] = [item[:3] for item in (structure.get("pages") or [])]
    return text, compact


def _a2aj_document_structure(
    document: a2aj_client.A2AJDocument, source_kind: str
) -> Dict[str, Any]:
    return _a2aj_document_evidence(document, source_kind)[1]


_A2AJ_CANLII_LAW_JURISDICTIONS = {
    "FED": "ca",
    "AB": "ab",
    "BC": "bc",
    "MB": "mb",
    "NB": "nb",
    "NL": "nl",
    "NS": "ns",
    "NT": "nt",
    "ON": "on",
    "YT": "yk",
}

_A2AJ_CANLII_STATUTE_PATTERNS = {
    "AB": r"(?:RSA|SA) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*",
    "BC": r"(?:RSBC|SBC) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*",
    "FED": (
        r"(?:RSC|SC) \d{4}(?:-\d{2}){0,3}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*"
        r"(?: \((?:\d+(?:st|nd|rd|th) )?Supp\))?(?:, s \d+)?"
    ),
    "MB": r"CCSM c [A-Z0-9]+(?:[.-][A-Z0-9]+)*",
    "NB": r"(?:RSNB|SNB) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*(?:, s \d+)?",
    "NL": r"(?:RSNL|SNL) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*",
    "NS": (
        r"(?:RSNS|SNS) \d{4}(?:-\d{2})?(?: \(\d+(?:st|nd|rd|th) Sess\))?, "
        r"c [A-Z0-9]+(?:[.-][A-Z0-9]+)*(?:, Sch [A-Z])?"
    ),
    "NT": r"(?:RSNWT|SNWT) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*(?: \(Supp\))?",
    "ON": (
        r"(?:RSO|SO) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*"
        r"(?:, Sched (?:[A-Z]|\d+))?"
    ),
    "YT": r"(?:RSY|SY) \d{4}, c [A-Z0-9]+(?:[.-][A-Z0-9]+)*(?:, Sch [A-Z0-9]+)?",
}

_A2AJ_CANLII_REGULATION_PATTERNS = {
    "AB": r"Alta Reg \d+/(?:\d{2}|\d{4})",
    "BC": r"BC Reg \d+[AB]?/(?:\d{2}|\d{4})",
    "FED": r"(?:(?:SOR|SI)/(?:\d{2}|\d{4})-\d+[AB]?|CRC, c \d+)",
    "MB": r"Man Reg \d+/(?:\d{2}|\d{4})(?: R)?",
    "NB": r"NB Reg \d{2,4}-\d+(?:, r \d+(?:\.\d+)*)?",
    "NL": r"(?:NLR|CNLR) \d+/(?:\d{2}|\d{4})",
    "NS": r"NS Reg \d+[A]?/(?:\d{2}|\d{4})",
    "NT": r"(?:NWT Reg \d{3}-(?:\d{2}|\d{4})|RRNWT \d{4}, c [A-Z0-9.-]+(?: \(Supp\))?)",
    "ON": r"(?:O Reg \d+/(?:\d{2}|\d{4})|RRO \d{4}, Reg \d+)",
    "YT": r"(?:YOIC|YMO|YCO|WCHSBO|WCBO|WCB|WSCBO) \d{4}/\d+[A-Z]?",
}

_A2AJ_CURRENT_LAW_SOURCE_HOSTS = {
    "LEGISLATION-AB": {"kings-printer.alberta.ca"},
    "REGULATIONS-AB": {"kings-printer.alberta.ca"},
    "LEGISLATION-BC": {"bclaws.gov.bc.ca", "www.bclaws.gov.bc.ca"},
    "REGULATIONS-BC": {"bclaws.gov.bc.ca", "www.bclaws.gov.bc.ca"},
    "LEGISLATION-FED": {"laws-lois.justice.gc.ca"},
    "REGULATIONS-FED": {"laws-lois.justice.gc.ca"},
    "LEGISLATION-MB": {"web2.gov.mb.ca"},
    "REGULATIONS-MB": {"web2.gov.mb.ca"},
    "LEGISLATION-NB": {"laws.gnb.ca"},
    "REGULATIONS-NB": {"laws.gnb.ca"},
    "LEGISLATION-NL": {"assembly.nl.ca", "www.assembly.nl.ca"},
    "REGULATIONS-NL": {"assembly.nl.ca", "www.assembly.nl.ca"},
    "LEGISLATION-NS": {"nslegislature.ca", "www.nslegislature.ca"},
    "REGULATIONS-NS": {"novascotia.ca", "www.novascotia.ca"},
    "LEGISLATION-NT": {"justice.gov.nt.ca", "www.justice.gov.nt.ca"},
    "REGULATIONS-NT": {"justice.gov.nt.ca", "www.justice.gov.nt.ca"},
    "LEGISLATION-ON": {"ontario.ca", "www.ontario.ca"},
    "REGULATIONS-ON": {"ontario.ca", "www.ontario.ca"},
    "LEGISLATION-YT": {"laws.yukon.ca"},
    "REGULATIONS-YT": {"laws.yukon.ca"},
}


def _a2aj_current_law_source_matches(dataset: str, source_url: str) -> bool:
    parsed = urlsplit(str(source_url or "").strip())
    host = (parsed.hostname or "").lower()
    if parsed.scheme.lower() not in {"http", "https"}:
        return False
    if host not in _A2AJ_CURRENT_LAW_SOURCE_HOSTS.get(dataset, set()):
        return False

    family, jurisdiction = dataset.split("-", 1)
    statute = family == "LEGISLATION"
    path = unquote(parsed.path).lower()
    if jurisdiction == "AB":
        params = {key.lower(): value.lower() for key, value in parse_qsl(parsed.query)}
        return path.endswith("/1266.cfm") and params.get("leg_type") == (
            "acts" if statute else "regs"
        )
    if jurisdiction == "BC":
        return "/civix/document/id/complete/statreg/" in path and path.endswith("/xml")
    if jurisdiction == "FED":
        return path.startswith("/eng/xml/") and path.endswith(".xml")
    if jurisdiction == "MB":
        shelf = "/laws/statutes/ccsm/" if statute else "/laws/regs/current/"
        return shelf in path and path.endswith(".php")
    if jurisdiction == "NB":
        shelf = "/en/document/cs/" if statute else "/en/document/cr/"
        return shelf in path
    if jurisdiction == "NL":
        shelf = "/legislation/sr/statutes/" if statute else "/legislation/sr/regulations/"
        return shelf in path and path.endswith(".htm")
    if jurisdiction == "NS":
        if statute:
            return "/sites/default/files/legc/" in path and path.endswith((".htm", ".pdf"))
        return "/just/regulations/regs/" in path and path.endswith((".htm", ".html"))
    if jurisdiction == "NT":
        if "/en/files/legislation/" not in path:
            return False
        return bool(re.search(r"\.a\.pdf$" if statute else r"\.r\d+\.pdf$", path))
    if jurisdiction == "ON":
        shelf = "statute" if statute else "regulation"
        return f"/laws/api/v2/legislation/en/doc-search/{shelf}/" in path
    if jurisdiction == "YT":
        shelves = (
            ("/legislation/principal/",)
            if statute
            else ("/legislation/subordinate/", "/legislation/regs/")
        )
        return any(shelf in path for shelf in shelves) and path.endswith(".pdf")
    return False


def _a2aj_canlii_law_url(
    dataset: str,
    citation: str,
    source_url: str,
    language: str,
) -> str:
    """Construct a current CanLII law URL only from independently locked fields."""
    dataset = str(dataset or "").strip().upper()
    citation = re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(citation or ""))).strip()
    if str(language or "").lower() != "en":
        return ""
    match = re.fullmatch(r"(LEGISLATION|REGULATIONS)-([A-Z]{2,3})", dataset)
    if not match or not _a2aj_current_law_source_matches(dataset, source_url):
        return ""

    family, jurisdiction = match.groups()

    # A2AJ has one current Ontario row whose otherwise canonical citation has
    # a stray space before the schedule comma.  Its official source identity
    # and independently confirmed CanLII target make this exact repair safe;
    # do not normalize the other anomalous schedule rows wholesale.
    if (
        dataset == "LEGISLATION-ON"
        and citation.lower() == "so 2010, c 16 , sched 4"
        and urlsplit(source_url).path.lower().endswith("/10c16b")
    ):
        citation = "SO 2010, c 16, Sched 4"

    # Three distinct federal documents share this normalized citation.  A
    # citation-derived URL cannot select one of them with identity certainty.
    if (
        dataset == "LEGISLATION-FED"
        and citation.lower() == "rsc 1985, c 41 (4th supp)"
    ):
        return ""

    canlii_jurisdiction = _A2AJ_CANLII_LAW_JURISDICTIONS.get(jurisdiction)
    patterns = (
        _A2AJ_CANLII_STATUTE_PATTERNS
        if family == "LEGISLATION"
        else _A2AJ_CANLII_REGULATION_PATTERNS
    )
    pattern = patterns.get(jurisdiction)
    if not canlii_jurisdiction or not pattern or not re.fullmatch(pattern, citation, flags=re.I):
        return ""

    slug_citation = citation
    if dataset == "REGULATIONS-AB":
        citation_match = re.fullmatch(
            r"Alta Reg (\d+)/(\d{2}|\d{4})", citation, flags=re.I
        )
        query = {key.lower(): value for key, value in parse_qsl(urlsplit(source_url).query)}
        page_match = re.fullmatch(r"(\d{4})_(\d+)\.cfm", query.get("page", ""), flags=re.I)
        if not citation_match or not page_match:
            return ""
        number, cited_year = citation_match.groups()
        source_year, source_number = page_match.groups()
        if int(number) != int(source_number) or not source_year.endswith(cited_year):
            return ""
        slug_citation = f"Alta Reg {number}/{source_year}"
    elif dataset == "REGULATIONS-NB":
        slug_citation = re.sub(r", r \d+(?:\.\d+)*$", "", citation, flags=re.I)
    elif dataset == "LEGISLATION-ON":
        slug_citation = re.sub(r"\bSched\b", "Sch", citation, flags=re.I)

    slug = unicodedata.normalize("NFKC", slug_citation).lower()
    slug = slug.replace("–", "-").replace("—", "-")
    slug = re.sub(r"[(),]", "", slug)
    slug = re.sub(r"[\s/]+", "-", slug).strip("-")
    slug = re.sub(r"-{2,}", "-", slug)
    if not re.fullmatch(r"[a-z0-9]+(?:[.-][a-z0-9]+)*", slug):
        return ""

    canlii_family = "stat" if family == "LEGISLATION" else "regu"
    return (
        f"https://www.canlii.org/en/{canlii_jurisdiction}/laws/{canlii_family}/"
        f"{slug}/latest/{slug}.html"
    )


def _a2aj_url_matches_document(
    url: str, document: a2aj_client.A2AJDocument, source_kind: str
) -> bool:
    """Whether A2AJ and the URL independently identify the same document."""
    if source_kind == "law":
        expected = _a2aj_canlii_law_url(
            document.dataset,
            document.citation,
            document.url,
            document.language,
        )
        base, _fragment = _split_url(_canlii_source_lookup_url(url))
        return bool(expected and base.rstrip("/").lower() == expected.lower())
    if source_kind != "case":
        return False
    url_citation = _canlii_doc_citation_from_url(url)
    if not url_citation:
        return False
    target = a2aj_client._citation_key(url_citation)
    return bool(target and target in {
        a2aj_client._citation_key(document.citation),
        a2aj_client._citation_key(document.alternate_citation),
    })


def _register_a2aj_document(
    base: str,
    document: a2aj_client.A2AJDocument,
    source_kind: str,
    *,
    structure: Optional[Dict[str, Any]] = None,
    evidence_text: Optional[str] = None,
) -> None:
    key = (base or "").lower()
    if not key:
        return
    if evidence_text is None:
        derived_text, derived_structure = _a2aj_document_evidence(
            document, source_kind
        )
        evidence_text = derived_text
        if structure is None:
            structure = derived_structure
    _A2AJ_LOCKED_DOCUMENTS[key] = document
    _A2AJ_LOCKED_STRUCTURES[key] = structure or {}
    _A2AJ_LOCKED_TEXTS[key] = evidence_text or ""


def _a2aj_identity_citation(text: str) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    patterns = (
        r"\b\d{4}\s+[A-Z][A-Z0-9-]{1,15}\s+\d+\b",
        r"\[\d{4}\]\s*(?:\d+\s+)?S\.?C\.?R\.?\s+\d+\b",
        r"\(\d{4}\)\s+\d+\s+S\.?C\.?R\.?\s+\d+\b",
    )
    for pattern in patterns:
        match = re.search(pattern, value, flags=re.I)
        if match:
            identity = match.group(0)
            if re.search(r"S\.?\s*C\.?\s*R\.?", identity, flags=re.I):
                identity = re.sub(
                    r"S\.?\s*C\.?\s*R\.?", "SCR", identity, flags=re.I
                )
            return identity
    return ""


def _a2aj_registered_section_text(url: str, pinpoint: str) -> str:
    """Return a mapped A2AJ section without refetching or parsing the Act."""
    section_match = re.match(
        r"^sec(\d{1,8}(?:[.-]\d{1,8}){0,3})",
        str(pinpoint or "").strip(),
        flags=re.IGNORECASE,
    )
    if not section_match:
        return ""
    key = _fragment_doc_key(url)
    structure = _A2AJ_LOCKED_STRUCTURES.get(key, {})
    document_text = _A2AJ_LOCKED_TEXTS.get(key, "")
    if structure.get("source") != "section_map" or not document_text:
        return ""
    locator = "sec" + section_match.group(1)
    for _section, block_locator, start, end in structure.get("blocks") or []:
        if str(block_locator).casefold() == locator.casefold():
            return re.sub(r"\s+", " ", document_text[start:end]).strip()
    return ""


def _a2aj_case_link(document: a2aj_client.A2AJDocument, language: str) -> str:
    from verifier_core.canlii_urls import COURT_MAP

    citations = [document.citation, document.alternate_citation]
    for citation in citations:
        match = re.search(r"\b(\d{4})\s+([A-Za-z][A-Za-z0-9-]{1,15})\s+(\d+)\b", citation or "")
        if not match:
            continue
        year, court, number = match.groups()
        court_key = court.upper()
        if document.dataset.upper() != court_key:
            continue
        route = COURT_MAP.get(court_key)
        if not route:
            continue
        slug_court = "canlii" if court_key == "CANLII" else court.lower()
        slug = f"{year}{slug_court}{number}"
        return f"https://www.canlii.org/{language}/{route}/doc/{year}/{slug}/{slug}.html"
    # Reporter-only identities may not contain a neutral citation from which we
    # can construct a CanLII URL.  A2AJ still supplies the source court URL.
    source_url = (document.url or "").strip()
    parsed = urlsplit(source_url)
    if parsed.scheme.lower() in {"http", "https"} and parsed.netloc:
        return source_url
    return ""


def _adapter_identity_citation(citation_text: str) -> str:
    """Reserved for installed identity providers; none ship in this build."""
    return ""




def _a2aj_resolve_case_before_browser(
    citation_text: str, link_candidate: str, pinpoint_fragments: Optional[List[str]]
) -> str:
    if not USE_A2AJ:
        return ""
    citation = _a2aj_identity_citation(citation_text)
    used_alias = False
    if not citation:
        # Reporter-style citations ("Hunter v Southam Inc, 11 DLR (4th) 641")
        # carry no neutral/SCR pattern; the consensus reporter-alias inventory
        # (or an installed identity provider) may still identify the document.
        citation = (
            a2aj_client.get_client().reporter_alias_canonical(citation_text)
            or _adapter_identity_citation(citation_text)
        )
        if not citation:
            return ""
        used_alias = True
    neutral = re.search(r"\b\d{4}\s+([A-Z][A-Z0-9-]{1,15})\s+\d+\b", citation, flags=re.I)
    expected_dataset = neutral.group(1).upper() if neutral else "SCC"
    coverage = a2aj_client.get_client().coverage("cases")
    if coverage and expected_dataset not in coverage:
        # Citations styled after an uncovered scheme (for example CanLII
        # numbers) may still map to a covered document via the alias
        # inventory.
        alias_canonical = "" if used_alias else (
            a2aj_client.get_client().reporter_alias_canonical(citation_text)
            or _adapter_identity_citation(citation_text)
        )
        alias_neutral = re.search(
            r"\b\d{4}\s+([A-Z][A-Z0-9-]{1,15})\s+\d+\b", alias_canonical, flags=re.I
        )
        alias_dataset = alias_neutral.group(1).upper() if alias_neutral else "SCC"
        if not alias_canonical or alias_dataset not in coverage:
            _ts_print(
                f"  A2AJ miss (coverage excludes {expected_dataset}): "
                f"{citation}; trying CanLII"
            )
            return ""
        citation = alias_canonical
        expected_dataset = alias_dataset
        used_alias = True
    if used_alias:
        _ts_print(
            f"  A2AJ reporter alias: {_a2aj_query_citation(citation_text)} -> {citation}"
        )
    language = "fr" if "/fr/" in (link_candidate or "").lower() else "en"
    _pause_gate()
    lookup = a2aj_client.lookup_document(citation, "case", language=language)
    if lookup.status != "found" or lookup.document is None:
        _ts_print(f"  A2AJ miss ({lookup.status}): {citation}; trying CanLII")
        return ""
    link = _a2aj_case_link(lookup.document, language)
    if not link:
        return ""
    if used_alias and "canlii.org" not in link.lower():
        # An alias-resolved document rarely carries the CanLII-styled
        # citation, so its A2AJ link is usually the court site.  Prefer a
        # CanLII URL (it supports paragraph anchors): keep an incoming
        # candidate that already names the same decision, else construct
        # one from the original CanLII citation plus the now-known court.
        from verifier_core.canlii_urls import COURT_MAP

        derived = _canlii_doc_citation_from_url(link_candidate or "")
        original = _a2aj_identity_citation(citation_text)
        if (
            derived and original
            and a2aj_client._citation_key(derived) == a2aj_client._citation_key(original)
        ):
            candidate_base, _candidate_frag = _split_url(
                _sanitize_url_candidate(link_candidate)
            )
            if candidate_base:
                link = candidate_base
        else:
            canlii_number = re.search(
                r"\b(\d{4})\s+CanLII\s+(\d+)\b", citation_text, flags=re.I
            )
            route = COURT_MAP.get(lookup.document.dataset.upper())
            if canlii_number and route:
                year, number = canlii_number.groups()
                slug = f"{year}canlii{number}"
                link = (
                    f"https://www.canlii.org/{language}/{route}"
                    f"/doc/{year}/{slug}/{slug}.html"
                )
    if (
        expected_dataset == "SCC"
        and "canlii.org" not in link.lower()
        and USE_DB_SEARCH
    ):
        citation_db = _provider_registry.get_citation_db()
        db_search = getattr(citation_db, "search_case_db", None)
        if callable(db_search):
            canlii_link = db_search(citation_text, None)
            if "canlii.org" in (canlii_link or "").lower():
                link = canlii_link
    link = _append_first_pinpoint_fragment(link, pinpoint_fragments)
    base, _fragment = _split_url(link)
    if base:
        _register_a2aj_document(base, lookup.document, "case")
    _ts_print(f"  A2AJ identity lock ({lookup.method}): {citation} -> {link}")
    return link


def _a2aj_has_law_before_browser(
    citation_text: str, link_candidate: str
) -> Optional[Dict[str, Any]]:
    """Probe A2AJ for the cited law. Returns the probe row (carrying the
    A2AJ evidence keys, including ``_a2aj_source_url``) when exact source
    text is available, else None."""
    if not USE_A2AJ:
        return None
    if re.search(r"/laws/(?:astat|hstat)/", link_candidate or "", flags=re.I):
        return None
    probe = {
        "bare_citation": citation_text,
        "citation_part_kind": "statute",
        "citation_part_link": link_candidate,
    }
    if not _fetch_a2aj_source_text_for_row(probe):
        return None
    _ts_print(f"  A2AJ source available: {_a2aj_query_citation(citation_text)}")
    return probe


_FED_LAW_XML_RE = re.compile(r"/(eng|fra)/XML/([^/]+?)\.xml$", re.IGNORECASE)
_FED_REGULATION_CODE_RE = re.compile(r"(?:SOR|SI|C\.?R\.?C\.?|DORS|TR)\b", re.IGNORECASE)
_ON_ELAWS_API_RE = re.compile(
    r"/laws/api/v2/legislation/en/doc-search/(statute|regulation)/([^/?#]+)$",
    re.IGNORECASE,
)


def _a2aj_official_law_url(probe: Dict[str, Any]) -> str:
    """Human-readable official-source URL for an A2AJ-locked law, or "".

    A2AJ source URLs are sometimes machine formats; the known dataset shapes
    normalize to their human pages (live-verified 2026-07): federal XML ->
    laws-lois acts/regulations pages, Ontario e-Laws API -> ontario.ca/laws
    pages, BC CiviX XML -> its HTML sibling. Anything still pointing at a
    machine format is withheld rather than shipped as a citation link.
    """
    url = str(probe.get("_a2aj_source_url") or "").strip()
    parsed = urlsplit(url)
    if parsed.scheme.lower() not in ("http", "https") or not parsed.netloc:
        return ""
    host = parsed.netloc.lower()
    if host.endswith("laws-lois.justice.gc.ca"):
        m = _FED_LAW_XML_RE.search(parsed.path)
        if m:
            lang, code = m.group(1).lower(), m.group(2)
            regulation = bool(_FED_REGULATION_CODE_RE.match(code))
            if lang == "fra":
                shelf = "reglements" if regulation else "lois"
            else:
                shelf = "regulations" if regulation else "acts"
            return f"https://laws-lois.justice.gc.ca/{lang}/{shelf}/{code}/"
    if host.endswith("ontario.ca"):
        m = _ON_ELAWS_API_RE.search(parsed.path)
        if m:
            return f"https://www.ontario.ca/laws/{m.group(1).lower()}/{m.group(2)}"
    if host.endswith("bclaws.gov.bc.ca") and parsed.path.rstrip("/").lower().endswith("/xml"):
        trimmed = parsed.path.rstrip("/")
        return f"https://{parsed.netloc}{trimmed[: -len('/xml')]}"
    if parsed.path.lower().endswith(".xml"):
        return ""
    return url


def _resolve_footnote_part_link(
    *,
    verbatim: str,
    citation_with_style: str,
    kind: str,
    link_candidate: str,
    pinpoint_fragments: Optional[List[str]],
    bare_citation: str = "",
) -> str:
    with _LINK_RESOLVER_LOCK:
        return _resolve_footnote_part_link_unlocked(
            verbatim=verbatim,
            citation_with_style=citation_with_style,
            kind=kind,
            link_candidate=link_candidate,
            pinpoint_fragments=pinpoint_fragments,
            bare_citation=bare_citation,
        )


def _resolve_footnote_part_link_unlocked(
    *,
    verbatim: str,
    citation_with_style: str,
    kind: str,
    link_candidate: str,
    pinpoint_fragments: Optional[List[str]],
    bare_citation: str = "",
) -> str:
    first_pf = (pinpoint_fragments or [None])[0]
    normalized_kind = (kind or "").strip().lower()
    if normalized_kind in ("case", "unreported"):
        a2aj_link = _a2aj_resolve_case_before_browser(
            citation_with_style or verbatim, link_candidate, pinpoint_fragments
        )
        if a2aj_link:
            return a2aj_link
    elif normalized_kind in ("statute", "gazette"):
        law_probe = _a2aj_has_law_before_browser(
            bare_citation or citation_with_style or verbatim, link_candidate
        )
        if law_probe:
            existing_link = (link_candidate or '').strip()
            if existing_link.lower() == 'other':
                existing_link = ''
            constructed_link = _a2aj_canlii_law_url(
                law_probe.get("_a2aj_dataset", ""),
                law_probe.get("_a2aj_citation", ""),
                law_probe.get("_a2aj_source_url", ""),
                law_probe.get("_a2aj_language", ""),
            )
            link = constructed_link or existing_link or _generate_fallback_url(
                citation_with_style or verbatim, first_pf, normalized_kind
            ) or _a2aj_official_law_url(law_probe)
            if link:
                if constructed_link:
                    _ts_print(
                        "  A2AJ current-law identity lock: "
                        f"{law_probe.get('_a2aj_citation', '')} -> {constructed_link}"
                    )
                return _append_first_pinpoint_fragment(link, pinpoint_fragments)

    citation_db = _provider_registry.get_citation_db()
    external_search = getattr(citation_db, "search_external_case_url", None)
    existing_candidate = (link_candidate or "").strip()
    existing_host = urlsplit(existing_candidate).netloc.lower()
    if (
        USE_DB_SEARCH
        and normalized_kind in ("case", "unreported")
        and callable(external_search)
        and (
            not existing_candidate
            or existing_candidate.lower() == "other"
            or existing_host == "canlii.org"
            or existing_host.endswith(".canlii.org")
        )
    ):
        external_link = external_search(citation_with_style or verbatim, first_pf)
        if external_link:
            return external_link
    if normalized_kind in ("case", "unreported") and _has_scr_reported_citation(verbatim or citation_with_style):
        fallback_override = _generate_fallback_url(verbatim or citation_with_style, first_pf, normalized_kind)
        if fallback_override:
            # TODO: hacky override, but we should probably try heading this off at the source:
            # would it be possible to detect where the LLM built the link off an SCR citation,
            # w/ no other citations in the footnote, and tried to build a link? Because the LLM,
            # in theory, would sort of always just be guessing at how to build an SCR link to begin with.
            current_base = _split_url(_canlii_source_lookup_url(link_candidate or ""))[0]
            override_base = _split_url(_canlii_source_lookup_url(fallback_override))[0]
            if override_base and override_base != current_base:
                _ts_print(f"  Override URL: {fallback_override}")
            _pause_gate()
            resolved_override = LINK_RESOLVER.resolve_url(_canlii_source_lookup_url(fallback_override))
            if resolved_override and override_base != current_base:
                _ts_print(f"  Resolved override URL: {resolved_override}")
            return _append_first_pinpoint_fragment(fallback_override, pinpoint_fragments)

    source_link_candidate = _canlii_source_lookup_url(link_candidate)
    _pause_gate()
    resolved_link = LINK_RESOLVER.resolve_url(source_link_candidate)  # returns "" on failure
    if resolved_link and _is_canlii_source_lookup_variant(link_candidate, source_link_candidate):
        link = link_candidate
    else:
        link = resolved_link

    if not link or link.lower() == "other":
        fallback = _generate_fallback_url(citation_with_style or verbatim, first_pf, normalized_kind)
        if fallback:
            _ts_print(f"  Fallback URL: {fallback}")
            _pause_gate()
            resolved_fallback = LINK_RESOLVER.resolve_url(_canlii_source_lookup_url(fallback))
            if resolved_fallback:
                _ts_print(f"  Resolved fallback URL: {resolved_fallback}")
            link = fallback

    return _append_first_pinpoint_fragment(link, pinpoint_fragments)


def _normalize_footnote_part_link(part: "FootnotePart") -> "FootnotePart":
    pre_provider_link = (
        part.pre_provider_link
        if part.pre_provider_link is not None
        else part.link
    )
    pinpoint_fragments, page_pinpoints = _pinpoints_for_source_kind(
        part.kind, part.pinpoint_fragments, part.page_pinpoints
    )
    link = _resolve_footnote_part_link(
        verbatim=part.verbatim,
        citation_with_style=part.citation_with_style,
        kind=part.kind,
        link_candidate=pre_provider_link,
        pinpoint_fragments=pinpoint_fragments,
        bare_citation=part.bare_citation,
    )
    return replace(
        part,
        link=link,
        pinpoint_fragments=pinpoint_fragments,
        page_pinpoints=page_pinpoints,
        pre_provider_link=pre_provider_link,
    )


# --- JSON Schema for Structured Outputs (Responses API) ---
FOOTNOTE_SPLIT_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "parts": {
            "type": "array",
            "minItems": 1,
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "verbatim": {"type": "string"},
                    "corrected": {"type": "string"},
                    "kind": {"type": "string"},
                    "link": {"type": "string"},
                    "pinpoint_fragments": {"type": "array", "items": {"type": "string"}},
                    "page_pinpoints": {"type": "array", "items": {"type": "integer"}},
                    "notes": {"type": "string"},
                    "short_form": {"type": "string"},
                    "bare_citation": {"type": "string"},
                    "citation_with_style": {"type": "string"},
                },
                "required": ["verbatim", "corrected", "kind", "link", "pinpoint_fragments", "page_pinpoints", "notes", "short_form", "bare_citation", "citation_with_style"],
            },
        },
        "joined_corrected": {"type": "string"},
    },
    "required": ["parts", "joined_corrected"],
}


SYSTEM_INSTRUCTIONS = """You split legal footnotes, correct citation formatting according to the McGill Guide, and return structured JSON including a link whenever possible.

Task:
1) Split the input footnote into citation parts. Default rule: split on top-level semicolons.
2) However, authors sometimes split incorrectly. You must fix splitting errors:
   - If a semicolon is missing between two distinct citations, split anyway.
    - If a semicolon wrongly splits a single citation (e.g., inside a citation element), do NOT split there.

3) For each part:
   - verbatim: the exact substring for that citation part (preserve spelling, punctuation, spacing except trimming outer whitespace).
   - corrected: the same citation, but corrected to conform to McGill-style ordering and punctuation/spacing where determinable from the text.
     * Do not invent missing bibliographic data.
     * If information is missing, keep it missing; only fix structure, separators, spacing, obvious punctuation.
     * If there are notes (other than the citation) in the footnote, they should be preserved.
   - kind: one of: statute, gazette, case, unreported, parliamentary_paper, non_parliamentary, journal, book, essay_collection, report, other.
   - link: according to the link patterns.
     - pinpoint_fragments: list of pinpoint fragments extracted from the citation part. For ranges (e.g., "paras 15–17"), output only the first fragment (par15). For comma-separated pinpoints (e.g., "paras 99, 101"), output each one separately (par99, par101). Examples: "Servatius BCSC, supra note 1 at paras 15–17" → ["par15"]; "Servatius BCCA, supra note 2 at paras 99, 101" → ["par99", "par101"]; "Criminal Code, RSC 1985, c C-46, s 718(c), 16, 672.54(b)" → ["sec718", "sec16", "sec672.54"]. Rules and articles are legislation provisions too: "r 11.10" → ["sec11.10"], "Rule 4-1" → ["sec4-1"], and "art 1457" → ["sec1457"]. Preserve the full dotted or hyphenated provision number; omit subsection parentheses from the fragment.
    - page_pinpoints: list of integer page numbers from the citation's page pinpoint. For "at 245" or "at p 245" → [245]; for "at 763-64" → [763, 764]; for "at pp 99-101" → [99, 100, 101]; for "at para 20" or "at paras 15-17" → [] (paragraphs are not pages); for section pinpoints → []. Empty list if no page pinpoint exists.
    - notes: brief description of what you changed (or "no change").
    - short_form: the short form label for this citation (e.g., "[Brown (SCC)]" or, for secondary sources without a bracket short form, the author surname(s) used to refer to this work in later supra references (e.g., "Roach" for "Kent Roach, Criminal Law...", or "Mishra, Logan and Prescott" for a joint work). If the citation text already has a [short form] in brackets, extract it without brackets. For secondary sources (journal articles, books, reports), infer the author name(s) as they would appear in a supra reference.
    - bare_citation: the citation itself with introductory signals, editorial notes, style of cause, and surrounding commentary stripped away — just the bare legal citation with pinpoint. For example, if verbatim is "See, e.g., The Queen v. King, [1962] SCR 746 at 763-64", bare_citation should be "[1962] SCR 746 at 763-64". If verbatim is "(R v Sullivan, [2016] OJ No 6847 at paras 47 and 80 [Sullivan (ONSC)])", bare_citation should be "[2016] OJ No 6847 at paras 47 and 80". If verbatim is "Criminal Code, RSC 1985, c C-46, s 718(c)", bare_citation should be "RSC 1985, c C-46, s 718(c)". This field is used for URL resolution and A2AJ lookup, and is not displayed to the user.
    - citation_with_style: the full citation text including style of cause if present, with introductory signals and editorial notes stripped. For example, if verbatim is "See, e.g., The Queen v. King, [1962] SCR 746 at 763-64", citation_with_style should be "The Queen v. King, [1962] SCR 746 at 763-64". If verbatim is "Criminal Code, RSC 1985, c C-46, s 718(c)", citation_with_style should be "Criminal Code, RSC 1985, c C-46, s 718(c)". This field is used for database search and fallback URL generation.

4) Split supra and ibid references into their own parts. Never leave a supra or ibid merged with another citation in the same part. If the whole footnote is only one supra or ibid, keep it as a single part.

Splitting examples (each → shows the correct parts):

"Jane Smith, (2024) 45 Journal 100. See also Short Form, supra note 3 at 45."
→ Part 1: "Jane Smith, (2024) 45 Journal 100."
→ Part 2: "See also Short Form, supra note 3 at 45."

"supra note 3; see also Case Name, 2025 SCC 10 at para 20."
→ Part 1: "supra note 3"
→ Part 2: "see also Case Name, 2025 SCC 10 at para 20."

"Short Form, supra note 3 at 45. Ibid at 50."
→ Part 1: "Short Form, supra note 3 at 45."
→ Part 2: "Ibid at 50."

"Citation A. See also supra note 3. See also supra note 7."
→ Part 1: "Citation A."
→ Part 2: "See also supra note 3."
→ Part 3: "See also supra note 7."

"Full Citation (2024) 45 Journal 100. See also Author, supra note 3 at 20-25. See further Other, supra note 8 at 5."
→ Part 1: "Full Citation (2024) 45 Journal 100."
→ Part 2: "See also Author, supra note 3 at 20-25."
→ Part 3: "See further Other, supra note 8 at 5."

Every part must contain unique text that does not overlap with any other part.

Link patterns:
- For cases, in the form: https://www.canlii.org/en/[jurisdiction]/[court/tribunal]/doc/[YYYY]/[citation]/[citation].html[#par[first pinpoint]] (e.g., https://www.canlii.org/en/bc/bcca/doc/2026/2026bcca2/2026bcca2.html#par34; https://www.canlii.org/en/ca/scc/doc/2025/2025scc42/2025scc42.html; https://www.canlii.org/en/ab/ablprt/doc/2025/2025ablprt809/2025ablprt809.html#par4).
- If a case citation has a CanLII citation, use the CanLII citation in the link.
- If a case court/tribunal is either QCCM, QCCQ, QCCS, or QCCA, `/en/` must be replaced with `/fr/` in the link.
- For legislation, in the form: https://www.canlii.org/en/ca/laws/[legislation type]/[citation]/latest/[citation].html[#sec[first pinpoint]] (e.g., https://www.canlii.org/en/ab/laws/regu/alta-reg-272-1996/latest/alta-reg-272-1996.html#sec2; https://www.canlii.org/en/on/laws/astat/so-2025-c-pr16/latest/so-2025-c-pr16.html; https://www.canlii.org/en/ca/laws/stat/rsc-1985-c-a-1/latest/rsc-1985-c-a-1.html).
- If there are paragraph pinpoints in a citation to a case, append the first pinpoint number to the link with `#par[first pinpoint]`.
- If legislation has a section, rule, or article pinpoint, append its normalized first provision fragment with `#sec[first pinpoint]`.
- If there is a URL or perma.cc link, you must isolate and return the link (preferring the URL link over the perma.cc link when both are available).
- If the footnote is a supra note to a previous case/statute citation, return the link to that citation. You should use the list of citations in "Previous citations and their links" to find the link to return.
- If the footnote is an ibid, it must use the link from the last citation in the "Previous citations and their links" list.

McGill Guide patterns (high-level):
- Statutes: Title, | volume | jurisdiction | year, | c | other elements | (session/supp) | pinpoint.
- Gazettes: Title (person/body), | (year) | Gazette abbr | part | page | (additional info).
- Jurisprudence: Style of cause, | main citation | pinpoint, | parallel citation | (jurisdiction/court) | [short form].
- Unreported: Style of cause | (date), | judicial district | docket no | (jurisdiction/court).
- Parliamentary Papers: Jurisdiction, | legislature, | title, | session, | volume | number | (date) | pinpoint | (speaker).
- Non-parliamentary documents: Jurisdiction, | issuing body, | title, | (type), | publication info | pinpoint.
- Journals: Author, | “title” | (year) | volume: | issue | journal abbr | first page | pinpoint.
- Books: Author, | title, | edition | (place: publisher, year) | pinpoint.
- Essay/entry in collection: Author, | “title” | in | editor, ed, | book title, | edition | (place: publisher, year) | first page | pinpoint.
- Reports: Author, | title, | (place: publisher, year) | pinpoint.
- Website: Author, | “title of the page/article” | (date of the page/article) | pinpoint, | online | (type of electronic source) | : | <URL> | [perma.cc URL].

Examples:
- Citation: CQLR c L-0.3, s 34 [Bill 21] --> Link: https://www.canlii.org/en/qc/laws/stat/cqlr-c-l-0.3/latest/cqlr-c-l-0.3.html#sec34
- Citation: Election Finances Act, RSO 1990, c E.7, ss 37.101(2), 53.1(1), as it appeared on 4 December 2024 [Election Finances Act] --> Link: https://www.canlii.org/en/on/laws/stat/rso-1990-c-e7/latest/rso-1990-c-e7.html#sec37
- Citation: The Education (Parents' Bill of Rights) Amendment Act, SS 2023, c 46, s 4 [Parents’ Bill of Rights] --> Link: https://www.canlii.org/en/sk/laws/astat/ss-2023-c-46/latest/ss-2023-c-46.html#sec4
- Citation: English Montreal School Board, et al v Attorney General of Quebec, 2025 CanLII 2818 (SCC) --> Link: https://www.canlii.org/en/ca/scc-l/doc/2025/2025canlii2818/2025canlii2818.html
- Citation: Constitution Act, 1982, s 52(1), being Schedule B to the Canada Act 1982 (UK), 1982, c 11 [Constitution Act, 1982] --> Link: https://www.canlii.org/en/ca/laws/stat/schedule-b-to-the-canada-act-1982-uk-1982-c-11/latest/schedule-b-to-the-canada-act-1982-uk-1982-c-11.html#sec52
- Citation: The Canadian Charter of Rights and Freedoms, Part I of the Constitution Act, 1982, being Schedule B to the Canada Act 1982 (UK), 1982, c 11 [the Charter] --> Link: https://www.canlii.org/en/ca/laws/stat/schedule-b-to-the-canada-act-1982-uk-1982-c-11/latest/schedule-b-to-the-canada-act-1982-uk-1982-c-11.html#sec1
- Citation: Multiple Access Ltd v McCutcheon, [1982] 2 SCR 161, 1982 CanLII 55 (SCC) --> Link: https://www.canlii.org/en/ca/scc/doc/1982/1982canlii1705/1982canlii1705.html
- Citation: Zylberberg v Sudbury Board of Education, 1988 CanLII 189, 65 OR (2d) 641 (ONCA) --> Link: https://www.canlii.org/en/on/onca/doc/1988/1988canlii189/1988canlii189.html
- Citation: R v Example, 2024 SCC 1 --> Link: https://www.canlii.org/en/ca/scc/doc/2024/2024scc1/2024scc1.html
- Citation: G, supra note 78 at para 98, citing Marbury v Madison, 5 US (1 Cranch) 137 (1803) at 177, “[i]t is emphatically the province and duty of [the courts] to say what the law is.” --> Link: other
- Citation: R v Katigakyok, 2019 NWTTC 12 at para 34 --> Link: https://www.canlii.org/en/nt/nttc/doc/2019/2019nwttc12/2019nwttc12.html#par34
- Citation: B(R) v Children’s Aid Society of Metropolitan Toronto, [1995] 1 SCR 315 at 346–347, [1994] SCJ No 24 (SCC) --> Link: other
- Citation: Dessureault c. Désaulniers, 2026 QCCA 45 --> Link: https://www.canlii.org/fr/qc/qcca/doc/2026/2026qcca45/2026qcca45.html

Output requirements:
- Return JSON strictly matching the schema.
- Preserve the order of citations as they appear.
- Ensure joined_corrected is the corrected parts joined by "; " (semicolon + space).
- Do not include any extra keys. Foreign case citations (e.g., US, UK, AUS) can be ignored.

Previous citations and their links (each entry has: verbatim text, resolved link, and the short_form):
"""

# Stable cache identity for the sole production prompt. Keeping the historical
# value preserves existing validated LLM cache entries.
FOOTNOTE_PROMPT_CACHE_ID = "original"


def _prompt_cache_fingerprint(system_prompt: str) -> str:
    prompt_hash = hashlib.sha256(system_prompt.encode("utf-8")).hexdigest()[:16]
    return f"{FOOTNOTE_PROMPT_CACHE_ID}:{prompt_hash}"


FOOTNOTE_SPLIT_CACHE_VERSION = "footnote_split_cache_v2"
FOOTNOTE_RESPONSE_FORMAT_NAME = "footnote_split"
FOOTNOTE_RESPONSE_REASONING = {"effort": "none"}
FOOTNOTE_RESPONSE_MAX_OUTPUT_TOKENS = 16000
_LLM_CACHE_LOOKUP_INDEX: Optional[Dict[str, str]] = None
_LLM_CACHE_LOOKUP_INDEX_DIR: str = ""


def _stable_json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False)


def _sha256_text(value: str) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8")).hexdigest()


def _footnote_request_config(
    *,
    system_prompt: str,
    prompt_fingerprint: str,
    text: str,
    previous_citations: str,
    schema: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    schema = schema or FOOTNOTE_SPLIT_SCHEMA
    schema_fingerprint = hashlib.sha256(_stable_json(schema).encode("utf-8")).hexdigest()
    return {
        "cache_version": FOOTNOTE_SPLIT_CACHE_VERSION,
        "api": "responses.create",
        "model": LLM_MODEL,
        "prompt_mode": FOOTNOTE_PROMPT_CACHE_ID,
        "prompt_fingerprint": prompt_fingerprint,
        "system_prompt_sha256": _sha256_text(system_prompt),
        "previous_citations_sha256": _sha256_text(previous_citations),
        "user_text_sha256": _sha256_text(text),
        "reasoning": FOOTNOTE_RESPONSE_REASONING,
        "max_output_tokens": FOOTNOTE_RESPONSE_MAX_OUTPUT_TOKENS,
        "text_format": {
            "type": "json_schema",
            "name": FOOTNOTE_RESPONSE_FORMAT_NAME,
            "strict": True,
            "schema_sha256": schema_fingerprint,
        },
        "required_part_fields": list(schema["properties"]["parts"]["items"]["required"]),
        "required_top_level_fields": list(schema["required"]),
    }


def _footnote_request_fingerprint(config: Dict[str, Any]) -> str:
    return hashlib.sha256(_stable_json(config).encode("utf-8")).hexdigest()


def _canonicalize_previous_citations_for_cache(previous_citations: str) -> str:
    text = str(previous_citations or "")
    return re.sub(
        r"(--> Link: ).*?( --> short_form:)",
        r"\1<link>\2",
        text,
    )


def _footnote_cache_lookup_fingerprint(config: Dict[str, Any], previous_citations: str) -> str:
    lookup_config = dict(config)
    lookup_config["previous_citations_sha256"] = _sha256_text(
        _canonicalize_previous_citations_for_cache(previous_citations)
    )
    lookup_config["previous_citations_cache_mode"] = "links-neutral-v1"
    return hashlib.sha256(_stable_json(lookup_config).encode("utf-8")).hexdigest()


def _footnote_cache_key(fingerprint: str) -> str:
    return hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()


def _llm_cache_dir() -> str:
    # All caches live together under one cache/ folder (next to the exe when
    # frozen, at the repo root from source).
    base = (os.path.dirname(os.path.abspath(sys.executable))
            if getattr(sys, "frozen", False)
            else os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, "cache", "llm")


def _cached_parts_from_payload(cached: Dict[str, Any]) -> Tuple[List[FootnotePart], str]:
    parts = [
        _normalize_footnote_part_link(FootnotePart(**p))
        for p in cached.get("parts", [])
        if (p.get("verbatim") or "").strip()
    ]
    return parts, _build_footnote_history_entries(parts)


def _load_footnote_cache_entry(
    cache_path: str,
    *,
    request_fingerprint: str,
    lookup_fingerprint: str,
) -> Optional[Tuple[List[FootnotePart], str, str]]:
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            cached = json.load(f)
        hit_kind = ""
        if cached.get("request_fingerprint") == request_fingerprint:
            hit_kind = "exact"
        elif cached.get("cache_lookup_fingerprint") == lookup_fingerprint:
            hit_kind = "link-neutral"
        else:
            cached_config = cached.get("request_config")
            cached_previous = cached.get("previous_citations", "")
            if isinstance(cached_config, dict) and (
                _footnote_cache_lookup_fingerprint(cached_config, cached_previous) == lookup_fingerprint
            ):
                hit_kind = "legacy-link-neutral"
        if not hit_kind:
            return None
        parts, history = _cached_parts_from_payload(cached)
        if not parts:
            return None
        return parts, history, hit_kind
    except Exception:
        return None


# Sidecar for the lookup index: maps cache filename -> [size, mtime_ns,
# fingerprint] so later runs only json-parse cache files that changed.
# Fingerprints derive purely from file content, so size+mtime validation
# is sound; a missing/corrupt sidecar just means one full rescan.
_LLM_CACHE_INDEX_SIDECAR = "_lookup_index_v1.json"
# Guards the one-time index scan/sidecar write and cache-entry writes:
# parallel documents share the cache directory, and two threads writing the
# same fingerprint file (or the sidecar tmp, which is keyed by pid alone)
# would tear it.
_LLM_CACHE_IO_LOCK = threading.Lock()


def _build_llm_cache_lookup_index(cache_dir: str) -> Dict[str, str]:
    with _LLM_CACHE_IO_LOCK:
        return _build_llm_cache_lookup_index_locked(cache_dir)


def _build_llm_cache_lookup_index_locked(cache_dir: str) -> Dict[str, str]:
    global _LLM_CACHE_LOOKUP_INDEX, _LLM_CACHE_LOOKUP_INDEX_DIR
    if _LLM_CACHE_LOOKUP_INDEX is not None and _LLM_CACHE_LOOKUP_INDEX_DIR == cache_dir:
        return _LLM_CACHE_LOOKUP_INDEX

    sidecar_path = os.path.join(cache_dir, _LLM_CACHE_INDEX_SIDECAR)
    known: Dict[str, list] = {}
    try:
        with open(sidecar_path, "r", encoding="utf-8") as f:
            loaded = json.load(f)
        if isinstance(loaded, dict):
            known = {
                k: v for k, v in loaded.items()
                if isinstance(v, list) and len(v) == 3
            }
    except Exception:
        known = {}

    index: Dict[str, str] = {}
    fresh: Dict[str, list] = {}
    dirty = False
    try:
        for name in os.listdir(cache_dir):
            if not name.endswith(".json") or name == _LLM_CACHE_INDEX_SIDECAR:
                continue
            path = os.path.join(cache_dir, name)
            try:
                st = os.stat(path)
            except Exception:
                continue
            entry = known.get(name)
            if entry and entry[0] == st.st_size and entry[1] == st.st_mtime_ns:
                lookup_fingerprint = entry[2]
            else:
                # New or changed file — parse it; failures are remembered
                # with an empty fingerprint so they aren't re-read forever.
                dirty = True
                lookup_fingerprint = ""
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        cached = json.load(f)
                    cached_config = cached.get("request_config")
                    if isinstance(cached_config, dict) and cached.get("parts"):
                        lookup_fingerprint = cached.get("cache_lookup_fingerprint") or _footnote_cache_lookup_fingerprint(
                            cached_config,
                            cached.get("previous_citations", ""),
                        )
                except Exception:
                    lookup_fingerprint = ""
            fresh[name] = [st.st_size, st.st_mtime_ns, lookup_fingerprint]
            if lookup_fingerprint:
                index.setdefault(lookup_fingerprint, path)
    except Exception:
        index = {}
        fresh = {}

    if fresh and (dirty or len(fresh) != len(known)):
        try:
            tmp = sidecar_path + f".tmp{os.getpid()}"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(fresh, f)
            os.replace(tmp, sidecar_path)
        except Exception:
            pass

    _LLM_CACHE_LOOKUP_INDEX = index
    _LLM_CACHE_LOOKUP_INDEX_DIR = cache_dir
    return index


def _parse_json_payload(raw_text: str) -> Optional[Dict[str, Any]]:
    if not isinstance(raw_text, str):
        return None
    candidate = raw_text.strip()
    if not candidate:
        return None
    if candidate.startswith("```"):
        candidate = re.sub(r"^```(?:json)?\s*", "", candidate, flags=re.IGNORECASE)
        candidate = re.sub(r"\s*```$", "", candidate)
    first = candidate.find("{")
    last = candidate.rfind("}")
    if first != -1 and last != -1 and last > first:
        candidate = candidate[first:last + 1]
    try:
        return json.loads(candidate)
    except Exception:
        return None


# --- Deterministic bare_citation derivation (fallback only) ---
# Used when the model returns an empty bare_citation. The field stays in the
# schema because removing it degrades the model's OTHER fields (same-tuple:
# corrected 88.3%->77.7%, all fields 92.9%->88.3%) — extracting the bare core
# evidently scaffolds the citation parsing. The derived value itself is safe:
# on the field's only production consumer (A2AJ lookup, case/unreported/
# statute/gazette kinds) it returned identical results to the model's on
# 113/113 benchmark parts.

_BARE_SUPRA_RE = re.compile(r"\b(supra\s+note\s+\d+|ibid)\b", re.IGNORECASE)
_BARE_CASE_CIT_RE = re.compile(
    r"(\[\d{4}\]\s|\(\d{4}\)[\s,]|\b[12]\d{3}\s+(?:CanLII|[A-Z]{2,7})\s+\d+"
    r"|\b\d+\s+US\s+\d+|\b\d+\s+F\s*\(?\d*[a-z]*\)?\s+\d+)"
)
_BARE_STATUTE_RE = re.compile(
    r"\b(?:RSC|RSO|RSA|RSS|RSM|RSQ|RSY|RSBC|RSNL|RSNB|RSNS|RSPEI|RSNWT|"
    r"SC|SO|SA|SS|SM|SQ|SY|SBC|SNL|SNB|SNS|SNWT|CQLR|CCSM)\b\s*[\s,]\s*\d{4}"
    r"|\bCQLR\b|\bCCSM\b"
)
_BARE_SIGNAL_RE = re.compile(
    r"^(?:but\s+see|see\s+e\.?g\.?,?|see\s+eg,?|see\s+also|see\s+further|"
    r"see\s+generally|see|contra|cf|compare|and|or)[,.]?\s+", re.IGNORECASE)
_BARE_TRAIL_SHORTFORM_RE = re.compile(r"\s*\[[^\[\]]{1,60}\]\s*\.?\s*$")
# trailing parenthetical that is a judge attribution or editorial commentary,
# not citation material like (HL), (Ont CA), (WL Can), (appeal denied, ...)
_BARE_JUDGE_PAREN_RE = re.compile(
    r"\s*\((?:[^()]*\b(?:J|JA|JJ|JJA|CJ|CJC|LC|LJ|dissenting|concurring|"
    r"discussing|suggesting|noting|holding|describing|emphasizing|quoting|"
    r"citing|stating|arguing|observing)\b[^()]*|Lord\s[^()]*|Lady\s[^()]*"
    r"|[“\"][^()]*)\)\s*\.?\s*$"
)
_BARE_SENTENCE_TAIL_RE = re.compile(r"\.\s+(?=[A-Z])")


def _derive_bare_citation(citation_with_style: str, kind: str) -> str:
    """Derive bare_citation from citation_with_style + kind: strip signals,
    short-form labels and judge/commentary parentheticals, then cut to the
    citation core for case/statute kinds."""
    s = (citation_with_style or "").strip()
    for _ in range(3):  # strip stacked signals
        s2 = _BARE_SIGNAL_RE.sub("", s)
        if s2 == s:
            break
        s = s2
    s = _BARE_TRAIL_SHORTFORM_RE.sub("", s).strip()
    for _ in range(2):
        s2 = _BARE_JUDGE_PAREN_RE.sub("", s).strip()
        if s2 == s:
            break
        s = s2

    def _final(x: str) -> str:
        x = _BARE_TRAIL_SHORTFORM_RE.sub("", x.strip())
        return x.rstrip(".").strip()

    if _BARE_SUPRA_RE.search(s):
        # keep the short-form prefix; the deterministic resolver and the
        # ref-chain origin fields handle these parts in production anyway
        return _final(s)
    if kind in ("case", "unreported"):
        m = _BARE_CASE_CIT_RE.search(s)
        if m:
            s = s[m.start():]
            # cut a trailing commentary sentence after the citation
            pieces = _BARE_SENTENCE_TAIL_RE.split(s, 1)
            if len(pieces) == 2 and len(pieces[1]) > 20 and not _BARE_CASE_CIT_RE.search(pieces[1]):
                s = pieces[0]
            return _final(s)
    if kind in ("statute", "gazette"):
        m = _BARE_STATUTE_RE.search(s)
        if m:
            return _final(s[m.start():])
    return _final(s)


def split_footnote_parts(
    text: str,
    previous_citations: str,
) -> Tuple[List[FootnotePart], str]:
    """
    Args:
        text: The footnote text to split.
        previous_citations: A string containing all previously processed citations for this doc.
    Returns:
        A tuple of (List of parts, String of new citation history to add).
    """
    if not isinstance(text, str) or not text.strip():
        return [], ""

    system_prompt = SYSTEM_INSTRUCTIONS
    response_schema = FOOTNOTE_SPLIT_SCHEMA
    prompt_fingerprint = _prompt_cache_fingerprint(system_prompt)
    current_system_prompt = system_prompt + previous_citations
    request_config = _footnote_request_config(
        system_prompt=system_prompt,
        prompt_fingerprint=prompt_fingerprint,
        text=text,
        previous_citations=previous_citations,
        schema=response_schema,
    )
    request_fingerprint = _footnote_request_fingerprint(request_config)
    lookup_fingerprint = _footnote_cache_lookup_fingerprint(request_config, previous_citations)
    cache_dir = _llm_cache_dir()
    cache_path = ""
    lookup_cache_path = ""

    # Check LLM cache if enabled
    if LLM_CACHE_ENABLED:
        cache_key = _footnote_cache_key(request_fingerprint)
        lookup_cache_key = _footnote_cache_key(lookup_fingerprint)
        cache_path = os.path.join(cache_dir, f"{cache_key}.json")
        lookup_cache_path = os.path.join(cache_dir, f"{lookup_cache_key}.json")
        candidate_paths = [cache_path]
        if lookup_cache_path != cache_path:
            candidate_paths.append(lookup_cache_path)
        for candidate_path in candidate_paths:
            if not os.path.exists(candidate_path):
                continue
            cached_entry = _load_footnote_cache_entry(
                candidate_path,
                request_fingerprint=request_fingerprint,
                lookup_fingerprint=lookup_fingerprint,
            )
            if cached_entry:
                parts, history, hit_kind = cached_entry
                suffix = "" if hit_kind == "exact" else f", {hit_kind}"
                _ts_print(f"  LLM cache hit ({len(parts)} parts{suffix})")
                return parts, history
        lookup_index = _build_llm_cache_lookup_index(cache_dir)
        indexed_path = lookup_index.get(lookup_fingerprint, "")
        if indexed_path and indexed_path not in candidate_paths and os.path.exists(indexed_path):
            cached_entry = _load_footnote_cache_entry(
                indexed_path,
                request_fingerprint=request_fingerprint,
                lookup_fingerprint=lookup_fingerprint,
            )
            if cached_entry:
                parts, history, hit_kind = cached_entry
                _ts_print(f"  LLM cache hit ({len(parts)} parts, {hit_kind})")
                return parts, history

    _pause_gate()

    response = _llm_call(
        model=LLM_MODEL,
        input=[
            {"role": "system", "content": current_system_prompt},
            {"role": "user", "content": text},
        ],
        reasoning=FOOTNOTE_RESPONSE_REASONING,
        max_output_tokens=FOOTNOTE_RESPONSE_MAX_OUTPUT_TOKENS,
        text={
            "format": {
                "type": "json_schema",
                "name": FOOTNOTE_RESPONSE_FORMAT_NAME,
                "strict": True,
                "schema": response_schema,
            }
        },
    )
    raw = response.output_text

    payload = _parse_json_payload(raw)
    if not payload or "parts" not in payload:
        _ts_print("Warning: GPT footnote splitter returned invalid JSON; falling back to unsplit footnote.")
        return [], ""

    parts: List[FootnotePart] = []
    new_history_entries = ""

    for item in payload["parts"]:
            verbatim = (item.get("verbatim") or "").strip()
            if not verbatim:
                # The model occasionally emits a degenerate empty part; keeping
                # it would produce a blank row in the exported workbook.
                _ts_print("  Skipping empty citation part from splitter output")
                continue
            corrected = (item.get("corrected") or "").strip()
            kind = (item.get("kind") or "other").strip()
            citation_with_style = (item.get("citation_with_style") or "").strip()
            if not citation_with_style:
                citation_with_style = verbatim
            bare_citation = (item.get("bare_citation") or "").strip()
            if not bare_citation:
                bare_citation = _derive_bare_citation(citation_with_style, kind) or verbatim
            link_candidate = (item.get("link") or "other").strip()
            pinpoint_fragments = _repair_decimal_section_fragments(
                item.get("pinpoint_fragments") or [], verbatim
            )
            page_pinpoints = item.get("page_pinpoints") or []
            pinpoint_fragments, page_pinpoints = _pinpoints_for_source_kind(
                kind, pinpoint_fragments, page_pinpoints
            )
            link = _resolve_footnote_part_link(
                verbatim=verbatim,
                citation_with_style=citation_with_style,
                kind=kind,
                link_candidate=link_candidate,
                pinpoint_fragments=pinpoint_fragments,
                bare_citation=bare_citation,
            )
            short_form = (item.get("short_form") or "").strip()

            part = FootnotePart(
                verbatim=verbatim,
                corrected=corrected,
                kind=kind,
                link=link,
                pinpoint_fragments=pinpoint_fragments,
                page_pinpoints=page_pinpoints,
                short_form=short_form,
                bare_citation=bare_citation,
                citation_with_style=citation_with_style,
                pre_provider_link=link_candidate,
            )
            parts.append(part)
            new_history_entries += _build_footnote_history_entries([part])

    # Save to LLM cache
    if LLM_CACHE_ENABLED:
        try:
            os.makedirs(cache_dir, exist_ok=True)
            cache_data = {
                "cache_version": FOOTNOTE_SPLIT_CACHE_VERSION,
                "request_config": request_config,
                "request_fingerprint": request_fingerprint,
                "cache_lookup_fingerprint": lookup_fingerprint,
                "prompt_mode": FOOTNOTE_PROMPT_CACHE_ID,
                "prompt_fingerprint": prompt_fingerprint,
                "model": LLM_MODEL,
                "footnote_text": text,
                "previous_citations": previous_citations,
                "response_payload": payload,
                "parts": [p.__dict__ if hasattr(p, "__dict__") else {"verbatim": p.verbatim, "corrected": p.corrected, "kind": p.kind, "link": p.link, "pinpoint_fragments": list(p.pinpoint_fragments), "page_pinpoints": list(p.page_pinpoints), "short_form": p.short_form, "bare_citation": p.bare_citation, "citation_with_style": p.citation_with_style} for p in parts],
                "history": new_history_entries,
            }
            with _LLM_CACHE_IO_LOCK:
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(cache_data, f, ensure_ascii=False)
                if lookup_cache_path and lookup_cache_path != cache_path:
                    with open(lookup_cache_path, "w", encoding="utf-8") as f:
                        json.dump(cache_data, f, ensure_ascii=False)
                if _LLM_CACHE_LOOKUP_INDEX is not None:
                    _LLM_CACHE_LOOKUP_INDEX[lookup_fingerprint] = lookup_cache_path or cache_path
        except Exception:
            pass

    return parts, new_history_entries



# -----------------------------
# Ibid / Supra resolution
# -----------------------------


def _is_ibid(s: str) -> bool:
    return bool(IBID_RE.search(s or ""))


def _extract_supra_note_number(s: str) -> Optional[int]:
    if not s:
        return None
    for rx in (SUPRA_NOTE_RE, SUPRA_N_RE, SUPRA_NN_RE):
        m = rx.search(s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
    return None


def _extract_supra_note_numbers(s: str) -> set[int]:
    if not s:
        return set()
    nums: set[int] = set()
    for rx in (SUPRA_NOTE_RE, SUPRA_N_RE, SUPRA_NN_RE):
        for m in rx.finditer(s):
            try:
                nums.add(int(m.group(1)))
            except Exception:
                continue
    return nums


def _previous_footnote_id(fid: int, footnote_map: Dict[int, str]) -> Optional[int]:
    previous = [candidate for candidate in footnote_map.keys() if int(candidate) < int(fid)]
    if not previous:
        return None
    return int(max(previous))


def _expand_footnote_ids_with_reference_targets(
    footnote_ids: set[int],
    footnote_map: Dict[int, str],
) -> Tuple[set[int], Dict[int, set[int]], Dict[int, set[int]]]:
    expanded = set(footnote_ids)
    supra_sources_by_target: Dict[int, set[int]] = {}
    ibid_sources_by_target: Dict[int, set[int]] = {}
    queue = list(footnote_ids)
    seen = set(queue)
    while queue:
        fid = queue.pop(0)
        text = footnote_map.get(fid, "")
        for target in sorted(_extract_supra_note_numbers(text)):
            if target not in footnote_map:
                continue
            supra_sources_by_target.setdefault(target, set()).add(fid)
            if target not in expanded:
                expanded.add(target)
                if target not in seen:
                    queue.append(target)
                    seen.add(target)
        if _is_ibid(text):
            target = _previous_footnote_id(fid, footnote_map)
            if target is not None and target in footnote_map:
                ibid_sources_by_target.setdefault(target, set()).add(fid)
                if target not in expanded:
                    expanded.add(target)
                    if target not in seen:
                        queue.append(target)
                        seen.add(target)
    return expanded, supra_sources_by_target, ibid_sources_by_target


def _extract_supra_hint(s: str) -> str:
    """Extract a short-form token that precedes 'supra' (e.g., 'Ipeelee' in 'Ipeelee, supra note 6').

    In aggressive mode (default), also searches globally when the anchored match fails,
    to catch mid-text supras like '... and gender considerations in Brown (SCC), supra note 2'.
    """
    if not s:
        return ""
    s2 = re.sub(r"^\s*(?:see\s+also\s+|see\s+|cf\.?\s+|but\s+see\s+|contra\s+|e\.g\.,\s+|eg\.,\s+|i\.e\.,\s+|ie\.,\s+)", "", s, flags=re.IGNORECASE)

    m = re.search(
        r"^\s*([^,;]+?)\s*,\s*supra\s+(?:note|n\.|nn\.)\s+\d+\b",
        s2,
        flags=re.IGNORECASE,
    )
    if m:
        hint = (m.group(1) or "").strip()
        hint = re.sub(r"^[\s\(\[\{\"\'“”]+", "", hint)
        hint = re.sub(r"[\s\)\]\}\"\'“”]+$", "", hint)
        hint = re.sub(r"\s+", " ", hint).strip()
        return hint

    if SUPRA_MODE == "aggressive":
        m = re.search(
            r"([^,;]+?)\s*,\s*supra\s+(?:note|n\.|nn\.)\s+\d+\b",
            s,
            flags=re.IGNORECASE,
        )
        if m:
            hint = (m.group(1) or "").strip()
            hint = re.sub(r"^[\s\(\[\{\"\'“”]+", "", hint)
            hint = re.sub(r"[\s\)\]\}\"\'“”]+$", "", hint)
            hint = re.sub(r"\s+", " ", hint).strip()
            if re.search(r"[A-Za-z]", hint):
                return hint

    return ""


def _choose_supra_target_part_index(
    target_footnote_id: int,
    citing_text: str,
    parts_by_footnote: Dict[int, List[Dict[str, Any]]],
    display_note_number: Optional[int] = None,
) -> Tuple[int, Optional[str]]:
    """Choose the intended citation-part within a multi-citation 'supra note N' target footnote.

    If the citing text includes a short form (e.g., 'Ipeelee, supra note 6'), prefer a target citation-part
    containing that token (including bracket labels like '[Ipeelee]' or the GPT-inferred short_form).
    Otherwise, default to part 1 and emit a warning when the target footnote has multiple citation-parts.
    """
    cands = list(parts_by_footnote.get(int(target_footnote_id), []) or [])
    if not cands:
        return (1, "Supra target footnote not found in this DOCX.")
    cands.sort(key=lambda p: int(p.get("citation_part_index") or 0))

    if len(cands) == 1:
        return (int(cands[0].get("citation_part_index") or 1), None)

    hint = _extract_supra_hint(citing_text or "")
    if not hint:
        return (int(cands[0].get("citation_part_index") or 1),
                "Supra target footnote has multiple references; could not infer intended citation-part (no short form). Defaulted to part 1.")

    hint_l = hint.lower()

    def score(p: Dict[str, Any], token: str) -> int:
        txt = (p.get("citation_part_text") or "").lower()
        if f"[{token}]" in txt:
            return 4
        sf = (p.get("short_form") or "").lower()
        if sf and sf == token:
            return 4
        if re.search(rf"\b{re.escape(token)}\b", txt):
            return 2
        if token in txt:
            return 1
        return 0

    best = max(cands, key=lambda p: score(p, hint_l))
    if score(best, hint_l) > 0:
        return (int(best.get("citation_part_index") or 1), None)

    # Fallback: use the last word of the hint (often an author surname or case short name)
    last = hint_l.split()[-1]
    best2 = max(cands, key=lambda p: score(p, last))
    if score(best2, last) > 0:
        return (int(best2.get("citation_part_index") or 1), None)

    note_label = display_note_number if display_note_number is not None else target_footnote_id
    available = "; ".join(
        f"part {p.get('citation_part_index')}='{p.get('short_form') or p.get('citation_part_text', '')[:80]}'"
        for p in cands
    )
    return (
        int(cands[0].get("citation_part_index") or 1),
        f"Supra short form ('{hint}') did not match any citation-part in note {note_label}. Available: {available}. Defaulted to part 1.",
    )


def compute_footnote_order(anchors: List[Dict[str, Any]], footnote_map: Dict[int, str]) -> List[int]:
    """Prefer first-appearance order from anchors; append any remaining footnotes numerically."""
    order: List[int] = []
    seen: set[int] = set()
    for a in sorted(anchors, key=lambda x: x["global_pos"]):
        fid = a.get("footnote_id")
        if isinstance(fid, int) and fid > 0 and fid in footnote_map and fid not in seen:
            seen.add(fid)
            order.append(fid)
    for fid in sorted(footnote_map.keys()):
        if fid not in seen:
            order.append(fid)
    return order


# ---------------------------------------------------------------------------
# Reference-link pipeline (M1b): a single accumulated-history pass feeds BOTH
# the user-facing parts and the supra/ibid registry; reference links are then
# resolved deterministically (registry det chain + windowed guard +
# candidate-list disambig for abstains, drop on abstain). Held-out benchmark
# (8 docs, n=423): supra exact 97.6%, wrong 7, hallucinated 0 — vs the
# legacy app's 84.2%, wrong 13, hallucinated 3 — at ~1.07x its token cost
# (the disambig calls are tiny). Rejected alternates from the 2026-07
# benchmark campaign (parallel empty-context splitting, dual-pass
# registries, dual-prompt stitch, history re-split fallback) were removed;
# see the repo history and _temp/ benchmark corpora if they are needed
# again.
# ---------------------------------------------------------------------------

DETERMINISTIC_REF_LINKS = True
# Candidate-list disambiguation for resolver abstains (A8b, revived by A30).
# Candidates are tagged with
# the footnote number where they appeared ([fn N]) — that plus a drift-aware
# system prompt makes bare "Supra note N." references resolvable, which the
# old numberless history never allowed. Accepted links pass the windowed
# note-number guard. Benchmark (3 docs, n=383, on top of the det chain):
# 93.2% exact, wrong 6, hallucinated 0 — vs the re-split fallback's 93.2%,
# wrong 8, hallucinated 1, at ~1/40th the model-stage cost.
REF_DISAMBIG_FALLBACK = True
# Drop the link when both the det chain and the disambig fallback abstain on
# a supra reference, instead of keeping the split model's own (history-drift-
# prone) link. Wrong links are worse than dropped links: keeping the model
# link benchmarked at +9 exact / +7 wrong over dropping (evaluation n=364 wrong 15
# vs 9); dropping wins under that rule. Registry entries inherit the drop, so
# unresolved origins cascade to drops (safe direction), never to wrong links.
REF_DROP_ON_ABSTAIN = True
# Safe uses only the strict supra resolver. Aggressive runs two benchmarked
# fallbacks after strict abstention: a bare note number that identifies exactly
# one earlier citation, and a named short form inferred from an earlier full
# citation. Inferred forms are kept separate from author-defined short forms.
SUPRA_LINKING_AGGRESSIVENESS = "safe"
# A2AJ statute-slug canonicalization (canlii_slug_repair.py): CanLII never
# redirects a near-miss statute slug, and the slug derives mechanically from
# the canonical citation, which the free A2AJ API returns for FED/ON/BC
# consolidated statutes. Repair-only (a miss never drops or rewrites), runs
# before the supra registry is built so references inherit repaired links.
# Live-verified dead-emission triage (07-09): fixes the top dead families —
# enacted-as-section (sc-2019-c-28 -> -s-1, x5 + inherited refs), missing
# supplement (rsc-1985-c-1 -> -2nd-supp), chapter-dot spelling
# (rso-1990-c-f-3 -> -f3).
A2AJ_STATUTE_LINK_REPAIR = True
# Pure-reference pre-filter: a footnote whose ENTIRE text is mechanical
# supra/ibid clauses skips the split model — parts are synthesized
# deterministically and Phase 2's resolver chain assigns the links. Grammar
# matched the model's partition on all 751 footnotes it classified (31.0% of
# the reference runs, zero non-reference parts). With the strict leading-ibid
# gate (_pref_allow_leading_ibid) it skips 20.8% of footnotes and replayed
# BIT-IDENTICAL to the full-model pipeline: same accepted evaluation set, same
# resolution methods, zero link changes corpus-wide. OFF until the Economy
# mode is wired (pre-wiring test program).
PURE_REF_PREFILTER = False
# Ultra Economy replaces the expensive splitter call only when deterministic
# boundaries and the complete output tuple are available. Free mode never
# calls an LLM and uses the recall-first splitter for every footnote.
DETERMINISTIC_SOURCE_SPLITTER = False
FREE_NO_LLM = False
# Run modes (validated by the 2026-07 pre-wiring test program + two-doc live
# validation):
#   high_accuracy — the M1b pipeline as-is (default).
#   economy      — identical pipeline with PURE_REF_PREFILTER on: pure-reference
#                  footnotes skip the split model (~0.85x cost; replayed
#                  bit-identical on gold, live diffs limited to fresh-draw
#                  origin variance in the safe drop direction).
#   ultra_economy: Economy plus benchmark-gated deterministic full-tuple
#                  substitutions; ordinary splitter calls remain the fallback.
#   free: deterministic parts with a lossless one-part fallback and no LLM
#         client or reference-disambiguation calls.
RUN_MODE = "high_accuracy"
VALID_RUN_MODES = ("high_accuracy", "economy", "ultra_economy", "free")
SNAP_VERBATIM_ENABLED = True

_SNAP_CHAR_TRANS = {"‘": "'", "’": "'", "“": '"', "”": '"', "–": "-", "—": "-"}


def _snap_norm_with_map(s: str) -> Tuple[str, List[int]]:
    """Normalized text (lowercase, unified quotes/dashes, collapsed whitespace)
    plus a map from normalized index back to the original string index."""
    out: List[str] = []
    idx: List[int] = []
    prev_space = True
    for i, ch in enumerate(s):
        if ch.isspace():
            if prev_space:
                continue
            out.append(" ")
            idx.append(i)
            prev_space = True
        else:
            out.append(_SNAP_CHAR_TRANS.get(ch, ch).lower())
            idx.append(i)
            prev_space = False
    while out and out[-1] == " ":
        out.pop()
        idx.pop()
    return "".join(out), idx


def _snap_verbatim_parts(footnote_text: str, split: List["FootnotePart"]) -> List["FootnotePart"]:
    """Re-align each part's verbatim to the exact footnote substring.

    The model occasionally drops characters (trailing commas), normalizes
    punctuation, or duplicates text across parts. Downstream consumers (supra
    hints, history lines, quote anchoring) want faithful substrings, so snap
    every verbatim back onto the source text when an order-preserving
    alignment exists; otherwise leave the parts untouched.
    """
    if not SNAP_VERBATIM_ENABLED or not footnote_text or not split:
        return split
    fn_norm, fn_idx = _snap_norm_with_map(footnote_text)
    spans: List[List[int]] = []
    cursor = 0
    for part in split:
        p_norm, _ = _snap_norm_with_map(part.verbatim or "")
        p_norm = p_norm.strip()
        if not p_norm:
            return split
        pos = fn_norm.find(p_norm, cursor)
        if pos < 0:
            # Overlap duplication: the span may start before the cursor.
            pos = fn_norm.find(p_norm)
            if pos < 0:
                return split
        spans.append([pos, pos + len(p_norm)])
        cursor = max(cursor, pos)
    # Overlapping spans: truncate the earlier part at the later one's start.
    for i in range(len(spans) - 1):
        if spans[i][1] > spans[i + 1][0]:
            spans[i][1] = spans[i + 1][0]
            if spans[i][1] <= spans[i][0]:
                return split
    # A comma immediately after a span belongs to it (authors split there).
    for sp in spans:
        while sp[1] < len(fn_norm) and fn_norm[sp[1]] == ",":
            sp[1] += 1
    repaired: List["FootnotePart"] = []
    for part, (a, b) in zip(split, spans):
        snapped = footnote_text[fn_idx[a]:fn_idx[b - 1] + 1].strip()
        repaired.append(part if snapped == part.verbatim else replace(part, verbatim=snapped))
    return repaired


def _apply_author_provided_links(
    footnote_text: str,
    split: List["FootnotePart"],
    source_links: Optional[List[Dict[str, Any]]],
) -> List["FootnotePart"]:
    """Bind source-document URLs only to the split part whose text contains them."""
    if not footnote_text or not split or not source_links:
        return split

    part_spans: List[Tuple[int, int]] = []
    cursor = 0
    for part in split:
        start = footnote_text.find(part.verbatim, cursor)
        if start < 0:
            start = footnote_text.find(part.verbatim)
        if start < 0:
            part_spans.append((-1, -1))
            continue
        end = start + len(part.verbatim)
        part_spans.append((start, end))
        cursor = end

    targets_by_part: Dict[int, List[str]] = {}
    for item in source_links:
        try:
            link_start, link_end = int(item["start"]), int(item["end"])
        except (KeyError, TypeError, ValueError):
            continue
        target = str(item.get("target") or "").strip()
        if not target:
            continue
        for index, (part_start, part_end) in enumerate(part_spans):
            if part_start >= 0 and max(part_start, link_start) < min(part_end, link_end):
                bucket = targets_by_part.setdefault(index, [])
                if target not in bucket:
                    bucket.append(target)
                break

    repaired = list(split)
    for index, targets in targets_by_part.items():
        part = repaired[index]
        author_link = targets[0]
        working_link = _resolve_footnote_part_link(
            verbatim=part.verbatim,
            citation_with_style=part.citation_with_style,
            kind=part.kind,
            link_candidate=author_link,
            pinpoint_fragments=part.pinpoint_fragments,
            bare_citation=part.bare_citation,
        )
        repaired[index] = replace(
            part,
            link=working_link,
            author_provided_link=author_link,
            author_provided_links=targets,
            pre_provider_link=author_link,
        )
    return repaired

_REF_HINT_STOPWORDS = {
    "see", "also", "but", "and", "the", "note", "supra", "ibid",
    "generally", "contra", "compare", "with", "above", "e.g", "eg",
}
_REF_SIGNAL_PREFIX_RE = re.compile(
    r"^(?:see(?:,?\s+e\.?g\.?,?)?(?:\s+also)?|but\s+see|contra|compare|cf\.?|see\s+generally)\s+",
    re.IGNORECASE,
)
_REF_PAR_PIN_RE = re.compile(r"(?:\bparas?\.?|¶)\s*(\d+)", re.IGNORECASE)
_REF_PROVISION_PIN_RE = re.compile(
    r"(?:(?:\b(?:sub)?sections?|\bss?\.?)\s*"
    r"(?P<section>\d+(?:\.\d+)*)|"
    r"(?:\brules?|\brr?\.?)\s*(?P<rule>\d+(?:[.-]\d+)*))|"
    r"(?:\barticles?|\barts?\.?)\s*(?P<article>\d+(?:\.\d+)*)",
    re.IGNORECASE,
)
_REF_SCOPE_SPLIT_RE = re.compile(
    r"\b(?:quot(?:ing|ed)|cit(?:ing|ed)|amend(?:ing|ed)|aff'?d|rev'?d)\b")
_REF_TOKEN_RE = re.compile(r"\b(?:supra|ibid)\b", re.IGNORECASE)
_REF_SUPRA_ANY_RE = re.compile(r"\bsupra\b", re.IGNORECASE)


# --- Pure-reference pre-filter (PURE_REF_PREFILTER) -------------------------
# Grammar for footnotes that are NOTHING but mechanical reference clauses:
# optional signal, "ibid" or "[Name,] supra [note N]", optional pinpoint.
# Anything the grammar can't fully consume (subsection parens, bracketed
# notes, "quoting …" tails, blank draft note numbers) goes to the model.
_PREF_NUMLIST = r"\d+(?:\.\d+)?[a-z]?(?:\s*[-–]\s*\d+(?:\.\d+)?[a-z]?)?"
_PREF_NUMSEQ = rf"{_PREF_NUMLIST}(?:\s*(?:,|and|&)\s*{_PREF_NUMLIST})*"
_PREF_PROVISION_LABEL = r"(?:ss?\.?|sections?|arts?\.?|articles?)"
_PREF_RULE_LABEL = r"(?:rr?\.?|rules?)"
_PREF_RULE_NUM = r"\d+(?:\.\d+){0,3}[a-z]?"
_PREF_RULE_NUMSEQ = rf"{_PREF_RULE_NUM}(?:\s*(?:,|and|&)\s*{_PREF_RULE_NUM})*"
_PREF_PIN = (
    rf"(?:at\s+(?:(?:{_PREF_RULE_LABEL})\s+{_PREF_RULE_NUMSEQ}"
    rf"|(?:(?:paras?\.?|pp?\.?|pages?|{_PREF_PROVISION_LABEL})\s+)?{_PREF_NUMSEQ})(?:ff)?"
    rf"|(?:paras?\.?|{_PREF_PROVISION_LABEL})\s+{_PREF_NUMSEQ}(?:ff)?"
    rf"|(?:{_PREF_RULE_LABEL})\s+{_PREF_RULE_NUMSEQ}(?:ff)?)"
)
_PREF_CLAUSE_RE = re.compile(
    r"^\s*"
    r"(?:(?:see(?:,?\s+e\.?g\.?,?)?(?:\s+also)?|but\s+see|contra|compare|cf\.?|see\s+generally)\s+)?"
    r"(?:ibid\.?|(?:(?P<name>[^,;.]{1,60})\s*,\s*)?supra(?:\s+(?:note|nn?\.?)\s+\d+)?)"
    rf"(?:\s*,)?(?:\s+(?P<pin>{_PREF_PIN}))?\s*[.;]?\s*$",
    re.IGNORECASE,
)
_PREF_NUM_RE = re.compile(r"(\d+(?:\.\d+)?)[a-z]?(?:\s*[-–]\s*(\d+(?:\.\d+)?)[a-z]?)?")


def _pref_pin_numbers(pin: str) -> List[str]:
    """Range endpoints in a pinpoint sequence, with short second endpoints
    completed by prefix ("245-46" -> 245, 246)."""
    out: List[str] = []
    for m in _PREF_NUM_RE.finditer(pin or ""):
        first, second = m.group(1), m.group(2)
        out.append(first)
        if second:
            f_int, s_int = first.split(".")[0], second.split(".")[0]
            if "." not in second and len(s_int) < len(f_int) and int(s_int) < int(f_int):
                second = f_int[: len(f_int) - len(s_int)] + s_int
            out.append(second)
    return out


def _pref_make_part(clause: str, m: "re.Match[str]") -> "FootnotePart":
    pin = m.group("pin") or ""
    pin_head = pin.lower().lstrip()
    fragments: List[str] = []
    pages: List[int] = []
    nums = _pref_pin_numbers(pin)
    if re.match(r"(?:at\s+)?paras?\b", pin_head):
        fragments = [f"par{n}" for n in nums if "." not in n]
    elif re.match(
        r"(?:at\s+)?(?:s\b|ss\b|s\.|ss\.|sections?\b|arts?\.?\b|"
        r"articles?\b|rr?\.?\b|rules?\b)",
        pin_head,
    ):
        fragments = [f"sec{n}" for n in nums]
    elif re.match(r"at\b", pin_head):
        pages = [int(n) for n in nums if "." not in n]
    return FootnotePart(
        verbatim=clause,
        corrected=clause,
        kind="other",
        link="",
        pinpoint_fragments=fragments,
        bare_citation=clause,
        citation_with_style=clause,
        short_form=(m.group("name") or "").strip(),
        page_pinpoints=pages,
    )


def _prefilter_pure_ref_parts(
    footnote_text: str, allow_leading_ibid: bool = True
) -> Optional[List["FootnotePart"]]:
    """Deterministic parts for a footnote that is purely supra/ibid clauses,
    or None when any clause falls outside the grammar (model handles it).
    Links are left empty for Phase 2's resolver chain; on the reference runs
    the clause partition matched the model's on all 751 classified footnotes.

    allow_leading_ibid=False sends footnotes whose FIRST clause is an ibid to
    the model: an ibid whose registry predecessor is unlinked has no
    deterministic origin, and only the model's sequential-history judgment
    recovers those (the caller passes False when the previous footnote's
    split ended unlinked)."""
    text = re.sub(r"\s+", " ", footnote_text or "").strip()
    if not text or len(text) > 400 or not _REF_TOKEN_RE.search(text):
        return None
    clauses = [c.strip() for c in text.split(";")]
    if not (1 <= len(clauses) <= 4) or not all(clauses):
        return None
    if not allow_leading_ibid and _is_ibid(clauses[0]):
        return None
    parts: List[FootnotePart] = []
    for clause in clauses:
        m = _PREF_CLAUSE_RE.match(clause)
        if not m:
            return None
        parts.append(_pref_make_part(clause, m))
    return _snap_verbatim_parts(footnote_text, parts)


def _pref_allow_leading_ibid(prev_split: Optional[List["FootnotePart"]]) -> bool:
    """Whether the next footnote's leading ibid has a deterministic origin:
    the previous footnote's Phase-1 split ended with a LINKED part, which the
    registry hands straight to the ibid. When it ended unlinked — including a
    prefiltered predecessor, whose links only arrive in Phase 2 — only the
    model's sequential-history judgment ever linked these, so the footnote
    goes to the model. This strict form benchmarked bit-identical to the
    full-model pipeline (accepted evaluation set, all resolution methods, zero
    link changes corpus-wide) at a 20.8% footnote skip rate; allowing
    prefiltered predecessors reached 31.0% but traded away live ibid links
    through drop cascades."""
    if not prev_split:
        return False
    previous = prev_split[-1]
    last_link = (
        previous.pre_provider_link
        if previous.pre_provider_link is not None
        else previous.link
    ).strip()
    return bool(last_link) and last_link.lower() != "other"


def _deterministic_footnote_parts(
    footnote_text: str, *, allow_unsplit_fallback: bool = False,
    allow_reference_parts: bool = True,
    recall_first: bool = False,
) -> Optional[Tuple[List["FootnotePart"], Tuple[str, ...]]]:
    """Build the model's full tuple, or abstain when the selected mode requires it."""
    result = (
        _deterministic_splitter.split_footnote_recall_first(footnote_text)
        if allow_unsplit_fallback or recall_first
        else _deterministic_splitter.split_footnote(footnote_text)
    )
    if result.status == "deterministic_complete":
        source_parts = list(result.parts)
        fields = [_deterministic_splitter.extract_fields(part) for part in source_parts]
    else:
        return None
    if not allow_reference_parts and any(
        _detect_ref_kind(source.text) for source in source_parts
    ):
        return None
    if not allow_unsplit_fallback and any(field.status != "complete" for field in fields):
        return None

    parts: List[FootnotePart] = []
    for source, item in zip(source_parts, fields):
        verbatim = source.text
        pinpoint_fragments, page_pinpoints = _pinpoints_for_source_kind(
            item.kind, list(item.pinpoint_fragments), list(item.page_pinpoints)
        )
        citation_with_style = item.citation_with_style or verbatim
        bare_citation = (
            _derive_bare_citation(citation_with_style, item.kind)
            or item.bare_citation
            or verbatim
        )
        link_candidate = item.link_candidate
        if _detect_ref_kind(verbatim):
            link_candidate = ""
        link = _resolve_footnote_part_link(
            verbatim=verbatim,
            citation_with_style=citation_with_style,
            kind=item.kind,
            link_candidate=link_candidate,
            pinpoint_fragments=pinpoint_fragments,
            bare_citation=bare_citation,
        )
        parts.append(FootnotePart(
            verbatim=verbatim,
            corrected=item.corrected or verbatim,
            kind=item.kind,
            link=link,
            pinpoint_fragments=pinpoint_fragments,
            page_pinpoints=page_pinpoints,
            short_form=item.short_form,
            bare_citation=bare_citation,
            citation_with_style=citation_with_style,
            pre_provider_link=link,
        ))
    if not allow_unsplit_fallback and any(
        part.kind.strip().lower() in {"case", "unreported", "statute"}
        and (not part.link.strip() or part.link.strip().lower() == "other")
        for part in parts
    ):
        # Linkable authorities need an identity that later references can inherit.
        return None
    return _snap_verbatim_parts(footnote_text, parts), result.reasons


def _norm_ref_text(value: Any) -> str:
    s = re.sub(r"\s+", " ", str(value or "")).strip()
    s = s.replace("‘", "'").replace("’", "'").replace("“", '"').replace("”", '"')
    s = s.replace("–", "-").replace("—", "-")
    return s.lower()


def _ref_base_url(link: str) -> str:
    return (link or "").split("#", 1)[0].strip()


def _detect_ref_kind(part_verbatim: str) -> str:
    """"ibid" / "supra" / "" — mirrors the benchmark's detect_reference."""
    if _is_ibid(part_verbatim):
        return "ibid"
    if _extract_supra_note_number(part_verbatim) is not None:
        return "supra"
    if _REF_SUPRA_ANY_RE.search(part_verbatim or ""):
        return "supra"
    return ""


def _reanchor_ref_link(origin_link: str, part_text: str) -> str:
    """Re-anchor an inherited supra/ibid link to THIS part's own first pinpoint.

    The model (and the registry) hand back the origin's link, which may carry
    the origin's #par/#sec anchor; the reference part must instead point at its
    own pinpoint. A SUPRA part with no pinpoint of its own gets the BARE base
    URL — supra refers to the work, never the origin's pinpoint. A bare IBID
    keeps the origin link verbatim (ibid without a pinpoint means the same
    pinpoint as the preceding citation). Pinpoints after a quoting/citing
    marker belong to an embedded secondary citation and are ignored.
    """
    link = re.sub(r"\s+", " ", str(origin_link or "")).strip()
    if not link or link.lower() == "other" or "canlii.org" not in link.lower():
        return link
    base = _ref_base_url(link)
    # The reference's own pinpoint lives in the same connective segment as
    # the supra/ibid token; pinpoints in other segments belong to embedded
    # secondary citations (quoting/citing/amending/aff'd/rev'd).
    segs = _REF_SCOPE_SPLIT_RE.split(part_text or "")
    scope = next((s for s in segs if _REF_TOKEN_RE.search(s)),
                 segs[0] if segs else "")
    par_m = _REF_PAR_PIN_RE.search(scope)
    provision_m = _REF_PROVISION_PIN_RE.search(scope)
    if "/doc/" in base.lower() and par_m:
        return f"{base}#par{par_m.group(1)}"
    if "/laws/" in base.lower() and provision_m:
        number = (
            provision_m.group("section")
            or provision_m.group("rule")
            or provision_m.group("article")
        )
        return f"{base}#sec{number}"
    if _detect_ref_kind(part_text or "") == "supra":
        return base
    return link


def _ref_fallback_supra_hint(verbatim: str) -> str:
    """Hint for supra forms _extract_supra_hint can't parse (e.g. drafts with
    'supra note ___'): the text immediately before 'supra', minus signals."""
    v = verbatim or ""
    m = _REF_SUPRA_ANY_RE.search(v)
    if not m:
        return ""
    pre = v[: m.start()].strip()
    while True:
        stripped = _REF_SIGNAL_PREFIX_RE.sub("", pre).strip()
        if stripped == pre:
            break
        pre = stripped
    return pre.strip(" ,;:. ")[-80:]


def _ref_hint_tokens(hint: str) -> List[str]:
    hint = _norm_ref_text(hint).strip("[]()")
    return [t for t in re.split(r"[^a-z0-9]+", hint)
            if len(t) >= 3 and t not in _REF_HINT_STOPWORDS]


_REF_REGISTRY_MATCH_HEAD_CHARS = 200


def _ref_registry_match_text(entry: Dict[str, str]) -> str:
    """The text a supra hint is allowed to match: the short form plus the
    head of the verbatim. Authors reference works by their names and defined
    short forms, which sit at the head of a citation; prose trailing a long
    citation ("... which proposed the first version of what became the
    Supply Chains Act") must not make an entry matchable."""
    return (
        (entry.get("short_form") or "")
        + " "
        + (entry.get("verbatim") or "")[:_REF_REGISTRY_MATCH_HEAD_CHARS]
    )


_REF_BRACKET_RE = re.compile(r"\[([^\[\]]{2,60})\]")
_REF_BRACKET_SKIP_RE = re.compile(
    r"^(sic|emphasis (added|in original)|citations? omitted|footnotes? omitted|"
    r"translated by.*|translation|italics.*|underline.*|ellipsis.*|\d{4}|\d+)$",
    re.IGNORECASE,
)


def _resolve_supra_from_registry(
    verbatim: str,
    registry: List[Dict[str, str]],
) -> Tuple[str, str]:
    """Resolve a supra reference against previously emitted parts.

    Tier 0 (note number, A18/A24): when the reference says "supra note N" and
    a LINKED entry from footnote N matches every hint token, take it. The
    double key (number AND name) keeps it safe against authors' wrong note
    numbers (~29% of checkable references in the corpus): a drifted N either
    finds no hint match (abstain, falls through) or finds the cited work
    anyway. Exact N only — widening the lookup to N±3 benchmarked as a net
    regression (asserting from a neighbor picks wrong works).

    Then: exact short-form match, all-hint-tokens-in-short-form,
    all-hint-tokens-in-verbatim — over LINKED entries only. An unlinked entry
    can never satisfy a link resolution, and keeping unlinked entries in the
    pools let one dropped reference "resolve" every later same-hint supra to
    "other", silently suppressing the fallback (one early drop then poisons
    the whole document). A tier emptied by the linked filter falls through to
    the next; a non-empty tier resolves only when every match points at the
    same work (same base link); otherwise abstain — precision over recall,
    the fallback handles the rest. Returns (link, method); link == "" means
    abstain.
    """
    hint = _extract_supra_hint(verbatim) or _ref_fallback_supra_hint(verbatim)
    hint_norm = _norm_ref_text(hint).strip("[]() ")
    tokens = _ref_hint_tokens(hint)
    if not hint_norm and not tokens:
        return "", "abstain_no_hint"
    linked = [
        e for e in registry
        if str(e.get("link") or "").strip()
        and str(e.get("link") or "").strip().lower() != "other"
    ]
    if tokens:
        note_n = _extract_supra_note_number(verbatim)
        if note_n is not None:
            for e in linked:
                if e.get("note") != str(note_n):
                    continue
                blob = _norm_ref_text(_ref_registry_match_text(e))
                if all(t in blob for t in tokens):
                    return e.get("link", ""), "note_number"
            # When an UNLINKED entry from footnote N matches every hint
            # token, the note number is verifiably correct and the cited
            # work is known — it just has no link. Borrowing a link from a
            # different footnote whose text happens to share the hint
            # tokens would assert a different work (an unlinked statute
            # once pulled every "supra note N" onto an unrelated bill's
            # page), so abstain instead. The pools below remain reachable
            # only for drifted note numbers, where no note-N entry matches.
            for e in registry:
                if e.get("note") != str(note_n):
                    continue
                blob = _norm_ref_text(_ref_registry_match_text(e))
                if all(t in blob for t in tokens):
                    return "", "abstain_unlinked_target"
    pools: List[Tuple[str, List[Dict[str, str]]]] = []
    if hint_norm:
        pools.append(("exact_sf", [
            e for e in linked
            if _norm_ref_text(e.get("short_form", "")).strip("[]() ") == hint_norm
        ]))
    if tokens:
        pools.append(("token_sf", [
            e for e in linked
            if e.get("short_form") and all(t in _norm_ref_text(e["short_form"]) for t in tokens)
        ]))
        pools.append(("token_verb", [
            e for e in linked
            if e.get("verbatim")
            and all(t in _norm_ref_text(_ref_registry_match_text(e)) for t in tokens)
        ]))
    abstain_method = "abstain_no_match"
    for method, pool in pools:
        if not pool:
            continue
        groups: Dict[str, List[Dict[str, str]]] = {}
        for e in pool:
            link = re.sub(r"\s+", " ", str(e.get("link") or "")).strip()
            groups.setdefault(_ref_base_url(link).lower(), []).append(e)
        if len(groups) == 1:
            return next(iter(groups.values()))[0].get("link", ""), method
        abstain_method = f"abstain_ambiguous_{method}"
        break
    # Bracket-definition tier (A28, display resolution only, last before the
    # model): authors DEFINE short forms in brackets at the origin citation
    # ("R v Brown, 2022 SCC 18 [Brown (SCC)]"), so a hint that exactly matches
    # a bracketed segment in a linked prior part's RAW verbatim resolves to
    # that part — independent of the note number and of the model's
    # short_form field. Abstains when the same bracket text points at
    # different works. Benchmarked ahead of the standard tiers it picks wrong
    # neighbors; as the last tier it only shrinks the model's workload.
    if hint_norm:
        bases = set()
        best = ""
        for e in linked:
            for br in _REF_BRACKET_RE.findall(e.get("verbatim") or ""):
                if _REF_BRACKET_SKIP_RE.match(br.strip()):
                    continue
                for piece in br.split(";"):
                    piece = re.sub(r"^\s*hereinafter\s+", "", piece, flags=re.IGNORECASE)
                    if _norm_ref_text(piece).strip("[]() ") == hint_norm:
                        link = re.sub(r"\s+", " ", str(e.get("link") or "")).strip()
                        bases.add(_ref_base_url(link).lower())
                        best = e.get("link", "")
        if len(bases) == 1:
            return best, "bracket_definition"
    return "", abstain_method


REF_NOTE_GUARD_WINDOW = 3


def _ref_fallback_link_guard_ok(
    verbatim: str,
    link: str,
    registry: List[Dict[str, str]],
) -> bool:
    """Windowed note-number veto for fallback model links (A22/A24).

    When the reference names a note number and the registry holds linked
    entries within ±REF_NOTE_GUARD_WINDOW footnotes of it, the fallback link
    must point at one of those works (same base URL); otherwise reject it and
    keep the model's own link. The window — rather than exact N — absorbs
    authors' wrong note numbers, which an exact check turns into vetoes of
    correct recoveries. Benchmarked: converts only genuinely wrong fallback
    links into drops (wrong 10→8, hallucinated 2→1, exact count unchanged).
    """
    note_n = _extract_supra_note_number(verbatim)
    if note_n is None:
        return True
    notes = {
        str(note_n + s)
        for s in range(-REF_NOTE_GUARD_WINDOW, REF_NOTE_GUARD_WINDOW + 1)
    }
    pool_bases = {
        _ref_base_url(re.sub(r"\s+", " ", str(e.get("link") or "")).strip()).lower()
        for e in registry
        if e.get("note") in notes
        and str(e.get("link") or "").strip()
        and str(e.get("link") or "").strip().lower() != "other"
    }
    if not pool_bases:
        return True
    return _ref_base_url(re.sub(r"\s+", " ", str(link or "")).strip()).lower() in pool_bases


def _registry_history_text(registry: List[Dict[str, str]]) -> str:
    """Render registry entries in the exact production history-line format."""
    return "".join(
        "- Citation: " + (e.get("verbatim") or "")
        + " --> Link: " + (e.get("link") or "")
        + " --> short_form: " + ((e.get("short_form") or "").strip() or "N/A") + "\n"
        for e in registry
    )


REF_DISAMBIG_MAX_CANDIDATES = 8

REF_DISAMBIG_SYSTEM = """You resolve supra/ibid references in Canadian legal footnotes.
You are given one reference part and a numbered list of candidate prior
citations from the same document (footnote number where it appeared, short
form, citation text, link). Authors' "supra note N" numbers are often wrong
by one to three; treat the name/subject match as primary and the footnote
number as a tie-breaker. When body text the footnote is attached to is
shown, it usually names the work the reference points to. Pick the candidate
the reference points to. If none of them is the referenced work, or the
referenced work has no usable link, answer 0.
Answer with JSON: {"choice": <number>} where 0 means none."""

_REF_DISAMBIG_SCHEMA = {
    "type": "object",
    "properties": {"choice": {"type": "integer"}},
    "required": ["choice"],
    "additionalProperties": False,
}

_REF_DISAMBIG_TOKEN_STOPWORDS = {
    "supra", "note", "notes", "ibid", "at", "para", "paras", "the", "of",
    "and", "in", "see", "also", "v", "r", "c", "et", "al", "this", "that",
    "a", "an", "to", "for", "is", "was", "on", "by", "with", "as", "or",
}


def _ref_disambig_tokens(text: str) -> set:
    return {t for t in re.findall(r"[a-z][a-z'’-]+", _norm_ref_text(text))
            if t not in _REF_DISAMBIG_TOKEN_STOPWORDS}


def _ref_disambig_candidates(
    part_verbatim: str,
    registry: List[Dict[str, str]],
    prop_text: str,
) -> Tuple[List[Dict[str, str]], str]:
    """Rank registry entries for one abstained reference. Returns the top
    candidates and the proposition text to show the model (only when the
    reference itself names nothing, e.g. "Supra note 125.")."""
    hint = _extract_supra_hint(part_verbatim) or _ref_fallback_supra_hint(part_verbatim) or ""
    has_hint = bool(_ref_disambig_tokens(hint))
    q = _ref_disambig_tokens(hint) or _ref_disambig_tokens(part_verbatim)
    q_prop = set() if has_hint else _ref_disambig_tokens(prop_text)
    hint_norm = _norm_ref_text(hint).strip("[]() ")
    scored: List[Tuple[int, Dict[str, str]]] = []
    seen = set()
    for e in registry:
        sf = e.get("short_form") or ""
        vb = e.get("verbatim") or ""
        link = re.sub(r"\s+", " ", str(e.get("link") or "")).strip()
        key = (_norm_ref_text(sf), _ref_base_url(link).lower())
        if key in seen:
            continue
        seen.add(key)
        etoks = _ref_disambig_tokens(sf) | _ref_disambig_tokens(vb[:160])
        score = 2 * len(q & etoks) + len(q_prop & etoks)
        if hint_norm and _norm_ref_text(sf).strip("[]() ") == hint_norm:
            score += 20
        scored.append((score, e))
    scored.sort(key=lambda x: -x[0])
    top = [e for s, e in scored if s > 0][:REF_DISAMBIG_MAX_CANDIDATES]
    if not top:
        top = [e for _, e in scored[:REF_DISAMBIG_MAX_CANDIDATES]]
    return top, ("" if has_hint else prop_text)


def _ref_disambig_choose(
    part_verbatim: str,
    candidates: List[Dict[str, str]],
    prop_text: str,
) -> str:
    """One tiny model call: pick the referenced work from the candidate list.
    Returns the chosen link, or "" when the model answers "none"."""
    lines = [f"Reference part: {part_verbatim}"]
    if prop_text:
        lines.append(f"Body text the footnote is attached to: {prop_text[:600]}")
    lines += ["", "Candidates:"]
    for i, e in enumerate(candidates, start=1):
        lines.append(
            f"{i}. [fn {e.get('note') or '?'}] short_form: {e.get('short_form') or '(none)'} | "
            f"citation: {(e.get('verbatim') or '')[:160]} | "
            f"link: {e.get('link') or 'other'}"
        )
    lines.append("0. none of these / no usable link")
    user = "\n".join(lines)

    choice: Optional[int] = None
    cache_path = ""
    if LLM_CACHE_ENABLED:
        key = hashlib.sha256(
            f"{LLM_MODEL}|none|{REF_DISAMBIG_SYSTEM}|{user}".encode("utf-8")
        ).hexdigest()
        cache_path = os.path.join(_llm_cache_dir(), f"refdisambig_{key}.json")
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    choice = int(json.load(f).get("choice", 0))
            except Exception:
                choice = None
    if choice is None:
        _pause_gate()
        resp = _llm_call(
            model=LLM_MODEL,
            input=[
                {"role": "system", "content": REF_DISAMBIG_SYSTEM},
                {"role": "user", "content": user},
            ],
            reasoning={"effort": "none"},
            max_output_tokens=2000,
            text={"format": {"type": "json_schema", "name": "supra_choice",
                             "strict": True, "schema": _REF_DISAMBIG_SCHEMA}},
        )
        try:
            choice = int(json.loads(resp.output_text).get("choice", 0))
        except Exception:
            choice = 0
        if cache_path:
            try:
                os.makedirs(_llm_cache_dir(), exist_ok=True)
                with _LLM_CACHE_IO_LOCK:
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump({"choice": choice, "user": user}, f, ensure_ascii=False)
            except Exception:
                pass
    if 1 <= choice <= len(candidates):
        link = (candidates[choice - 1].get("link") or "").strip()
        if link and link.lower() != "other":
            return link
    return ""


def _canonicalize_statute_links(*part_maps: Dict[int, List["FootnotePart"]]) -> int:
    """A2AJ statute-slug canonicalization (canlii_slug_repair) over every
    split part, before the supra registry is built so references inherit the
    repaired links. Repair-only — a miss or any failure leaves links as-is."""
    import canlii_slug_repair

    repaired = 0
    for part_map in part_maps:
        for fid, parts in part_map.items():
            for i, part in enumerate(parts or []):
                _pause_gate()
                try:
                    result = canlii_slug_repair.repair_statute_link(
                        part.link, part.verbatim or part.citation_with_style or part.bare_citation
                    )
                except Exception as exc:
                    _ts_print(f"    Statute link canonicalization skipped ({type(exc).__name__})")
                    result = None
                if result:
                    new_link, reason = result
                    parts[i] = replace(part, link=new_link)
                    repaired += 1
                    _ts_print(
                        f"    Footnote {fid}: statute link canonicalized ({reason}): "
                        f"{part.link[:70]} -> {new_link[:70]}"
                    )
    return repaired


def _resolve_footnote_reference_links(
    fid: int,
    full_text: str,
    split: List["FootnotePart"],
    registry: List[Dict[str, str]],
    prop_text: str = "",
    allow_fallback: bool = True,
    inferred_short_forms: Optional[List[Dict[str, str]]] = None,
) -> Dict[int, str]:
    """Resolve supra/ibid links for one footnote's parts, in place.

    `registry` holds every previously emitted part (document order, excluding
    this footnote). Deterministic resolution runs first; the small candidate
    chooser handles remaining supras when model use is allowed. Returns
    {part_index: method} for every reference part ("registry", "note_number",
    "sibling", "disambig", "disambig_rejected", "model", or
    "dropped_abstain").
    """
    methods: Dict[int, str] = {}
    abstained: List[int] = []
    for idx, part in enumerate(split):
        ref_kind = _detect_ref_kind(part.verbatim)
        if not ref_kind:
            continue
        if ref_kind == "ibid":
            if idx > 0:
                origin_link = split[idx - 1].link
            elif registry:
                origin_link = registry[-1].get("link", "")
            else:
                origin_link = ""
            if origin_link:
                split[idx] = replace(part, link=_reanchor_ref_link(origin_link, part.verbatim))
                methods[idx] = "sibling" if idx > 0 else "registry"
            else:
                split[idx] = replace(part, link=_reanchor_ref_link(part.link, part.verbatim))
                methods[idx] = "model"
        else:  # supra
            resolved, _method = _resolve_supra_from_registry(
                part.verbatim, registry
            )
            if resolved:
                split[idx] = replace(part, link=_reanchor_ref_link(resolved, part.verbatim))
                methods[idx] = "note_number" if _method == "note_number" else "registry"
            else:
                fallback_link, fallback_method = ("", "")
                if SUPRA_LINKING_AGGRESSIVENESS == "aggressive":
                    fallback_link, fallback_method = (
                        _supra_fallbacks.resolve_after_strict_abstention(
                            part.verbatim, registry, inferred_short_forms or []
                        )
                    )
                if fallback_link:
                    split[idx] = replace(
                        part, link=_reanchor_ref_link(fallback_link, part.verbatim)
                    )
                    methods[idx] = fallback_method
                    continue
                abstained.append(idx)
                # Keep the model's own answer (re-anchored) until the
                # fallback call improves on it.
                split[idx] = replace(part, link=_reanchor_ref_link(part.link, part.verbatim))
                methods[idx] = "model"

    if abstained and allow_fallback and REF_DISAMBIG_FALLBACK:
        _ts_print(
            f"    Footnote {fid}: deterministic resolver abstained on "
            f"{len(abstained)} reference part(s); disambiguating from candidates ..."
        )
        for idx in abstained:
            part = split[idx]
            candidates, prop_shown = _ref_disambig_candidates(
                part.verbatim, registry, prop_text
            )
            if not candidates:
                continue
            try:
                link = _ref_disambig_choose(part.verbatim, candidates, prop_shown)
            except Exception as exc:
                _ts_print(f"    Disambiguation failed ({type(exc).__name__}).")
                continue
            if link and _ref_fallback_link_guard_ok(part.verbatim, link, registry):
                split[idx] = replace(part, link=_reanchor_ref_link(link, part.verbatim))
                methods[idx] = "disambig"
            elif link:
                methods[idx] = "disambig_rejected"
        return _drop_unresolved_supra_links(split, abstained, methods, allow_fallback)

    return _drop_unresolved_supra_links(split, abstained, methods, allow_fallback)


def _drop_unresolved_supra_links(
    split: List["FootnotePart"],
    abstained: List[int],
    methods: Dict[int, str],
    allow_fallback: bool,
) -> Dict[int, str]:
    """Clear the split model's link when every resolver stage abstains.

    Deterministic-only prepasses set ``allow_fallback=False`` and defer this
    decision to the normal Phase 2 pass.
    """
    if not (REF_DROP_ON_ABSTAIN and allow_fallback):
        return methods
    for idx in abstained:
        if methods.get(idx) in ("model", "disambig_rejected"):
            split[idx] = replace(split[idx], link="")
            methods[idx] = "dropped_abstain"
    return methods


def _prefetch_local_a2aj_sources(
    footnote_map: Dict[int, str], footnote_order: List[int]
) -> None:
    """Warm exact local-A2AJ lookups with one Parquet scan per partition."""
    if not (USE_A2AJ and DETERMINISTIC_SOURCE_SPLITTER):
        return
    client = a2aj_client.get_client()
    corpus = client.local_corpus
    if corpus is None:
        return

    candidates: Dict[str, List[str]] = {"cases": [], "laws": []}
    for fid in footnote_order:
        split = _deterministic_splitter.split_footnote_recall_first(
            footnote_map.get(fid, "")
        )
        if split.status != "deterministic_complete":
            continue
        for source in split.parts:
            item = _deterministic_splitter.extract_fields(source)
            kind = item.kind.strip().lower()
            styled = item.citation_with_style or source.text
            bare = (
                _derive_bare_citation(styled, kind)
                or item.bare_citation
                or source.text
            )
            raw_bare = re.sub(r"\s+", " ", str(bare or "")).strip()
            if not raw_bare:
                continue
            if kind in {"case", "unreported"}:
                direct = _a2aj_query_citation(raw_bare)
                identity = _a2aj_identity_citation(raw_bare)
                alias = client.reporter_alias_canonical(raw_bare)
                candidates["cases"].extend(
                    value for value in (direct, identity, alias) if value
                )
            elif kind in {"statute", "gazette"}:
                bare_query = _a2aj_query_citation(raw_bare)
                query = raw_bare.rstrip(".;") if re.search(
                    r",\s*(?:r(?:ule)?s?\.?)\s*\d+(?:\.\d+)*\s*$",
                    raw_bare,
                    flags=re.I,
                ) else bare_query
                candidates["laws"].extend(
                    value for value in (query, bare_query) if value
                )

    totals = {"requested": 0, "cached": 0, "partitions": 0, "rows": 0}

    def show_progress(update: Any) -> None:
        _pause_gate()
        label = "cases" if update.kind == "cases" else "legislation"
        _ts_print(
            f"    Loading local A2AJ {label}: partition "
            f"{update.completed}/{update.total} ..."
        )

    started = time.perf_counter()
    for doc_type in ("cases", "laws"):
        if not candidates[doc_type]:
            continue
        try:
            stats = corpus.prefetch_exact_citations(
                candidates[doc_type], doc_type, progress=show_progress
            )
        except (OSError, RuntimeError, ValueError) as exc:
            if client.local_only:
                _ts_print(
                    f"    Local A2AJ {doc_type} prefetch unavailable "
                    f"({type(exc).__name__}); continuing with ordinary lookup."
                )
            continue
        for name in totals:
            totals[name] += int(stats.get(name, 0))
    if totals["requested"]:
        _ts_print(
            "    Local A2AJ ready: "
            f"{totals['requested']} exact lookup(s), "
            f"{totals['rows']} row(s) from {totals['partitions']} new partition scan(s) "
            f"in {time.perf_counter() - started:.1f}s."
        )


def build_footnote_parts(
    footnote_map: Dict[int, str],
    footnote_order: List[int],
    prop_texts: Optional[Dict[int, str]] = None,
    author_links_by_fid: Optional[Dict[int, List[Dict[str, Any]]]] = None,
) -> Tuple[List[Dict[str, Any]], Dict[int, int]]:
    parts: List[Dict[str, Any]] = []
    fn_num_parts: Dict[int, int] = {}

    llm_start = time.time()

    total_fns = len(footnote_order)
    analyzer = "deterministically (no LLM)" if FREE_NO_LLM else "with GPT"
    _ts_print(f"  Analyzing {total_fns} footnotes {analyzer} ...")
    _prefetch_local_a2aj_sources(footnote_map, footnote_order)

    # Phase 1: split every footnote sequentially, each call seeing the
    # accumulated citation history of everything before it.
    split_by_fid: Dict[int, List[FootnotePart]] = {}
    doc_history = ""
    prev_split: Optional[List[FootnotePart]] = None
    display_ids, _, _ = _compute_footnote_display_ids(footnote_order, footnote_map)
    history_registry: List[Dict[str, str]] = []
    history_inferred_short_forms: List[Dict[str, str]] = []

    def remember_for_history(fid: int, split: List[FootnotePart]) -> None:
        note = display_ids.get(fid, str(fid))
        for part in split:
            history_registry.append({
                "verbatim": part.verbatim,
                "link": part.link,
                "short_form": part.short_form,
                "note": note,
            })
            if (
                SUPRA_LINKING_AGGRESSIVENESS == "aggressive"
                and part.link and part.link.lower() != "other"
            ):
                for short_form in _supra_fallbacks.infer_short_forms(
                    part.verbatim, part.kind
                ):
                    history_inferred_short_forms.append({
                        "short_form": short_form.value,
                        "short_form_norm": _supra_fallbacks.normalize_short_form(
                            short_form.value
                        ),
                        "rule": short_form.rule,
                        "link": part.link,
                        "note": note,
                        "origin": part.verbatim,
                    })

    for fn_idx, fid in enumerate(footnote_order, 1):
        _pause_gate()
        _ts_print(f"    Footnote {fn_idx}/{total_fns} (source id {fid}) ...")
        full = footnote_map.get(fid, "")
        pre_parts = None
        if PURE_REF_PREFILTER and full:
            pre_parts = _prefilter_pure_ref_parts(
                full,
                allow_leading_ibid=_pref_allow_leading_ibid(prev_split),
            )
        if pre_parts is not None:
            _ts_print(f"      Pure-reference footnote: {len(pre_parts)} part(s), no model call")
            if DETERMINISTIC_SOURCE_SPLITTER and DETERMINISTIC_REF_LINKS:
                _resolve_footnote_reference_links(
                    fid, full, pre_parts, history_registry,
                    allow_fallback=False,
                    inferred_short_forms=history_inferred_short_forms,
                )
            doc_history += _build_footnote_history_entries(pre_parts)
            split_by_fid[fid] = pre_parts
            prev_split = pre_parts
            remember_for_history(fid, pre_parts)
            _timing_event(
                "build_footnote:split",
                footnote_id=fid,
                ordinal=fn_idx,
                parts=len(pre_parts),
                prefiltered=True,
                elapsed_s=0.0,
            )
            continue
        deterministic = None
        if DETERMINISTIC_SOURCE_SPLITTER and full:
            deterministic = _deterministic_footnote_parts(
                full,
                allow_unsplit_fallback=FREE_NO_LLM,
                allow_reference_parts=FREE_NO_LLM,
            )
        if deterministic is not None:
            split, deterministic_reasons = deterministic
            split = _apply_author_provided_links(
                full, split, (author_links_by_fid or {}).get(fid)
            )
            if DETERMINISTIC_REF_LINKS:
                _resolve_footnote_reference_links(
                    fid, full, split, history_registry,
                    allow_fallback=False,
                    inferred_short_forms=history_inferred_short_forms,
                )
            doc_history += _build_footnote_history_entries(split)
            split_by_fid[fid] = split
            prev_split = split
            remember_for_history(fid, split)
            _ts_print(
                f"      Deterministic footnote: {len(split)} part(s), no model call "
                f"[{', '.join(deterministic_reasons) or 'lossless_unsplit'}]"
            )
            _timing_event(
                "build_footnote:split",
                footnote_id=fid,
                ordinal=fn_idx,
                parts=len(split),
                deterministic=True,
                deterministic_reasons=list(deterministic_reasons),
                elapsed_s=0.0,
            )
            continue
        if full:
            split_started = time.perf_counter()
            split, new_entries = split_footnote_parts(full, doc_history)
            split = _snap_verbatim_parts(full, split)
            split = _apply_author_provided_links(
                full, split, (author_links_by_fid or {}).get(fid)
            )
            if any(part.author_provided_link for part in split):
                new_entries = _build_footnote_history_entries(split)
            _timing_event(
                "build_footnote:split",
                footnote_id=fid,
                ordinal=fn_idx,
                parts=len(split),
                elapsed_s=round(time.perf_counter() - split_started, 3),
            )
            # Update the local history so the NEXT footnote in the loop sees it
            doc_history += new_entries
        else:
            split = []
        split_by_fid[fid] = split
        if split:
            prev_split = split
            remember_for_history(fid, split)

    # Phase 1b: canonicalize statute slugs against A2AJ before the supra
    # registry is built, so reference parts inherit the repaired links.
    if A2AJ_STATUTE_LINK_REPAIR:
        n_repaired = _canonicalize_statute_links(split_by_fid)
        if n_repaired:
            _ts_print(f"    Statute links canonicalized via A2AJ: {n_repaired}")

    # Phase 2: deterministic supra/ibid resolution against previously emitted
    # parts, followed by the small candidate chooser for remaining abstentions.
    ref_link_methods: Dict[Tuple[int, int], str] = {}
    if DETERMINISTIC_REF_LINKS:
        registry: List[Dict[str, str]] = []
        inferred_short_forms: List[Dict[str, str]] = []
        for fid in footnote_order:
            split = split_by_fid.get(fid) or []
            methods = _resolve_footnote_reference_links(
                fid, footnote_map.get(fid, ""), split, registry,
                prop_text=(prop_texts or {}).get(fid, ""),
                inferred_short_forms=inferred_short_forms,
            )
            for idx, method in methods.items():
                ref_link_methods[(fid, idx + 1)] = method
            for part in split:
                note = display_ids.get(fid, str(fid))
                registry.append({
                    "verbatim": part.verbatim,
                    "link": part.link,
                    "short_form": part.short_form,
                    "note": note,
                })
                if (
                    SUPRA_LINKING_AGGRESSIVENESS == "aggressive"
                    and part.link and part.link.lower() != "other"
                ):
                    for short_form in _supra_fallbacks.infer_short_forms(
                        part.verbatim, part.kind
                    ):
                        inferred_short_forms.append({
                            "short_form": short_form.value,
                            "short_form_norm": _supra_fallbacks.normalize_short_form(
                                short_form.value
                            ),
                            "rule": short_form.rule,
                            "link": part.link,
                            "note": note,
                            "origin": part.verbatim,
                        })

    # Phase 3: build the per-part rows.
    for fn_idx, fid in enumerate(footnote_order, 1):
        _pause_gate()
        full = footnote_map.get(fid, "")
        split = split_by_fid.get(fid) or []
        _timing_event("build_footnote:start", footnote_id=fid, ordinal=fn_idx, total=total_fns)

        # If splitting failed or was empty but text existed, handle fallback
        if not split and full:
            split = [FootnotePart(verbatim=full, corrected=full, kind="other", link="", pinpoint_fragments=[], page_pinpoints=[])]
            split = _apply_author_provided_links(
                full, split, (author_links_by_fid or {}).get(fid)
            )

        fn_num_parts[fid] = len(split)

        for idx, part in enumerate(split, start=1):
            source_started = time.perf_counter()
            citation_part_link = _canlii_pdf_to_html_sibling(part.link)
            link_base, _link_frag = _split_url(citation_part_link or "")
            locked_structure = _A2AJ_LOCKED_STRUCTURES.get((link_base or "").lower(), {})
            locked_document = _A2AJ_LOCKED_DOCUMENTS.get((link_base or "").lower())
            a2aj_source_available = bool(locked_document and locked_document.text)
            a2aj_probe: Dict[str, Any] = {}
            has_quote_context = bool(find_inline_quotes((prop_texts or {}).get(fid, "")))
            if (
                USE_A2AJ
                and has_quote_context
                and locked_structure.get("status") != "usable"
                and (part.kind or "").strip().lower()
                in {"case", "unreported", "statute", "gazette"}
            ):
                a2aj_probe = {
                    "bare_citation": part.bare_citation,
                    "citation_part_kind": part.kind,
                    "citation_part_link": citation_part_link,
                }
                a2aj_source_available = bool(
                    _fetch_a2aj_source_text_for_row(a2aj_probe)
                )
                locked_structure = a2aj_probe.get("_a2aj_structure") or {}
            effective_pinpoint_fragments = list(part.pinpoint_fragments or [])
            effective_page_pinpoints = list(part.page_pinpoints or [])
            scr_anchor_text = ""
            _timing_event(
                "build_footnote_part_source:start",
                footnote_id=fid,
                citation_part_index=idx,
                kind=part.kind,
                host=urlsplit(link_base).netloc.lower() if link_base else "",
                has_link=bool(citation_part_link and citation_part_link.lower() != "other"),
            )
            anchor_segments: List[Dict[str, str]] = []
            if a2aj_source_available:
                anchor_text = ""
            elif scr_anchor_text:
                first_fragment = (
                    _normalize_anchor_fragment((effective_pinpoint_fragments or [""])[0])
                    if effective_pinpoint_fragments
                    else ""
                )
                if first_fragment:
                    anchor_segments = [{"fragment": first_fragment, "text": scr_anchor_text}]
                anchor_text = scr_anchor_text
            else:
                anchor_segments = _extract_provider_anchor_text_segments(
                    citation_part_link,
                    effective_pinpoint_fragments,
                )
                anchor_text = " ".join(
                    seg.get("text", "") for seg in anchor_segments if seg.get("text")
                ).strip()
            full_source_text = ""
            _timing_event(
                "build_footnote_part_source:end",
                footnote_id=fid,
                citation_part_index=idx,
                kind=part.kind,
                anchor_chars=len(anchor_text or ""),
                full_source_chars=len(full_source_text or ""),
                elapsed_s=round(time.perf_counter() - source_started, 3),
            )
            row: Dict[str, Any] = {
                "footnote_id": fid,
                "footnote_full": full,
                "citation_part_index": idx,
                "citation_part_kind": part.kind,
                "bare_citation": part.bare_citation,
                "citation_with_style": part.citation_with_style,
                "citation_part_text": part.verbatim,
                "citation_part_corrected": part.corrected,
                "citation_part_link": citation_part_link,
                "_author_provided_link": part.author_provided_link,
                "_author_provided_links": list(part.author_provided_links),
                "citation_part_anchor_text": anchor_text,
                "_citation_part_anchor_segments": anchor_segments,
                "_citation_part_full_source_text": full_source_text,
                "short_form": part.short_form,
                "pinpoint_fragments": list(effective_pinpoint_fragments),
                "page_pinpoints": list(effective_page_pinpoints),
                "_ref_link_resolution": ref_link_methods.get((fid, idx), ""),
            }
            row.update({
                key: value for key, value in a2aj_probe.items()
                if key.startswith("_a2aj_")
            })
            row.update(_blank_ref_fields())
            parts.append(row)
        _timing_event("build_footnote:end", footnote_id=fid, ordinal=fn_idx, parts=len(split))

    llm_elapsed = time.time() - llm_start
    if llm_elapsed >= 1 and not FREE_NO_LLM:
        _ts_print(f"  LLM calls took {llm_elapsed:.0f}s")
    return parts, fn_num_parts


def resolve_reference_chains(
    parts: List[Dict[str, Any]],
    footnote_map: Dict[int, str],
    fn_num_parts: Dict[int, int],
    *,
    display_num_to_internal: Optional[Dict[int, int]] = None,
    internal_to_display_id: Optional[Dict[int, str]] = None,
) -> Dict[Tuple[int, int], Dict[str, Any]]:
    """Populate REF_FIELDS for each part, including chained supra/ibid resolution."""

    by_key: Dict[Tuple[int, int], Dict[str, Any]] = {
        (p["footnote_id"], p["citation_part_index"]): p for p in parts
    }

    # Also index parts by global order (the input list is already in appearance order).
    order_keys: List[Tuple[int, int]] = [(p["footnote_id"], p["citation_part_index"]) for p in parts]

    # Index citation parts by footnote for smarter supra resolution when a target footnote contains multiple citations.
    parts_by_footnote: Dict[int, List[Dict[str, Any]]] = {}
    for p in parts:
        try:
            fid = int(p.get("footnote_id") or 0)
        except Exception:
            continue
        parts_by_footnote.setdefault(fid, []).append(p)
    for fid in list(parts_by_footnote.keys()):
        parts_by_footnote[fid].sort(key=lambda x: int(x.get("citation_part_index") or 0))


    def prev_part_key(i: int) -> Optional[Tuple[int, int]]:
        return order_keys[i - 1] if i > 0 else None

    def _map_display_to_internal(n: int) -> Optional[int]:
        if display_num_to_internal:
            return display_num_to_internal.get(int(n))
        return int(n)

    def _display_id(fid: int) -> str:
        if internal_to_display_id and int(fid) in internal_to_display_id:
            return internal_to_display_id[int(fid)]
        return str(fid)

    def follow_chain(start_i: int) -> Tuple[Optional[Tuple[int, int]], List[str], List[str]]:
        """Return (origin_key, path, warnings)."""

        visited: set[Tuple[int, int]] = set()
        path: List[str] = []
        warnings: List[str] = []

        cur_key = order_keys[start_i]
        while cur_key is not None:
            if cur_key in visited:
                warnings.append("Reference chain loop detected.")
                return None, path, warnings
            visited.add(cur_key)

            cur = by_key.get(cur_key)
            if not cur:
                warnings.append("Reference chain target not found in part index.")
                return None, path, warnings

            txt = (cur.get("citation_part_text") or "").strip()
            supra_n = _extract_supra_note_number(txt)
            is_ibid = _is_ibid(txt)

            if supra_n is not None:
                target_fid = _map_display_to_internal(supra_n)
                if not target_fid:
                    warnings.append(f"Supra target footnote (note {supra_n}) not found in this DOCX.")
                    return None, path, warnings
                target_part, warn = _choose_supra_target_part_index(
                    target_fid,
                    txt,
                    parts_by_footnote,
                    display_note_number=supra_n,
                )
                path.append(f"SUPRA→{supra_n}:{target_part}")
                if warn:
                    warnings.append(warn)
                next_key = (target_fid, int(target_part))
                if next_key not in by_key:
                    warnings.append(f"Supra target footnote (note {supra_n}) not found in this DOCX.")
                    return None, path, warnings
                cur_key = next_key
                continue
            if is_ibid:
                path.append("IBID")
                # ibid resolves to immediately preceding citation-part in the document
                prev_key = prev_part_key(order_keys.index(cur_key))
                if prev_key is None:
                    warnings.append("Ibid appears as the first reference; no preceding reference exists.")
                    return None, path, warnings
                cur_key = prev_key
                continue

            # reached a non-ibid/supra origin
            return cur_key, path, warnings

        return None, path, warnings

    lookup: Dict[Tuple[int, int], Dict[str, Any]] = {}

    for i, p in enumerate(parts):
        txt = (p.get("citation_part_text") or "").strip()
        supra_n = _extract_supra_note_number(txt)
        is_ibid = _is_ibid(txt)

        # immediate target (for diagnostics)
        if supra_n is not None:
            p["ref_kind"] = "SUPRA"
            target_fid = _map_display_to_internal(supra_n)
            if target_fid is not None:
                p["ref_target_footnote_id"] = _display_id(target_fid)
                target_part, _ = _choose_supra_target_part_index(
                    target_fid,
                    txt,
                    parts_by_footnote,
                    display_note_number=supra_n,
                )
                p["ref_target_citation_part_index"] = int(target_part)
                p["ref_target_footnote_full"] = footnote_map.get(target_fid, "")
                target_text = (by_key.get((target_fid, int(target_part))) or {}).get(
                    "citation_part_text", ""
                )
                hint = _extract_supra_hint(txt or "")
                if hint:
                    target_text = f"{hint}, {target_text}"
                p["ref_target_citation_part_text"] = target_text
        elif is_ibid:
            p["ref_kind"] = "IBID"
            pk = order_keys[i - 1] if i > 0 else None
            if pk is not None:
                prev = by_key.get(pk, {})
                prev_fid = prev.get("footnote_id", "")
                try:
                    p["ref_target_footnote_id"] = _display_id(int(prev_fid))
                except Exception:
                    p["ref_target_footnote_id"] = prev_fid
                p["ref_target_citation_part_index"] = prev.get("citation_part_index", "")
                p["ref_target_citation_part_text"] = prev.get("citation_part_text", "")
                try:
                    p["ref_target_footnote_full"] = footnote_map.get(int(prev_fid), "")
                except Exception:
                    p["ref_target_footnote_full"] = ""

                if prev.get("footnote_id") != p.get("footnote_id") and fn_num_parts.get(
                    int(prev.get("footnote_id") or 0), 0
                ) > 1:
                    p["ref_resolution_notes"] = (
                        (p.get("ref_resolution_notes") or "")
                        + " Previous footnote has multiple references; style guidance prefers supra over ibid in this situation."
                    ).strip()

        origin_key, path, warnings = follow_chain(i)
        p["ref_chain_path"] = " → ".join(path)

        if origin_key is not None:
            origin = by_key[origin_key]
            origin_fid = origin.get("footnote_id", "")
            try:
                p["ref_chain_origin_footnote_id"] = _display_id(int(origin_fid))
            except Exception:
                p["ref_chain_origin_footnote_id"] = origin_fid
            p["ref_chain_origin_citation_part_index"] = origin.get("citation_part_index", "")
            p["ref_chain_origin_citation_part_text"] = origin.get("citation_part_text", "")
            # Propagate first_page through chain so supra/ibid know the source's starting page
            origin_fp = origin.get("first_page", "")
            if origin_fp:
                p["first_page"] = origin_fp

        if warnings:
            p["ref_resolution_notes"] = (
                (p.get("ref_resolution_notes") or "") + " " + " ".join(warnings)
            ).strip()

        lookup[(p["footnote_id"], p["citation_part_index"])] = {k: p.get(k, "") for k in REF_FIELDS}

    return lookup


def _effective_citation_text(row: Dict[str, Any]) -> str:
    """Prefer the chain origin for ibid/supra."""
    origin = row.get("ref_chain_origin_citation_part_text") or row.get("ref_target_citation_part_text")
    return origin or row.get("citation_part_text") or ""


# -----------------------------
# Anchor proposition text (between footnotes)
# -----------------------------


def _tail_sentences(text: str, n: int, max_chars: int) -> str:
    t = re.sub(r"\s+", " ", (text or "").strip())
    if not t:
        return ""
    # naive sentence split
    sents = re.split(r"(?<=[\.\!\?])\s+", t)
    tail = " ".join(sents[-n:]) if len(sents) >= n else t
    tail = tail.strip()
    if len(tail) > max_chars:
        tail = tail[-max_chars:]
    return tail


def _sentence_bounds_smart(clean_text: str, pos: int, max_window: int = 1200) -> Tuple[int, int]:
    """Same as _sentence_bounds but filters out periods in numbers
    (digit.digit like 33.1) and known legal abbreviations (s., R. v., etc.).
    Scans every character — not just regex-matched positions — to catch
    periods followed by closing quotes, footnote markers, etc."""
    _ABBR_RE = re.compile(
        r"\b(?:Dr|Mr|Mrs|Ms|Jr|Sr|Hon|Prof|Rev|St|No|Nos|pp|para|paras|vol|vols"
        r"|art|arts|pt|ch|cl|sch|sec|ss|ed|eds|e\.g|i\.e|etc|cf|viz|seq|vs"
        r"|al|Ltd|Inc|Co|Corp|Bros|Assn|Dept|Univ|Intl|Natl|Ct"
        r"|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sept?|Oct|Nov|Dec"
        r"|[A-Za-z])\.?$",
        re.IGNORECASE,
    )

    def _is_real_sentence_end(segment: str, p: int) -> bool:
        if p > 0 and segment[p - 1].isdigit() and p + 1 < len(segment) and segment[p + 1].isdigit():
            return False
        # Ellipsis runs ("..." / "[...]") are not sentence ends.
        if segment[p - 1:p] == "." or segment[p + 1:p + 2] == ".":
            return False
        if _ABBR_RE.search(segment[:p + 1]):
            return False
        return True

    pos = max(0, min(len(clean_text), int(pos)))
    left_limit = max(0, pos - max_window)
    right_limit = min(len(clean_text), pos + max_window)
    segment = clean_text[left_limit:right_limit]
    rel = pos - left_limit

    # Left boundary — scan backwards for . ! ?
    left_idx = -1
    for i in range(rel - 1, -1, -1):
        if segment[i] in ".!?" and (segment[i] != "." or _is_real_sentence_end(segment, i)):
            left_idx = i
            break
    left_nl = segment.rfind("\n", 0, rel)
    if left_nl > left_idx:
        left_idx = left_nl
    start = left_limit + (0 if left_idx == -1 else left_idx + 1)

    # skip closing quote after sentence end
    if start > 0 and start < len(clean_text):
        if clean_text[start - 1] in ".!?" and clean_text[start] in _CLOSE_DOUBLE_QUOTES:
            start += 1
    while start < right_limit and clean_text[start].isspace():
        start += 1

    # Right boundary — scan forward for . ! ?
    right_idx = -1
    for i in range(rel, len(segment)):
        if segment[i] in ".!?" and (segment[i] != "." or _is_real_sentence_end(segment, i)):
            right_idx = i
            break
    right_nl = segment.find("\n", rel)
    if right_nl >= 0 and (right_idx < 0 or right_nl < right_idx):
        right_idx = right_nl
    end = left_limit + (right_idx + 1 if right_idx >= 0 else len(segment))

    # Include trailing closing double quotes
    while 0 < end < len(clean_text) and clean_text[end - 1] in ".!?" and clean_text[end] in _CLOSE_DOUBLE_QUOTES:
        end += 1
    return (start, end)


def build_anchor_propositions(
    clean_global: str,
    anchors: List[Dict[str, Any]],
    raw_to_clean: List[int],
) -> Dict[int, Dict[str, Any]]:
    """Build sentence-bounded proposition text for each footnote anchor."""
    anchors_sorted = sorted(anchors, key=lambda a: a["global_pos"])
    out: Dict[int, Dict[str, Any]] = {}

    intro_cut = _introduction_cutoff(clean_global)

    prev_end = 0

    for a in anchors_sorted:
        fid = int(a["footnote_id"])
        if fid in out:
            continue

        curr_raw = int(a["global_pos"])
        anchor_c = raw_to_clean[curr_raw]

        pos = _anchor_sentence_pos(clean_global, anchor_c)
        start_c, end_c = _sentence_bounds_smart(clean_global, pos)

        # If the start boundary lands inside an unmatched quotation (open quote before
        # start_c that isn't closed until after start_c), extend to before the opening
        # quote so the introductory clause is included. Only search within the territory
        # of this proposition (prev_end .. start_c) to avoid reaching into earlier text.
        for q_open, q_close in [('\u201c', '\u201d'), ('"', '"')]:
            open_idx = clean_global.rfind(q_open, prev_end, start_c)
            if open_idx >= 0:
                close_before = clean_global.find(q_close, open_idx + 1, start_c)
                if close_before >= 0:
                    continue  # properly closed before start_c — not inside
                close_after = clean_global.find(q_close, start_c, start_c + 200)
                if close_after >= 0:
                    start_c, _ = _sentence_bounds_smart(clean_global, max(0, open_idx - 1))
                    break

        if fid == 1 and intro_cut:
            start_c = max(start_c, intro_cut)

        prop = clean_global[start_c:end_c]
        prop = re.sub(r"\s+", " ", prop).strip()

        out[fid] = {
            "proposition_text": prop,
        }
        prev_end = end_c

    return out


# -----------------------------
#
# -----------------------------


def _compute_footnote_display_ids(
    footnote_order: List[int],
    footnote_map: Dict[int, str],
) -> Tuple[Dict[int, str], Dict[int, int], Dict[int, Optional[int]]]:
    display_map: Dict[int, str] = {}
    display_num_to_internal: Dict[int, int] = {}
    internal_to_display_num: Dict[int, Optional[int]] = {}
    next_num = 1
    for fid in footnote_order:
        text = footnote_map.get(fid, "")
        m = re.match(r"^\s*(\*+)", text or "")
        if m:
            display_map[fid] = m.group(1)
            internal_to_display_num[fid] = None
        else:
            display_map[fid] = str(next_num)
            display_num_to_internal[next_num] = fid
            internal_to_display_num[fid] = next_num
            next_num += 1
    return display_map, display_num_to_internal, internal_to_display_num

_WORD_TOKEN_PATTERN = r"[^\W_]+(?:['\u2019][^\W_]+)*"
_QUOTE_TOKEN_RE = re.compile(
    rf"\.\.\.|\[[^\]]+\]|{_WORD_TOKEN_PATTERN}|[\"\u201c\u201d\u2018\u2019]|[^\w\s]"
)

_DOUBLE_QUOTE_TOKENS = {"\"", "\u201c", "\u201d", "\u00ab", "\u00bb", "\u201e"}
_SINGLE_QUOTE_TOKENS = {"'", "\u2018", "\u2019", "\u201a"}
_QUOTE_TOKENS = _DOUBLE_QUOTE_TOKENS | _SINGLE_QUOTE_TOKENS
_MATCH_IGNORE_TRAILING_PUNCT = {".", ",", ";", ":", "!", "?", "…", "..."}
_LONG_SOURCE_CHAR_LIMIT = 80000
_LONG_SOURCE_TOKEN_LIMIT = 8000
_ALT_PINPOINT_RECHECK_SCORE = 0.90
_SHORT_QUOTE_LONG_SOURCE_NOTE = (
    "Quote too short for full-document fuzzy checking in a long source; exact search found no match."
)
_MIXED_SHORT_QUOTE_LONG_SOURCE_NOTE = (
    "One or more quotes were too short for full-document fuzzy checking in a long source; exact search found no match."
)
_WORD_TOKEN_RE = re.compile(_WORD_TOKEN_PATTERN)
_DASH_EQUIV_TOKENS = {
    "-",
    "\u00ad",
    "\u2010",
    "\u2011",
    "\u2012",
    "\u2013",
    "\u2014",
    "\u2015",
    "\u2212",
}
_DASH_EQUIV_TRANS = str.maketrans({ch: "-" for ch in _DASH_EQUIV_TOKENS})


def _is_dash_quote_token(tok: str) -> bool:
    return tok in _DASH_EQUIV_TOKENS


def _normalize_quote_dash_chars(text: str) -> str:
    return (text or "").translate(_DASH_EQUIV_TRANS)


_INTRA_WORD_HYPHEN_RE = re.compile(r"(?<=[^\W_])\s*-\s*(?=[^\W_])")


def _normalize_intra_word_hyphen_for_compare(text: str) -> str:
    s = _normalize_quote_dash_chars(text)
    return _INTRA_WORD_HYPHEN_RE.sub(" ", s)


def _normalize_intra_word_hyphen_for_display(text: str) -> str:
    s = _normalize_quote_dash_chars(text)
    return _INTRA_WORD_HYPHEN_RE.sub("-", s)


def _mergeable_quote_token(tok: str) -> bool:
    return bool(_WORD_TOKEN_RE.search(tok or "")) or (tok.startswith("[") and tok.endswith("]"))


def _is_wordlike_quote_token(tok: str) -> bool:
    return bool(_WORD_TOKEN_RE.search(tok or ""))


@lru_cache(maxsize=32)
def _tokenize_quote_text_spans_cached(text: str) -> Tuple[Tuple[str, int, int], ...]:
    tokens: List[Tuple[str, int, int]] = []

    for m in _QUOTE_TOKEN_RE.finditer(text):
        tokens.append((m.group(0), m.start(), m.end()))

    merged: List[Tuple[str, int, int]] = []
    i = 0
    while i < len(tokens):
        tok, start, end = tokens[i]
        j = i + 1
        while (
            j < len(tokens)
            and tokens[j - 1][2] == tokens[j][1]
            and _mergeable_quote_token(tokens[j - 1][0])
            and _mergeable_quote_token(tokens[j][0])
        ):
            tok += tokens[j][0]
            end = tokens[j][2]
            j += 1
        merged.append((tok, start, end))
        i = j
    return tuple(merged)


def _tokenize_quote_text_with_spans(text: str) -> List[Tuple[str, int, int]]:
    return list(_tokenize_quote_text_spans_cached(text or ""))


def _tokenize_quote_text(text: str) -> List[str]:
    return [tok for tok, _start, _end in _tokenize_quote_text_spans_cached(text or "")]


_BRACKET_INITIAL_WORD_RE = re.compile(r"\[([A-Za-z])\]([A-Za-z]+)")
_PLAIN_WORD_TOKEN_RE = re.compile(r"[^\W_]+(?:'[^\W_]+)*")


@lru_cache(maxsize=65536)
def _quote_diff_equiv_token(tok: str) -> str:
    if tok in _DOUBLE_QUOTE_TOKENS:
        return '"'
    if tok in _SINGLE_QUOTE_TOKENS:
        return "'"
    if _is_dash_quote_token(tok):
        return "-"
    norm = _normalize_quote_dash_chars((tok or "").replace("\u2019", "'").replace("\u2018", "'"))
    m = _BRACKET_INITIAL_WORD_RE.fullmatch(norm)
    if m:
        norm = m.group(1) + m.group(2)
    if _PLAIN_WORD_TOKEN_RE.fullmatch(norm):
        return norm.lower()
    return norm


_WORD_BRACKET_SUFFIX_GROUPS_RE = re.compile(r"([A-Za-z]+)((?:\[[A-Za-z]+\])+)")
_BRACKET_LETTERS_RE = re.compile(r"\[([A-Za-z]+)\]")
_WORD_BRACKET_SUFFIX_RE = re.compile(r"[A-Za-z]+(?:\[[A-Za-z]+\])+")


def _quote_source_alignment_aliases(quote_tokens: List[str]) -> Dict[str, str]:
    aliases: Dict[str, str] = {}
    for tok in quote_tokens:
        norm = _quote_diff_equiv_token(tok)
        m = _WORD_BRACKET_SUFFIX_GROUPS_RE.fullmatch(norm)
        if not m:
            continue
        base = m.group(1).lower()
        suffix = "".join(_BRACKET_LETTERS_RE.findall(m.group(2))).lower()
        expanded = f"{base}{suffix}"
        canonical = f"{base}|{expanded}"
        aliases[base] = canonical
        aliases[expanded] = canonical
    return aliases


def _quote_source_alignment_token(tok: str, aliases: Optional[Dict[str, str]] = None) -> str:
    norm = _quote_diff_equiv_token(tok)
    if _WORD_BRACKET_SUFFIX_RE.fullmatch(norm):
        norm = _BRACKET_LETTERS_RE.sub("", norm)
    lowered = norm.lower()
    return (aliases or {}).get(lowered, lowered)


def _bracket_suffix_quote_exact_in_source(quote: str, source: str) -> bool:
    quote_tokens = _strip_trailing_match_punct_tokens(_tokenize_quote_text(quote or ""))
    aliases = _quote_source_alignment_aliases(quote_tokens)
    if not aliases:
        return False
    quote_words = [
        _quote_source_alignment_token(tok, aliases)
        for tok in quote_tokens
        if _is_wordlike_quote_token(tok)
    ]
    source_words = [
        _quote_source_alignment_token(tok, aliases)
        for tok in _tokenize_quote_text(source or "")
        if _is_wordlike_quote_token(tok)
    ]
    if not quote_words or len(quote_words) > len(source_words):
        return False
    last = len(source_words) - len(quote_words) + 1
    for idx in range(last):
        if source_words[idx:idx + len(quote_words)] == quote_words:
            return True
    return False


def _has_outer_quotes(text: str) -> bool:
    s = (text or "").strip()
    return len(s) >= 2 and s[0] in _QUOTE_TOKENS and s[-1] in _QUOTE_TOKENS


def _strip_trailing_match_punct_tokens(tokens: List[str]) -> List[str]:
    i = len(tokens)
    while i > 0 and tokens[i - 1] in _MATCH_IGNORE_TRAILING_PUNCT:
        i -= 1
    return tokens[:i]


_SIMPLE_QUOTE_WORD_RE = re.compile(rf"^{_WORD_TOKEN_PATTERN}$")


_EDITORIAL_INITIAL_CASE_MARKER_RE = re.compile(r"^\[([A-Za-z])\](?=[A-Za-z])")


def _collapse_editorial_initial_case_marker(text: str) -> str:
    # Narrow for now; we may broaden this to cover other correct authorial
    # bracket edits that should still count as perfect quote matches.
    return _EDITORIAL_INITIAL_CASE_MARKER_RE.sub(r"\1", text or "")


def _find_exact_quote_text(needle: str, haystack: str) -> int:
    needle = (needle or "").strip()
    haystack = haystack or ""
    if not needle or not haystack:
        return -1
    if _SIMPLE_QUOTE_WORD_RE.fullmatch(needle):
        m = re.search(
            rf"(?<![^\W_]){re.escape(needle)}(?![^\W_])",
            haystack,
            flags=re.IGNORECASE,
        )
        return m.start() if m else -1
    return haystack.lower().find(needle.lower())


def _exact_quote_text_in_source(quote: str, source: str) -> bool:
    if _find_exact_quote_text(quote, source) >= 0:
        return True
    quote_collapsed = _collapse_editorial_initial_case_marker(quote)
    source_collapsed = _collapse_editorial_initial_case_marker(source)
    if quote_collapsed == quote and source_collapsed == source:
        return False
    return _find_exact_quote_text(quote_collapsed, source_collapsed) >= 0


def _normalize_quote_exact_compare(text: str) -> str:
    return _collapse_editorial_initial_case_marker(_normalize_quote_compare(text))


def _meaningful_quote_tokens(text: str) -> List[str]:
    tokens = _strip_trailing_match_punct_tokens(_tokenize_quote_text(text or ""))
    return [
        tok
        for tok in tokens
        if tok not in _QUOTE_TOKENS
        and tok not in _MATCH_IGNORE_TRAILING_PUNCT
        and tok.strip()
    ]


def _is_only_bracketed_quote_token(text: str) -> bool:
    tokens = _meaningful_quote_tokens(text)
    return len(tokens) == 1 and bool(re.fullmatch(r"\[[^\]]+\]", tokens[0]))


def _quote_original_display(q: Dict[str, Any]) -> str:
    qt = (q.get("inner") or q.get("raw") or "").strip()
    open_q, close_q = _infer_outer_quote_chars(q.get("raw") or "", q.get("style"))
    return _ensure_outer_quotes(qt, open_q, close_q)


def _quote_dedupe_key(text: str) -> str:
    return _normalize_quote_exact_compare(text or "").lower()


def _trim_extra_display_trailing_punctuation(text: str, original: str) -> str:
    if re.search(r"[.,;:!?\u2026]\s*$", _normalize_quote_compare(original or "")):
        return text
    s = (text or "").strip()
    if len(s) >= 2 and s[0] in _QUOTE_TOKENS and s[-1] in _QUOTE_TOKENS:
        inner = re.sub(r"[.,;:!?\u2026]+$", "", s[1:-1].strip()).strip()
        return f"{s[0]}{inner}{s[-1]}"
    return re.sub(r"[.,;:!?\u2026]+$", "", s).strip()


def _is_long_source_text(source: str) -> bool:
    return len(source or "") > _LONG_SOURCE_CHAR_LIMIT


def _is_short_quote_for_long_source(quote: str, source: str) -> bool:
    if not _is_long_source_text(source):
        return False
    return len(_quote_word_tokens(quote)) < 3


def _quote_not_found_note(quote_entries: List[Dict[str, Any]], source: str) -> str:
    quotes = [
        (q.get("inner") or q.get("raw") or "").strip()
        for q in quote_entries
        if (q.get("inner") or q.get("raw") or "").strip()
    ]
    if quotes and any(_is_short_quote_for_long_source(qt, source) for qt in quotes):
        if all(_is_short_quote_for_long_source(qt, source) for qt in quotes):
            return _SHORT_QUOTE_LONG_SOURCE_NOTE
        return _MIXED_SHORT_QUOTE_LONG_SOURCE_NOTE
    return "Quote not found in source text."


@lru_cache(maxsize=16)
def _quote_word_tokens_cached(text: str) -> Tuple[str, ...]:
    return tuple(m.group(0).lower() for m in _WORD_TOKEN_RE.finditer(text))


def _quote_word_tokens(text: str) -> List[str]:
    return list(_quote_word_tokens_cached(text or ""))


_OBVIOUS_NON_CONTENT_QUOTE_WORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "been",
    "being",
    "by",
    "for",
    "from",
    "in",
    "is",
    "it",
    "its",
    "of",
    "on",
    "or",
    "that",
    "the",
    "their",
    "to",
    "was",
    "were",
    "with",
}


def _content_words_excluding_obvious_noncontent(text: str) -> List[str]:
    words = _quote_word_tokens(text)
    return [
        word
        for word in words
        if word not in _OBVIOUS_NON_CONTENT_QUOTE_WORDS
        and not (len(word) == 1 and word.isalpha())
    ]


def _has_plausible_partial_content_overlap(quote: str, source: str) -> bool:
    quote_words = list(dict.fromkeys(_content_words_excluding_obvious_noncontent(quote)))
    if not quote_words:
        return True
    source_words = set(_content_words_excluding_obvious_noncontent(source))
    if not source_words:
        return False

    overlap = sum(1 for word in quote_words if word in source_words)
    if overlap <= 0:
        return False
    if len(quote_words) <= 2:
        return True
    if len(quote_words) <= 4:
        return overlap >= 2
    return (overlap / len(quote_words)) >= 0.35


def _quote_anchor_phrases(words: List[str]) -> List[List[str]]:
    phrases: List[List[str]] = []
    seen: set[Tuple[str, ...]] = set()
    for n in (8, 6, 5, 4, 3):
        if len(words) < n:
            continue
        starts = [0, len(words) // 4, len(words) // 2, max(0, len(words) - n)]
        for start in starts:
            start = max(0, min(start, len(words) - n))
            phrase = tuple(words[start:start + n])
            if phrase and phrase not in seen:
                seen.add(phrase)
                phrases.append(list(phrase))
    return phrases


def _find_word_phrase(words: List[str], phrase: List[str]) -> Optional[int]:
    if not words or not phrase or len(phrase) > len(words):
        return None
    last = len(words) - len(phrase) + 1
    first = phrase[0]
    n = len(phrase)
    i = -1
    while True:
        try:
            i = words.index(first, i + 1, last)
        except ValueError:
            return None
        if words[i:i + n] == phrase:
            return i


def _long_source_quote_score(q_tokens: List[str], source: str) -> float:
    import difflib

    q_words = [tok.lower() for tok in q_tokens if _is_wordlike_quote_token(tok)]
    if not q_words:
        return 0.0
    s_words = _quote_word_tokens(source)
    if not s_words:
        return 0.0

    source_word_set = set(s_words)
    overlap = sum(1 for word in q_words if word in source_word_set) / max(1, len(q_words))
    if overlap < 0.6:
        return overlap

    best = 0.0
    seen_windows: set[Tuple[int, int]] = set()
    for phrase in _quote_anchor_phrases(q_words):
        idx = _find_word_phrase(s_words, phrase)
        if idx is None:
            continue
        start = max(0, idx - len(q_words))
        end = min(len(s_words), idx + len(phrase) + len(q_words) * 3)
        key = (start, end)
        if key in seen_windows:
            continue
        seen_windows.add(key)
        window_words = s_words[start:end]
        sm = difflib.SequenceMatcher(a=window_words, b=q_words, autojunk=False)
        matched = sum(block.size for block in sm.get_matching_blocks() if block.size)
        best = max(best, matched / max(1, len(q_words)))
        if best >= 0.98:
            break

    if best:
        return best
    return min(overlap, 0.59)


def _long_source_quote_region(quote: str, full_text: str, window: int = 300) -> str:
    q_words = _quote_word_tokens(quote)
    if not q_words:
        return ""
    for phrase in _quote_anchor_phrases(q_words):
        pattern = r"\b" + r"\W+".join(re.escape(word) for word in phrase) + r"\b"
        m = re.search(pattern, full_text or "", flags=re.IGNORECASE)
        if not m:
            continue
        start = max(0, m.start() - window)
        end = min(len(full_text), m.end() + len(quote) + window)
        return full_text[start:end]
    return ""


def _infer_outer_quote_chars(raw: str, style: Optional[str]) -> Tuple[str, str]:
    s = (raw or "").strip()
    if len(s) >= 2 and s[0] in _QUOTE_TOKENS and s[-1] in _QUOTE_TOKENS:
        return s[0], s[-1]
    style_norm = (style or "").strip().upper()
    if style_norm == "SMART":
        return OPEN_SMART_DQ, CLOSE_SMART_DQ
    if style_norm in {"STRAIGHT", "MIXED"}:
        return STRAIGHT_DQ, STRAIGHT_DQ
    return "", ""


def _ensure_outer_quotes(text: str, open_q: str, close_q: str) -> str:
    s = (text or "").strip()
    if not s:
        return s
    if not open_q and not close_q:
        return s
    if not open_q:
        open_q = close_q
    if not close_q:
        close_q = open_q

    has_open = s[0] in _QUOTE_TOKENS
    has_close = s[-1] in _QUOTE_TOKENS
    if has_open and has_close:
        return s
    if has_open and not has_close:
        return s + close_q
    if not has_open and has_close:
        return open_q + s
    return f"{open_q}{s}{close_q}"


def _quote_role(tok: str, state: Dict[str, bool]) -> Optional[str]:
    if tok == "\"":
        role = "open" if not state.get("double") else "close"
        state["double"] = not state.get("double")
        return role
    if tok == "'":
        role = "open" if not state.get("single") else "close"
        state["single"] = not state.get("single")
        return role
    if tok in {"\u201c", "\u00ab", "\u201e"}:
        return "open"
    if tok in {"\u201d", "\u00bb"}:
        return "close"
    if tok in {"\u2018", "\u201a"}:
        return "open"
    if tok in {"\u2019"}:
        return "close"
    return None


def _join_quote_tokens(tokens: List[str], prefer_single_quotes: bool) -> str:
    out = ""
    state = {"double": False, "single": False}
    prev_tok = ""
    prev_role = None
    open_punct = {"(", "[", "{"}
    close_punct = {")", "]", "}", ",", ".", ";", ":", "!", "?"}

    for tok in tokens:
        if prefer_single_quotes and tok in _DOUBLE_QUOTE_TOKENS:
            tok = "'"
        if _is_dash_quote_token(tok):
            tok = "-"

        role = _quote_role(tok, state) if tok in _QUOTE_TOKENS else None
        if not out:
            out = tok
        else:
            if _is_dash_quote_token(tok):
                out = out.rstrip() + tok
            elif _is_dash_quote_token(prev_tok):
                out += tok
            elif role == "close" or tok in close_punct:
                out += tok
            elif prev_tok in open_punct or prev_role == "open":
                out += tok
            else:
                out += " " + tok

        prev_tok = tok
        prev_role = role

    return _normalize_intra_word_hyphen_for_display(re.sub(r"\s+", " ", out).strip())


def _format_insert_segment(
    segment: List[str],
    source_segment: Optional[List[str]] = None,
) -> List[str]:
    if not segment:
        return []

    if source_segment and len(segment) == 1 and len(source_segment) == 1:
        q_tok = segment[0]
        s_tok = source_segment[0]
        if _is_wordlike_quote_token(q_tok) and _is_wordlike_quote_token(s_tok):
            if (
                len(q_tok) == len(s_tok)
                and q_tok[1:] == s_tok[1:]
                and q_tok[0].lower() == s_tok[0].lower()
                and q_tok[0] != s_tok[0]
            ):
                return [f"[{q_tok[0]}]{q_tok[1:]}"]

            char_level = _format_word_internal_quote_insertion(q_tok, s_tok)
            if char_level:
                return [char_level]

    if source_segment and len(segment) == 1:
        q_tok = segment[0]
        if _is_wordlike_quote_token(q_tok):
            for s_tok in source_segment:
                if not _is_wordlike_quote_token(s_tok):
                    continue
                char_level = _format_word_internal_quote_insertion(q_tok, s_tok)
                if char_level:
                    return [char_level]
                break

    if len(segment) == 1:
        tok = segment[0]
        if "[" in tok and "]" in tok:
            return [tok]
        if _is_wordlike_quote_token(tok):
            return [f"[{tok}]"]
        return [tok]

    if any("[" in t and "]" in t for t in segment):
        return segment

    if any(_is_wordlike_quote_token(t) for t in segment):
        seg_text = _join_quote_tokens(segment, prefer_single_quotes=False)
        return [f"[{seg_text}]"]

    return segment


def _format_equal_quote_segment(
    source_segment: List[str],
    quote_segment: List[str],
) -> List[str]:
    if len(source_segment) != len(quote_segment):
        return source_segment

    out: List[str] = []
    for s_tok, q_tok in zip(source_segment, quote_segment):
        if s_tok == q_tok:
            out.append(s_tok)
            continue
        if _quote_diff_equiv_token(s_tok) != _quote_diff_equiv_token(q_tok):
            out.append(s_tok)
            continue

        if _BRACKET_INITIAL_WORD_RE.fullmatch(q_tok or ""):
            out.append(q_tok)
            continue
        if (
            _is_wordlike_quote_token(q_tok)
            and _is_wordlike_quote_token(s_tok)
            and len(q_tok) == len(s_tok)
            and q_tok[1:] == s_tok[1:]
            and q_tok[0].lower() == s_tok[0].lower()
            and q_tok[0] != s_tok[0]
        ):
            out.append(f"[{q_tok[0]}]{q_tok[1:]}")
            continue
        out.append(s_tok)
    return out


def _format_word_internal_quote_insertion(q_tok: str, s_tok: str) -> str:
    """Render small one-word quote-side insertions at character granularity.

    Example: source "test" vs quote "testing" -> "test[ing]".
    Keep this narrow so unrelated word replacements still fall back to whole-word
    bracket display.
    """
    import difflib

    if not q_tok or not s_tok:
        return ""
    if "[" in q_tok or "]" in q_tok or "[" in s_tok or "]" in s_tok:
        return ""

    sm = difflib.SequenceMatcher(a=s_tok, b=q_tok, autojunk=False)
    opcodes = sm.get_opcodes()
    insertions = [op for op in opcodes if op[0] == "insert"]
    if not insertions or any(tag not in {"equal", "insert"} for tag, *_ in opcodes):
        return ""
    if len(insertions) > 1:
        return ""

    equal_chars = sum(i2 - i1 for tag, i1, i2, _j1, _j2 in opcodes if tag == "equal")
    inserted_chars = sum(j2 - j1 for tag, _i1, _i2, j1, j2 in insertions)
    if equal_chars < min(3, len(s_tok)):
        return ""
    if inserted_chars > max(4, len(s_tok) // 2 + 1):
        return ""

    out: List[str] = []
    for tag, i1, i2, j1, j2 in opcodes:
        if tag == "equal":
            out.append(s_tok[i1:i2])
        elif tag == "insert":
            out.append(f"[{q_tok[j1:j2]}]")
    return "".join(out)


def _build_corrected_citation(
    quote: str,
    excerpt: str,
    *,
    prefer_single_quotes: bool = True,
) -> str:
    """Build a corrected quote suitable to directly replace the original quote string.

    Rules (McGill Guide):
      - Spelling, capitalization, and internal punctuation should match the source.
      - Additions or changes are wrapped in brackets.
      - Omissions are indicated with ellipses, but avoid leading/trailing ellipses by default.
    """
    import difflib

    quote = (quote or "").strip()
    excerpt = (excerpt or "").strip()
    if not quote:
        return excerpt
    if not excerpt:
        return quote

    quote_tokens = _tokenize_quote_text(quote)
    source_tokens = _tokenize_quote_text(excerpt)
    source_tokens = _strip_unmatched_source_wrapper_quotes(source_tokens, quote_tokens)

    source_cmp = [_quote_diff_equiv_token(tok) for tok in source_tokens]
    quote_cmp = [_quote_diff_equiv_token(tok) for tok in quote_tokens]
    sm = difflib.SequenceMatcher(a=source_cmp, b=quote_cmp, autojunk=False)

    out: List[str] = []

    def add_ellipsis():
        if not out or out[-1] != "...":
            out.append("...")

    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            out.extend(_format_equal_quote_segment(source_tokens[i1:i2], quote_tokens[j1:j2]))
        elif tag == "delete":
            deleted = source_tokens[i1:i2]
            if any(_is_wordlike_quote_token(t) for t in deleted):
                if out and j1 < len(quote_tokens):
                    add_ellipsis()
            else:
                out.extend(deleted)
        elif tag == "insert":
            inserted = quote_tokens[j1:j2]
            out.extend(_format_insert_segment(inserted))
        elif tag == "replace":
            replaced = quote_tokens[j1:j2]
            out.extend(_format_insert_segment(replaced, source_tokens[i1:i2]))

    return _join_quote_tokens(out, prefer_single_quotes=prefer_single_quotes)


def _strip_unmatched_source_wrapper_quotes(
    source_tokens: List[str],
    quote_tokens: List[str],
) -> List[str]:
    source = list(source_tokens)
    if not source or not quote_tokens:
        return source

    quote_starts_with_wrapper = quote_tokens[0] in _DOUBLE_QUOTE_TOKENS
    quote_core = _strip_trailing_match_punct_tokens(quote_tokens)
    quote_ends_with_wrapper = bool(quote_core and quote_core[-1] in _DOUBLE_QUOTE_TOKENS)

    if not quote_starts_with_wrapper:
        while source and source[0] in _DOUBLE_QUOTE_TOKENS:
            source.pop(0)

    if not quote_ends_with_wrapper:
        while source and source[-1] in _DOUBLE_QUOTE_TOKENS:
            source.pop()
        idx = len(source) - 1
        while idx >= 0 and source[idx] in _MATCH_IGNORE_TRAILING_PUNCT:
            idx -= 1
        if idx >= 0 and source[idx] in _DOUBLE_QUOTE_TOKENS:
            del source[idx]

    return source


_WHITESPACE_RUN_RE = re.compile(r"\s+")
_PRE_PUNCT_STRAY_DQUOTE_RE = re.compile(r'(?<=[A-Za-z0-9])"\s*(?=[.,;:!?…])')
_TRAILING_COMPARE_PUNCT_RE = re.compile(r"[.,;:!?…]+$")


@lru_cache(maxsize=4096)
def _normalize_quote_compare(text: str) -> str:
    s = (text or "").strip()
    if not s:
        return ""
    s = s.replace("“", '"').replace("”", '"').replace("’", "'").replace("‘", "'")
    s = _normalize_intra_word_hyphen_for_compare(s)
    s = _WHITESPACE_RUN_RE.sub(" ", s).strip()
    if len(s) >= 2 and s[0] in _QUOTE_TOKENS and s[-1] in _QUOTE_TOKENS:
        s = s[1:-1].strip()
    s = _PRE_PUNCT_STRAY_DQUOTE_RE.sub("", s)
    s = _TRAILING_COMPARE_PUNCT_RE.sub("", s).strip()
    return s


def _quote_match_score(quote: str, source: str) -> float:
    import difflib

    q = (quote or "").strip()
    s = (source or "").strip()
    if not q or not s:
        return 0.0

    q_tokens = _strip_trailing_match_punct_tokens(_tokenize_quote_text(q))
    if not q_tokens:
        return 0.0

    q_norm = _normalize_quote_exact_compare(q).lower()
    if q_norm and _exact_quote_text_in_source(q_norm, _normalize_quote_exact_compare(s).lower()):
        return 1.0
    if _bracket_suffix_quote_exact_in_source(q, s):
        return 1.0

    if len(s) > _LONG_SOURCE_CHAR_LIMIT:
        return _long_source_quote_score(q_tokens, s)

    s_tokens = _tokenize_quote_text(s)
    if not s_tokens:
        return 0.0
    if len(s_tokens) > _LONG_SOURCE_TOKEN_LIMIT:
        return _long_source_quote_score(q_tokens, s)

    sm = difflib.SequenceMatcher(a=s_tokens, b=q_tokens, autojunk=False)
    matched = sum(block.size for block in sm.get_matching_blocks() if block.size)
    return matched / max(1, len(q_tokens))






def _search_full_document(
    root, quote_texts: List[str], result: Dict[str, Any],
    min_match: float, strong_match: float,
) -> None:
    """Search full document text via #originalDocument for legacy CanLII pages
    that lack modern paragraph anchors. Sets result but no pinpoint data."""
    import re as _re
    try:
        nodes = root.xpath('//*[@id="originalDocument"]')
    except Exception:
        return
    if not nodes:
        return
    text = "".join(nodes[0].itertext())
    text = _re.sub(r"\s+", " ", text).strip()
    if not text:
        return

    best_score = 0.0
    best_quote = ""
    for qt in quote_texts:
        score = _quote_match_score(qt, text)
        if score > best_score:
            best_score = score
            best_quote = qt

    if best_score >= strong_match:
        result["found"] = True
        result["exact"] = True
        result["source_text"] = text
        result["source_tag"] = "CANLII"
    elif best_score >= min_match and _has_plausible_partial_content_overlap(
        best_quote,
        _trim_to_text_region(best_quote, text, window=400) or text,
    ):
        result["found"] = True
        result["exact"] = False
        result["source_text"] = text
        result["source_tag"] = "CANLII"


def _search_a2aj_fallback(
    row: Dict[str, Any], quote_texts: List[str], result: Dict[str, Any],
    min_match: float, strong_match: float,
) -> None:
    """Fall back to A2AJ full-text search when CanLII HTML isn't available."""
    if not USE_A2AJ:
        return

    a2aj_text = _fetch_a2aj_source_text_for_row(row)
    if not a2aj_text:
        link_cite = _canlii_doc_citation_from_url((row.get("citation_part_link") or "").strip())
        if link_cite:
            a2aj_text = _fetch_a2aj_source_text_for_row(row, link_cite, "case")
    if not a2aj_text:
        return

    # Score against full text
    best_score = 0.0
    best_quote = ""
    for qt in quote_texts:
        score = _quote_match_score(qt, a2aj_text)
        if score > best_score:
            best_score = score
            best_quote = qt

    if best_score >= strong_match:
        result["found"] = True
        result["exact"] = True
        result["source_tag"] = "A2AJ"
    elif best_score >= min_match and _has_plausible_partial_content_overlap(
        best_quote,
        _trim_to_text_region(best_quote, a2aj_text, window=400) or a2aj_text,
    ):
        result["found"] = True
        result["exact"] = False
        result["source_tag"] = "A2AJ"


_CHAR_LIMIT_GAP = 300


def _trim_to_text_region(quote: str, full_text: str, window: int = _CHAR_LIMIT_GAP) -> str:
    """Find the closest match region for the quote within full_text and return
    a window around it. This prevents the difflib from matching tokens across
    distant locations in large documents."""
    qt = (quote or "").strip()
    ft = (full_text or "").strip()
    if not qt or not ft:
        return full_text

    # Exact match — fastest path
    idx = _find_exact_quote_text(qt[:40], ft)
    if idx >= 0:
        start = max(0, idx - window // 2)
        end = min(len(ft), idx + len(qt) + window // 2)
        return ft[start:end]

    if len(ft) > _LONG_SOURCE_CHAR_LIMIT:
        return _long_source_quote_region(qt, ft, window=window)

    # Sliding window: find the best-matching region, then slice from raw string
    qt_tokens = _tokenize_quote_text(qt)
    if not qt_tokens:
        return ft
    step = max(1, len(qt_tokens) // 2)
    ft_tokens = _tokenize_quote_text(ft)
    # Build a position map: token index → original string position
    tok_to_pos = []
    search_from = 0
    ft_lower = ft.lower()
    for tok in ft_tokens:
        pos = ft_lower.find(tok.lower(), search_from)
        if pos < 0:
            pos = search_from
        tok_to_pos.append(pos)
        search_from = pos + max(1, len(tok))
    best_score = 0.0
    best_pos = 0
    for i in range(0, len(ft_tokens) - len(qt_tokens) + 1, step):
        chunk = " ".join(ft_tokens[i:i + len(qt_tokens) * 2])
        score = _quote_match_score(qt, chunk)
        if score > best_score:
            best_score = score
            best_pos = i
    start_tok = max(0, best_pos - len(qt_tokens))
    end_tok = min(len(ft_tokens) - 1, best_pos + len(qt_tokens) * 3)
    raw_start = tok_to_pos[start_tok]
    raw_end = tok_to_pos[end_tok] + len(ft_tokens[end_tok])
    return ft[raw_start:raw_end]


def _strip_source_leading_marker(fragment: str, quote: str = "") -> str:
    text = re.sub(r"\s+", " ", str(fragment or "")).strip()
    if not text:
        return ""
    quote_words = _quote_word_tokens(quote or "")
    if quote_words and quote_words[0].isdigit():
        return text
    stripped = re.sub(
        r"^(?:\[\s*\d{1,4}\s*\]|\d{1,4}\s*\(\d{1,4}\)|\d{1,4})\s+",
        "",
        text,
        count=1,
    ).strip()
    return stripped or text


def _source_side_quote_fragment_text(source_region: str, quote: str, max_chars: int = 900) -> str:
    """Return the source-side text span that best corresponds to the quote.

    The manuscript quote is only the search key; text fragments should point at
    what actually appears in the source.
    """
    import difflib

    source_region = source_region or ""
    quote = quote or ""
    source_spans = _tokenize_quote_text_with_spans(source_region)
    quote_tokens = _tokenize_quote_text(quote)
    if not source_spans or not quote_tokens:
        return ""

    source_tokens = [tok for tok, _start, _end in source_spans]
    alignment_aliases = _quote_source_alignment_aliases(quote_tokens)
    source_cmp = [_quote_source_alignment_token(tok, alignment_aliases) for tok in source_tokens]
    quote_cmp = [_quote_source_alignment_token(tok, alignment_aliases) for tok in quote_tokens]
    sm = difflib.SequenceMatcher(a=source_cmp, b=quote_cmp, autojunk=False)

    aligned_ranges: List[Tuple[int, int]] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag not in {"equal", "replace"} or i2 <= i1 or j2 <= j1:
            continue
        source_word_count = sum(1 for tok in source_tokens[i1:i2] if _is_wordlike_quote_token(tok))
        quote_word_count = sum(1 for tok in quote_tokens[j1:j2] if _is_wordlike_quote_token(tok))
        if tag == "replace" and (
            source_word_count > quote_word_count + 2
            or quote_word_count > source_word_count + 2
        ):
            continue
        source_slice = source_tokens[i1:i2]
        quote_slice = quote_tokens[j1:j2]
        if not any(_is_wordlike_quote_token(tok) for tok in source_slice):
            continue
        if not any(_is_wordlike_quote_token(tok) for tok in quote_slice):
            continue
        aligned_ranges.append((i1, i2))

    if not aligned_ranges:
        return ""

    clusters: List[List[Tuple[int, int]]] = []
    for rng in aligned_ranges:
        if not clusters or rng[0] - clusters[-1][-1][1] > 8:
            clusters.append([rng])
        else:
            clusters[-1].append(rng)

    def _cluster_score(cluster: List[Tuple[int, int]]) -> Tuple[int, int]:
        start = cluster[0][0]
        end = cluster[-1][1]
        word_count = sum(
            1
            for tok in source_tokens[start:end]
            if _is_wordlike_quote_token(tok)
        )
        return (word_count, -(end - start))

    best_cluster = max(clusters, key=_cluster_score)
    start_idx = best_cluster[0][0]
    end_idx = best_cluster[-1][1]
    if end_idx <= start_idx:
        return ""

    raw_start = source_spans[start_idx][1]
    raw_end = source_spans[end_idx - 1][2]
    fragment = re.sub(r"\s+", " ", source_region[raw_start:raw_end]).strip()
    fragment = _strip_source_leading_marker(fragment, quote)
    if not fragment or _is_only_bracketed_quote_token(fragment):
        return ""
    if len(fragment) > max_chars:
        return ""
    return fragment


def _build_source_side_text_fragment_url(
    url: str,
    source_fragment: str,
    source_context: str = "",
) -> Tuple[str, int]:
    context = str(source_context or "").strip() or str(source_fragment or "").strip()
    built = _build_source_side_text_fragment_directive(source_fragment, context, url)
    if built:
        return (_append_text_fragment_directives(url, [built.directive]), 1)
    return (url, 0)


def _is_canlii_html_link(url: str) -> bool:
    candidate = _canlii_source_lookup_url(_sanitize_url_candidate(url))
    base, _frag = _split_url(candidate)
    if "canlii.org" not in (base or "").lower():
        return False
    return urlsplit(base).path.lower().endswith(".html")




def _build_preferred_source_fragment_url(
    url: str,
    source_fragment: str,
    source_context: str = "",
    corrected_text: str = "",
) -> Tuple[str, int, str]:
    source_target, source_count = _build_source_side_text_fragment_url(
        url,
        source_fragment,
        source_context,
    )
    return (source_target, source_count, "source")



def _format_pinpoint_summary(pinpoints: List[str], limit: int = 2) -> str:
    pp = [str(p).strip() for p in pinpoints if str(p).strip()]
    pp_str = ", ".join(pp[:limit])
    if len(pp) > limit:
        pp_str += f" [+{len(pp) - limit} more instances]"
    return pp_str


def _first_pinpoint_from_summary(summary: str) -> str:
    m = re.search(r"\b(?:par\d+[A-Za-z]?|sec[^\s,\[\]]+)\b", summary or "", flags=re.IGNORECASE)
    if not m:
        return ""
    return _normalize_anchor_fragment(m.group(0))


def _build_alternate_pinpoint_fragment_url(
    link: str,
    pinpoint_summary: str,
    corrected: str,
    source_fragment: str = "",
    *,
    prefer_range: bool = False,
) -> str:
    target_pinpoint = _first_pinpoint_from_summary(pinpoint_summary)
    if not target_pinpoint:
        return ""
    base, _frag = _split_url(_sanitize_url_candidate(link))
    if not base or base.lower() == "other":
        return ""
    target = _a2aj_link_for_pinpoint(base, target_pinpoint)
    if not target:
        host = urlsplit(base).netloc.casefold()
        if (
            target_pinpoint.casefold().startswith("par")
            and host == "decisions.scc-csc.ca"
        ):
            target = _recombine_url(base, target_pinpoint)
        else:
            return ""
    if source_fragment:
        # A directive is only trustworthy when independent source text is
        # available to disambiguate it (a repeated phrase highlights
        # whichever copy the browser meets first); without it the bare
        # paragraph anchor is the honest link.
        source_text = (
            _a2aj_registered_paragraph_text(target, target_pinpoint)
            or _a2aj_registered_section_text(target, target_pinpoint)
        )
        if source_text:
            fragment_result = _build_quote_check_fragment_url(
                target,
                source_fragment,
                source_text,
                corrected,
                prefer_range=prefer_range,
            )
            if fragment_result.fragment_count:
                return fragment_result.url
        return target
    return target




def _has_textual_pinpoint_link(row: Dict[str, Any]) -> bool:
    link = (row.get("citation_part_link") or "").strip()
    if not link or link.lower() == "other":
        return False
    _base, frag = _split_url(link)
    return bool(frag and not frag.lower().startswith("page="))


def _coerce_pinpoint_fragments(value: Any) -> List[str]:
    if not value:
        return []
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except Exception:
            parsed = [value]
    else:
        parsed = value
    if not isinstance(parsed, list):
        parsed = [parsed]
    return [str(item).strip() for item in parsed if str(item).strip()]


def _textual_pinpoint_from_link_or_fragments(link: str, fragments: Any = None) -> str:
    _base, frag = _split_url((link or "").strip())
    frag = frag.split(":~:text=", 1)[0]
    norm_frag = _normalize_anchor_fragment(frag)
    if norm_frag and not norm_frag.lower().startswith("page="):
        return norm_frag
    for raw in _coerce_pinpoint_fragments(fragments):
        norm = _normalize_anchor_fragment(raw)
        if norm and not norm.lower().startswith("page="):
            return norm
    return ""


def _first_page_pinpoint(row: Dict[str, Any]) -> Optional[int]:
    raw = row.get("page_pinpoints")
    if not raw:
        return None
    try:
        values = json.loads(raw) if isinstance(raw, str) else list(raw)
    except Exception:
        return None
    if not isinstance(values, list) or not values:
        return None
    try:
        return int(values[0])
    except (TypeError, ValueError):
        return None


def _journal_page_pinpoint_for_row(row: Dict[str, Any]) -> str:
    page_label = _first_page_pinpoint(row)
    if page_label is None:
        return ""
    article_id = _journal_article_id_for_row(row)
    if article_id and journal_search.pdf_page_for_label(article_id, page_label) is not None:
        return f"page {page_label}"
    return ""


_JOURNAL_DB_PAGE_MARKER_RE = re.compile(r"\[\s*page\s+([^\]\r\n]+?)\s*\]", re.IGNORECASE)


def _journal_db_page_spans(text: str) -> List[Tuple[str, int, int]]:
    markers = list(_JOURNAL_DB_PAGE_MARKER_RE.finditer(text or ""))
    spans: List[Tuple[str, int, int]] = []
    for idx, marker in enumerate(markers):
        label = re.sub(r"\s+", " ", marker.group(1)).strip()
        if not label:
            continue
        start = marker.start()
        end = markers[idx + 1].start() if idx + 1 < len(markers) else len(text)
        spans.append((label, start, end))
    return spans


def _journal_db_page_text(text: str, page_label: Any) -> str:
    target = re.sub(r"\s+", " ", str(page_label or "")).strip()
    if not target:
        return ""
    for label, start, end in _journal_db_page_spans(text):
        if label == target:
            return (text or "")[start:end].strip()
    return ""


def _journal_db_page_for_region(text: str, source_region: str) -> str:
    if not text or not source_region:
        return ""
    needle = source_region.strip()
    idx = text.find(needle)
    spans = _journal_db_page_spans(text)
    if idx >= 0:
        for label, start, end in spans:
            if start <= idx < end:
                return f"page {label}"

    # Match regions are sometimes whitespace-normalized before reaching this
    # function.  Recover the page only when exactly one page contains the same
    # normalized region; ambiguity should continue to display as unknown.
    normalized_needle = re.sub(r"\s+", " ", needle).strip()
    if normalized_needle:
        matching_labels = [
            label
            for label, start, end in spans
            if normalized_needle in re.sub(r"\s+", " ", text[start:end]).strip()
        ]
        if len(matching_labels) == 1:
            return f"page {matching_labels[0]}"
    return ""


def _journal_db_pages_for_quote(
    text: str, quote_text: str, min_score: float = 0.0
) -> List[str]:
    """Return mapped journal pages that satisfy the quote match threshold."""
    needle = re.sub(r"\s+", " ", quote_text or "").strip().casefold()
    if not needle:
        return []
    pages: List[str] = []
    for label, start, end in _journal_db_page_spans(text):
        page_text = text[start:end]
        normalized_page = re.sub(r"\s+", " ", page_text).strip().casefold()
        if needle in normalized_page or (
            min_score > 0 and _quote_match_score(quote_text, page_text) >= min_score
        ):
            pages.append(f"page {label}")
    return pages


def _page_label_from_match_pinpoint(value: str) -> Optional[int]:
    m = re.search(r"\bpage\s+(\d{1,5})\b", value or "", flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(1))
    except (TypeError, ValueError):
        return None


def _journal_db_page_link_for_row(row: Dict[str, Any], page_label: Any) -> str:
    try:
        label = int(page_label)
    except (TypeError, ValueError):
        return ""
    article_id = _journal_article_id_for_row(row)
    pdf_page = journal_search.pdf_page_for_label(article_id, label)
    if pdf_page is None:
        return ""
    link = (row.get("citation_part_link") or _ref_chain_origin_value(row, "_ref_chain_origin_citation_part_link") or "").strip()
    if not link or link.lower() == "other":
        return ""
    base, _frag = _split_url(link)
    if not base:
        return ""
    return f"{base}#page={pdf_page}"


def _quote_match_pinpoint_for_source(row: Dict[str, Any], source_tag: str) -> str:
    if source_tag == "anchor":
        return _textual_pinpoint_from_link_or_fragments(
            row.get("citation_part_link", ""),
            row.get("pinpoint_fragments"),
        )
    if source_tag == "origin_anchor":
        return _textual_pinpoint_from_link_or_fragments(
            row.get("_ref_chain_origin_citation_part_link", ""),
            row.get("_ref_chain_origin_pinpoint_fragments"),
        )
    if source_tag in _A2AJ_SOURCE_TAGS:
        link = (
            row.get("_ref_chain_origin_citation_part_link", "")
            if source_tag.startswith("origin_")
            else row.get("citation_part_link", "")
        )
        return _textual_pinpoint_from_link_or_fragments(link, None)
    return ""


def _coerce_anchor_segments(value: Any) -> List[Dict[str, str]]:
    if not value:
        return []
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except Exception:
            return []
    if not isinstance(value, list):
        return []

    segments: List[Dict[str, str]] = []
    for item in value:
        if isinstance(item, dict):
            fragment = _normalize_anchor_fragment(item.get("fragment") or item.get("pinpoint") or "")
            text = str(item.get("text") or "")
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            fragment = _normalize_anchor_fragment(item[0])
            text = str(item[1] or "")
        else:
            continue
        if fragment and text.strip():
            segments.append({"fragment": fragment, "text": text.strip()})
    return segments


def _best_anchor_segment_for_quotes(
    segments: List[Dict[str, str]],
    quote_entries: List[Dict[str, Any]],
    min_match: float,
) -> Optional[Dict[str, str]]:
    best: Optional[Tuple[float, int, Dict[str, str]]] = None
    for idx, segment in enumerate(segments):
        text = (segment.get("text") or "").strip()
        if not text:
            continue
        scores = [
            _quote_match_score((q.get("inner") or q.get("raw") or "").strip(), text)
            for q in quote_entries
            if (q.get("inner") or q.get("raw") or "").strip()
        ]
        if not scores:
            continue
        score = max(scores)
        if best is None or score > best[0]:
            best = (score, idx, segment)
    if best and best[0] >= min_match:
        return best[2]
    return None


def _pinpoint_link_for_source(row: Dict[str, Any], source_tag: str, pinpoint: str) -> str:
    pinpoint = _normalize_anchor_fragment(pinpoint)
    if not pinpoint or pinpoint.lower().startswith("page="):
        return ""
    if source_tag == "origin_anchor":
        link = str(row.get("_ref_chain_origin_citation_part_link") or "").strip()
    else:
        link = str(row.get("citation_part_link") or "").strip()
    if not link or link.lower() == "other":
        return ""
    base, _frag = _split_url(link)
    if not base:
        return ""
    return _recombine_url(base, pinpoint)


def _canlii_document_link_for_row(row: Dict[str, Any], source_tag: str) -> str:
    """Return the row's CanLII decision URL (a ``…/*.html`` document link) that
    the quote points at — from ``quote_match_link`` or ``citation_part_link``,
    origin-aware for reference rows. Empty if the row has no CanLII link. This is
    the base for both CanLII paragraph-block recovery and the A2AJ ``#parN``
    anchor, regardless of which source_tag verified the quote."""
    candidates: List[str] = []
    quote_match_link = str(row.get("quote_match_link") or "").strip()
    if quote_match_link:
        candidates.append(quote_match_link)
    if source_tag.startswith("origin_"):
        candidates.append(str(_ref_chain_origin_value(row, "_ref_chain_origin_citation_part_link") or "").strip())
    else:
        candidates.append(str(row.get("citation_part_link") or "").strip())
    for link in candidates:
        base, _frag = _split_url(link)
        if base and "canlii.org" in base.lower() and urlsplit(base).path.lower().endswith(".html"):
            return link
    return ""


def _canlii_node_anchor(root_url: str, node: Any, text: str = "") -> str:
    base, _frag = _split_url(_canlii_source_lookup_url(_sanitize_url_candidate(root_url)))
    path = urlsplit(base).path.lower() if base else ""

    nodes_to_scan: List[Any] = []
    if node is not None and isinstance(getattr(node, "tag", None), str):
        nodes_to_scan.append(node)
        nodes_to_scan.extend(
            child for child in node.iterdescendants()
            if isinstance(getattr(child, "tag", None), str)
        )
    for current in nodes_to_scan:
        for attr in ("id", "name", "data-lbh-p-anchor"):
            value = _normalize_anchor_fragment(current.get(attr) or "")
            if _CANLII_ANCHOR_RE.fullmatch(value or ""):
                return value

    current = node
    for _ in range(8):
        if current is None or not isinstance(getattr(current, "tag", None), str):
            break
        for attr in ("id", "name", "data-lbh-p-anchor"):
            value = _normalize_anchor_fragment(current.get(attr) or "")
            if _CANLII_ANCHOR_RE.fullmatch(value or ""):
                return value
        data_par = (current.get("data-viibes-parag") or "").strip()
        if data_par:
            if re.fullmatch(r"par\d+", data_par, flags=re.IGNORECASE):
                return _normalize_anchor_fragment(data_par)
            if re.fullmatch(r"\d{1,4}", data_par):
                return f"par{data_par}" if "/doc/" in path else data_par
        current = current.getparent()

    if "/doc/" in path:
        m = re.match(r"^\s*(?:\[\s*)?(\d{1,4})(?:\s*\])?(?:\s+|$)", text or "")
        if m:
            return f"par{m.group(1)}"
    return ""


def _canlii_pinpoint_label(anchor: str) -> str:
    anchor = _normalize_anchor_fragment(anchor)
    if not anchor:
        return ""
    if re.fullmatch(r"par\d+", anchor, flags=re.IGNORECASE):
        m = re.search(r"\d+", anchor)
        return f"par{m.group(0)}" if m else anchor
    if _SEC_ANCHOR_RE.fullmatch(anchor):
        return anchor
    return anchor






_A2AJ_SOURCE_TAGS = frozenset({"a2aj", "link_a2aj", "origin_a2aj", "a2aj_locked"})

_A2AJ_PARAGRAPH_MARK_RE = _a2aj_structure.PARAGRAPH_MARK_RE
_A2AJ_PAGE_MARK_RE = _a2aj_structure.PAGE_MARK_RE


def _a2aj_paragraph_index(text: str, *, min_run: int = 5) -> List[Tuple[int, int, int, str]]:
    return _a2aj_structure.paragraph_index(text, min_run=min_run)


def _a2aj_paragraph_structure(text: str) -> List[Tuple[int, int, int, str]]:
    return _a2aj_structure.paragraph_index(text)


def _locate_a2aj_paragraph(
    text: str,
    quote: str,
    link: str = "",
    *,
    paragraphs: Optional[List[Tuple[Any, ...]]] = None,
    min_score: float = 0.98,
) -> Optional[Dict[str, str]]:
    """Recover a paragraph pinpoint for a quote matched in flat A2AJ case text,
    by finding the numbered paragraph that contains it. Mirrors the CanLII
    uniqueness guard: return a pinpoint only when exactly ONE paragraph contains
    the quote, so short or repeated quotes stay unresolved on purpose.

    A2AJ paragraph numbers ARE the decision's own numbering, which CanLII exposes
    as ``#parN`` anchors — so when the quote links to a CanLII page we attach the
    anchor. The A2AJ->CanLII text-fragment path then
    builds ``#parN:~:text=…`` from it, and if the directive is unbuildable the
    link still lands at the paragraph via the bare anchor."""
    index = paragraphs if paragraphs is not None else _a2aj_paragraph_structure(text)
    if not index:
        return None
    qt = re.sub(r"\s+", " ", str(quote or "")).strip()
    if not qt:
        return None
    hits = []
    for item in index:
        num, start, end = item[:3]
        para_text = item[3] if len(item) > 3 else text[start:end]
        if _quote_match_score(qt, para_text) >= min_score:
            hits.append((num, para_text))
    distinct = {num for num, _text in hits}
    if len(distinct) != 1:
        return None  # zero or ambiguous — leave the pinpoint unknown
    num, para_text = hits[0]
    result: Dict[str, str] = {"text": para_text, "label": f"par{num}"}
    base, _frag = _split_url(_canlii_source_lookup_url(_sanitize_url_candidate(link)))
    if base and "canlii.org" in base.lower() and urlsplit(base).path.lower().endswith(".html"):
        result["link"] = _recombine_url(base, f"par{num}")
    return result


def _a2aj_page_index(text: str) -> List[Tuple[int, int, int, str]]:
    return _a2aj_structure.page_index(text)


def _a2aj_page_structure(text: str) -> List[Tuple[int, int, int, str]]:
    return _a2aj_structure.page_structure(text)


def _locate_a2aj_page(
    text: str,
    quote: str,
    *,
    pages: Optional[List[Tuple[Any, ...]]] = None,
    min_score: float = 0.98,
) -> Optional[Dict[str, str]]:
    hits = []
    for item in pages if pages is not None else _a2aj_page_structure(text):
        number, start, end = item[:3]
        page_text = item[3] if len(item) > 3 else text[start:end]
        if _quote_match_score(quote, page_text) >= min_score:
            hits.append((number, page_text))
    if len({number for number, _text in hits}) != 1:
        return None
    number, page_text = hits[0]
    return {"text": page_text, "label": f"page {number}"}


def _a2aj_section_structure(text: str) -> List[Tuple[str, int, int, str]]:
    return _a2aj_structure.section_structure(text)


def _locate_a2aj_section(
    text: str,
    quote: str,
    link: str = "",
    *,
    blocks: Optional[List[Tuple[str, str, int, int]]] = None,
    min_score: float = 0.98,
) -> Optional[Dict[str, str]]:
    hits = [
        (section, locator, text[start:end])
        for section, locator, start, end in (
            blocks if blocks is not None else _a2aj_structure.legislation_blocks(text)
        )
        for block_text in (text[start:end],)
        if _quote_match_score(quote, block_text) >= min_score
    ]
    if len({section for section, _locator, _text in hits}) != 1:
        return None
    section = hits[0][0]
    top_locator = f"sec{section}"
    specific = {locator for _section, locator, _text in hits if locator != top_locator}
    if len(specific) == 1:
        locator = next(iter(specific))
        _section, _locator, block_text = min(
            (item for item in hits if item[1] == locator), key=lambda item: len(item[2])
        )
    else:
        locator = top_locator
        _section, _locator, block_text = next(item for item in hits if item[1] == top_locator)
    result = {"text": block_text, "label": locator}
    base, _fragment = _split_url(link)
    if base and "canlii.org" in base.lower() and "/laws/" in urlsplit(base).path.lower():
        result["link"] = _recombine_url(base, f"sec{section}")
    return result


def _locate_a2aj_pinpoint(
    text: str, quote: str, link: str, structure: Dict[str, Any], *, min_score: float
) -> Optional[Dict[str, str]]:
    """Dispatch only to structures validated for this locked source."""
    kind = str(structure.get("type") or "")
    if kind == "section":
        return _locate_a2aj_section(
            text, quote, link, blocks=structure.get("blocks"), min_score=min_score
        )
    if kind == "paragraph":
        located = _locate_a2aj_paragraph(
            text,
            quote,
            link,
            paragraphs=structure.get("paragraphs"),
            min_score=min_score,
        )
        return located or _locate_a2aj_page(
            text, quote, pages=structure.get("pages"), min_score=min_score
        )
    if kind == "page":
        return _locate_a2aj_page(
            text, quote, pages=structure.get("pages"), min_score=min_score
        )
    return None


def _a2aj_link_for_pinpoint(link: str, pinpoint: str) -> str:
    """Attach an A2AJ-derived paragraph or top-level provision anchor."""
    base, _fragment = _split_url(_canlii_source_lookup_url(_sanitize_url_candidate(link)))
    if not base or "canlii.org" not in base.lower() or not urlsplit(base).path.lower().endswith(".html"):
        return ""
    label = str(pinpoint or "").strip()
    if re.fullmatch(r"par\d{1,4}", label, flags=re.IGNORECASE):
        return _recombine_url(base, label)
    provision = re.fullmatch(r"sec(\d{1,8}(?:[.-]\d{1,8}){0,3})(?:\([^)]+\))*", label, flags=re.IGNORECASE)
    if provision and "/laws/" in urlsplit(base).path.lower():
        return _recombine_url(base, "sec" + provision.group(1))
    return ""


def _a2aj_cited_law_structure(
    row: Dict[str, Any],
    scopes: _a2aj_pinpoint_scope.CitedScopes,
    anchor_text: str = "",
    locked_structure: Optional[Dict[str, Any]] = None,
) -> Tuple[str, Dict[str, Any]]:
    """Use mapped section text first; fetch only when no map is available."""
    citation = str(row.get("_a2aj_citation") or "").strip()
    language = str(row.get("_a2aj_language") or "en")
    sections = []
    for label in scopes.sections:
        match = re.match(r"sec(\d{1,8}(?:[.-]\d{1,8}){0,3})", label, re.IGNORECASE)
        if match and match.group(1) not in sections:
            sections.append(match.group(1))
    if not sections:
        return "", {}

    structure = locked_structure or {}
    if structure.get("source") == "section_map":
        source_blocks = structure.get("blocks") or []
        pieces: List[str] = []
        blocks: List[Tuple[str, str, int, int]] = []
        position = 0
        found_sections = 0
        for section in sections:
            locator = f"sec{section}"
            root = next(
                (
                    block
                    for block in source_blocks
                    if str(block[1]).casefold() == locator.casefold()
                ),
                None,
            )
            if root is None:
                continue
            _root_section, _root_locator, root_start, root_end = root
            section_text = anchor_text[root_start:root_end]
            if not section_text:
                continue
            if pieces:
                pieces.append("\n")
                position += 1
            pieces.append(section_text)
            for block_section, block_locator, start, end in source_blocks:
                if start < root_start or end > root_end:
                    continue
                blocks.append((
                    block_section,
                    block_locator,
                    position + start - root_start,
                    position + end - root_start,
                ))
            position += len(section_text)
            found_sections += 1
        if not blocks:
            return "", {}
        return "".join(pieces), {
            "status": "usable",
            "type": "section",
            "source": "section_map",
            "blocks": blocks,
            "count": found_sections,
        }

    if not citation:
        return "", {}

    pieces: List[str] = []
    blocks: List[Tuple[str, str, int, int]] = []
    position = 0
    for section in sections:
        lookup = a2aj_client.lookup_document(
            citation,
            "statute",
            section=section,
            language=language,
            search=False,
        )
        if lookup.status != "found" or not lookup.document or not lookup.document.text:
            continue
        section_text = _normalize_a2aj_source_text(lookup.document.text)
        if pieces:
            pieces.append("\n")
            position += 1
        pieces.append(section_text)
        blocks.extend(_a2aj_structure.single_section_blocks(
            section_text, section, start=position
        ))
        position += len(section_text)
    if not blocks:
        return "", {}
    return "".join(pieces), {
        "status": "usable",
        "type": "section",
        "blocks": blocks,
        "count": len(sections),
    }


def _should_report_source_text_failure(row: Dict[str, Any]) -> bool:
    """Whether a missing-source warning is actionable for this row."""
    link = _sanitize_url_candidate(row.get("citation_part_link") or "")
    parsed = urlsplit(link)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return False
    kind = str(row.get("citation_part_kind") or "").strip().lower()
    if kind in {"case", "unreported", "statute", "regulation", "legislation", "gazette"}:
        return True
    if kind == "journal":
        return bool(row.get("_journal_link_resolved"))
    return False


def _apply_quote_checks(
    rows: List[Dict[str, Any]],
    quotes_by_footnote: Dict[int, List[Dict[str, Any]]],
    *,
    min_match: float = 0.6,
    strong_match: float = 0.98,
) -> None:
    def _log_quote_row_timing(
        row: Dict[str, Any],
        started_at: float,
        quote_count: int,
        source_tag: str = "",
    ) -> None:
        if not TIMING_LOG_PATH:
            return
        link = (row.get("citation_part_link") or "").strip()
        base, _frag = _split_url(link)
        host = urlsplit(base).netloc.lower() if base else ""
        _timing_event(
            "quote_check_row",
            elapsed_s=round(time.perf_counter() - started_at, 3),
            footnote_id=row.get("footnote_id", ""),
            citation_part_index=row.get("citation_part_index", ""),
            citation_part_kind=row.get("citation_part_kind", ""),
            quote_count=quote_count,
            host=host,
            has_link=bool(link and link.lower() != "other"),
            source_tag=(row.get("_quote_source_tag") or source_tag or ""),
            quote_check_status=row.get("quote_check_status", ""),
        )

    for row in rows:
        _pause_gate()
        row_started_at = time.perf_counter()
        fid = int(row.get("footnote_id") or 0)
        quotes = quotes_by_footnote.get(fid, [])
        if not quotes:
            continue
        if TIMING_LOG_PATH:
            link = (row.get("citation_part_link") or "").strip()
            base, _frag = _split_url(link)
            _timing_event(
                "quote_check_row:start",
                footnote_id=row.get("footnote_id", ""),
                citation_part_index=row.get("citation_part_index", ""),
                citation_part_kind=row.get("citation_part_kind", ""),
                host=urlsplit(base).netloc.lower() if base else "",
                has_link=bool(link and link.lower() != "other"),
            )

        quote_entries: List[Dict[str, Any]] = []
        seen_quote_entry_keys: set[str] = set()
        for q in quotes:
            qt_inner = (q.get("quote_inner") or "").strip()
            qt_raw = (q.get("quote_raw") or "").strip()
            qt = qt_inner or qt_raw
            if qt:
                key = _quote_dedupe_key(qt)
                if key and key in seen_quote_entry_keys:
                    continue
                if key:
                    seen_quote_entry_keys.add(key)
                quote_entries.append(
                    {
                        "inner": qt_inner,
                        "raw": qt_raw,
                        "style": q.get("quote_delimiter_style"),
                    }
                )

        if not quote_entries:
            continue

        anchor_text = ""
        source_tag = ""
        source_pinpoint = ""
        locked_structure: Dict[str, Any] = {}
        a2aj_scopes = _a2aj_pinpoint_scope.CitedScopes()
        a2aj_resolutions: Dict[str, _a2aj_pinpoint_scope.QuoteResolution] = {}
        journal_db_full_text = ""
        article_id = _journal_article_id_for_row(row)
        if article_id:
            db_text = journal_search.get_article_text(article_id).strip()
            if db_text:
                journal_db_full_text = db_text
                cited_page = _first_page_pinpoint(row)
                cited_page_text = _journal_db_page_text(db_text, cited_page)
                if cited_page_text:
                    cited_page_score = max(
                        _quote_match_score((q.get("inner") or q.get("raw") or "").strip(), cited_page_text)
                        for q in quote_entries
                    )
                else:
                    cited_page_score = 0.0
                if cited_page_text and cited_page_score >= min_match:
                    anchor_text = cited_page_text
                    source_pinpoint = f"page {cited_page}"
                    row["quote_match_link"] = _journal_db_page_link_for_row(row, cited_page)
                    _ts_print(
                        f"  Journal DB text: using cited page {cited_page} "
                        f"({len(cited_page_text)} chars, score={cited_page_score:.2f})"
                    )
                else:
                    anchor_text = db_text
                source_tag = "journal_db"
                _ts_print(f"  Journal DB text: fetched {len(db_text)} chars for article_id={article_id}")

        if not anchor_text and USE_A2AJ:
            row_link = (row.get("citation_part_link") or "").strip()
            base, _fragment = _split_url(_canlii_source_lookup_url(row_link))
            locked = _A2AJ_LOCKED_DOCUMENTS.get((base or "").lower())
            if locked and locked.text:
                locked_structure = _A2AJ_LOCKED_STRUCTURES.get((base or "").lower(), {})
                if not locked_structure:
                    _register_a2aj_document(base, locked, "law" if "/laws/" in (base or "").lower() else "case")
                    locked_structure = _A2AJ_LOCKED_STRUCTURES.get((base or "").lower(), {})
                anchor_text = _A2AJ_LOCKED_TEXTS.get((base or "").lower(), "")
                if not anchor_text:
                    anchor_text, locked_structure = _a2aj_document_evidence(
                        locked,
                        "law" if "/laws/" in (base or "").lower() else "case",
                    )
                    _A2AJ_LOCKED_TEXTS[(base or "").lower()] = anchor_text
                    _A2AJ_LOCKED_STRUCTURES[(base or "").lower()] = locked_structure
                source_tag = "a2aj_locked"
                row["_a2aj_identity_locked"] = True
                row["_a2aj_url_reconciled"] = True
                row["_a2aj_dataset"] = locked.dataset
                row["_a2aj_citation"] = locked.citation
                row["_a2aj_source_url"] = locked.url
                row["_a2aj_structure"] = locked_structure
                row["_a2aj_language"] = locked.language
                row["_a2aj_structure_status"] = locked_structure.get("status", "unavailable")
                row["_a2aj_structure_type"] = locked_structure.get("type", "")
                row["_a2aj_structure_count"] = locked_structure.get("count", 0)
                _register_fragment_document_text(row_link, anchor_text)
                _register_fragment_document_text(locked.url, anchor_text)
                _ts_print(f"  A2AJ locked source: reused {len(anchor_text)} chars")

        if not anchor_text:
            current_anchor = (row.get("citation_part_anchor_text") or "").strip()
            if current_anchor:
                anchor_text = current_anchor
                source_tag = "anchor"
                source_pinpoint = _quote_match_pinpoint_for_source(row, source_tag)
                best_segment = _best_anchor_segment_for_quotes(
                    _coerce_anchor_segments(row.get("_citation_part_anchor_segments")),
                    quote_entries,
                    min_match,
                )
                if best_segment:
                    anchor_text = best_segment.get("text", "").strip() or anchor_text
                    source_pinpoint = best_segment.get("fragment", "") or source_pinpoint
                    segment_link = _pinpoint_link_for_source(row, source_tag, source_pinpoint)
                    if segment_link:
                        row["quote_match_link"] = segment_link
        if not anchor_text:
            origin_anchor = str(_ref_chain_origin_value(row, "_ref_chain_origin_citation_part_anchor_text") or "").strip()
            if origin_anchor:
                anchor_text = origin_anchor
                source_tag = "origin_anchor"
                source_pinpoint = _quote_match_pinpoint_for_source(row, source_tag)
                best_segment = _best_anchor_segment_for_quotes(
                    _coerce_anchor_segments(_ref_chain_origin_value(row, "_ref_chain_origin_anchor_segments")),
                    quote_entries,
                    min_match,
                )
                if best_segment:
                    anchor_text = best_segment.get("text", "").strip() or anchor_text
                    source_pinpoint = best_segment.get("fragment", "") or source_pinpoint
                    segment_link = _pinpoint_link_for_source(row, source_tag, source_pinpoint)
                    if segment_link:
                        row["quote_match_link"] = segment_link
        if not anchor_text:
            full_source_text = (row.get("_citation_part_full_source_text") or "").strip()
            if full_source_text:
                anchor_text = full_source_text
                source_tag = "full_source"
                _ts_print(f"  Full source text fallback: {len(full_source_text)} chars")
        if not anchor_text:
            origin_full_source_text = str(_ref_chain_origin_value(row, "_ref_chain_origin_full_source_text") or "").strip()
            if origin_full_source_text:
                anchor_text = origin_full_source_text
                source_tag = "origin_full_source"
                _ts_print(f"  Origin full source text fallback: {len(origin_full_source_text)} chars")
        if not anchor_text:
            # Try A2AJ fallback when browser text is unavailable and A2AJ is enabled
            if USE_A2AJ:
                a2aj_text = ""
                row_link = (row.get("citation_part_link") or "").strip()
                a2aj_candidates = [
                    (
                        (row.get("bare_citation") or "").strip(),
                        (row.get("citation_part_kind") or "").strip(),
                        "a2aj",
                        row_link,
                    ),
                    # The resolved link is the ground truth for which document
                    # this row points at (reference rows' bare is a supra/ibid
                    # form, and origin part indexes can point at a sibling
                    # citation) — derive the citation from the link slug.
                    (
                        _canlii_doc_citation_from_url(row_link),
                        "case",
                        "link_a2aj",
                        row_link,
                    ),
                    (
                        str(_ref_chain_origin_value(row, "_ref_chain_origin_bare_citation") or "").strip(),
                        str(_ref_chain_origin_value(row, "_ref_chain_origin_citation_part_kind") or "").strip(),
                        "origin_a2aj",
                        str(_ref_chain_origin_value(row, "_ref_chain_origin_citation_part_link") or "").strip(),
                    ),
                ]
                seen_a2aj: set[Tuple[str, str]] = set()
                for bare, kind, candidate_tag, register_url in a2aj_candidates:
                    if not bare or not kind or (bare, kind) in seen_a2aj:
                        continue
                    seen_a2aj.add((bare, kind))
                    _ts_print(f"  A2AJ from bare citation: {bare[:60]} ({kind})")
                    a2aj_text = _fetch_a2aj_source_text_for_row(
                        row, bare, kind, register_url=register_url
                    )
                    if a2aj_text:
                        anchor_text = a2aj_text
                        source_tag = candidate_tag
                        locked_structure = row.get("_a2aj_structure") or {}
                        _ts_print(f"  A2AJ: fetched {len(a2aj_text)} chars")
                        break

            if not anchor_text:
                row["quote_check_status"] = "NO_MATCH"
                row["quote_corrected_citation"] = "\n\n".join(
                    _quote_original_display(q) for q in quote_entries
                ).strip()
                if _should_report_source_text_failure(row):
                    row["quote_check_notes"] = "No source text found for this citation."
                _log_quote_row_timing(row, row_started_at, len(quote_entries), source_tag)
                continue
        if source_tag in _A2AJ_SOURCE_TAGS:
            a2aj_scopes = _a2aj_pinpoint_scope.cited_scopes(row)
            scoped_law_text = ""
            scoped_law_structure: Dict[str, Any] = {}
            if a2aj_scopes.sections:
                scoped_law_text, scoped_law_structure = _a2aj_cited_law_structure(
                    row,
                    a2aj_scopes,
                    anchor_text=anchor_text,
                    locked_structure=locked_structure,
                )

            def _plausible_a2aj_location(candidate_quote: str, candidate_text: str) -> bool:
                return _has_plausible_partial_content_overlap(
                    candidate_quote,
                    _trim_to_text_region(candidate_quote, candidate_text, window=400)
                    or candidate_text,
                )

            for q in quote_entries:
                qt = (q.get("inner") or q.get("raw") or "").strip()
                key = _quote_dedupe_key(qt)
                if key:
                    scoped_resolution = None
                    if scoped_law_text:
                        scoped_resolution = _a2aj_pinpoint_scope.resolve_quote(
                            scoped_law_text,
                            qt,
                            scoped_law_structure,
                            a2aj_scopes,
                            _quote_match_score,
                            minimum=min_match,
                            pinpoint_minimum=strong_match,
                            plausible=_plausible_a2aj_location,
                        )
                    if scoped_resolution and scoped_resolution.score >= strong_match:
                        resolution = scoped_resolution
                    else:
                        resolution = _a2aj_pinpoint_scope.resolve_quote(
                            anchor_text,
                            qt,
                            locked_structure,
                            a2aj_scopes,
                            _quote_match_score,
                            minimum=min_match,
                            pinpoint_minimum=strong_match,
                            plausible=_plausible_a2aj_location,
                        )
                        if (
                            (
                                scoped_law_text
                                or locked_structure.get("source") == "section_map"
                            )
                            and resolution.location.startswith("scope_unavailable")
                        ):
                            resolution = replace(
                                resolution,
                                location="alternate_document",
                                cited_scope_available=True,
                            )
                        if (
                            resolution.score < strong_match
                            and scoped_resolution
                            and scoped_resolution.location != "unmatched"
                        ):
                            resolution = scoped_resolution
                    a2aj_resolutions[key] = resolution
            row["_a2aj_cited_scopes"] = list(a2aj_scopes.labels)
            row["_a2aj_quote_locations"] = [
                a2aj_resolutions[_quote_dedupe_key((q.get("inner") or q.get("raw") or "").strip())].location
                for q in quote_entries
                if _quote_dedupe_key((q.get("inner") or q.get("raw") or "").strip()) in a2aj_resolutions
            ]

        row["_quote_source_tag"] = source_tag

        scored: List[Tuple[float, Dict[str, Any]]] = []
        for q in quote_entries:
            qt = (q.get("inner") or q.get("raw") or "").strip()
            resolution = a2aj_resolutions.get(_quote_dedupe_key(qt))
            scored.append((resolution.score if resolution else _quote_match_score(qt, anchor_text), q))

        scored.sort(key=lambda x: x[0], reverse=True)
        matched = [(score, q) for score, q in scored if score >= min_match]

        if not matched:
            row["quote_check_status"] = "NO_MATCH"
            row["quote_corrected_citation"] = "\n\n".join(
                _quote_original_display(q) for q in quote_entries
            ).strip()
            row["quote_check_notes"] = _quote_not_found_note(quote_entries, anchor_text)
            _log_quote_row_timing(row, row_started_at, len(quote_entries), source_tag)
            continue

        corrected_list: List[str] = []
        unchanged_flags: List[bool] = []
        accepted_scores: List[float] = []
        accepted_a2aj_locations: List[str] = []
        matched_regions: List[str] = []
        matched_source_fragments: List[str] = []
        resolved_pinpoints: List[str] = []
        resolved_links: List[str] = []
        alternate_pinpoints: List[str] = []
        alternate_links: List[str] = []
        initial_source_pinpoint = source_pinpoint
        seen_corrected_keys: set[str] = set()
        for score, q in matched:
            qt = (q.get("inner") or q.get("raw") or "").strip()
            a2aj_resolution = a2aj_resolutions.get(_quote_dedupe_key(qt))
            match_text = a2aj_resolution.text if a2aj_resolution and a2aj_resolution.text else anchor_text
            source_region = _trim_to_text_region(qt, match_text, window=400)
            if (
                score < strong_match
                and not _has_plausible_partial_content_overlap(qt, source_region or match_text)
            ):
                continue
            correction_source = source_region or match_text
            if not source_region and len(match_text or "") > _LONG_SOURCE_CHAR_LIMIT:
                correction_source = qt
            source_fragment = _source_side_quote_fragment_text(source_region, qt) if source_region else ""
            if source_fragment and _quote_match_score(qt, source_fragment) >= min_match:
                correction_source = source_fragment
            corrected = _build_corrected_citation(qt, correction_source, prefer_single_quotes=False)
            open_q, close_q = _infer_outer_quote_chars(q.get("raw") or "", q.get("style"))
            corrected = _ensure_outer_quotes(corrected, open_q, close_q)
            if _is_only_bracketed_quote_token(corrected):
                continue
            corrected_key = _quote_dedupe_key(corrected)
            if corrected_key and corrected_key in seen_corrected_keys:
                continue
            if corrected_key:
                seen_corrected_keys.add(corrected_key)
            corrected_list.append(corrected)
            accepted_scores.append(score)
            if a2aj_resolution:
                accepted_a2aj_locations.append(a2aj_resolution.location)
            unchanged_flags.append(
                _normalize_quote_exact_compare(corrected) == _normalize_quote_exact_compare(qt)
            )
            matched_region_for_output = source_region
            def _apply_promoted(promoted):
                if not promoted:
                    return
                if promoted.get("link"):
                    row["quote_match_link"] = promoted["link"]
                    resolved_links.append(promoted["link"])
                if promoted.get("label"):
                    nonlocal_pin["v"] = promoted["label"]
                if promoted.get("text"):
                    nonlocal_region["v"] = promoted["text"]

            # Small boxes so the local helper can write back the recovered
            # pinpoint/region without a nonlocal declaration on loop variables.
            nonlocal_pin = {"v": initial_source_pinpoint}
            nonlocal_region = {"v": matched_region_for_output}

            if (
                a2aj_resolution
                and a2aj_resolution.labels
                and not a2aj_resolution.location.startswith("scope_unavailable")
            ):
                labels = list(a2aj_resolution.labels)
                resolved_pinpoints.extend(labels)
                is_alternate = a2aj_resolution.location.startswith("alternate")
                if is_alternate:
                    alternate_pinpoints.extend(labels)
                promoted_link = ""
                if row.get("_a2aj_url_reconciled"):
                    source_link = _canlii_document_link_for_row(row, source_tag)
                    if not source_link:
                        source_link = (row.get("citation_part_link") or "").strip()
                    promoted_link = _a2aj_link_for_pinpoint(source_link, labels[0])
                if is_alternate and promoted_link:
                    alternate_links.append(promoted_link)
                _apply_promoted({
                    "label": _format_pinpoint_summary(labels),
                    "link": promoted_link,
                    "text": a2aj_resolution.text,
                })
            elif (
                a2aj_resolution
                and a2aj_resolution.location.startswith("alternate")
                and row.get("_a2aj_url_reconciled")
            ):
                source_link = _canlii_document_link_for_row(row, source_tag)
                if not source_link:
                    source_link = (row.get("citation_part_link") or "").strip()
                base_link, _base_fragment = _split_url(source_link)
                if base_link:
                    alternate_links.append(base_link)

            # A2AJ serves flat text, but its paragraph numbers survive as
            # line-leading markers — recover a pinpoint from the numbered
            # paragraph that unambiguously contains the quote, and attach the
            # CanLII #parN anchor (the quote link points at CanLII).
            if not nonlocal_pin["v"] and source_tag in _A2AJ_SOURCE_TAGS:
                source_link = ""
                if row.get("_a2aj_url_reconciled"):
                    source_link = _canlii_document_link_for_row(row, source_tag)
                    if not source_link:
                        source_link = (row.get("citation_part_link") or "").strip()
                if not a2aj_resolution:
                    _apply_promoted(_locate_a2aj_pinpoint(
                        anchor_text, qt, source_link, locked_structure, min_score=strong_match))

            current_pinpoint = nonlocal_pin["v"]
            matched_region_for_output = nonlocal_region["v"]
            if source_region:
                matched_regions.append(matched_region_for_output.replace("\n", " "))
                if source_tag == "journal_db" and not current_pinpoint:
                    current_pinpoint = _journal_db_page_for_region(journal_db_full_text or anchor_text, source_region)
            if source_tag == "journal_db" and not current_pinpoint:
                current_pinpoint = _format_pinpoint_summary(
                    _journal_db_pages_for_quote(
                        journal_db_full_text or anchor_text,
                        qt,
                        min_score=strong_match,
                    )
                )
            if source_tag == "journal_db" and current_pinpoint:
                page_label = _page_label_from_match_pinpoint(current_pinpoint)
                if page_label is not None:
                    row["quote_match_link"] = _journal_db_page_link_for_row(row, page_label)
            if current_pinpoint and not a2aj_resolution:
                resolved_pinpoints.append(current_pinpoint)
            if source_fragment:
                matched_source_fragments.append(source_fragment)

        if not corrected_list:
            row["quote_check_status"] = "NO_MATCH"
            row["quote_corrected_citation"] = "\n\n".join(
                _quote_original_display(q) for q in quote_entries
            ).strip()
            row["quote_check_notes"] = _quote_not_found_note(quote_entries, anchor_text)
            _log_quote_row_timing(row, row_started_at, len(quote_entries), source_tag)
            continue

        source_pinpoint = _format_pinpoint_summary(list(dict.fromkeys(resolved_pinpoints))) or initial_source_pinpoint
        if resolved_links:
            row["quote_match_link"] = resolved_links[0]
        row["quote_corrected_citation"] = "\n\n".join(corrected_list).strip()
        row["quote_match_pinpoint"] = source_pinpoint
        if matched_regions:
            row["matched_source"] = "\n\n".join(matched_regions).strip()
        if matched_source_fragments:
            row["matched_source_fragment"] = "\n\n".join(matched_source_fragments).strip()
        best_score = max(accepted_scores)
        perfect = (
            len(corrected_list) == len(quote_entries)
            and all(score >= strong_match for score in accepted_scores)
            and all(unchanged_flags)
        )
        alternate_locations = [
            location for location in accepted_a2aj_locations
            if location.startswith("alternate")
        ]
        limited_scope = any(
            location.startswith("scope_unavailable") or location == "cited_parent"
            for location in accepted_a2aj_locations
        )
        if a2aj_scopes.has_any and alternate_locations:
            source_pinpoint = _format_pinpoint_summary(list(dict.fromkeys(alternate_pinpoints)))
            row["quote_match_pinpoint"] = source_pinpoint
            row["quote_match_link"] = alternate_links[0] if alternate_links else ""
            all_alternate = len(alternate_locations) == len(accepted_a2aj_locations)
            level = "MATCH" if perfect and all_alternate else "PARTIAL"
            if source_pinpoint:
                row["quote_check_status"] = f"ALT_PINPOINT_{level}_A2AJ"
                row["quote_check_notes"] = source_pinpoint
            else:
                row["quote_check_status"] = f"ALT_PINPOINTLESS_{level}_A2AJ"
                row["quote_check_notes"] = ""
        elif perfect and not limited_scope:
            row["quote_check_status"] = "OG_PINPOINT_MATCH"
        else:
            row["quote_check_status"] = "OG_PINPOINT_PARTIAL"
            if limited_scope:
                row["quote_check_notes"] = "The cited pinpoint could not be isolated in the available source structure."
        _log_quote_row_timing(row, row_started_at, len(quote_entries), source_tag)

    for row in rows:
        existing = str(row.get("quote_match_link") or "").strip()
        candidate = existing
        if not candidate and str(row.get("quote_check_status") or "").startswith(
            "OG_PINPOINT"
        ):
            candidate = str(row.get("citation_part_link") or "").strip()
        if not candidate:
            continue
        preferred = _prefer_scc_official_quote_link(row, candidate)
        if existing or preferred != candidate:
            row["quote_match_link"] = preferred

# -----------------------------
# Data build
# -----------------------------


def build_audit_data(
    docx_path: str,
    *,
    max_lookahead: int = MAX_LOOKAHEAD_CHARS_FOR_QUOTE_TO_FOOTNOTE,
    footnote_ids: Optional[set[int]] = None,
) -> Dict[str, Any]:
    with zipfile.ZipFile(docx_path) as zf:
        doc = _get_zip_xml(zf, "word/document.xml")
        fns = _get_zip_xml(zf, "word/footnotes.xml")
        fn_rels = _get_zip_xml(zf, "word/_rels/footnotes.xml.rels")
        styles = _get_zip_xml(zf, "word/styles.xml")
        if doc is None:
            raise ValueError("Missing word/document.xml in DOCX")

        paragraphs = extract_doc_stream_with_styles(doc, styles)
        global_text, para_starts, anchors = build_global_text(paragraphs)
        raw_author_links: Dict[int, List[Dict[str, Any]]] = {}
        footnote_map = extract_footnotes(fns, fn_rels, raw_author_links)

    _ts_print(f"  Parsed DOCX: {len(footnote_map)} footnotes, {len(paragraphs)} paragraphs")

    # Remap DOCX-internal footnote ids to display-order ids (1..N), matching anchor appearance order.
    # Word typically starts real footnote ids at 2, which otherwise shifts ids by +1.
    footnote_order_raw = compute_footnote_order(anchors, footnote_map)
    raw_to_display = {raw_fid: i for i, raw_fid in enumerate(footnote_order_raw, start=1)}

    # Remap anchors
    anchors = [
        {**a, "footnote_id": raw_to_display.get(int(a.get("footnote_id") or 0), int(a.get("footnote_id") or 0))}
        for a in anchors
        if int(a.get("footnote_id") or 0) in raw_to_display
    ]

    # Remap footnote text map
    footnote_map = {raw_to_display[raw_fid]: footnote_map.get(raw_fid, "") for raw_fid in footnote_order_raw}
    author_links_by_fid = {
        raw_to_display[raw_fid]: raw_author_links[raw_fid]
        for raw_fid in footnote_order_raw
        if raw_fid in raw_author_links
    }

    # Display-order list (1..N)
    footnote_order = list(range(1, len(footnote_order_raw) + 1))

    if footnote_ids is not None:
        requested_ids = set(footnote_ids)
        expanded_ids, _supra_targets, _ibid_targets = _expand_footnote_ids_with_reference_targets(requested_ids, footnote_map)
        if expanded_ids != requested_ids:
            added = sorted(expanded_ids - requested_ids)
            _ts_print(
                "  Added reference target footnotes to filtered batch: "
                + ",".join(str(fid) for fid in added)
            )
        footnote_ids = expanded_ids
        before = len(footnote_order)
        footnote_order = [fid for fid in footnote_order if fid in footnote_ids]
        _ts_print(f"  Filtered to {len(footnote_order)}/{before} footnotes (ids: {','.join(str(f) for f in sorted(footnote_order))})")

    clean_global, raw_to_clean = build_clean_text_and_index_map(global_text)

    footnote_display_ids, display_num_to_internal, _internal_to_display_num = _compute_footnote_display_ids(
        footnote_order, footnote_map
    )

    # Proposition text per footnote (built first: the reference disambiguation
    # fallback shows the body text to the model for hint-less supras)
    _ts_print(f"  Building proposition text for {len(footnote_order)} footnotes ...")
    prop_map = build_anchor_propositions(clean_global, anchors, raw_to_clean)
    prop_texts = {
        fid: (prop_map.get(fid, {}).get("proposition_text") or "").strip()
        for fid in footnote_order
    }

    # Footnotes (in appearance order)
    footnote_parts, fn_num_parts = build_footnote_parts(
        footnote_map, footnote_order, prop_texts, author_links_by_fid
    )

    # Quotes from the same text shown in "Quotes and proposition"
    quotes_by_footnote: Dict[int, List[Dict[str, Any]]] = {}
    for fid in footnote_order:
        prop_text = (prop_map.get(fid, {}).get("proposition_text") or "").strip()
        if not prop_text:
            continue
        qlist = []
        for q in find_inline_quotes(prop_text):
            qlist.append(
                {
                    "quote_id": None,
                    "quote_raw": q["raw"],
                    "quote_inner": q["inner"],
                    "quote_type": q["quote_type"],
                    "context_sentence": prop_text,
                }
            )
        if qlist:
            quotes_by_footnote[fid] = qlist

    # Master rows: 1 row per footnote citation-part
    footnote_rows: List[Dict[str, Any]] = []
    for p in footnote_parts:
        fid = int(p["footnote_id"])
        pm = prop_map.get(fid, {})
        qlist = quotes_by_footnote.get(fid, [])

        row = {
            "footnote_id": fid,
            "footnote_internal_id": fid,
            "footnote_display_id": footnote_display_ids.get(fid, str(fid)),
            "citation_part_index": int(p.get("citation_part_index") or 0),
            "citation_part_kind": p.get("citation_part_kind", ""),
            "citation_part_text": p.get("citation_part_text", ""),
            "citation_part_corrected": p.get("citation_part_corrected", ""),
            "citation_part_link": p.get("citation_part_link", ""),
            "_author_provided_link": p.get("_author_provided_link", ""),
            "_author_provided_links": p.get("_author_provided_links", []),
            "citation_part_anchor_text": p.get("citation_part_anchor_text", ""),
            "_citation_part_anchor_segments": p.get("_citation_part_anchor_segments", []),
            "_citation_part_full_source_text": p.get("_citation_part_full_source_text", ""),
            "short_form": p.get("short_form", ""),
            "bare_citation": p.get("bare_citation", ""),
            "citation_with_style": p.get("citation_with_style", ""),
            "pinpoint_fragments": p.get("pinpoint_fragments", []),
            "page_pinpoints": p.get("page_pinpoints", []),
            "first_page": p.get("first_page", ""),
            "footnote_full": p.get("footnote_full", ""),
            "proposition_text": pm.get("proposition_text", ""),
            "has_quotes": "YES" if qlist else "NO",
            "quote_count": len(qlist),
            "quotes_list_json": json.dumps(
                [((q.get("quote_inner") or q.get("quote_raw") or "").strip()) for q in qlist],
                ensure_ascii=False
            ),
            "quote_check_status": "",
            "quote_corrected_citation": "",
            "quote_check_notes": "",
            "quote_match_pinpoint": "",
            "quote_match_link": "",
            "matched_source": "",
            "matched_source_fragment": "",
            "alternate_matched_source_fragment": "",
            "journal_match_info": "",
        }

        footnote_rows.append(row)

    for row in footnote_rows:
        pf = row.get("pinpoint_fragments")
        if isinstance(pf, list):
            row["pinpoint_fragments"] = json.dumps(pf)
        pp = row.get("page_pinpoints")
        if isinstance(pp, list):
            row["page_pinpoints"] = json.dumps(pp)
        # Infer first page from citation_with_style if not already set
        if not row.get("first_page"):
            cit = row.get("citation_with_style") or row.get("bare_citation") or ""
            row["first_page"] = _extract_first_page(cit)

    return {
        "footnote_rows": footnote_rows,
        "footnote_parts": footnote_parts,
        "fn_num_parts": fn_num_parts,
        "footnote_display_ids": footnote_display_ids,
        "display_num_to_internal": display_num_to_internal,
        "footnote_map": footnote_map,
        "_quotes_by_footnote": quotes_by_footnote,
    }


def _collect_docx_files(input_folder: str | Path, recursive: bool = False) -> List[Path]:
    in_dir = Path(input_folder)
    pattern = "**/*.docx" if recursive else "*.docx"
    return sorted(
        p for p in in_dir.glob(pattern)
        if p.is_file() and not p.name.startswith("~$")
    )


def _load_quote_discovery_context(docx_path: str | Path) -> Dict[str, Any]:
    with zipfile.ZipFile(docx_path) as zf:
        doc = _get_zip_xml(zf, "word/document.xml")
        fns = _get_zip_xml(zf, "word/footnotes.xml")
        styles = _get_zip_xml(zf, "word/styles.xml")
        if doc is None:
            raise ValueError("Missing word/document.xml in DOCX")

        paragraphs = extract_doc_stream_with_styles(doc, styles)
        global_text, _para_starts, anchors = build_global_text(paragraphs)
        footnote_map = extract_footnotes(fns)

    footnote_order_raw = compute_footnote_order(anchors, footnote_map)
    raw_to_display = {raw_fid: i for i, raw_fid in enumerate(footnote_order_raw, start=1)}
    anchors = [
        {**a, "footnote_id": raw_to_display.get(int(a.get("footnote_id") or 0), int(a.get("footnote_id") or 0))}
        for a in anchors
        if int(a.get("footnote_id") or 0) in raw_to_display
    ]
    footnote_map = {raw_to_display[raw_fid]: footnote_map.get(raw_fid, "") for raw_fid in footnote_order_raw}
    footnote_order = list(range(1, len(footnote_order_raw) + 1))
    clean_global, raw_to_clean = build_clean_text_and_index_map(global_text)
    prop_map = build_anchor_propositions(clean_global, anchors, raw_to_clean)
    return {
        "footnote_order": footnote_order,
        "footnote_map": footnote_map,
        "prop_map": prop_map,
    }


def _quote_manifest_rows_for_doc(
    docx_path: Path,
    input_root: Path,
    *,
    max_lookahead: int = MAX_LOOKAHEAD_CHARS_FOR_QUOTE_TO_FOOTNOTE,
) -> Tuple[List[Dict[str, Any]], set[int]]:
    _ = max_lookahead  # quote association uses the same proposition builder as the app.
    context = _load_quote_discovery_context(docx_path)
    footnote_order: List[int] = context["footnote_order"]
    footnote_map: Dict[int, str] = context["footnote_map"]
    prop_map: Dict[int, Dict[str, Any]] = context["prop_map"]

    quotes_by_footnote: Dict[int, List[Dict[str, Any]]] = {}
    for fid in footnote_order:
        prop_text = (prop_map.get(fid, {}).get("proposition_text") or "").strip()
        if not prop_text:
            continue
        qlist = []
        for q in find_inline_quotes(prop_text):
            qlist.append(
                {
                    "quote_raw": q["raw"],
                    "quote_inner": q["inner"],
                    "quote_type": q["quote_type"],
                    "context_sentence": prop_text,
                }
            )
        if qlist:
            quotes_by_footnote[fid] = qlist

    quote_ids = set(quotes_by_footnote)
    selected_ids, supra_sources, ibid_sources = _expand_footnote_ids_with_reference_targets(quote_ids, footnote_map)
    rows: List[Dict[str, Any]] = []
    doc_key = _doc_key_for_manifest(docx_path, input_root)
    for fid in sorted(selected_ids):
        qlist = quotes_by_footnote.get(fid, [])
        reasons = []
        if qlist:
            reasons.append("quote")
        if fid in supra_sources:
            reasons.append("supra_target")
        if fid in ibid_sources:
            reasons.append("ibid_target")
        source_footnotes = sorted(set(supra_sources.get(fid, set())) | set(ibid_sources.get(fid, set())))
        prop_text = (prop_map.get(fid, {}).get("proposition_text") or "").strip()
        rows.append(
            {
                "doc_key": doc_key,
                "doc_name": docx_path.name,
                "doc_path": str(docx_path),
                "footnote_id": fid,
                "include_reason": ",".join(reasons) or "selected",
                "source_footnotes": ",".join(str(v) for v in source_footnotes),
                "supra_source_footnotes": ",".join(str(v) for v in sorted(supra_sources.get(fid, set()))),
                "ibid_source_footnotes": ",".join(str(v) for v in sorted(ibid_sources.get(fid, set()))),
                "quote_count": len(qlist),
                "quotes_json": json.dumps(
                    [((q.get("quote_inner") or q.get("quote_raw") or "").strip()) for q in qlist],
                    ensure_ascii=False,
                ),
                "proposition_text": prop_text,
                "footnote_text": footnote_map.get(fid, ""),
            }
        )
    return rows, selected_ids


def build_quote_footnote_manifest(
    input_folder: str,
    *,
    recursive: bool = False,
    max_lookahead: int = MAX_LOOKAHEAD_CHARS_FOR_QUOTE_TO_FOOTNOTE,
) -> Tuple[List[Dict[str, Any]], Dict[str, set[int]]]:
    input_root = Path(input_folder).expanduser().resolve()
    if not input_root.exists() or not input_root.is_dir():
        raise NotADirectoryError(f"Input folder not found or not a directory: {input_root}")
    rows: List[Dict[str, Any]] = []
    selection: Dict[str, set[int]] = {}
    for docx_path in _collect_docx_files(input_root, recursive=recursive):
        _ts_print(f"  Discovering quote footnotes: {docx_path.name}")
        doc_rows, selected_ids = _quote_manifest_rows_for_doc(
            docx_path,
            input_root,
            max_lookahead=max_lookahead,
        )
        rows.extend(doc_rows)
        if selected_ids:
            selection.setdefault(_normalize_selection_key(_doc_key_for_manifest(docx_path, input_root)), set()).update(selected_ids)
    return rows, selection


def write_quote_footnote_manifest(rows: List[Dict[str, Any]], output_path: str | Path) -> None:
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "doc_key",
        "doc_name",
        "doc_path",
        "footnote_id",
        "include_reason",
        "source_footnotes",
        "supra_source_footnotes",
        "ibid_source_footnotes",
        "quote_count",
        "quotes_json",
        "proposition_text",
        "footnote_text",
    ]
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def _response_text(response: Any) -> str:
    text = getattr(response, "output_text", "")
    if text:
        return str(text).strip()
    try:
        parts: List[str] = []
        for item in getattr(response, "output", []) or []:
            if isinstance(item, dict) and item.get("type") == "message":
                for c in item.get("content", []) or []:
                    if isinstance(c, dict) and c.get("type") == "output_text":
                        parts.append(c.get("text", ""))
        return "\n".join(parts).strip()
    except Exception:
        return ""


# -----------------------------
# Workbook writing
# -----------------------------


def write_workbook(data: Dict[str, Any], output_path: str) -> None:
    _ensure_openpyxl()
    # 1. Compile regex for illegal XML characters
    # Matches low ASCII control chars except \t (0x09), \n (0x0A), \r (0x0D)
    illegal_char_re = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

    def clean_val(v: Any) -> Any:
        """Remove Excel-illegal characters from strings."""
        if isinstance(v, str):
            return illegal_char_re.sub("", v)
        return v

    wb = Workbook()
    wb.remove(wb.active)


    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    def _compact_pinpoint_summary(summary: str) -> str:
        s = re.sub(r"\s*\[\+(\d+)\s+more instances\]", r" +\1 more", summary or "")
        return s.strip()

    def _status_prefix(status: str, row: dict) -> str:
        s = (status or "").strip().upper()

        def _match_location_label(r: dict) -> str:
            pinpoint = _compact_pinpoint_summary(r.get("quote_match_pinpoint") or "")
            if not pinpoint or "unknown" in pinpoint.lower():
                return " at [unknown]"
            return f" at {pinpoint}"

        if s in ("OG_PINPOINT_MATCH",):
            return f"✓Perfect Match{_match_location_label(row)}✓"
        if s in ("OG_PINPOINT_PARTIAL",):
            alt_notes = (row.get("_alternate_quote_check_notes") or "").strip()
            alt_suffix = f" - ✓Perfect Match found at: {_compact_pinpoint_summary(alt_notes)}✓" if alt_notes else ""
            return f"~Partial Match{_match_location_label(row)}~{alt_suffix}"
        if s == "NO_MATCH":
            notes = (row.get("quote_check_notes") or "")
            link = (row.get("citation_part_link") or "").strip()
            if not link:
                return ""
            if "too short for full-document fuzzy checking" in notes:
                return "❌Short quote in long source; exact search found no match❌"
            if "Quote not found" in notes:
                return "❌No match found❌"
            if "No source text found" in notes:
                return "❌Unable to grab source text❌"
            base, frag = _split_url(link)
            has_requested_pinpoint = bool(
                _coerce_pinpoint_fragments(row.get("pinpoint_fragments"))
            ) or _first_page_pinpoint(row) is not None
            if base and not frag and not has_requested_pinpoint:
                return "❌No pinpoint provided❌"
            return "❌No match found❌"
        if s in ("ALT_PINPOINT_MATCH_CANLII", "ALT_PINPOINT_MATCH_A2AJ"):
            return f"✓Perfect Match{_match_location_label(row)}✓"
        if s in ("ALT_PINPOINT_PARTIAL_CANLII", "ALT_PINPOINT_PARTIAL_A2AJ"):
            return f"~Partial Match{_match_location_label(row)}~"
        if s in ("ALT_PINPOINTLESS_MATCH_CANLII", "ALT_PINPOINTLESS_MATCH_A2AJ"):
            return f"✓Perfect Match{_match_location_label(row)}✓"
        if s in ("ALT_PINPOINTLESS_PARTIAL_CANLII", "ALT_PINPOINTLESS_PARTIAL_A2AJ"):
            return f"~Partial Match{_match_location_label(row)}~"
        return ""

    def _corrected_quote_display(r: dict) -> str:
        # If no quote in proposition, blank.
        if (r.get("has_quotes") or "").upper() != "YES":
            return ""
        prefix = _status_prefix(r.get("quote_check_status"), r)
        if not prefix:
            return ""
        # Pure-problem messages (❌...❌) don't show the corrected body
        if prefix.startswith("❌") and prefix.endswith("❌"):
            return prefix
        body = (r.get("quote_corrected_citation") or "").strip()
        if (
            (r.get("quote_check_status") or "").strip().upper() == "OG_PINPOINT_PARTIAL"
            and (r.get("_alternate_quote_check_notes") or "").strip()
        ):
            partial_label = re.sub(r"^~|~$", "", prefix.split(" - ", 1)[0]).replace("Match", "match")
            alt_notes = _compact_pinpoint_summary(r.get("_alternate_quote_check_notes") or "")
            alt_body = (r.get("_alternate_quote_corrected_citation") or "").strip()
            lines = [partial_label]
            if body:
                lines.append(body)
            lines.append("")
            lines.append(f"Perfect match found at {alt_notes}")
            if alt_body:
                lines.append(alt_body)
            return "\n".join(lines).strip()
        return (prefix + "\n" + body).strip()

    BASE_DISPLAY_COLS = [
        ("Footnote #",                lambda r: r.get("footnote_display_id", r.get("footnote_id", ""))),
        ("Footnote Text",             lambda r: r.get("footnote_full", "")),
        ("Quotes and proposition",    lambda r: r.get("proposition_text", "")),  # rich text applied later
        ("Citation",                  lambda r: r.get("citation_part_text", "")),  # hyperlink applied later
        ("Automatic  Checking  System ►", lambda r: ""),
        ("Corrected quote",           _corrected_quote_display),
    ]

    # Raw keys you DO NOT want duplicated after the display columns
    DISPLAY_SOURCE_KEYS = {
        "footnote_display_id",
        "footnote_full",
        "proposition_text",
        "citation_part_text",
        "source_doc",
    }

    def add_footnote_references_sheet(rows: list[dict]):
        ws = wb.create_sheet(title="FootnoteReferences")
        if not rows:
            ws.append(["(no rows)"])
            return ws

        display_cols = list(BASE_DISPLAY_COLS)
        if any(r.get("source_doc") for r in rows):
            display_cols = [("Document", lambda r: r.get("source_doc", ""))] + display_cols

        raw_headers = [h for h in rows[0].keys() if not str(h).startswith("_")]
        diagnostic_headers = [h for h in raw_headers if h not in DISPLAY_SOURCE_KEYS]

        headers = [h for (h, _) in display_cols] + diagnostic_headers
        ws.append(headers)

        header_font = Font(bold=True, size=13)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        prev_footnote_text = None
        prev_prop_text = None
        prev_footnote_key = None
        prev_prop_key = None
        # write rows
        for r in rows:
            display_vals = [clean_val(fn(r)) for (_, fn) in display_cols]
            display_by_header = {h: display_vals[i] for i, (h, _) in enumerate(display_cols)}
            footnote_text = display_by_header.get("Footnote Text") or ""
            prop_text = display_by_header.get("Quotes and proposition") or ""
            source_doc = display_by_header.get("Document") or ""
            footnote_key = (source_doc, footnote_text)
            prop_key = (source_doc, prop_text)

            if prev_footnote_text is not None and footnote_key == prev_footnote_key:
                display_vals[[h for h, _ in display_cols].index("Footnote Text")] = ""  # blank Footnote Text
            else:
                prev_footnote_text = footnote_text
                prev_footnote_key = footnote_key
            if prev_prop_text is not None and prop_key == prev_prop_key:
                display_vals[[h for h, _ in display_cols].index("Quotes and proposition")] = ""  # blank Quotes and proposition
            else:
                prev_prop_text = prop_text
                prev_prop_key = prop_key

            diagnostic_vals = [clean_val(r.get(h, "")) for h in diagnostic_headers]
            ws.append(display_vals + diagnostic_vals)

        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # column widths (tune as desired)
        widths = {
            "Document": 34,
            "Footnote #": 12,
            "Footnote Text": 60,
            "Quotes and proposition": 70,
            "Citation": 60,
            "Automatic  Checking  System â–º": 12,
            "Corrected quote": 55,
            "ref_resolution_notes": 25,
        }
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 15)
        # Diagnostic columns stay visible only in diagnostic export mode; display
        # modes strip them after formatting has used their raw values.
        for col_idx in range(len(display_cols) + 1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        return ws


    def add_sheet(name: str, rows: List[Dict[str, Any]]):
        ws = wb.create_sheet(title=name)
        if not rows:
            ws.append(["(no rows)"])
            return ws

        headers = [h for h in rows[0].keys() if not str(h).startswith('_')]
        ws.append(headers)

        header_font = Font(bold=True, size=13)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for r in rows:
            # 2. Apply cleaning to every cell value before writing
            row_values = [clean_val(r.get(h, "")) for h in headers]
            ws.append(row_values)

        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        max_width = 70
        for col_idx, h in enumerate(headers, start=1):
            sample = [str(h)]
            # Sample first 200 rows for auto-width to save time
            for r_i in range(2, min(ws.max_row, 201) + 1):
                v = ws.cell(row=r_i, column=col_idx).value
                if v is not None:
                    sample.append(str(v))
            width = min(max(len(s) for s in sample), max_width)
            width = max(12, width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        return ws

    ws_main = add_footnote_references_sheet(data.get("footnote_rows", []))

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def apply_cell_formatting(xlsx_path: str, *, save: bool = True) -> Optional[Any]:
    _ensure_openpyxl()
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter


    def _apply_richtext_bolding_and_color(ws, row_idx: int, col_idx: int, text: str, quotes: list[str], quote_argb: str) -> None:
        from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

        if not text or not quotes:
            return

        spans = []
        for q in quotes:
            q = (q or "").strip()
            if not q:
                continue
            start = 0
            while True:
                i = text.find(q, start)
                if i == -1:
                    break
                spans.append((i, i + len(q)))
                start = i + len(q)

        if not spans:
            return

        spans.sort(key=lambda t: (t[0], -(t[1] - t[0])))
        chosen = []
        last_end = -1
        for s, e in spans:
            if s >= last_end:
                chosen.append((s, e))
                last_end = e

        blocks = []
        cursor = 0
        quote_font = InlineFont(b=True, color=quote_argb, size=13)  # ARGB string like "FFF6B26B"

        for s, e in chosen:
            if cursor < s:
                blocks.append(TextBlock(text[cursor:s]))
            blocks.append(TextBlock(text[s:e], quote_font))
            cursor = e
        if cursor < len(text):
            blocks.append(TextBlock(text[cursor:]))

        ws.cell(row=row_idx, column=col_idx).value = CellRichText(blocks)


    def _apply_corrected_quote_status_richtext(cell, text: str) -> bool:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont

        if not text or "Perfect match found at" not in text or "Partial match at" not in text:
            return False

        link_marker = "Perfect match found at"
        marker_idx = text.find(link_marker)
        if marker_idx <= 0:
            return False

        plain_text = text[:marker_idx]
        link_text = text[marker_idx:]
        normal_font = InlineFont(color="FF000000", sz=13)
        link_font = InlineFont(color="FF0000FF", u="single", sz=13)
        cell.value = CellRichText(
            TextBlock(normal_font, plain_text),
            TextBlock(link_font, link_text),
        )
        cell.font = Font(color="000000", underline=None, size=13)
        return True


    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception:
        return

    if "FootnoteReferences" not in wb.sheetnames:
        return

    ws = wb["FootnoteReferences"]
    headers = [c.value for c in ws[1]]
    col_map = {h: (i + 1) for i, h in enumerate(headers) if h}

    def _col(name: str):
        return col_map.get(name)

    def _font_with_size(font: Font | None, size: int) -> Font:
        if font is None:
            return Font(size=size)
        return Font(
            name=font.name,
            size=size,
            bold=font.bold,
            italic=font.italic,
            vertAlign=font.vertAlign,
            underline=font.underline,
            strike=font.strike,
            color=font.color,
            outline=font.outline,
            shadow=font.shadow,
            condense=font.condense,
            extend=font.extend,
            charset=font.charset,
            family=font.family,
            scheme=font.scheme,
        )

    # --- Existing fills you already had ---
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red   = PatternFill("solid", fgColor="FFC7CE")
    fill_grey  = PatternFill("solid", fgColor="C3C3C3")

    c_quote = _col("quote_check_status")   # diagnostic column
    c_ref   = _col("ref_kind")             # diagnostic column

    # --- New alternating row fills for display columns A-G ---
    # Use strong contrast; adjust hex to your exact accessibility-tested palette.
    fill_orange = PatternFill("solid", fgColor="FF9000")  # orange for footnote number
    fill_row_light = PatternFill("solid", fgColor="F3F3F3")  # light grey for row background
    fill_blue   = PatternFill("solid", fgColor="0C343D")  # dark teal
    font_white  = Font(color="FFFFFF", size=13)
    font_black  = Font(color="000000", size=13)

    # Display columns (fixed positions from your new layout)
    COL_FOOTNOTE_ID = _col("Footnote #") or 1
    COL_FOOTNOTE_INTERNAL = _col("footnote_internal_id")
    COL_CITATION_DISPLAY = _col("Citation") or 4
    COL_DIVIDER = _col("Automatic  Checking  System ►") or 5
    COL_CORRECTED = _col("Corrected quote") or 6

    display_col_names = [
        "Document",
        "Footnote #",
        "Footnote Text",
        "Quotes and proposition",
        "Citation",
        "Automatic  Checking  System â–º",
        "Corrected quote",
    ]
    display_col_indexes = [idx for idx in (_col(name) for name in display_col_names) if idx]
    DISPLAY_FIRST = min(display_col_indexes) if display_col_indexes else 1
    DISPLAY_LAST = max(display_col_indexes) if display_col_indexes else 7

    # Find diagnostic raw link column for citation hyperlinks
    COL_LINK_RAW = _col("citation_part_link")
    COL_CORR_RAW = _col("quote_corrected_citation")
    COL_MATCHED_SOURCE = _col("matched_source")
    COL_MATCHED_SOURCE_FRAGMENT = _col("matched_source_fragment")
    COL_ALT_MATCHED_SOURCE_FRAGMENT = _col("alternate_matched_source_fragment")
    COL_QUOTE_MATCH_LINK = _col("quote_match_link")
    COL_QUOTE_NOTES = _col("quote_check_notes")

    # Diagnostic column widths
    COL_ANCHOR_TEXT = _col("citation_part_anchor_text")
    COL_REF_CHAIN_ORIGIN_TEXT = _col("ref_chain_origin_citation_part_text")

    # Optional: make the divider column visually distinct
    divider_fill = PatternFill("solid", fgColor="A6A6A6")

    # Set base font size for the entire sheet (preserve existing attributes).
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = _font_with_size(cell.font, 13)

    # Style the header row (row 1)
    header_fill = PatternFill("solid", fgColor="0C343D")
    header_font = Font(color="FFFFFF", bold=True, size=13)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _alternate_notes_from_corrected_display(value: Any) -> str:
        text = str(value or "")
        m = re.search(r"(?:Perfect Match found at:?|Found at:)\s*([^\r\n~]+)", text, flags=re.IGNORECASE)
        if not m:
            return ""
        return m.group(1).split("✓", 1)[0].strip().strip("~").strip()

    def _alternate_quote_from_corrected_display(value: Any) -> str:
        text = str(value or "")
        m = re.search(r"Perfect match found at[^\r\n]*\r?\n(.+)\s*$", text, flags=re.IGNORECASE | re.DOTALL)
        return m.group(1).strip() if m else ""

    for r in range(2, ws.max_row + 1):
        quote_status_raw = ws.cell(row=r, column=c_quote).value if c_quote else ""
        quote_status = str(quote_status_raw or "")
        corrected_display_cell = ws.cell(row=r, column=COL_CORRECTED)
        corrected_display_text = corrected_display_cell.value or ""

        # 1) Add hyperlink for display "Citation" using diagnostic citation_part_link
        if COL_LINK_RAW:
            link = ws.cell(row=r, column=COL_LINK_RAW).value
            link_str = str(link).strip() if link is not None else ""
            text = ws.cell(row=r, column=COL_CITATION_DISPLAY).value
            if link_str and link_str.lower() != "other":
                cell = ws.cell(row=r, column=COL_CITATION_DISPLAY)
                citation_target = link_str
                quote_match_link = (
                    str(ws.cell(row=r, column=COL_QUOTE_MATCH_LINK).value or "").strip()
                    if COL_QUOTE_MATCH_LINK
                    else ""
                )
                verification_link = quote_match_link or link_str
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=citation_target, location=None)
                if not text or not str(text).strip():
                    cell.value = link_str
                cell.font = Font(color="0000FF", underline="single", size=13)

                corr_raw = ws.cell(row=r, column=COL_CORR_RAW).value if COL_CORR_RAW else ""
                alt_notes = ""
                if (
                    quote_status.startswith("ALT_PINPOINT")
                    and not quote_status.startswith("ALT_PINPOINTLESS")
                    and COL_QUOTE_NOTES
                ):
                    alt_notes = str(ws.cell(row=r, column=COL_QUOTE_NOTES).value or "").strip()
                if not alt_notes:
                    alt_notes = _alternate_notes_from_corrected_display(corrected_display_text)
                corrected_target = ""
                if alt_notes and COL_CORR_RAW:
                    alt_corr = _alternate_quote_from_corrected_display(corrected_display_text) or str(corr_raw or "")
                    alt_source_frag = (
                        ws.cell(row=r, column=COL_ALT_MATCHED_SOURCE_FRAGMENT).value
                        if COL_ALT_MATCHED_SOURCE_FRAGMENT
                        else ""
                    )
                    if not alt_source_frag and COL_MATCHED_SOURCE_FRAGMENT:
                        alt_source_frag = ws.cell(row=r, column=COL_MATCHED_SOURCE_FRAGMENT).value
                    alt_target = _build_alternate_pinpoint_fragment_url(
                        verification_link,
                        alt_notes,
                        alt_corr,
                        str(alt_source_frag or ""),
                        prefer_range="PARTIAL" in quote_status,
                    )
                    if alt_target:
                        corrected_target = alt_target
                    elif "[unknown pinpoint]" in alt_notes.lower():
                        corrected_target = verification_link
                if not corrected_target and alt_notes and quote_status.startswith("ALT_PINPOINT"):
                    match_link, _match_fragment = _split_url(verification_link)
                    source_raw = ws.cell(row=r, column=COL_MATCHED_SOURCE).value if COL_MATCHED_SOURCE else ""
                    source_frag = (
                        ws.cell(row=r, column=COL_MATCHED_SOURCE_FRAGMENT).value
                        if COL_MATCHED_SOURCE_FRAGMENT
                        else ""
                    )
                    if match_link and source_raw and source_frag:
                        fragment_result = _build_quote_check_fragment_url(
                            match_link,
                            str(source_frag),
                            str(source_raw),
                            str(corr_raw or ""),
                            prefer_range="PARTIAL" in quote_status,
                        )
                        corrected_target = fragment_result.url if fragment_result.fragment_count else match_link
                    else:
                        corrected_target = match_link
                if not corrected_target and quote_status.startswith("ALT_PINPOINTLESS"):
                    match_link, _match_fragment = _split_url(verification_link)
                    source_raw = ws.cell(row=r, column=COL_MATCHED_SOURCE).value if COL_MATCHED_SOURCE else ""
                    source_frag = (
                        ws.cell(row=r, column=COL_MATCHED_SOURCE_FRAGMENT).value
                        if COL_MATCHED_SOURCE_FRAGMENT
                        else ""
                    )
                    if match_link and source_raw and source_frag:
                        fragment_result = _build_quote_check_fragment_url(
                            match_link,
                            str(source_frag),
                            str(source_raw),
                            str(corr_raw or ""),
                            prefer_range="PARTIAL" in quote_status,
                        )
                        corrected_target = fragment_result.url if fragment_result.fragment_count else match_link
                    else:
                        corrected_target = match_link
                if not corrected_target and quote_status in ("OG_PINPOINT_MATCH", "OG_PINPOINT_PARTIAL"):
                    match_link = quote_match_link or link_str
                    source_raw = ws.cell(row=r, column=COL_MATCHED_SOURCE).value if COL_MATCHED_SOURCE else ""
                    source_frag = (
                        ws.cell(row=r, column=COL_MATCHED_SOURCE_FRAGMENT).value
                        if COL_MATCHED_SOURCE_FRAGMENT
                        else ""
                    )
                    if match_link and source_raw and source_frag:
                        fragment_result = _build_quote_check_fragment_url(
                            match_link,
                            str(source_frag),
                            str(source_raw),
                            str(corr_raw or ""),
                            prefer_range=quote_status == "OG_PINPOINT_PARTIAL",
                        )
                        if fragment_result.fragment_count:
                            corrected_target = fragment_result.url
                            _timing_event(
                                "quote_check_text_fragment_built",
                                builder=fragment_result.builder,
                                status=quote_status,
                                fragments=fragment_result.fragment_count,
                                reason=fragment_result.reason,
                            )
                        else:
                            corrected_target = match_link
                            _timing_event(
                                "quote_check_text_fragment_unresolved",
                                status=quote_status,
                                reason=fragment_result.reason,
                            )
                    else:
                        corrected_target = match_link
                if not corrected_target and quote_status in ("OG_PINPOINT_MATCH", "OG_PINPOINT_PARTIAL"):
                    corrected_target = quote_match_link or link_str
                if corrected_target:
                    corrected_display_cell.hyperlink = Hyperlink(
                        ref=corrected_display_cell.coordinate,
                        target=corrected_target,
                        location=None,
                    )
                    richtext_applied = False
                    if quote_status == "OG_PINPOINT_PARTIAL":
                        richtext_applied = _apply_corrected_quote_status_richtext(
                            corrected_display_cell,
                            str(corrected_display_text or ""),
                        )
                    if not richtext_applied:
                        corrected_display_cell.font = Font(color="0000FF", underline="single", size=13)

        # 2) Apply alternating row color by Footnote # value
        fid_val = None
        if COL_FOOTNOTE_INTERNAL:
            fid_val = ws.cell(row=r, column=COL_FOOTNOTE_INTERNAL).value
        if fid_val is None:
            fid_val = ws.cell(row=r, column=COL_FOOTNOTE_ID).value
        try:
            fid = int(fid_val)
        except Exception:
            fid = None

        is_odd = (fid is not None and (fid % 2 == 1))
        row_fill = fill_blue if is_odd else fill_orange
        row_font = font_white if is_odd else font_black

        for c in range(DISPLAY_FIRST, DISPLAY_LAST + 1):
            cell = ws.cell(row=r, column=c)
            if c == COL_FOOTNOTE_ID:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if cell.font and cell.font.underline:
                    cell.font = Font(color=row_font.color, underline=cell.font.underline, size=20)
                else:
                    cell.font = Font(color=row_font.color, size=20)

            elif not is_odd and c != COL_DIVIDER:
                cell.fill = fill_row_light

            # Keep divider column dark teal with white text
            if c == COL_DIVIDER:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")


        # 3) Preserve status coloring on diagnostic columns (does not touch display columns)
        if c_quote:
            v = ws.cell(row=r, column=c_quote).value
            if v and v.startswith("ALT_"):
                ws.cell(row=r, column=c_quote).fill = fill_green
            elif v in ("OG_PINPOINT_MATCH", "OG_PINPOINT_PARTIAL"):
                ws.cell(row=r, column=c_quote).fill = fill_green
            elif v == "NO_MATCH":
                ws.cell(row=r, column=c_quote).fill = fill_red

        if c_ref:
            v = ws.cell(row=r, column=c_ref).value
            if v in ("IBID", "SUPRA"):
                ws.cell(row=r, column=c_ref).fill = fill_grey


        import json

        COL_QUOTES_PROP = _col("Quotes and proposition") or 3
        COL_QUOTES_JSON = _col("quotes_list_json")  # diagnostic source for quote highlighting

        quote_argb = "FF0B5394" if is_odd else "FFFF9000"  # match the column A fill colours

        if COL_QUOTES_JSON:
            raw = ws.cell(row=r, column=COL_QUOTES_JSON).value
            try:
                quotes = json.loads(raw) if raw else []
            except Exception:
                quotes = []

            cell_text = ws.cell(row=r, column=COL_QUOTES_PROP).value or ""
            if cell_text:  # if you blanked it as duplicate, skip
                try:
                    _apply_richtext_bolding_and_color(ws, r, COL_QUOTES_PROP, cell_text, quotes, quote_argb)
                except Exception:
                    # If rich text isn't supported in the runtime, degrade silently.
                    pass


    # 4) Size diagnostic columns. Display-only exports strip these columns after
    # formatting, while diagnostic exports leave them visible.
    for c in range(DISPLAY_LAST + 1, ws.max_column + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].hidden = False
        if COL_ANCHOR_TEXT and c == COL_ANCHOR_TEXT:
            ws.column_dimensions[letter].width = 80
        elif COL_REF_CHAIN_ORIGIN_TEXT and c == COL_REF_CHAIN_ORIGIN_TEXT:
            ws.column_dimensions[letter].width = 80
        else:
            ws.column_dimensions[letter].width = 15

    if save:
        wb.save(xlsx_path)
    return wb


def _json_ready(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): _json_ready(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_ready(v) for v in value]
    if isinstance(value, tuple):
        return [_json_ready(v) for v in value]
    if isinstance(value, set):
        return [_json_ready(v) for v in sorted(value, key=str)]
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _workbook_sidecar_path(xlsx_path: str | Path) -> Path:
    path = Path(xlsx_path)
    return path.with_name(f"{path.stem}.diagnostics.json")


def _workbook_sidecar_signature(xlsx_path: str | Path, rows: list[dict]) -> str:
    seed = "|".join(
        [
            "alr-sidecar-v1",
            Path(xlsx_path).name,
            str(len(rows)),
            str(time.time_ns()),
        ]
    )
    return hashlib.sha256(seed.encode("utf-8", errors="ignore")).hexdigest()[:20]


def _write_workbook_json_sidecar(xlsx_path: str | Path, data: Dict[str, Any], signature: str) -> Path:
    sidecar_path = _workbook_sidecar_path(xlsx_path)
    sidecar = {
        "schema": "alr_quote_verifier_diagnostics_v1",
        "xlsx_file": Path(xlsx_path).name,
        "xlsx_signature": signature,
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "footnote_rows": _json_ready(data.get("footnote_rows", [])),
    }
    with open(sidecar_path, "w", encoding="utf-8") as f:
        json.dump(sidecar, f, ensure_ascii=False, indent=2)
    return sidecar_path


def _diagnostic_start_column(ws: Any) -> Optional[int]:
    diagnostic_headers = {
        "footnote_id",
        "footnote_internal_id",
        "citation_part_index",
        "citation_part_kind",
        "citation_part_corrected",
        "citation_part_link",
        "quote_check_status",
    }
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value in diagnostic_headers:
            return idx
    return None


def _hide_unused_columns(ws: Any) -> None:
    """Hide Excel's empty right-side grid without materializing extra cells."""
    max_col = max(1, int(ws.max_column or 1))
    if max_col >= 16384:
        return
    ws.column_dimensions.group(
        start=get_column_letter(max_col + 1),
        end="XFD",
        hidden=True,
    )


def finalize_workbook_export(
    xlsx_path: str,
    data: Dict[str, Any],
    *,
    workbook: Optional[Any] = None,
) -> None:
    """Apply final export-detail policy after formatting has used raw columns."""
    _ensure_openpyxl()
    import openpyxl

    rows = data.get("footnote_rows", [])
    mode = EXPORT_DETAIL_MODE
    signature = _workbook_sidecar_signature(xlsx_path, rows) if mode == "display-json" else ""

    wb = workbook or openpyxl.load_workbook(xlsx_path, rich_text=True)
    if "FootnoteReferences" in wb.sheetnames:
        ws = wb["FootnoteReferences"]
        for dim in ws.column_dimensions.values():
            dim.hidden = False
        diagnostic_start = _diagnostic_start_column(ws)
        if mode in {"display", "display-json"}:
            if diagnostic_start:
                ws.delete_cols(diagnostic_start, ws.max_column - diagnostic_start + 1)
        elif mode == "diagnostic-hidden" and diagnostic_start:
            for col_idx in range(diagnostic_start, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True
        _hide_unused_columns(ws)

    if signature:
        existing = (wb.properties.identifier or "").strip()
        marker = f"alr-sidecar:{signature}"
        wb.properties.identifier = f"{existing}; {marker}" if existing else marker

    wb.save(xlsx_path)

    if signature:
        sidecar_path = _write_workbook_json_sidecar(xlsx_path, data, signature)
        _ts_print(f"  Diagnostic JSON: {sidecar_path}")


# -----------------------------
# -----------------------------
# First-page extraction
# -----------------------------

_FIRST_PAGE_RE = re.compile(r"""
    (?:
        \[                # e.g. [1962] SCR 746  → 746 is first page
            (?:\d{4})\]    # year in brackets
            \s+
            (?:\w+(?:\s+\w+){0,3})  # reporter name (1-4 words)
            \s+
            (\d+)          # first page
        |
        \(                # e.g. (2024) 45:3 Alberta Law Review 234 → 234 is first page
            \d{4}\)
            \s+
            \d+(?::\d+)?  # volume[:issue]
            \s+
            (?:\w+(?:\s+\w+){0,3})  # journal name (1-4 words)
            \s+
            (\d+)          # first page
        |
        \[                # e.g. [2023] 2 SCR 45 → 45 is first page
            \d{4}\]
            \s+
            \d+
            \s+
            (?:\w+(?:\s+\w+){0,3})
            \s+
            (\d+)
    )
""", re.VERBOSE)


def _extract_first_page(text: str) -> str:
    """Extract the first page number from a citation text.
    Returns empty string if no page number can be inferred."""
    if not text:
        return ""
    text = text.strip().rstrip(".")
    # Strip pinpoint suffix (e.g. " at 763-64", " at para 20") to avoid matching pinpoint numbers
    text = re.sub(r"\s+at\s+(?:p\.?\s*)?(?:\d+[\d,\s\-–]*\d*|para(?:s)?\.?\s+\d+).*$", "", text, flags=re.IGNORECASE)
    m = _FIRST_PAGE_RE.search(text)
    if m:
        for g in m.groups():
            if g:
                return g
    return ""


# -----------------------------
# Batch runner (folder -> sister folder)
# -----------------------------

from pathlib import Path
from datetime import datetime


def _append_page_pinpoint_links(rows: List[Dict[str, Any]]) -> None:
    """Post-processing: append #page=N to citation links using page_pinpoints."""
    for row in rows:
        pp_raw = row.get("page_pinpoints", "")
        link = row.get("citation_part_link", "") or ""
        cleaned_link = _strip_invalid_page_fragment(link)
        if cleaned_link != link:
            row["citation_part_link"] = cleaned_link
            link = cleaned_link
        if not pp_raw or not link or link.lower() == "other":
            continue
        try:
            pp_list = json.loads(pp_raw) if isinstance(pp_raw, str) else list(pp_raw)
        except Exception:
            continue
        if not pp_list:
            continue
        try:
            first_pp = int(pp_list[0])
        except (ValueError, TypeError):
            continue
        pdf_page = journal_search.pdf_page_for_label(row.get("_journal_article_id"), first_pp)
        if pdf_page is None:
            first_page_raw = row.get("first_page", "")
            try:
                first_page_val = int(first_page_raw)
            except (ValueError, TypeError):
                continue
            pdf_page = first_pp - first_page_val + 1
        if pdf_page < 1:
            continue
        base = link.split("#")[0]
        row["citation_part_link"] = f"{base}#page={pdf_page}"


def _restore_author_provided_links(rows: List[Dict[str, Any]]) -> None:
    for row in rows:
        author_link = str(row.get("_author_provided_link") or "").strip()
        if author_link:
            row["_resolved_citation_part_link"] = row.get("citation_part_link", "")
            row["citation_part_link"] = author_link


def _resolve_journal_links(rows: List[Dict[str, Any]]) -> None:
    """Post-processing step: match journal articles against public_endpoint.db for galley URLs.
    Runs after all critical pipeline work to avoid blocking the GUI with CPU-intensive
    SequenceMatcher calls against the full article database."""
    JOURNAL_KINDS = {"journal", "book", "report", "essay_collection"}
    journal_count = 0
    match_count = 0
    search_cache: Dict[str, Tuple[Optional[Dict[str, Any]], bool, str]] = {}
    for row in rows:
        _pause_gate()
        kind = row.get("citation_part_kind", "") or ""
        verbatim = row.get("citation_part_text", "") or ""
        fn_id = row.get("footnote_display_id", row.get("footnote_id", ""))
        pt_idx = row.get("citation_part_index", "")
        if kind in JOURNAL_KINDS:
            journal_count += 1
        if kind in JOURNAL_KINDS and verbatim:
            try:
                cached = search_cache.get(verbatim)
                if cached is None:
                    meta = journal_search.extract_citation_metadata(verbatim)
                    has_anchor = bool(
                        meta.get("year")
                        or meta.get("journal_hint")
                        or meta.get("volume")
                        or meta.get("issue")
                    )
                    deadline = None
                    if not has_anchor:
                        deadline = time.perf_counter() + getattr(
                            journal_search,
                            "UNFILTERED_SEARCH_TIMEOUT_S",
                            30.0,
                        )
                    hit = journal_search.search_by_title(verbatim, deadline=deadline)
                    timed_out = deadline is not None and time.perf_counter() >= deadline
                    extracted = journal_search.extract_title(verbatim) if not hit else ""
                    cached = (hit, timed_out, extracted)
                    search_cache[verbatim] = cached
                hit, timed_out, extracted = cached
                if hit:
                    db_fp = hit.get("first_page", "") or ""
                    if db_fp and db_fp != row.get("first_page", ""):
                        _ts_print(f"    first_page: citation={row.get('first_page','')} db={db_fp} (using db)")
                        row["first_page"] = db_fp
                    row["_journal_article_id"] = hit.get("article_id", "")
                    hit_url = hit.get("galley_url") or ""
                    if hit_url:
                        match_count += 1
                        _ts_print(f"  [FN {fn_id}.{pt_idx}] Journal match: {hit['title'][:60]} -> {hit_url}")
                        row["citation_part_link"] = hit_url
                        row["_journal_link_resolved"] = True
                        row["journal_match_info"] = f"Journal: {hit['title'][:60]} [{(hit.get('journal_name') or '')[:20]}]"
                    else:
                        _ts_print(f"  [FN {fn_id}.{pt_idx}] Journal match missing galley_url: {hit['title'][:60]}")
                        row["journal_match_info"] = f"Matched, no galley_url: {hit['title'][:60]} [{(hit.get('journal_name') or '')[:20]}]"
                else:
                    if timed_out:
                        shown = extracted or verbatim
                        _ts_print(f"  [FN {fn_id}.{pt_idx}] Journal search timed out for unanchored citation: \"{shown[:80]}\"")
                        row["journal_match_info"] = f"Timed out: \"{shown[:80]}\""
                    elif extracted:
                        _ts_print(f"  [FN {fn_id}.{pt_idx}] Journal no DB match for: \"{extracted[:80]}\"")
                        row["journal_match_info"] = f"No match: \"{extracted[:80]}\""
                    else:
                        _ts_print(f"  [FN {fn_id}.{pt_idx}] Journal no title extracted for: {verbatim[:60]}...")
                        row["journal_match_info"] = "No title extracted"
            except Exception as e:
                _ts_print(f"  [FN {fn_id}.{pt_idx}] Journal search error: {e}")
    _ts_print(f"  Journal search: {match_count}/{journal_count} matches")


def _int_or_none(value: Any) -> Optional[int]:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(str(value).strip())
    except Exception:
        return None


def _attach_ref_chain_origin_sources(rows: List[Dict[str, Any]]) -> None:
    rows_by_internal_key: Dict[Tuple[int, int], Dict[str, Any]] = {}
    rows_by_display_key: Dict[Tuple[str, int], Dict[str, Any]] = {}
    for row in rows:
        fid = _int_or_none(row.get("footnote_id"))
        idx = _int_or_none(row.get("citation_part_index"))
        if fid is not None and idx is not None:
            rows_by_internal_key[(fid, idx)] = row
        display_id = str(row.get("footnote_display_id") or "").strip()
        if display_id and idx is not None:
            rows_by_display_key[(display_id, idx)] = row

    for row in rows:
        origin_id = str(row.get("ref_chain_origin_footnote_id") or "").strip()
        origin_fid = _int_or_none(origin_id)
        origin_idx = _int_or_none(row.get("ref_chain_origin_citation_part_index"))
        if origin_idx is None or not origin_id:
            continue
        origin = rows_by_display_key.get((origin_id, origin_idx))
        if not origin and origin_fid is not None:
            origin = rows_by_internal_key.get((origin_fid, origin_idx))
        if not origin or origin is row:
            continue
        row["_ref_chain_origin_citation_part_kind"] = origin.get("citation_part_kind", "")
        row["_ref_chain_origin_bare_citation"] = origin.get("bare_citation", "")
        row["_ref_chain_origin_citation_part_link"] = origin.get("citation_part_link", "")
        row["_ref_chain_origin_short_form"] = origin.get("short_form", "")
        row["_ref_chain_origin_pinpoint_fragments"] = origin.get("pinpoint_fragments", "")
        row["_ref_chain_origin_citation_part_anchor_text"] = origin.get("citation_part_anchor_text", "")
        row["_ref_chain_origin_anchor_segments"] = origin.get("_citation_part_anchor_segments", [])
        row["_ref_chain_origin_full_source_text"] = origin.get("_citation_part_full_source_text", "")
        row["_ref_chain_origin_journal_article_id"] = origin.get("_journal_article_id", "")
        row["_ref_chain_origin_journal_link_resolved"] = origin.get(
            "_journal_link_resolved", False
        )


def _effective_ref_origin_link(row: Dict[str, Any], origin_link: str) -> str:
    origin_link = (origin_link or "").strip()
    if not origin_link or origin_link.lower() == "other":
        return ""

    pinpoint_fragments = _coerce_pinpoint_fragments(row.get("pinpoint_fragments"))
    if pinpoint_fragments:
        base, _frag = _split_url(origin_link)
        return _append_first_pinpoint_fragment(base or origin_link, pinpoint_fragments)

    if _first_page_pinpoint(row) is not None:
        base, _frag = _split_url(origin_link)
        return base or origin_link

    return origin_link


def _ref_origin_matches_supra_hint(row: Dict[str, Any]) -> bool:
    if (row.get("ref_kind") or "").strip().upper() != "SUPRA":
        return True
    hint = _extract_supra_hint(row.get("citation_part_text") or row.get("citation_part_corrected") or "")
    if not hint:
        return True

    origin_text = (
        row.get("_ref_chain_origin_citation_part_text")
        or row.get("ref_chain_origin_citation_part_text")
        or row.get("_ref_chain_origin_bare_citation")
        or ""
    )
    hint_tokens = [t for t in re.findall(r"[A-Za-z][A-Za-z'-]*", hint.lower()) if len(t) >= 3]
    origin_tokens = set(re.findall(r"[A-Za-z][A-Za-z'-]*", str(origin_text).lower()))
    if not hint_tokens:
        return True
    return any(token in origin_tokens for token in hint_tokens)


def _normalize_ref_short_form(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = re.sub(r"^[\[\(\{\"'\s]+|[\]\)\}\"'\s]+$", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\.+$", "", text).strip()
    return text.casefold()


def _supra_reference_short_form(row: Dict[str, Any]) -> str:
    return (
        str(row.get("short_form") or "").strip()
        or _extract_supra_hint(row.get("citation_part_text") or row.get("citation_part_corrected") or "")
    )


def _supra_short_form_exactly_matches_origin(row: Dict[str, Any]) -> bool:
    ref_short = _normalize_ref_short_form(_supra_reference_short_form(row))
    if not ref_short:
        return True
    origin_short = _normalize_ref_short_form(row.get("_ref_chain_origin_short_form"))
    return bool(origin_short and origin_short == ref_short)


def _link_fragment_matches_requested_pinpoint(link: str, row: Dict[str, Any]) -> bool:
    _base, frag = _split_url((link or "").strip())
    if not frag:
        return False
    norm_frag = _normalize_anchor_fragment(frag)
    requested_textual = {
        _normalize_anchor_fragment(raw)
        for raw in _coerce_pinpoint_fragments(row.get("pinpoint_fragments"))
        if str(raw or "").strip()
    }
    if requested_textual:
        return norm_frag in requested_textual
    requested_page = _first_page_pinpoint(row)
    if requested_page is not None:
        return norm_frag.startswith("page=")
    return False


def _supra_link_disagrees_with_requested_pinpoint(row: Dict[str, Any]) -> bool:
    has_requested_pinpoint = bool(_coerce_pinpoint_fragments(row.get("pinpoint_fragments"))) or (
        _first_page_pinpoint(row) is not None
    )
    if not has_requested_pinpoint:
        return False
    link = (row.get("citation_part_link") or "").strip()
    if not link or link.lower() == "other":
        return True
    return not _link_fragment_matches_requested_pinpoint(link, row)


def _should_use_ref_chain_origin(row: Dict[str, Any]) -> bool:
    if (row.get("ref_kind") or "").strip().upper() != "SUPRA":
        return True

    link = (row.get("citation_part_link") or "").strip()
    has_existing_link = bool(link and link.lower() != "other")
    has_ref_short_form = bool(_normalize_ref_short_form(_supra_reference_short_form(row)))

    if has_existing_link:
        if not _supra_link_disagrees_with_requested_pinpoint(row):
            return False
        return _supra_short_form_exactly_matches_origin(row)

    if has_ref_short_form:
        return _supra_short_form_exactly_matches_origin(row)

    return _ref_origin_matches_supra_hint(row)


def _ref_chain_origin_value(row: Dict[str, Any], key: str) -> Any:
    if not _should_use_ref_chain_origin(row):
        return ""
    return row.get(key)


def _journal_article_id_for_row(row: Dict[str, Any]) -> Any:
    return row.get("_journal_article_id") or _ref_chain_origin_value(row, "_ref_chain_origin_journal_article_id")


def _apply_ref_chain_origin_sources(rows: List[Dict[str, Any]]) -> None:
    for row in rows:
        if (row.get("ref_kind") or "").strip().upper() not in {"IBID", "SUPRA"}:
            continue
        if not _should_use_ref_chain_origin(row):
            continue
        origin_link = (row.get("_ref_chain_origin_citation_part_link") or "").strip()
        effective_link = _effective_ref_origin_link(row, origin_link)
        # Registry/sibling-resolved links are higher-precision than the
        # note-number chain (draft note numbers go stale); keep them.
        if effective_link and row.get("_ref_link_resolution") not in ("registry", "sibling"):
            row["citation_part_link"] = _reanchor_ref_link(
                effective_link, row.get("citation_part_text") or ""
            )

        origin_article_id = str(row.get("_ref_chain_origin_journal_article_id") or "").strip()
        if origin_article_id:
            row["_journal_article_id"] = origin_article_id
        if row.get("_ref_chain_origin_journal_link_resolved"):
            row["_journal_link_resolved"] = True

        origin_anchor = str(row.get("_ref_chain_origin_citation_part_anchor_text") or "").strip()
        if origin_anchor:
            row["citation_part_anchor_text"] = origin_anchor

        origin_full_source = str(row.get("_ref_chain_origin_full_source_text") or "").strip()
        if origin_full_source:
            row["_citation_part_full_source_text"] = origin_full_source


def build_verified_audit_data(
    input_docx: str,
    *,
    max_lookahead: int = 400,
    footnote_ids: Optional[set[int]] = None,
    exclude_web_citations: bool = False,
    source_matchable_only: bool = False,
) -> Dict[str, Any]:
    if not os.path.exists(input_docx):
        raise FileNotFoundError(f"Input DOCX not found: {input_docx}")
    if not input_docx.lower().endswith(".docx"):
        raise ValueError("Input must be a .docx file")

    _ts_print(f"  Building audit data ...")
    with _timing_span(
        "build_audit_data",
        doc=os.path.basename(input_docx),
        selected_footnotes=len(footnote_ids or []),
    ):
        data = build_audit_data(
            input_docx,
            max_lookahead=max_lookahead,
            footnote_ids=footnote_ids,
        )

    footnote_rows = data.get("footnote_rows", [])
    _timing_event(
        "build_audit_data:counts",
        doc=os.path.basename(input_docx),
        footnote_rows=len(footnote_rows),
        footnote_parts=len(data.get("footnote_parts", [])),
        footnotes=len(data.get("footnote_map", {})),
    )
    analyzed_footnotes = len(data.get("fn_num_parts", {}))
    total_footnotes = len(data.get("footnote_map", {}))
    if analyzed_footnotes != total_footnotes:
        _ts_print(
            f"  Split {analyzed_footnotes} footnotes in "
            f"{len(data.get('footnote_parts', []))} parts (Partial Run)"
        )
    else:
        _ts_print(
            f"  Split {total_footnotes} footnotes into "
            f"{len(data.get('footnote_parts', []))} citation parts"
        )

    _ts_print(f"  Resolving journal article links ...")
    with _timing_span(
        "resolve_journal_links",
        doc=os.path.basename(input_docx),
        rows=len(footnote_rows),
    ):
        _resolve_journal_links(footnote_rows)

    if exclude_web_citations:
        before = len(footnote_rows)
        footnote_rows = [row for row in footnote_rows if not _is_prefilled_web_citation(row)]
        data["footnote_rows"] = footnote_rows
        removed = before - len(footnote_rows)
        if removed:
            _ts_print(f"  Excluded {removed} pre-filled web citation rows")

    _ts_print(f"  Resolving ibid/supra reference chains ...")
    with _timing_span(
        "resolve_reference_chains",
        doc=os.path.basename(input_docx),
        rows=len(footnote_rows),
    ):
        resolve_reference_chains(
            data.get("footnote_parts", []),
            data.get("footnote_map", {}),
            data.get("fn_num_parts", {}),
            display_num_to_internal=data.get("display_num_to_internal"),
            internal_to_display_id=data.get("footnote_display_ids"),
        )
        # Copy reference chain fields from parts into rows
        parts_by_key: Dict[Tuple[int, int], Dict[str, Any]] = {
            (p["footnote_id"], p["citation_part_index"]): p
            for p in data.get("footnote_parts", [])
        }
        for row in footnote_rows:
            key = (int(row.get("footnote_id", 0)), int(row.get("citation_part_index", 0)))
            p = parts_by_key.get(key)
            if p:
                for k in REF_FIELDS:
                    row[k] = p.get(k, "")
        _attach_ref_chain_origin_sources(footnote_rows)
        _apply_ref_chain_origin_sources(footnote_rows)
        ref_rows = [r for r in footnote_rows if str(r.get("ref_kind") or "").strip()]
        if ref_rows:
            resolved = sum(
                1 for r in ref_rows
                if str(r.get("ref_target_footnote_id") or "").strip()
            )
            _ts_print(f"  References: {resolved} of {len(ref_rows)} ibid/supra links resolved")

    _ts_print(f"  Running quote checks against source text ...")
    with _timing_span(
        "apply_quote_checks",
        doc=os.path.basename(input_docx),
        rows=len(footnote_rows),
        quote_footnotes=len(data.get("_quotes_by_footnote", {})),
    ):
        _apply_quote_checks(footnote_rows, data.get("_quotes_by_footnote", {}))
    checked = [
        r for r in footnote_rows
        if str(r.get("quote_check_status") or "").strip()
    ]
    if checked:
        no_match = sum(1 for r in checked if "NO_MATCH" in r["quote_check_status"])
        partial = sum(1 for r in checked if "PARTIAL" in r["quote_check_status"])
        verified = len(checked) - no_match - partial
        _ts_print(
            f"  Quote checks: {verified} verified, {partial} partial, "
            f"{no_match} not found"
        )

    if source_matchable_only:
        before = len(footnote_rows)
        footnote_rows = [row for row in footnote_rows if _has_quote_matchable_source(row)]
        data["footnote_rows"] = footnote_rows
        removed = before - len(footnote_rows)
        if removed:
            _ts_print(f"  Excluded {removed} rows without a quote-checkable source")

    _ts_print(f"  Appending #page= pinpoints to links ...")
    with _timing_span(
        "append_page_pinpoint_links",
        doc=os.path.basename(input_docx),
        rows=len(footnote_rows),
    ):
        _append_page_pinpoint_links(footnote_rows)

    # Processing uses a derived working link; the workbook's Citation link is
    # restored to the source-document URL after all checking is complete.
    _restore_author_provided_links(footnote_rows)
    _restore_author_provided_links(data.get("footnote_parts", []))

    return data


PREFILLED_WEB_CITATION_KINDS = {"other", "non_parliamentary", "website"}


def _is_prefilled_web_citation(row: Dict[str, Any]) -> bool:
    kind = (row.get("citation_part_kind") or "").strip().lower()
    if kind not in PREFILLED_WEB_CITATION_KINDS:
        return False
    link = _canlii_source_lookup_url((row.get("citation_part_link") or "").strip())
    if not link or link.lower() == "other":
        return False
    base, _frag = _split_url(link)
    if not base:
        return False
    host = urlsplit(base).netloc.lower()
    if host.endswith("canlii.org"):
        return False
    return True


def _has_quote_matchable_source(row: Dict[str, Any]) -> bool:
    if (row.get("has_quotes") or "").strip().upper() != "YES":
        return False
    if (row.get("_quote_source_tag") or "").strip():
        return True
    status = (row.get("quote_check_status") or "").strip().upper()
    if status.startswith("ALT_"):
        return True
    return False


def run_audit(
    input_docx: str,
    output_xlsx: str,
    *,
    max_lookahead: int = 400,
    footnote_ids: Optional[set[int]] = None,
    exclude_web_citations: bool = False,
    source_matchable_only: bool = False,
) -> str:
    started_at = time.time()

    global _ACTIVE_RUNS
    with _ACTIVE_RUNS_LOCK:
        _ACTIVE_RUNS += 1
        if _ACTIVE_RUNS == 1:
            # Per-document derived caches (parsed lxml trees, full source
            # texts): cross-document hit rates are low and the trees are
            # large, so holding them for the process lifetime just grows GUI
            # memory run after run. When documents verify in parallel, only
            # the first run of a batch clears — yanking the caches out from
            # under a concurrent document would lose its in-flight state.
            _FRAGMENT_DOC_TEXT_CACHE.clear()
            _A2AJ_LOCKED_DOCUMENTS.clear()
            _A2AJ_LOCKED_STRUCTURES.clear()
            _A2AJ_LOCKED_TEXTS.clear()

    try:
        data = build_verified_audit_data(
            input_docx,
            max_lookahead=max_lookahead,
            footnote_ids=footnote_ids,
            exclude_web_citations=exclude_web_citations,
            source_matchable_only=source_matchable_only,
        )

        _ts_print(f"  Writing workbook ...")
        with _timing_span("write_workbook", output=output_xlsx, rows=len(data.get("footnote_rows", []))):
            write_workbook(data, output_xlsx)
        _ts_print(f"  Formatting workbook ...")
        with _timing_span("format_workbook", output=output_xlsx):
            formatted_workbook = apply_cell_formatting(output_xlsx, save=False)
        with _timing_span("finalize_workbook_export", output=output_xlsx, mode=EXPORT_DETAIL_MODE):
            finalize_workbook_export(
                output_xlsx, data, workbook=formatted_workbook,
            )
    finally:
        with _ACTIVE_RUNS_LOCK:
            _ACTIVE_RUNS -= 1

    elapsed = time.time() - started_at
    _ts_print(f"  Total time: {elapsed:.0f}s")

    return output_xlsx


def run_audit_folder(
    input_folder: str,
    output_sister_folder_name: str = "CHECKER_OUTPUT",
    recursive: bool = False,
    *,
    max_lookahead: int = 400,
    footnote_ids: Optional[set[int]] = None,
    footnote_selection: Optional[Dict[str, set[int]]] = None,
    single_workbook_path: Optional[str] = None,
    max_selected_footnotes: Optional[int] = None,
    max_output_rows: Optional[int] = None,
    exclude_web_citations: bool = False,
    source_matchable_only: bool = False,
) -> dict:
    """
    Runs run_audit() for every .docx in input_folder and writes outputs
    into a sister folder named output_sister_folder_name.

    Output naming: "[CHECKER] <input_stem>.xlsx"
    Returns a summary dict with counts and per-file statuses.
    """
    in_dir = Path(input_folder).expanduser().resolve()
    if not in_dir.exists() or not in_dir.is_dir():
        raise NotADirectoryError(f"Input folder not found or not a directory: {in_dir}")

    # Sister folder = same parent, different child folder name
    out_dir = in_dir.parent / output_sister_folder_name
    out_dir.mkdir(parents=True, exist_ok=True)
    combined_path: Optional[Path] = None
    combined_rows: List[Dict[str, Any]] = []
    if single_workbook_path:
        combined_path = Path(single_workbook_path).expanduser()
        if not combined_path.is_absolute():
            combined_path = (Path.cwd() / combined_path).resolve()

    # Collect .docx files (skip temporary Word lock files: "~$*.docx")
    docx_files = _collect_docx_files(in_dir, recursive=recursive)

    summary = {
        "input_folder": str(in_dir),
        "output_folder": str(out_dir),
        "single_workbook": str(combined_path) if combined_path else "",
        "recursive": recursive,
        "found": len(docx_files),
        "succeeded": 0,
        "skipped": 0,
        "failed": 0,
        "results": [],  # list of dicts: {input, output, status, error}
        "started_at": datetime.now().isoformat(timespec="seconds"),
    }

    selected_count = 0
    for docx_path in sorted(docx_files):
        doc_specific_ids = _selection_ids_for_doc(footnote_selection, docx_path, in_dir)
        effective_footnote_ids: Optional[set[int]] = None
        if footnote_ids is not None:
            effective_footnote_ids = set(footnote_ids)
        if doc_specific_ids is not None:
            if doc_specific_ids:
                effective_footnote_ids = set(effective_footnote_ids or set())
                effective_footnote_ids.update(doc_specific_ids)
            elif footnote_selection and footnote_ids is None:
                summary["skipped"] += 1
                summary["results"].append({
                    "input": str(docx_path),
                    "output": None,
                    "status": "skipped",
                    "error": "No selected footnotes for this document.",
                })
                _ts_print(f"SKIP {docx_path.name}  (no selected footnotes)")
                continue

        if max_selected_footnotes is not None and effective_footnote_ids is not None:
            remaining = int(max_selected_footnotes) - selected_count
            if remaining <= 0:
                summary["skipped"] += 1
                summary["results"].append({
                    "input": str(docx_path),
                    "output": str(combined_path) if combined_path else None,
                    "status": "skipped",
                    "error": "Selected-footnote limit reached.",
                })
                _ts_print(f"SKIP {docx_path.name}  (selected-footnote limit reached)")
                continue
            limited = set(sorted(effective_footnote_ids)[:remaining])
            selected_count += len(limited)
            effective_footnote_ids = limited

        # Preserve base name; force .xlsx extension; prepend [CHECKED]
        base_stem = f"[CHECKED] {docx_path.stem}"
        output_name = Path(f"{base_stem}.xlsx")
        output_xlsx = out_dir / output_name
        suffix = 0
        while output_xlsx.exists():
            suffix += 1
            if suffix > 99:
                break
            stem_trunc = f"_{suffix}_{base_stem}"
            max_stem = 245 - len(str(out_dir)) - len(".xlsx")
            if len(stem_trunc) > max_stem:
                stem_trunc = stem_trunc[:max_stem]
            output_name = Path(f"{stem_trunc}.xlsx")
            output_xlsx = out_dir / output_name

        _ts_print(f"Working on {output_name}...")

        try:
            doc_started_at = time.perf_counter()
            if effective_footnote_ids:
                _ts_print(
                    "  FN selection: "
                    + ",".join(str(fid) for fid in sorted(effective_footnote_ids))
                )
            _timing_event(
                "doc:start",
                doc=docx_path.name,
                selected_footnotes=len(effective_footnote_ids or []),
                combined=bool(combined_path),
            )
            if combined_path:
                data = build_verified_audit_data(
                    str(docx_path),
                    max_lookahead=max_lookahead,
                    footnote_ids=effective_footnote_ids,
                    exclude_web_citations=exclude_web_citations,
                    source_matchable_only=source_matchable_only,
                )
                doc_key = _doc_key_for_manifest(docx_path, in_dir)
                for row in data.get("footnote_rows", []):
                    row["source_doc"] = docx_path.name
                    row["source_doc_key"] = doc_key
                combined_rows.extend(data.get("footnote_rows", []))
                out = str(combined_path)
                _timing_event(
                    "combined_rows:extend",
                    doc=docx_path.name,
                    added_rows=len(data.get("footnote_rows", [])),
                    total_rows=len(combined_rows),
                    max_output_rows=max_output_rows,
                )
            else:
                out = run_audit(
                    str(docx_path),
                    str(output_xlsx),
                    max_lookahead=max_lookahead,
                    footnote_ids=effective_footnote_ids,
                    exclude_web_citations=exclude_web_citations,
                    source_matchable_only=source_matchable_only,
                )
            summary["succeeded"] += 1
            summary["results"].append({
                "input": str(docx_path),
                "output": str(out),
                "status": "ok",
                "error": None,
            })
            _timing_event(
                "doc:end",
                doc=docx_path.name,
                elapsed_s=round(time.perf_counter() - doc_started_at, 3),
                status="ok",
                output=str(out),
                combined_rows=len(combined_rows),
            )
            _ts_print(f"OK   {docx_path.name}  ->  {output_name}")
            if (
                combined_path
                and max_output_rows is not None
                and int(max_output_rows) >= 0
                and len(combined_rows) >= int(max_output_rows)
            ):
                _timing_event(
                    "combined_row_limit_reached",
                    total_rows=len(combined_rows),
                    max_output_rows=max_output_rows,
                    doc=docx_path.name,
                )
                _ts_print(
                    f"Combined output row limit reached ({len(combined_rows)}/{max_output_rows}); stopping batch early"
                )
                break
        except Exception as e:
            summary["failed"] += 1
            summary["results"].append({
                "input": str(docx_path),
                "output": str(output_xlsx),
                "status": "fail",
                "error": repr(e),
            })
            _timing_event(
                "doc:end",
                doc=docx_path.name,
                elapsed_s=round(time.perf_counter() - doc_started_at, 3),
                status="fail",
                error=repr(e),
            )
            _ts_print(f"FAIL {docx_path.name}  ({e})")


    if combined_path:
        if max_output_rows is not None and int(max_output_rows) >= 0:
            before = len(combined_rows)
            combined_rows = combined_rows[:int(max_output_rows)]
            if len(combined_rows) != before:
                _ts_print(f"Trimmed combined workbook rows to {len(combined_rows)}")
        _ts_print(f"Writing combined workbook: {combined_path}")
        with _timing_span("write_combined_workbook", output=str(combined_path), rows=len(combined_rows)):
            write_workbook({"footnote_rows": combined_rows}, str(combined_path))
        _ts_print(f"Formatting combined workbook: {combined_path}")
        with _timing_span("format_combined_workbook", output=str(combined_path), rows=len(combined_rows)):
            apply_cell_formatting(str(combined_path))
        with _timing_span("finalize_combined_workbook_export", output=str(combined_path), mode=EXPORT_DETAIL_MODE):
            finalize_workbook_export(str(combined_path), {"footnote_rows": combined_rows})

    summary["finished_at"] = datetime.now().isoformat(timespec="seconds")
    return summary


# -----------------------------
# Entrypoint
# -----------------------------

def _parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate an ALR footnote/citation audit workbook for a folder of DOCX files."
    )
    p.add_argument("--input", required=True, help="Folder containing .docx files to process.")
    p.add_argument(
        "--output-name",
        default="CHECKED_EDITS",
        help='Name of the output sister folder (default: "CHECKED_EDITS").',
    )
    p.add_argument(
        "--single-workbook",
        default=None,
        help="Write all processed batch rows into this one workbook instead of one workbook per DOCX.",
    )
    p.add_argument(
        "--max-selected-footnotes",
        type=int,
        default=None,
        help="Cap selected footnotes across the batch; useful with --footnote-selection-file or --quote-footnotes-only.",
    )
    p.add_argument(
        "--max-output-rows",
        type=int,
        default=None,
        help="Cap rows written to a combined --single-workbook after any row-level filters.",
    )
    p.add_argument(
        "--exclude-web-citations",
        action="store_true",
        help="Drop pre-filled ordinary website citation rows from generated workbooks.",
    )
    p.add_argument(
        "--source-matchable-only",
        action="store_true",
        help="Keep only quote rows with an actual source text path: CanLII/full text, A2AJ, or journal DB.",
    )
    p.add_argument(
        "--timing-log",
        default=None,
        help="Write stage and quote-row timing events to a JSONL file.",
    )
    p.add_argument("--recursive", action="store_true", help="Include subfolders under --input.")
    p.add_argument(
        "--max-lookahead",
        type=int,
        default=400,
        help="Max characters to look ahead when assigning quotes to footnotes (default: 400).",
    )
    p.add_argument(
        "--dry-fire",
        action="store_true",
        help="Skip all CanLII/network access; resolve URLs as-is and skip quote verification against source text.",
    )
    p.add_argument(
        "--footnote-ids",
        default=None,
        help='Comma/range-separated footnote display IDs to process (e.g. "1,3,30-35"). Default: all footnotes.',
    )
    p.add_argument(
        "--footnote-selection-file",
        default=None,
        help="CSV with doc_key/doc_name/doc_path and footnote_id columns for per-document batch selection.",
    )
    p.add_argument(
        "--quote-footnotes-out",
        default=None,
        help="Write a CSV manifest of footnotes whose associated proposition text contains quotes.",
    )
    p.add_argument(
        "--quote-footnotes-only",
        action="store_true",
        help="Run only quote-associated footnotes discovered in this batch, including their supra targets.",
    )
    p.add_argument(
        "--quote-footnotes-list-only",
        action="store_true",
        help="Write --quote-footnotes-out and exit without running verification.",
    )
    p.add_argument(
        "--supra-mode",
        default="aggressive",
        choices=("aggressive", "safe"),
        help='Supra short-form extraction mode. "aggressive" (default) searches globally for supras; "safe" only matches at the start of the citation text.',
    )
    p.add_argument(
        "--supra-linking",
        default="safe",
        choices=("safe", "aggressive"),
        help=(
            'Supra linking aggressiveness: "safe" uses the strict resolver; '
            '"aggressive" also resolves unambiguous bare-note references and '
            'matches inferred case, legislation, and author short forms.'
        ),
    )
    p.add_argument(
        "--no-a2aj",
        action="store_true",
        help="Disable A2AJ source lookup for quote verification (skips API calls).",
    )
    p.add_argument(
        "--local-only",
        action="store_true",
        help="Use installed local data only; disables AI and all network source lookups.",
    )
    p.add_argument(
        "--search-alt-pinpoints",
        action="store_true",
        help="When quote check fails at specified pinpoint, search entire document for the quote and report alternate locations.",
    )
    p.add_argument(
        "--text-fragment-mode",
        default="off",
        choices=("all", "pinpointless", "off"),
        help='Text fragment links: "all" (every match), "pinpointless" (only when no pinpoint), "off" (default).',
    )
    p.add_argument(
        "--export-detail",
        default="diagnostic-hidden",
        choices=("display", "display-json", "diagnostic-hidden", "diagnostic"),
        help='Workbook export detail: "diagnostic-hidden" keeps raw diagnostic columns hidden (default), "display" strips them, "display-json" strips them plus writes a diagnostics JSON sidecar, or "diagnostic" keeps them visible.',
    )
    p.add_argument(
        "--no-hidden-columns",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    p.add_argument(
        "--no-llm-cache",
        action="store_true",
        help="Disable LLM response caching (default: cache enabled).",
    )
    p.add_argument(
        "--run-mode",
        default="high_accuracy",
        choices=VALID_RUN_MODES,
        help=(
            'Pipeline mode: "high_accuracy" (current production LLM), "economy" '
            '(pure references bypass the split call), "ultra_economy" '
            '(complete deterministic tuples also bypass eligible split calls), '
            'or "free" (no LLM calls).'
        ),
    )
    p.add_argument(
        "--no-alt-pinpoints",
        action="store_true",
        help="Disable the alternate pinpoint location search (enabled by default).",
    )
    return p.parse_args(argv)


def _configure_from_args(args: argparse.Namespace) -> None:
    global LINK_RESOLVER, SUPRA_MODE, USE_DB_SEARCH, USE_A2AJ, SEARCH_ALT_PINPOINTS, TEXT_FRAGMENT_MODE, EXPORT_DETAIL_MODE, LLM_CACHE_ENABLED
    global RUN_MODE, PURE_REF_PREFILTER, DETERMINISTIC_SOURCE_SPLITTER, FREE_NO_LLM, LOCAL_ONLY
    global REF_DISAMBIG_FALLBACK, SUPRA_LINKING_AGGRESSIVENESS
    global LLM_MODEL, LLM_API_KEY, client
    SUPRA_MODE = getattr(args, "supra_mode", "aggressive") or "aggressive"
    SUPRA_LINKING_AGGRESSIVENESS = (
        getattr(args, "supra_linking", "safe") or "safe"
    )
    if SUPRA_LINKING_AGGRESSIVENESS not in {"safe", "aggressive"}:
        SUPRA_LINKING_AGGRESSIVENESS = "safe"
    USE_DB_SEARCH = bool(getattr(args, "use_db_search", True))
    USE_A2AJ = bool(getattr(args, "use_a2aj", True))
    SEARCH_ALT_PINPOINTS = not bool(getattr(args, "no_alt_pinpoints", False))
    TEXT_FRAGMENT_MODE = getattr(args, "text_fragment_mode", "off") or "off"
    EXPORT_DETAIL_MODE = getattr(args, "export_detail", "diagnostic-hidden") or "diagnostic-hidden"
    LLM_CACHE_ENABLED = not bool(getattr(args, "no_llm_cache", False))
    LOCAL_ONLY = bool(getattr(args, "local_only", False))
    RUN_MODE = getattr(args, "run_mode", "high_accuracy") or "high_accuracy"
    if LOCAL_ONLY:
        RUN_MODE = "free"
    if RUN_MODE not in VALID_RUN_MODES:
        RUN_MODE = "high_accuracy"
    PURE_REF_PREFILTER = RUN_MODE in ("economy", "ultra_economy", "free")
    DETERMINISTIC_SOURCE_SPLITTER = RUN_MODE in ("ultra_economy", "free")
    FREE_NO_LLM = RUN_MODE == "free"
    REF_DISAMBIG_FALLBACK = RUN_MODE != "free"
    # --no-a2aj CLI flag overrides
    if getattr(args, "no_a2aj", False):
        USE_A2AJ = False
    a2aj_client.set_local_only(LOCAL_ONLY)
    if LOCAL_ONLY:
        citation_db = _provider_registry.get_citation_db()
        setter = getattr(citation_db, "set_external_enabled", None)
        if callable(setter):
            setter(False)

    LLM_MODEL = "gpt-5.2"
    # Leave LLM_API_KEY as an explicit override only; _ensure_llm_client
    # resolves the effective key (env/keys.py, then the per-user encrypted
    # store). Reset the client so it re-creates with new config.
    client = None
    if not FREE_NO_LLM:
        _ensure_llm_client()

    if args.dry_fire:
        USE_DB_SEARCH = False
        USE_A2AJ = False
        LINK_RESOLVER = UrlResolver()
        return

    if LOCAL_ONLY:
        LINK_RESOLVER = UrlResolver()
        return

    LINK_RESOLVER = UrlResolver()


def _main(argv: Optional[List[str]] = None) -> None:
    args = _parse_args(argv)

    if args.timing_log:
        _set_timing_log_path(args.timing_log)
        _timing_event(
            "run:start",
            input=args.input,
            output_name=args.output_name,
            single_workbook=args.single_workbook or "",
            max_selected_footnotes=args.max_selected_footnotes,
            max_output_rows=args.max_output_rows,
            exclude_web_citations=bool(args.exclude_web_citations),
            source_matchable_only=bool(args.source_matchable_only),
            export_detail=args.export_detail,
        )

    _ts_print()

    quote_selection: Dict[str, set[int]] = {}
    if args.quote_footnotes_out or args.quote_footnotes_only or args.quote_footnotes_list_only:
        if args.quote_footnotes_list_only and not args.quote_footnotes_out:
            raise SystemExit("--quote-footnotes-list-only requires --quote-footnotes-out")
        rows, quote_selection = build_quote_footnote_manifest(
            args.input,
            recursive=bool(args.recursive),
            max_lookahead=int(args.max_lookahead),
        )
        if args.quote_footnotes_out:
            write_quote_footnote_manifest(rows, args.quote_footnotes_out)
            _ts_print(f"Quote footnote list: wrote {len(rows)} rows -> {args.quote_footnotes_out}")
        if args.quote_footnotes_list_only:
            return
        if not args.quote_footnotes_only:
            quote_selection = {}

    file_selection: Dict[str, set[int]] = {}
    if args.footnote_selection_file:
        file_selection = _load_footnote_selection_file(args.footnote_selection_file)
        _ts_print(f"Loaded footnote selection file: {args.footnote_selection_file}")

    footnote_selection = _merge_selection_maps(file_selection, quote_selection)

    _configure_from_args(args)

    footnote_ids = _parse_footnote_ids(args.footnote_ids)

    report = run_audit_folder(
        input_folder=args.input,
        output_sister_folder_name=args.output_name,
        recursive=bool(args.recursive),
        max_lookahead=int(args.max_lookahead),
        footnote_ids=footnote_ids,
        footnote_selection=footnote_selection or None,
        single_workbook_path=args.single_workbook,
        max_selected_footnotes=args.max_selected_footnotes,
        max_output_rows=args.max_output_rows,
        exclude_web_citations=bool(args.exclude_web_citations),
        source_matchable_only=bool(args.source_matchable_only),
    )

    _ts_print()
    _ts_print("--- Batch summary ---")
    _ts_print(f"Input:   {report['input_folder']}")
    _ts_print(f"Output:  {report['output_folder']}")
    if report.get("single_workbook"):
        _ts_print(f"Workbook:{report['single_workbook']}")
    _ts_print(f"Found:   {report['found']}")
    _ts_print(f"OK:      {report['succeeded']}")
    if report.get("skipped"):
        _ts_print(f"Skipped: {report['skipped']}")
    _ts_print(f"Failed:  {report['failed']}")
    elapsed = (datetime.now() - datetime.fromisoformat(report['started_at'])).total_seconds()
    if elapsed >= 1:
        _ts_print(f"Total:   {elapsed:.0f}s")
    _timing_event(
        "run:end",
        found=report.get("found"),
        succeeded=report.get("succeeded"),
        skipped=report.get("skipped"),
        failed=report.get("failed"),
        elapsed_s=round(elapsed, 3),
    )


if __name__ == "__main__":
    _main()
